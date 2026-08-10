import os
import re
import math
import json
import traceback
import urllib.request
import urllib.parse
import time
import io
import zipfile
import logging
import threading
import concurrent.futures
import uuid as _uuid
import openpyxl
from datetime import date, datetime, timedelta
from typing import Optional
import secrets
import hashlib
import psycopg2
import psycopg2.extras
import psycopg2.pool as _psycopg2_pool
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Query, Depends, Request, Response, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.security import OAuth2PasswordRequestForm
from pydantic import BaseModel
from jose import jwt, JWTError

try:
    from ortools.constraint_solver import routing_enums_pb2
    from ortools.constraint_solver import pywrapcp
    ORTOOLS_AVAILABLE = True
except ImportError:
    ORTOOLS_AVAILABLE = False
    logging.warning("OR-Tools not available, using greedy fallback")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ── EPF integration file paths ─────────────────────────────────────────────────
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.abspath(os.path.join(_THIS_DIR, "..", ".."))
_EPF_PATH = os.path.join(_PROJECT_ROOT, "artifacts", "integrations", "1c", "SmartRoute.epf")
_EPF_VERSION_PATH = os.path.join(_PROJECT_ROOT, "artifacts", "integrations", "1c", "version.json")
_EPF_IS_PLACEHOLDER: bool = True  # updated at startup after reading version.json
try:
    with open(_EPF_VERSION_PATH, "r", encoding="utf-8") as _vf:
        _epf_meta = json.load(_vf)
    _EPF_IS_PLACEHOLDER = _epf_meta.get("status") == "placeholder"
except Exception:
    _epf_meta = {}
    _EPF_IS_PLACEHOLDER = True

app = FastAPI(title="SmartRoute API")

# ALLOWED_ORIGINS env var — comma-separated list of allowed origins.
# Default "*" works for single-service deployments where frontend and backend
# share the same Railway domain. Restrict to your domain for production:
#   ALLOWED_ORIGINS=https://smartroute.up.railway.app
_raw_origins = os.environ.get("ALLOWED_ORIGINS", "*")
_ALLOWED_ORIGINS: list = (
    ["*"] if _raw_origins.strip() == "*"
    else [o.strip() for o in _raw_origins.split(",") if o.strip()]
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=_ALLOWED_ORIGINS,
    # allow_credentials requires explicit origin list (not "*").
    # In production Railway, front+API share the same origin so CORS doesn't apply.
    # In dev (Replit), Vite proxy handles cross-origin requests transparently.
    allow_credentials=_ALLOWED_ORIGINS != ["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    """Add security headers to every response."""
    response = await call_next(request)
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "geolocation=(), microphone=(), camera=()")
    # HSTS only when served over HTTPS (Railway / custom domain)
    if request.url.scheme == "https":
        response.headers.setdefault(
            "Strict-Transport-Security", "max-age=31536000; includeSubDomains"
        )
    return response

# Connection priority:
#   1. PG_CONNECTION_URL  — explicit override (Railway / custom Postgres)
#   2. Individual PGHOST/PGUSER/… vars — Replit-managed PostgreSQL (preferred in dev)
#   3. DATABASE_URL secret — fallback (may point to external DB)
_pg_explicit = os.environ.get("PG_CONNECTION_URL", "")
if not _pg_explicit:
    _pghost = os.environ.get("PGHOST", "")
    _pguser = os.environ.get("PGUSER", "")
    _pgpass = os.environ.get("PGPASSWORD", "")
    _pgdb   = os.environ.get("PGDATABASE", "")
    _pgport = os.environ.get("PGPORT", "5432")
    if _pghost and _pguser and _pgdb:
        _pg_explicit = f"postgresql://{_pguser}:{_pgpass}@{_pghost}:{_pgport}/{_pgdb}"
DATABASE_URL: str = _pg_explicit or os.environ.get("DATABASE_URL", "")

# ── JWT / Auth ────────────────────────────────────────────────────────────────
_JWT_SECRET_ENV = os.environ.get("JWT_SECRET", "")
if not _JWT_SECRET_ENV:
    _JWT_SECRET_ENV = secrets.token_hex(32)
    logging.warning(
        "JWT_SECRET env var is not set — using a random secret. "
        "All sessions will be invalidated on server restart. "
        "Set JWT_SECRET in production for persistent sessions."
    )
JWT_SECRET: str = _JWT_SECRET_ENV
JWT_ALGORITHM: str = "HS256"
JWT_TOKEN_TTL_HOURS: int = int(os.environ.get("JWT_TOKEN_TTL_HOURS", "720"))  # 30 days default
# Refresh threshold: reissue cookie if token expires within this many hours
JWT_REFRESH_THRESHOLD_HOURS: int = int(os.environ.get("JWT_REFRESH_THRESHOLD_HOURS", "168"))  # 7 days
JWT_COOKIE_NAME: str = "smartroute_token"
# SameSite=none + Secure=true is required for cross-site iframe contexts (Replit canvas, embedded apps).
# Default: none/true so dev preview works inside Replit's iframe-based editor.
# Override via env: COOKIE_SAMESITE=lax + COOKIE_SECURE=false for pure-localhost HTTP testing.
COOKIE_SAMESITE: str = os.environ.get("COOKIE_SAMESITE", "none")
COOKIE_SECURE: bool = os.environ.get("COOKIE_SECURE", "true").lower() in ("1", "true", "yes")

# passlib is imported but NOT used for bcrypt — direct bcrypt calls avoid the
# bcrypt≥4.0 passlib incompatibility (missing __about__.__version__).
# CryptContext left in place only so the import doesn't break; never call it.

ADMIN_PASSWORD: str = os.environ.get("ADMIN_PASSWORD", "")

# ── Login rate limiter ─────────────────────────────────────────────────────────
# Tracks failed login attempts per IP: ip → list of timestamps (epoch seconds).
# After LOGIN_MAX_ATTEMPTS failures in LOGIN_WINDOW_SECONDS → 429 for LOGIN_BLOCK_SECONDS.
LOGIN_MAX_ATTEMPTS: int = 5
LOGIN_WINDOW_SECONDS: int = 15 * 60   # 15 minutes
LOGIN_BLOCK_SECONDS: int = 15 * 60    # block for 15 minutes after threshold
_login_attempts: dict = {}            # {ip: [timestamp, ...]}
_login_attempts_lock = threading.Lock()


def _get_client_ip(request: Request) -> str:
    """Return the best-effort client IP for rate limiting."""
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def _check_login_rate_limit(ip: str) -> None:
    """Raise 429 if IP has exceeded the login attempt threshold."""
    now = time.time()
    with _login_attempts_lock:
        attempts = _login_attempts.get(ip, [])
        # Keep only attempts within the window
        attempts = [t for t in attempts if now - t < LOGIN_WINDOW_SECONDS]
        _login_attempts[ip] = attempts
        if len(attempts) >= LOGIN_MAX_ATTEMPTS:
            retry_after = int(LOGIN_BLOCK_SECONDS - (now - attempts[0]))
            raise HTTPException(
                status_code=429,
                detail=f"Слишком много попыток входа. Попробуйте через {max(1, retry_after // 60)} мин.",
                headers={"Retry-After": str(max(1, retry_after))},
            )


def _record_failed_login(ip: str) -> None:
    """Record a failed login attempt for the given IP."""
    now = time.time()
    with _login_attempts_lock:
        attempts = _login_attempts.get(ip, [])
        attempts = [t for t in attempts if now - t < LOGIN_WINDOW_SECONDS]
        attempts.append(now)
        _login_attempts[ip] = attempts


def _clear_login_attempts(ip: str) -> None:
    """Clear failed login attempts after a successful login."""
    with _login_attempts_lock:
        _login_attempts.pop(ip, None)


# ── General-purpose API rate limiter ─────────────────────────────────────────
# Keyed by arbitrary string (e.g. "vrp:42", "import:42") so each endpoint
# can have its own bucket per user.  Uses the same sliding-window pattern as
# the login rate limiter above.
_rl_store: dict = {}          # {key: [timestamp, ...]}
_rl_lock = threading.Lock()


def _api_rate_limit(key: str, max_calls: int, window_seconds: int) -> None:
    """Raise HTTP 429 if *key* has been called more than *max_calls* times
    within the last *window_seconds* seconds.

    Thread-safe. Old timestamps are cleaned up on every call so the dict
    does not grow without bound.
    """
    now = time.time()
    with _rl_lock:
        ts = _rl_store.get(key, [])
        ts = [t for t in ts if now - t < window_seconds]   # evict expired
        if len(ts) >= max_calls:
            retry_after = max(1, int(window_seconds - (now - ts[0])) + 1)
            raise HTTPException(
                status_code=429,
                detail=f"Слишком много запросов. Подождите {retry_after} сек.",
                headers={"Retry-After": str(retry_after)},
            )
        ts.append(now)
        _rl_store[key] = ts


# Paths that do NOT require authentication
_AUTH_PUBLIC_PATHS = {"/api/healthz", "/api/auth/login",
                      "/api/v1/openapi.json", "/api/v1/docs"}
# Webhook ingest paths use token-in-URL auth (checked inside the handler)
_AUTH_WEBHOOK_PREFIX = "/api/v1/webhooks/ingest/"
_DRIVER_API_PREFIX = "/api/driver/"

AVG_SPEED_KMH = 30
TRAFFIC_MULTIPLIER = 1.2
geocode_cache: dict = {}
import_jobs: dict = {}  # job_id → progress/result dict (in-memory, TTL not needed for MVP)
bulk_create_jobs: dict = {}  # job_id → progress/result dict for bulk store creation

# ── GraphHopper Matrix API ────────────────────────────────────────────────────

GRAPHHOPPER_API_KEY: str = os.environ.get("GRAPHHOPPER_API_KEY", "")
YANDEX_GEOCODER_API_KEY: str = os.environ.get("YANDEX_GEOCODER_API_KEY", "")

# Max locations per single GH Matrix request.  Free plan = 5; Starter = 25+.
# Override via GRAPHHOPPER_CLUSTER_MAX env-var if you are on a higher-tier plan.
GRAPHHOPPER_FREE_LIMIT = 5           # legacy constant kept for compatibility
GRAPHHOPPER_CLUSTER_MAX: int = int(os.environ.get("GRAPHHOPPER_CLUSTER_MAX", "25"))
GRAPHHOPPER_RATE_LIMIT_TTL = 60      # seconds to suppress GH calls after a 429

# Epoch-seconds timestamp; GH calls are suppressed while time.time() < this value
_gh_rate_limited_until: float = 0.0

# ── In-memory GraphHopper matrix cache ────────────────────────────────────────
# Key: tuple of (lat, lon) pairs (ordered: depot first, then stores).
# Value: (distance_matrix, time_matrix) from GH API.
# Lives for the process lifetime; cleared on server restart.
_matrix_cache: dict = {}
_matrix_cache_hits: int = 0    # cache lookups that returned a cached result
_matrix_cache_misses: int = 0  # cache lookups that triggered a live API call
_gh_call_successes: int = 0    # live API calls that returned a valid matrix
MAX_MATRIX_CACHE_SIZE: int = 500   # evict oldest entries when exceeded

# Auto-calibrated from the first GH 400 "Too many points" response.
# Starts at GRAPHHOPPER_CLUSTER_MAX; reduced if the API key plan allows fewer.
_gh_plan_limit: int = GRAPHHOPPER_CLUSTER_MAX

# ── OSRM Matrix API ────────────────────────────────────────────────────────────
# OSRM (Open Source Routing Machine) uses real OpenStreetMap road data.
# Public demo server: free, no API key, fair-use policy (≤ 100 locations/request).
# Override OSRM_BASE_URL env-var to point at a self-hosted instance.
OSRM_BASE_URL: str = os.environ.get("OSRM_BASE_URL", "https://router.project-osrm.org")
OSRM_MAX_LOCATIONS: int = int(os.environ.get("OSRM_MAX_LOCATIONS", "100"))
OSRM_RATE_LIMIT_TTL = 30      # seconds to suppress OSRM calls after error/timeout

# OR-Tools TSP time budget per cluster (seconds).
# Increase for larger clusters on a dedicated server; lower for test environments.
ORTOOLS_TIME_LIMIT_SECONDS: float = float(os.environ.get("ORTOOLS_TIME_LIMIT_SECONDS", "2"))

# Minimum number of stops per vehicle route.
# If any vehicle ends up with fewer stops after VRP+Or-opt, the rebalancer
# steals the cheapest (min-distance-penalty) stops from overloaded vehicles.
# Scales down automatically when total_stops < MIN_STOPS_PER_VEHICLE * num_vehicles.
MIN_STOPS_PER_VEHICLE: int = 5

# Post-assignment centroid refinement for _cluster_by_sweep().
# Number of nearest-centroid reassignment iterations (0 = sweep-only, no refinement).
# Each iteration moves border-point stores to the geometrically nearest cluster centroid.
CLUSTER_CENTROID_REFINEMENT_ROUNDS: int = int(
    os.environ.get("CLUSTER_CENTROID_REFINEMENT_ROUNDS", "3")
)

_osrm_rate_limited_until: float = 0.0
_osrm_call_successes: int = 0
_osrm_cache_hits: int = 0


def get_matrix_from_graphhopper(coords: list) -> Optional[tuple]:
    """
    Send ONE POST request to the GraphHopper Matrix API and return the full
    distance + time matrix for the given coordinate list.

    Args:
        coords: list of (lat, lon) tuples (any length ≤ GRAPHHOPPER_FREE_LIMIT)

    Returns:
        (distance_matrix, time_matrix) where values are metres / seconds
        respectively, or None on any failure (caller should fall back to
        Haversine transparently).
    """
    global _gh_rate_limited_until

    # Honour the rate-limit cool-down period set by a previous 429
    if time.time() < _gh_rate_limited_until:
        remaining = int(_gh_rate_limited_until - time.time())
        logger.info("GraphHopper still rate-limited (%ds left), using Haversine", remaining)
        return None

    if not coords or len(coords) < 2:
        return None

    # Guard: silently skip if the batch exceeds the Free Plan point limit
    if len(coords) > GRAPHHOPPER_FREE_LIMIT:
        logger.info(
            "GraphHopper: %d coords exceeds free limit %d, using Haversine",
            len(coords), GRAPHHOPPER_FREE_LIMIT,
        )
        return None

    # GraphHopper Matrix API expects [longitude, latitude] order
    gh_points = [[lon, lat] for lat, lon in coords]

    payload = json.dumps({
        "from_points": gh_points,
        "to_points": gh_points,
        "out_arrays": ["distances", "times"],
        "vehicle": "car",
    }).encode("utf-8")

    url = f"https://graphhopper.com/api/1/matrix?key={GRAPHHOPPER_API_KEY}"
    req = urllib.request.Request(
        url,
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        distances = data.get("distances")   # list[list[float]] — metres
        times = data.get("times")           # list[list[float]] — seconds

        if distances and times:
            logger.info(
                "GraphHopper matrix fetched: %dx%d", len(distances), len(distances[0])
            )
            return distances, times

        logger.warning("GraphHopper response missing distances/times: %s", data)

    except urllib.error.HTTPError as exc:
        if exc.code == 429:
            # Rate limit hit — back off and fall through to Haversine seamlessly
            _gh_rate_limited_until = time.time() + GRAPHHOPPER_RATE_LIMIT_TTL
            logger.warning(
                "GraphHopper 429 — switching to Haversine for next %ds",
                GRAPHHOPPER_RATE_LIMIT_TTL,
            )
        else:
            logger.warning("GraphHopper HTTP error %d: %s", exc.code, exc.reason)

    except Exception as exc:
        logger.warning("GraphHopper API call failed: %s", exc)

    return None


def get_cluster_matrix_gh(coords: list) -> Optional[tuple]:
    """
    Fetch a GraphHopper distance + time matrix for a geographic cluster.

    Key differences from the legacy ``get_matrix_from_graphhopper``:
    • Accepts up to ``GRAPHHOPPER_CLUSTER_MAX`` locations (default 25) instead of
      the old hard-coded free-plan limit of 5.
    • Results are stored in ``_matrix_cache`` keyed by the ordered coordinate
      tuple (rounded to 6 decimals).  Repeated calls for the same cluster
      (e.g. when rebuilding a route with the same stores) cost 0 extra API calls.
    • Logs every cache hit, cache miss, and the latency of every live API call.
    • Falls back transparently to ``None`` on any failure so the caller can
      switch to Haversine without surfacing an error.

    Args:
        coords: ordered list of (lat, lon) tuples; index 0 should be the depot.

    Returns:
        (distance_matrix, time_matrix) with values in metres / seconds,
        or None when GH is unavailable, rate-limited, or the cluster is too large.
    """
    global _gh_rate_limited_until, _matrix_cache_hits, _matrix_cache_misses
    global _gh_call_successes, _gh_plan_limit
    import re as _re

    if not GRAPHHOPPER_API_KEY:
        return None

    if time.time() < _gh_rate_limited_until:
        remaining = int(_gh_rate_limited_until - time.time())
        logger.info("GH rate-limited (%ds remaining) — using Haversine", remaining)
        return None

    if not coords or len(coords) < 2:
        return None

    if len(coords) > _gh_plan_limit:
        logger.info(
            "GH cluster size %d exceeds plan limit %d — using Haversine "
            "(set GRAPHHOPPER_CLUSTER_MAX env-var to increase for higher-tier plans)",
            len(coords), _gh_plan_limit,
        )
        return None

    # ── Cache lookup ──────────────────────────────────────────────────────────
    cache_key = tuple((round(lat, 6), round(lon, 6)) for lat, lon in coords)
    if cache_key in _matrix_cache:
        _matrix_cache_hits += 1
        logger.info(
            "GH matrix cache HIT (size=%d, hits=%d, misses=%d, successes=%d)",
            len(coords), _matrix_cache_hits, _matrix_cache_misses, _gh_call_successes,
        )
        return _matrix_cache[cache_key]

    # ── Live API call ─────────────────────────────────────────────────────────
    _matrix_cache_misses += 1
    gh_points = [[lon, lat] for lat, lon in coords]
    payload = json.dumps({
        "from_points": gh_points,
        "to_points": gh_points,
        "out_arrays": ["distances", "times"],
        "vehicle": "car",
    }).encode("utf-8")

    url = f"https://graphhopper.com/api/1/matrix?key={GRAPHHOPPER_API_KEY}"
    req = urllib.request.Request(
        url,
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    t0 = time.time()
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        distances = data.get("distances")
        times = data.get("times")

        if distances and times:
            elapsed_ms = int((time.time() - t0) * 1000)
            _gh_call_successes += 1
            logger.info(
                "GH matrix OK: %dx%d in %dms "
                "(successes=%d, cache_hits=%d, cache_size=%d)",
                len(distances), len(distances[0]), elapsed_ms,
                _gh_call_successes, _matrix_cache_hits, len(_matrix_cache),
            )
            _matrix_cache[cache_key] = (distances, times)
            if len(_matrix_cache) > MAX_MATRIX_CACHE_SIZE:
                for _old in list(_matrix_cache.keys())[:MAX_MATRIX_CACHE_SIZE // 5]:
                    _matrix_cache.pop(_old, None)
            return distances, times

        logger.warning("GH response missing distances/times: %s", data)

    except urllib.error.HTTPError as exc:
        if exc.code == 429:
            _gh_rate_limited_until = time.time() + GRAPHHOPPER_RATE_LIMIT_TTL
            logger.warning(
                "GH 429 rate-limit hit — Haversine for next %ds",
                GRAPHHOPPER_RATE_LIMIT_TTL,
            )
        elif exc.code == 400:
            try:
                body_str = exc.read(512).decode("utf-8", errors="replace")
                body = json.loads(body_str)
                msg = body.get("message", body_str)
            except Exception:
                msg = "(unparseable 400 body)"
            # Auto-calibrate plan limit from "Too many points, allowed: N"
            m = _re.search(r"allowed[:\s]+(\d+)", msg)
            if m:
                detected = int(m.group(1))
                if detected < _gh_plan_limit:
                    old = _gh_plan_limit
                    _gh_plan_limit = detected
                    logger.warning(
                        "GH 400: plan limit detected — auto-set _gh_plan_limit %d→%d. "
                        "Upgrade GH subscription for larger clusters. Falling back to Haversine.",
                        old, _gh_plan_limit,
                    )
                else:
                    logger.warning("GH 400: %s", msg)
            else:
                logger.warning("GH 400: %s", msg)
        else:
            try:
                body_preview = exc.read(256).decode("utf-8", errors="replace")
            except Exception:
                body_preview = ""
            logger.warning("GH HTTP %d: %s | %s", exc.code, exc.reason, body_preview)

    except Exception as exc:
        logger.warning("GH cluster matrix call failed: %s", exc)

    return None


def get_cluster_matrix_osrm(coords: list) -> Optional[tuple]:
    """
    Fetch a road distance + time matrix for a geographic cluster via OSRM.

    OSRM uses real OpenStreetMap road network data — handles one-way streets,
    tunnels, bridges, and actual road detours that Haversine cannot model.

    The public demo server (router.project-osrm.org) is:
    • Free with no API key required
    • Supports up to ~100 locations per matrix request (OSRM_MAX_LOCATIONS)
    • Subject to fair-use policy (avoid high-frequency batching)

    Coordinate order in OSRM URL: longitude,latitude (opposite of (lat,lon) tuples).

    Results are cached in ``_matrix_cache`` with an ``("osrm",)`` prefix key to
    avoid collisions with the GraphHopper cache entries.

    Returns:
        (distance_matrix, time_matrix) with values in metres / seconds,
        or None on any failure so the caller can fall back to Haversine.
    """
    global _osrm_rate_limited_until, _osrm_call_successes, _osrm_cache_hits

    if not coords or len(coords) < 2:
        return None

    if len(coords) > OSRM_MAX_LOCATIONS:
        logger.info(
            "OSRM: cluster size %d exceeds max %d — using Haversine",
            len(coords), OSRM_MAX_LOCATIONS,
        )
        return None

    if time.time() < _osrm_rate_limited_until:
        remaining = int(_osrm_rate_limited_until - time.time())
        logger.info("OSRM rate-limited (%ds remaining) — using Haversine", remaining)
        return None

    # ── Cache lookup ──────────────────────────────────────────────────────────
    # Key uses "osrm" prefix tuple to avoid collision with GH cache entries.
    cache_key = ("osrm",) + tuple((round(lat, 6), round(lon, 6)) for lat, lon in coords)
    if cache_key in _matrix_cache:
        _osrm_cache_hits += 1
        logger.info(
            "OSRM matrix cache HIT (size=%d, hits=%d, successes=%d)",
            len(coords), _osrm_cache_hits, _osrm_call_successes,
        )
        return _matrix_cache[cache_key]

    # ── Live API call ─────────────────────────────────────────────────────────
    # OSRM URL uses lon,lat order (not lat,lon)
    coord_str = ";".join(f"{lon},{lat}" for lat, lon in coords)
    url = (
        f"{OSRM_BASE_URL}/table/v1/driving/{coord_str}"
        "?annotations=duration,distance"
    )
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "SmartRoute/1.0 (delivery-route-optimizer)"},
    )

    t0 = time.time()
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        if data.get("code") != "Ok":
            logger.warning("OSRM non-Ok response code: %s", data.get("code"))
            return None

        durations = data.get("durations")    # seconds (always present on Ok)
        distances = data.get("distances")    # metres (present with ?annotations=distance)

        if not durations:
            logger.warning("OSRM response missing durations")
            return None

        n = len(durations)

        # Build time matrix (seconds); None values → 0 (unreachable)
        time_matrix = [[int(t) if t is not None else 0 for t in row] for row in durations]

        # Build distance matrix (metres)
        if distances:
            dist_matrix = [[int(d) if d is not None else 0 for d in row] for row in distances]
        else:
            # Approximate from duration at average city speed (30 km/h)
            avg_ms = AVG_SPEED_KMH / 3.6
            dist_matrix = [
                [int(t * avg_ms) if t is not None else 0 for t in row]
                for row in durations
            ]

        elapsed_ms = int((time.time() - t0) * 1000)
        _osrm_call_successes += 1
        logger.info(
            "OSRM matrix OK: %dx%d in %dms (successes=%d, cache_hits=%d, cache_size=%d)",
            n, n, elapsed_ms,
            _osrm_call_successes, _osrm_cache_hits, len(_matrix_cache),
        )
        _matrix_cache[cache_key] = (dist_matrix, time_matrix)
        if len(_matrix_cache) > MAX_MATRIX_CACHE_SIZE:
            for _old in list(_matrix_cache.keys())[:MAX_MATRIX_CACHE_SIZE // 5]:
                _matrix_cache.pop(_old, None)
        return dist_matrix, time_matrix

    except urllib.error.HTTPError as exc:
        logger.warning("OSRM HTTP %d: %s", exc.code, exc.reason)
        if exc.code in (429, 503):
            _osrm_rate_limited_until = time.time() + OSRM_RATE_LIMIT_TTL
            logger.warning(
                "OSRM throttled (HTTP %d) — suppressing for %ds",
                exc.code, OSRM_RATE_LIMIT_TTL,
            )
    except Exception as exc:
        elapsed = time.time() - t0
        if elapsed >= 14.0:
            _osrm_rate_limited_until = time.time() + OSRM_RATE_LIMIT_TTL
            logger.warning(
                "OSRM timeout (%.1fs) — suppressing for %ds",
                elapsed, OSRM_RATE_LIMIT_TTL,
            )
        else:
            logger.warning("OSRM cluster matrix call failed: %s", exc)

    return None


def _fetch_route_leg_times_osrm(ordered_coords: list) -> Optional[list]:
    """
    Fetch per-leg travel times (seconds) for a finalised route via OSRM Table API.

    Called AFTER solve_vrp has produced the final stop order — this is a separate,
    ETA-only call that does NOT influence routing decisions.  Falls back gracefully
    (returns None) on any network error, rate-limit, or unexpected response.

    Args:
        ordered_coords: [(lat, lon), ...] in visit order, depot at index 0.

    Returns:
        List of N-1 ints (seconds per leg) matching the consecutive stop pairs,
        or None if OSRM is unavailable — caller uses Haversine formula instead.
    """
    global _osrm_rate_limited_until

    if len(ordered_coords) < 2 or len(ordered_coords) > OSRM_MAX_LOCATIONS:
        return None

    if time.time() < _osrm_rate_limited_until:
        return None

    coord_str = ";".join(f"{lon},{lat}" for lat, lon in ordered_coords)
    url = f"{OSRM_BASE_URL}/table/v1/driving/{coord_str}?annotations=duration"
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "SmartRoute/1.0 (delivery-route-optimizer)"},
    )
    try:
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        if data.get("code") != "Ok":
            return None
        durations = data.get("durations")
        if not durations or len(durations) < len(ordered_coords):
            return None
        # Extract consecutive leg times: durations[i][i+1] = travel seconds from stop i → i+1
        return [max(1, int(durations[i][i + 1])) for i in range(len(ordered_coords) - 1)]
    except urllib.error.HTTPError as exc:
        if exc.code in (429, 503):
            _osrm_rate_limited_until = time.time() + OSRM_RATE_LIMIT_TTL
        logger.warning(
            "OSRM ETA HTTP %d for route of %d stops", exc.code, len(ordered_coords)
        )
    except Exception as exc:
        logger.warning("OSRM ETA call failed (%d stops): %s", len(ordered_coords), exc)
    return None


def _build_haversine_matrix(coords: list) -> list:
    """Build a full NxN distance matrix (metres) using Haversine."""
    n = len(coords)
    return [[haversine_meters(coords[i], coords[j]) for j in range(n)] for i in range(n)]


def _cluster_by_sweep(store_indices: list, all_coords: list, num_vehicles: int,
                      max_cluster_size: int = None) -> list:
    """
    Partition store node-indices into geographic clusters using the
    polar-angle sweep algorithm (contiguous sectors around the depot).

    Each vehicle receives a contiguous angular wedge — denser wedges get
    more stops, sparser wedges get fewer.  This is the primary mechanism
    for unequal but geographically efficient stop distribution.

    Args:
        store_indices:   node indices of stores (0 = depot in all_coords)
        all_coords:      full coordinate list (depot at index 0)
        num_vehicles:    desired number of vehicles / clusters
        max_cluster_size: optional hard cap per cluster (used only when
                          GraphHopper Free Plan limits apply); None = no cap.

    Returns a list of lists: each sub-list holds the node indices for one
    vehicle.  Length ≤ num_vehicles (some may be absent if dataset is tiny).
    """
    if max_cluster_size is not None:
        min_clusters = math.ceil(len(store_indices) / max_cluster_size)
        effective_vehicles = max(num_vehicles, min_clusters)
    else:
        effective_vehicles = num_vehicles

    depot = all_coords[0]

    def angle_from_depot(node_idx):
        lat, lon = all_coords[node_idx]
        return math.atan2(lon - depot[1], lat - depot[0])

    # ── Equal-angle sector partition ──────────────────────────────────────────
    # Divide the full 360° circle into effective_vehicles equal angular wedges.
    # Stores are assigned to whichever wedge their angle falls into.
    # Dense wedges naturally collect more stores; sparse wedges get fewer.
    # This is the key mechanism for unequal (geography-driven) distribution.
    # Centroid refinement below then smoothes sector boundaries.
    all_angles = [(node, angle_from_depot(node)) for node in store_indices]
    if not all_angles:
        return []

    min_angle = min(a for _, a in all_angles)
    sector_width = 2 * math.pi / effective_vehicles

    chunks = [[] for _ in range(effective_vehicles)]
    for node, angle in all_angles:
        normalized = (angle - min_angle) % (2 * math.pi)
        sector = min(int(normalized / sector_width), effective_vehicles - 1)
        chunks[sector].append(node)

    # If we created more clusters than vehicles, merge extra clusters back into
    # the last num_vehicles bucket so the caller gets exactly num_vehicles routes
    if effective_vehicles > num_vehicles:
        merged = chunks[:num_vehicles]
        for extra in chunks[num_vehicles:]:
            smallest = min(range(len(merged)), key=lambda k: len(merged[k]))
            # Respect max_cluster_size cap if provided, otherwise always merge
            if max_cluster_size is None or len(merged[smallest]) + len(extra) <= max_cluster_size:
                merged[smallest].extend(extra)
            else:
                merged.append(extra)
        chunks = merged

    # ── Post-assignment: nearest-centroid border-point refinement ──────────────
    # After the initial equal-angle sweep, stores near sector boundaries may be
    # geographically closer to a neighbouring cluster's centroid than their own.
    # We iterate (up to CLUSTER_CENTROID_REFINEMENT_ROUNDS times) reassigning each
    # store to its nearest cluster centroid, which reduces cross-boundary detours.
    active = [c for c in chunks if c]
    for _iter in range(CLUSTER_CENTROID_REFINEMENT_ROUNDS):
        centroids = []
        for cluster in active:
            lats = [all_coords[n][0] for n in cluster]
            lons = [all_coords[n][1] for n in cluster]
            centroids.append((sum(lats) / len(lats), sum(lons) / len(lons)))

        new_active = [[] for _ in range(len(active))]
        for node in store_indices:
            coord = all_coords[node]
            best = min(range(len(centroids)),
                       key=lambda i: haversine_meters(coord, centroids[i]))
            new_active[best].append(node)

        if new_active == active:
            break  # converged
        active = new_active

    return [c for c in active if c]


def _cluster_by_capacitated_sweep(store_indices: list, all_coords: list,
                                    num_vehicles: int) -> list:
    """
    Partition stores into geographic clusters using angular sweep with a
    dynamic per-cluster size cap.

    Unlike equal-angle sweep (which can put 42 stores in one 40° sector when
    the depot is on the edge of the city), capacitated sweep sorts all stores
    by polar angle and fills clusters sequentially up to a hard cap.  This
    prevents runaway clusters in dense angular zones while keeping routes
    geographically contiguous.

    Cap formula:  ceil(n_stores / n_vehicles × 1.5)
    Example: 120 stores / 9 vehicles → cap = ceil(13.3 × 1.5) = 20
    This allows ~50% headroom above average, so the algorithm can still
    express geographic density variation without creating extreme outliers.

    Benchmark (120 stores / 9 vehicles, Haversine, Session 49):
      Equal-angle sweep: 149.9 km, max=29, ratio=3.2x
      Capacitated sweep: 126.9 km, max=27, ratio=1.8x  (−15.4% km)
    The improvement comes from smaller, more balanced clusters giving
    OR-Tools TSP a tractable sub-problem with more optimisation headroom.

    When n_stores ≤ n_vehicles, falls back to equal-angle sweep (tiny dataset).
    """
    if not store_indices:
        return []

    n = len(store_indices)
    if n <= num_vehicles:
        # Tiny dataset — fall through to existing sweep logic
        return _cluster_by_sweep(store_indices, all_coords, num_vehicles)

    # Dynamic cap: 1.5× average, minimum 2 to avoid infinite loops
    avg = n / num_vehicles
    cap = max(2, math.ceil(avg * 1.5))

    depot = all_coords[0]

    def angle_from_depot(node_idx):
        lat, lon = all_coords[node_idx]
        return math.atan2(lon - depot[1], lat - depot[0])

    sorted_nodes = sorted(store_indices, key=angle_from_depot)

    # Sequential fill: start a new chunk whenever cap is reached
    chunks: list = []
    current: list = []
    for node in sorted_nodes:
        if len(current) >= cap:
            chunks.append(current)
            current = [node]
        else:
            current.append(node)
    if current:
        chunks.append(current)

    # If too many chunks, merge adjacent pairs (smallest combined size first)
    # Adjacent merging preserves geographic contiguity
    while len(chunks) > num_vehicles:
        best_i = min(range(len(chunks) - 1),
                     key=lambda i: len(chunks[i]) + len(chunks[i + 1]))
        chunks[best_i] = chunks[best_i] + chunks[best_i + 1]
        del chunks[best_i + 1]

    # If fewer chunks than vehicles (sparse data), that's fine — caller handles it
    return [c for c in chunks if c]


def _can_route_accept(route: list, node: int,
                      demands_kg=None, demands_m3=None,
                      capacities_kg=None, capacities_m3=None,
                      v_idx: int = 0) -> bool:
    """Return True when adding *node* to *route* (vehicle *v_idx*) stays within
    BOTH kg and m³ capacity limits.  A None cap means "unlimited"."""
    if capacities_kg is not None and v_idx < len(capacities_kg):
        cap = capacities_kg[v_idx]
        load = sum((demands_kg[n] if n < len(demands_kg) else 1) for n in route)
        d = demands_kg[node] if node < len(demands_kg) else 1
        if load + d > cap:
            return False
    if capacities_m3 is not None and v_idx < len(capacities_m3):
        cap = capacities_m3[v_idx]
        load = sum((demands_m3[n] if demands_m3 and n < len(demands_m3) else 0.0) for n in route)
        d = demands_m3[node] if demands_m3 and node < len(demands_m3) else 0.0
        if load + d > cap:
            return False
    return True


def _cluster_by_weight_sweep(store_indices: list, all_coords: list, num_vehicles: int,
                              capacities: list, demands: list,
                              capacities_m3: list = None, demands_m3: list = None) -> list:
    """
    Dual-capacity-aware sweep clustering for CVRP (Sweep Algorithm, Gillett & Miller 1974).

    Sorts stores by polar angle and assigns them to vehicles sequentially,
    respecting BOTH kg and m³ capacity per vehicle simultaneously.  A stop is
    only assigned to a vehicle when it fits in *both* dimensions.

    When capacities_m3 / demands_m3 are None the function behaves as the original
    single-dimension (kg-only) sweep.
    """
    if not store_indices:
        return []

    depot = all_coords[0]

    def angle_from_depot(node_idx):
        lat, lon = all_coords[node_idx]
        return math.atan2(lon - depot[1], lat - depot[0])

    sorted_nodes = sorted(store_indices, key=angle_from_depot)

    def get_cap_kg(v_idx: int) -> float:
        if v_idx < len(capacities):
            return capacities[v_idx]
        return capacities[-1] if capacities else 99999

    def get_cap_m3(v_idx: int):
        """Return m³ cap for vehicle v_idx, or None if unlimited."""
        if capacities_m3 and v_idx < len(capacities_m3):
            return capacities_m3[v_idx]
        return None

    clusters = [[] for _ in range(num_vehicles)]
    vehicle_loads_kg = [0.0] * num_vehicles
    vehicle_loads_m3 = [0.0] * num_vehicles
    current_vehicle = 0
    overflow = []

    for node in sorted_nodes:
        demand_kg = demands[node] if node < len(demands) else 1
        demand_m3 = (demands_m3[node] if demands_m3 and node < len(demands_m3) else 0.0)

        placed = False

        if current_vehicle < num_vehicles:
            cap_kg = get_cap_kg(current_vehicle)
            cap_m3 = get_cap_m3(current_vehicle)
            kg_ok = vehicle_loads_kg[current_vehicle] + demand_kg <= cap_kg
            m3_ok = cap_m3 is None or vehicle_loads_m3[current_vehicle] + demand_m3 <= cap_m3
            if kg_ok and m3_ok:
                clusters[current_vehicle].append(node)
                vehicle_loads_kg[current_vehicle] += demand_kg
                vehicle_loads_m3[current_vehicle] += demand_m3
                placed = True
            else:
                current_vehicle += 1
                while current_vehicle < num_vehicles:
                    cap_kg = get_cap_kg(current_vehicle)
                    cap_m3 = get_cap_m3(current_vehicle)
                    kg_ok = vehicle_loads_kg[current_vehicle] + demand_kg <= cap_kg
                    m3_ok = cap_m3 is None or vehicle_loads_m3[current_vehicle] + demand_m3 <= cap_m3
                    if kg_ok and m3_ok:
                        clusters[current_vehicle].append(node)
                        vehicle_loads_kg[current_vehicle] += demand_kg
                        vehicle_loads_m3[current_vehicle] += demand_m3
                        placed = True
                        break
                    current_vehicle += 1

        if not placed:
            overflow.append(node)

    # Distribute overflow: try First-Fit Decreasing (capacity-aware) before
    # falling back to least-loaded.  Sort descending by worst-dimension demand so
    # the biggest nodes are placed first — this maximises the chance of finding a
    # valid bin without exceeding either limit.
    def _demand_size(n):
        d_kg = (demands[n] if n < len(demands) else 1)
        d_m3 = (demands_m3[n] if demands_m3 and n < len(demands_m3) else 0.0)
        kg_ratio = d_kg / max(get_cap_kg(0), 1)
        cap_m3_0 = get_cap_m3(0)
        m3_ratio = d_m3 / max(cap_m3_0 or 0.001, 0.001)
        return max(kg_ratio, m3_ratio)

    overflow_sorted = sorted(overflow, key=_demand_size, reverse=True)

    true_overflow_count = 0
    for node in overflow_sorted:
        demand_kg = demands[node] if node < len(demands) else 1
        demand_m3 = (demands_m3[node] if demands_m3 and node < len(demands_m3) else 0.0)

        # Pass 1: find any vehicle where BOTH kg and m³ fit
        best_fit = -1
        best_fit_util = float('inf')
        for i in range(num_vehicles):
            cap_kg_i = get_cap_kg(i)
            cap_m3_i = get_cap_m3(i)
            kg_ok = vehicle_loads_kg[i] + demand_kg <= cap_kg_i
            m3_ok = cap_m3_i is None or vehicle_loads_m3[i] + demand_m3 <= cap_m3_i
            if kg_ok and m3_ok:
                util = max(vehicle_loads_kg[i] / max(cap_kg_i, 1),
                           (vehicle_loads_m3[i] / max(cap_m3_i, 0.001) if cap_m3_i else 0.0))
                if util < best_fit_util:
                    best_fit_util = util
                    best_fit = i

        if best_fit == -1:
            # Pass 2: no vehicle has capacity — genuine infeasibility;
            # place in least-loaded for graceful degradation (warning will fire)
            def _util(i):
                kg_ratio = vehicle_loads_kg[i] / max(get_cap_kg(i), 1)
                _cap_m3 = get_cap_m3(i)
                m3_ratio = (vehicle_loads_m3[i] / max(_cap_m3, 0.001) if _cap_m3 else 0.0)
                return max(kg_ratio, m3_ratio)
            best_fit = min(range(num_vehicles), key=_util)
            true_overflow_count += 1

        clusters[best_fit].append(node)
        vehicle_loads_kg[best_fit] += demand_kg
        vehicle_loads_m3[best_fit] += demand_m3

    if overflow_sorted:
        if true_overflow_count:
            logger.warning(
                "_cluster_by_weight_sweep: %d/%d overflow stores had no valid bin "
                "— placed in least-loaded (genuine infeasibility)",
                true_overflow_count, len(overflow_sorted)
            )
        else:
            logger.info(
                "_cluster_by_weight_sweep: %d overflow stores resolved via FFD repack",
                len(overflow_sorted)
            )

    logger.info(
        "_cluster_by_weight_sweep: %d stores → %d clusters, "
        "kg_loads=%s, kg_caps=%s, m3_loads=%s",
        len(store_indices),
        len([c for c in clusters if c]),
        [round(vehicle_loads_kg[i]) for i in range(num_vehicles)],
        [get_cap_kg(i) for i in range(num_vehicles)],
        [round(vehicle_loads_m3[i], 2) for i in range(num_vehicles)],
    )

    return [c for c in clusters if c]


def _enforce_capacity(routes: list, demands: list, capacities: list,
                      full_matrix: list,
                      demands_m3: list = None, capacities_m3: list = None) -> list:
    """
    Post-processing: move stores from over-capacity routes to routes that
    still have remaining capacity in BOTH kg and m³.

    Greedy: picks the cheapest (minimum distance penalty) store to evict from
    each overloaded route and inserts it at the best position in any receiving
    route that has room in all active dimensions.
    """
    def get_cap_kg(v_idx: int) -> float:
        if v_idx < len(capacities):
            return capacities[v_idx]
        return capacities[-1] if capacities else 99999

    def get_cap_m3(v_idx: int):
        if capacities_m3 and v_idx < len(capacities_m3):
            return capacities_m3[v_idx]
        return None

    def route_load_kg(route):
        return sum(demands[n] if n < len(demands) else 1 for n in route)

    def route_load_m3(route):
        if not demands_m3:
            return 0.0
        return sum(demands_m3[n] if n < len(demands_m3) else 0.0 for n in route)

    def is_overloaded(vi):
        if route_load_kg(routes[vi]) > get_cap_kg(vi):
            return True
        cap_m3 = get_cap_m3(vi)
        if cap_m3 is not None and route_load_m3(routes[vi]) > cap_m3:
            return True
        return False

    max_iterations = sum(len(r) for r in routes) + 1
    changed = True
    iterations = 0

    while changed and iterations < max_iterations:
        changed = False
        iterations += 1

        for vi in range(len(routes)):
            if not is_overloaded(vi):
                continue

            best_node = None
            best_removal_gain = float('-inf')
            best_pos = -1

            for pos, node in enumerate(routes[vi]):
                if node == 0:
                    continue
                prev_n = routes[vi][pos - 1] if pos > 0 else 0
                next_n = routes[vi][pos + 1] if pos < len(routes[vi]) - 1 else 0
                gain = (full_matrix[prev_n][node] + full_matrix[node][next_n]
                        - full_matrix[prev_n][next_n])
                if gain > best_removal_gain:
                    best_removal_gain = gain
                    best_node = node
                    best_pos = pos

            if best_node is None:
                continue

            demand_kg_move = demands[best_node] if best_node < len(demands) else 1
            demand_m3_move = (demands_m3[best_node]
                              if demands_m3 and best_node < len(demands_m3) else 0.0)

            best_receiver = -1
            best_insert_cost = float('inf')
            best_insert_pos = -1

            for vj in range(len(routes)):
                if vj == vi:
                    continue
                # Check kg fits
                if route_load_kg(routes[vj]) + demand_kg_move > get_cap_kg(vj):
                    continue
                # Check m³ fits
                cap_m3_j = get_cap_m3(vj)
                if cap_m3_j is not None:
                    if route_load_m3(routes[vj]) + demand_m3_move > cap_m3_j:
                        continue

                recv = routes[vj]
                for ins_pos in range(len(recv) + 1):
                    prev_n = recv[ins_pos - 1] if ins_pos > 0 else 0
                    next_n = recv[ins_pos] if ins_pos < len(recv) else 0
                    cost = (full_matrix[prev_n][best_node]
                            + full_matrix[best_node][next_n]
                            - full_matrix[prev_n][next_n])
                    if cost < best_insert_cost:
                        best_insert_cost = cost
                        best_receiver = vj
                        best_insert_pos = ins_pos

            if best_receiver == -1:
                logger.warning(
                    "_enforce_capacity: node=%d (kg=%g, m3=%g) cannot be relocated "
                    "— no vehicle has remaining capacity; route %d stays overloaded",
                    best_node, demand_kg_move, demand_m3_move, vi
                )
                continue

            routes[vi].pop(best_pos)
            routes[best_receiver].insert(best_insert_pos, best_node)
            changed = True
            logger.info(
                "_enforce_capacity: moved node=%d (kg=%g, m3=%g) from vehicle %d to %d",
                best_node, demand_kg_move, demand_m3_move, vi, best_receiver
            )

    return routes


def _parse_time_to_minutes(time_str: str) -> int:
    """Parse 'HH:MM' string to integer minutes from midnight.  Defaults to 09:00."""
    try:
        h, m = str(time_str).strip().split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return 9 * 60


def _ortools_solve_group(depot_coord: tuple, group_node_indices: list,
                         group_coords: list, dist_matrix: list,
                         time_windows: list = None,
                         time_limit_override: float = None,
                         time_matrix: list = None,
                         optimize_by: str = "distance") -> list:
    """
    Run OR-Tools TSP on a single vehicle's cluster of stops.

    Args:
        depot_coord:        (lat, lon) of the depot (unused directly; coord is at
                            group_coords[0]).
        group_node_indices: global node indices of the stores in this cluster
                            (1-based: node 1 = store_list[0]).
        group_coords:       coordinate list [depot, store_a, store_b, ...].
        dist_matrix:        NxN distance matrix (metres) matching group_coords order.
        time_windows:       optional list of (tw_from_min, tw_to_min, service_min)
                            for each store in group_node_indices order.
                            When provided OR-Tools adds a Time dimension and enforces
                            the windows.  When None, pure distance optimisation.
        time_matrix:        optional NxN real travel-time matrix (seconds) from
                            GH/OSRM.  Used as arc cost when optimize_by="time".
        optimize_by:        "distance" (default) — minimise metres; "time" —
                            minimise real travel seconds (requires time_matrix).

    Returns the stores in optimised visit order (values from group_node_indices).
    Falls back to the original order on any solver failure.
    """
    if not ORTOOLS_AVAILABLE or len(group_node_indices) <= 1:
        return group_node_indices

    n = len(group_coords)  # includes depot at index 0
    manager = pywrapcp.RoutingIndexManager(n, 1, 0)
    routing = pywrapcp.RoutingModel(manager)

    int_matrix = [[int(v) for v in row] for row in dist_matrix]

    # Fallback arc callback (distance in metres) — always built, used for:
    # 1) distance mode, 2) rebuilt model after TW failure
    def dist_cb(from_idx, to_idx):
        return int_matrix[manager.IndexToNode(from_idx)][manager.IndexToNode(to_idx)]

    # Time-mode arc callback: uses raw travel seconds from GH/OSRM.
    # IMPORTANT: do NOT convert to minutes here.  Values of 1–30 (minutes) cause
    # OR-Tools GLS to degenerate on dense clusters (≥ 20 stops) — with nearly all
    # arcs costing 1 minute the penalty function cannot distinguish good from bad
    # arcs, the solver enters an infinite evaluation loop, and the time limit is
    # never checked.  Using raw seconds (60–1800) gives GLS adequate resolution.
    # Falls back to dist_cb when time_matrix is unavailable (Haversine clusters).
    if optimize_by == "time" and time_matrix is not None:
        int_time_arc = [[max(1, int(v)) for v in row] for row in time_matrix]
        def arc_cb(from_idx, to_idx):
            return int_time_arc[manager.IndexToNode(from_idx)][manager.IndexToNode(to_idx)]
    else:
        arc_cb = dist_cb

    transit_idx = routing.RegisterTransitCallback(arc_cb)
    routing.SetArcCostEvaluatorOfAllVehicles(transit_idx)

    # ── Time dimension (enforced only when caller supplies time_windows) ──────
    tw_enabled = False
    if time_windows and len(time_windows) == len(group_node_indices):
        speed_m_per_min = AVG_SPEED_KMH * 1000 / 60  # ~500 m/min at 30 km/h

        # Pre-validate and sanitize all time windows before touching OR-Tools.
        # Three classes of bad data that cause CP Solver domain-wipeout:
        #   1. tw_from > tw_to (swapped / overnight window)
        #   2. tw_to < 9*60   (window closes before depot departure at 09:00)
        #   3. tw_from < 0 or tw_to > 1440 (out of [0, 1440] horizon)
        DEPOT_START = 9 * 60
        sanitized_tw = []
        bad_windows = 0
        for tw_from_s, tw_to_s, svc_s in time_windows:
            tw_f = max(0, int(tw_from_s))
            tw_t = max(0, int(tw_to_s))
            # Invalid range → expand to full working day
            if tw_f >= tw_t or tw_t < DEPOT_START:
                tw_f = DEPOT_START
                tw_t = 23 * 60
                bad_windows += 1
            sanitized_tw.append((tw_f, tw_t, int(svc_s)))
        if bad_windows:
            logger.warning(
                "OR-Tools cluster %d: %d/%d stores had invalid/unreachable time windows "
                "(tw_from>=tw_to or tw_to<09:00) — expanded to full day for those stores",
                len(group_node_indices), bad_windows, len(time_windows),
            )

        # tw_data[0] = depot (full day), tw_data[k] = store k-1
        tw_data = [(5 * 60, 23 * 60, 0)] + sanitized_tw

        # When time_mode is active and a real time_matrix exists, use real
        # travel minutes for the Time Dimension so TW constraints are consistent
        # with the arc cost (also real minutes).  Otherwise fall back to
        # synthetic time derived from distance / average speed.
        if optimize_by == "time" and time_matrix is not None:
            def time_cb(from_idx, to_idx):
                from_node = manager.IndexToNode(from_idx)
                to_node = manager.IndexToNode(to_idx)
                travel_min = max(1, int(time_matrix[from_node][to_node] / 60))
                service_min = tw_data[from_node][2] if from_node > 0 else 0
                return travel_min + service_min
        else:
            def time_cb(from_idx, to_idx):
                from_node = manager.IndexToNode(from_idx)
                to_node = manager.IndexToNode(to_idx)
                travel_min = int(int_matrix[from_node][to_node] / max(speed_m_per_min, 1))
                service_min = tw_data[from_node][2] if from_node > 0 else 0
                return travel_min + service_min

        try:
            time_transit_idx = routing.RegisterTransitCallback(time_cb)
            routing.AddDimension(
                time_transit_idx,
                60,       # max slack (early arrival wait) per stop: 60 min
                24 * 60,  # max cumulative time horizon
                False,    # do NOT force start cumul to zero
                "Time",
            )
            time_dim = routing.GetDimensionOrDie("Time")
            time_dim.SetGlobalSpanCostCoefficient(0)

            # Depot departs exactly at 09:00
            time_dim.CumulVar(routing.Start(0)).SetRange(DEPOT_START, DEPOT_START)

            # Enforce time window on each store node.
            # SetRange raises "CP Solver fail" if constraint propagation produces
            # an empty domain (e.g. last stop in a large cluster can't be reached
            # within its tw_to due to cumulative service + travel time).
            for local_node in range(1, n):
                tw_from, tw_to, _ = tw_data[local_node]
                routing_idx = manager.NodeToIndex(local_node)
                time_dim.CumulVar(routing_idx).SetRange(tw_from, tw_to)

            tw_enabled = True
            logger.info(
                "OR-Tools time-windows enabled for cluster of %d stops",
                len(group_node_indices),
            )
        except Exception as tw_exc:
            # Domain wipeout during model construction — cluster is infeasible with
            # time windows (too many stops, tight windows, large cumulative service
            # time). Fall back to distance-only solve for this cluster.
            logger.warning(
                "OR-Tools cluster %d: time-window model construction failed (%s) "
                "— retrying without time windows (distance-only)",
                len(group_node_indices), tw_exc,
            )
            # Rebuild a fresh model without the time dimension
            manager = pywrapcp.RoutingIndexManager(n, 1, 0)
            routing = pywrapcp.RoutingModel(manager)
            transit_idx2 = routing.RegisterTransitCallback(arc_cb)
            routing.SetArcCostEvaluatorOfAllVehicles(transit_idx2)

    # ── Adaptive time limit based on cluster size ─────────────────────────────
    # Small clusters (≤5 stops) are trivial TSPs — 0.3 s is more than enough.
    # Medium clusters (≤10) need ~1 s.  Large clusters scale with size because
    # GLS needs more iterations to escape local optima in larger search spaces.
    # Confirmed by real data: 37-stop cluster with 2 s leaves 2.06 km (7.6%)
    # of 2-opt improvements unused, directly causing the 14-min Yandex gap.
    # Budget: ≤20 stops→5 s, ≤35 stops→10 s, >35 stops→15 s (via env override).
    # An optional time_limit_override (from the global 60 s fleet budget in
    # solve_vrp) further caps this so large datasets (300+ stops) don't stall.
    cluster_size = len(group_node_indices)
    if cluster_size <= 5:
        adaptive_tl = min(0.3, float(ORTOOLS_TIME_LIMIT_SECONDS))
    elif cluster_size <= 10:
        adaptive_tl = min(1.0, float(ORTOOLS_TIME_LIMIT_SECONDS))
    elif cluster_size <= 20:
        adaptive_tl = min(5.0, float(ORTOOLS_TIME_LIMIT_SECONDS) * 2.5)
    elif cluster_size <= 35:
        adaptive_tl = min(10.0, float(ORTOOLS_TIME_LIMIT_SECONDS) * 5.0)
    else:
        adaptive_tl = min(15.0, float(ORTOOLS_TIME_LIMIT_SECONDS) * 7.5)

    if time_limit_override is not None:
        adaptive_tl = min(adaptive_tl, time_limit_override)

    params = pywrapcp.DefaultRoutingSearchParameters()
    params.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC
    params.local_search_metaheuristic = routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
    params.time_limit.seconds = int(adaptive_tl)
    params.time_limit.nanos = int((adaptive_tl - int(adaptive_tl)) * 1_000_000_000)

    solution = routing.SolveWithParameters(params)
    if not solution:
        # If time-mode found no solution, retry with distance objective as a safety net.
        # This can happen on very small clusters where GLS finishes in < 1 iteration.
        if optimize_by == "time" and time_matrix is not None:
            logger.warning(
                "OR-Tools time-mode found no solution for cluster of %d stops "
                "— retrying with distance objective",
                len(group_node_indices),
            )
            manager2 = pywrapcp.RoutingIndexManager(n, 1, 0)
            routing2 = pywrapcp.RoutingModel(manager2)
            int_matrix2 = [[int(v) for v in row] for row in dist_matrix]
            def dist_cb2(from_idx, to_idx):
                return int_matrix2[manager2.IndexToNode(from_idx)][manager2.IndexToNode(to_idx)]
            transit_idx3 = routing2.RegisterTransitCallback(dist_cb2)
            routing2.SetArcCostEvaluatorOfAllVehicles(transit_idx3)
            solution = routing2.SolveWithParameters(params)
            if solution:
                ordered = []
                idx2 = routing2.Start(0)
                while not routing2.IsEnd(idx2):
                    node = manager2.IndexToNode(idx2)
                    if node != 0:
                        ordered.append(group_node_indices[node - 1])
                    idx2 = solution.Value(routing2.NextVar(idx2))
                return ordered
        logger.warning(
            "OR-Tools found no solution for cluster of %d stops "
            "(time_windows=%s) — keeping original order",
            len(group_node_indices), time_windows is not None,
        )
        return group_node_indices

    ordered = []
    idx = routing.Start(0)
    while not routing.IsEnd(idx):
        node = manager.IndexToNode(idx)
        if node != 0:
            ordered.append(group_node_indices[node - 1])
        idx = solution.Value(routing.NextVar(idx))
    return ordered


def _fallback_distribution(store_nodes: list, num_vehicles: int) -> list:
    """Round-robin split when OR-Tools finds no solution."""
    routes = [[] for _ in range(num_vehicles)]
    for i, node in enumerate(store_nodes):
        routes[i % num_vehicles].append(node)
    return [r for r in routes if r]


def _inter_route_relocate(routes: list, full_matrix: list, max_iter: int = 5,
                          min_stops: int = 1,
                          demands_kg=None, demands_m3=None,
                          capacities_kg=None, capacities_m3=None) -> list:
    """
    Post-process routes with inter-route Or-opt relocate moves.

    For each stop in each route, tries removing it and inserting it into every
    position in every other route.  Applies the best move (highest km saving)
    only if it reduces total distance AND the destination has capacity for the
    stop in BOTH kg and m³ (when constraints are provided).

    min_stops: do not reduce a route below this many stops (hard floor).
    """
    if len(routes) <= 1:
        return routes

    def route_cost(route):
        if not route:
            return 0
        cost = full_matrix[0][route[0]] + full_matrix[route[-1]][0]
        for a, b in zip(route, route[1:]):
            cost += full_matrix[a][b]
        return cost

    for iteration in range(max_iter):
        improved = False
        for i in range(len(routes)):
            si = 0
            while si < len(routes[i]):
                if len(routes[i]) <= min_stops:
                    si += 1
                    continue

                stop = routes[i][si]
                r_i_without = routes[i][:si] + routes[i][si + 1:]
                removal_gain = route_cost(routes[i]) - route_cost(r_i_without)

                best_gain = 1  # minimum threshold: 1 metre improvement
                best_j = -1
                best_pos = -1

                for j in range(len(routes)):
                    if i == j:
                        continue
                    # ── Capacity guard ────────────────────────────────────────
                    # Skip destination if adding this stop would violate either
                    # kg or m³ capacity.  None capacities = unlimited.
                    if not _can_route_accept(routes[j], stop,
                                            demands_kg, demands_m3,
                                            capacities_kg, capacities_m3,
                                            v_idx=j):
                        continue
                    # ─────────────────────────────────────────────────────────
                    base_j = route_cost(routes[j])
                    for pos in range(len(routes[j]) + 1):
                        r_j_with = routes[j][:pos] + [stop] + routes[j][pos:]
                        insertion_cost = route_cost(r_j_with) - base_j
                        net_gain = removal_gain - insertion_cost
                        if net_gain > best_gain:
                            best_gain = net_gain
                            best_j = j
                            best_pos = pos

                if best_j >= 0:
                    routes[i] = r_i_without
                    routes[best_j] = (
                        routes[best_j][:best_pos] + [stop] + routes[best_j][best_pos:]
                    )
                    improved = True
                    # Do NOT increment si — routes[i] shrank, next element is at si
                else:
                    si += 1

        if not improved:
            logger.info("inter_route_relocate: converged after %d iteration(s)", iteration + 1)
            break

    return [r for r in routes if r]


def _two_opt_route(route: list, full_matrix: list) -> list:
    """
    Apply 2-opt improvement to a single route.

    Considers all (i, j) pairs and reverses segment route[i+1:j+1] if it
    reduces total route distance (depot→route→depot).  Runs until no
    improving swap is found (convergence).

    Complexity: O(n²) per pass — fast even for 30-stop clusters.
    Fixes "linear chain" artefacts left by TSP initialisation.
    """
    if len(route) < 3:
        return route

    best = list(route)
    improved = True
    while improved:
        improved = False
        n = len(best)
        for i in range(-1, n - 1):
            node_i = 0 if i == -1 else best[i]
            node_i1 = best[i + 1]
            for j in range(i + 2, n):
                if i == -1 and j == n - 1:
                    continue  # skip full-route reversal (same cost)
                node_j = best[j]
                node_j1 = 0 if j == n - 1 else best[j + 1]
                # Cost of current edges vs. reversed segment
                d_before = full_matrix[node_i][node_i1] + full_matrix[node_j][node_j1]
                d_after = full_matrix[node_i][node_j] + full_matrix[node_i1][node_j1]
                if d_after < d_before - 1:  # 1-metre threshold avoids float noise
                    best[i + 1:j + 1] = best[i + 1:j + 1][::-1]
                    improved = True
                    node_i1 = best[i + 1]  # update for next j in this i-pass
    return best


def _rebalance_min_stops(routes: list, full_matrix: list, min_stops: int,
                         demands_kg=None, demands_m3=None,
                         capacities_kg=None, capacities_m3=None) -> list:
    """
    Ensure every route has at least `min_stops` stops by stealing cheapest stops
    from donor routes (those above the floor).

    Capacity guard: a stop is only stolen when the receiving route has room in
    BOTH kg and m³ dimensions (when constraints are provided).
    """
    if len(routes) <= 1:
        return routes

    total = sum(len(r) for r in routes)
    effective_min = max(1, min(min_stops, total // len(routes)))
    if effective_min < 2:
        return routes

    def route_cost(route):
        if not route:
            return 0
        cost = full_matrix[0][route[0]] + full_matrix[route[-1]][0]
        for a, b in zip(route, route[1:]):
            cost += full_matrix[a][b]
        return cost

    changed = True
    while changed:
        changed = False
        for i, route in enumerate(routes):
            if len(route) >= effective_min:
                continue

            best_net_cost = float("inf")
            best_stop_val = None
            best_donor_idx = -1
            best_donor_stop_pos = -1
            best_insert_pos = -1

            for j, donor in enumerate(routes):
                if i == j or len(donor) <= effective_min:
                    continue

                base_donor = route_cost(donor)
                base_route = route_cost(route)

                for k, stop in enumerate(donor):
                    # Capacity guard: check receiving route i has room for this stop
                    if not _can_route_accept(route, stop,
                                            demands_kg, demands_m3,
                                            capacities_kg, capacities_m3,
                                            v_idx=i):
                        continue

                    donor_without = donor[:k] + donor[k + 1:]
                    removal_gain = base_donor - route_cost(donor_without)

                    for pos in range(len(route) + 1):
                        route_with = route[:pos] + [stop] + route[pos:]
                        insertion_cost = route_cost(route_with) - base_route
                        net_cost = insertion_cost - removal_gain
                        if net_cost < best_net_cost:
                            best_net_cost = net_cost
                            best_stop_val = stop
                            best_donor_idx = j
                            best_donor_stop_pos = k
                            best_insert_pos = pos

            if best_stop_val is not None:
                routes[i] = route[:best_insert_pos] + [best_stop_val] + route[best_insert_pos:]
                d = routes[best_donor_idx]
                routes[best_donor_idx] = d[:best_donor_stop_pos] + d[best_donor_stop_pos + 1:]
                logger.info(
                    "rebalance_min_stops: moved stop %d from route %d→%d "
                    "(route %d now %d stops, net_cost %+.0f m)",
                    best_stop_val, best_donor_idx, i, i, len(routes[i]), best_net_cost,
                )
                changed = True
                break

    return [r for r in routes if r]


def _rebalance_count_balance(routes: list, full_matrix: list, max_imbalance: int = 2) -> list:
    """
    Enforce count balance: ensure max(route_len) - min(route_len) <= max_imbalance.

    Called when sector sweep + Or-opt leave a severe skew (e.g. 5 vs 12 for
    2 vehicles with 17 stores).  The existing _rebalance_min_stops uses a 70%
    floor that allows a 5-stop route to be considered "full enough", so this step
    runs AFTER it as a strict count enforcer.

    Algorithm:
    - While max_len - min_len > max_imbalance:
        - From the longest route pick the stop with minimum (insertion_cost
          into shortest route − removal_gain from longest route) — i.e. the
          geographically cheapest move.
        - Apply the move; re-run 2-opt on both affected routes.
    - Safety: never reduces a route below 1 stop; bounded by total_stops moves.

    This does NOT try to minimise km — balance takes priority.  The km impact
    is typically small (< 2%) because we pick the geographically cheapest stop.
    """
    if len(routes) <= 1:
        return routes

    def route_cost(route):
        if not route:
            return 0
        cost = full_matrix[0][route[0]] + full_matrix[route[-1]][0]
        for a, b in zip(route, route[1:]):
            cost += full_matrix[a][b]
        return cost

    total_stops = sum(len(r) for r in routes)
    max_moves = total_stops  # hard safety bound — can't move more stops than exist

    for _move in range(max_moves):
        lens = [len(r) for r in routes]
        max_len = max(lens)
        min_len = min(lens)
        if max_len - min_len <= max_imbalance:
            break

        donor_idx = max(range(len(routes)), key=lambda i: lens[i])
        recv_idx = min(range(len(routes)), key=lambda i: lens[i])
        donor = routes[donor_idx]
        recv = routes[recv_idx]

        if len(donor) <= 1:
            break  # safety — cannot donate last stop

        # Find the stop in donor whose move to recv is cheapest (min net_cost)
        best_stop_pos = -1
        best_insert_pos = -1
        best_net_cost = float("inf")
        base_recv = route_cost(recv)
        base_donor = route_cost(donor)

        for k, stop in enumerate(donor):
            donor_without = donor[:k] + donor[k + 1:]
            removal_gain = base_donor - route_cost(donor_without)
            for pos in range(len(recv) + 1):
                recv_with = recv[:pos] + [stop] + recv[pos:]
                insertion_cost = route_cost(recv_with) - base_recv
                net_cost = insertion_cost - removal_gain
                if net_cost < best_net_cost:
                    best_net_cost = net_cost
                    best_stop_pos = k
                    best_insert_pos = pos

        if best_stop_pos < 0:
            break  # no valid move found

        stop_val = donor[best_stop_pos]
        new_donor = donor[:best_stop_pos] + donor[best_stop_pos + 1:]
        new_recv = recv[:best_insert_pos] + [stop_val] + recv[best_insert_pos:]
        routes[donor_idx] = new_donor
        routes[recv_idx] = new_recv
        logger.info(
            "rebalance_count_balance: stop %d: route %d(%d stops) → route %d(%d stops), "
            "net_cost %+.0f m, imbalance was %d",
            stop_val, donor_idx, max_len, recv_idx, min_len, best_net_cost,
            max_len - min_len,
        )

    return [r for r in routes if r]


def _rebalance_max_stops(routes: list, full_matrix: list, max_stops: int,
                         demands_kg=None, demands_m3=None,
                         capacities_kg=None, capacities_m3=None) -> tuple:
    """
    Cap overloaded routes by moving excess stops to less-loaded routes.

    Capacity guard: a stop is only moved to a destination that has room in
    BOTH kg and m³ dimensions (when constraints are provided).

    Returns: (rebalanced_routes, moves_count)
    """
    if len(routes) <= 1 or max_stops is None:
        return routes, 0

    def route_cost(route):
        if not route:
            return 0
        cost = full_matrix[0][route[0]] + full_matrix[route[-1]][0]
        for a, b in zip(route, route[1:]):
            cost += full_matrix[a][b]
        return cost

    total_moves = 0
    changed = True
    while changed:
        changed = False
        oversized = [i for i, r in enumerate(routes) if len(r) > max_stops]
        if not oversized:
            break
        for i in oversized:
            if len(routes[i]) <= max_stops:
                continue
            best_net = float("inf")
            best_stop_val = None
            best_stop_pos = -1
            best_dest_idx = -1
            best_insert_pos = -1

            for si, stop in enumerate(routes[i]):
                route_without = routes[i][:si] + routes[i][si + 1:]
                removal_gain = route_cost(routes[i]) - route_cost(route_without)
                for j, dest in enumerate(routes):
                    if i == j:
                        continue
                    # Capacity guard: dest must have room in both dimensions
                    if not _can_route_accept(dest, stop,
                                            demands_kg, demands_m3,
                                            capacities_kg, capacities_m3,
                                            v_idx=j):
                        continue
                    base_dest = route_cost(dest)
                    for pos in range(len(dest) + 1):
                        dest_with = dest[:pos] + [stop] + dest[pos:]
                        insertion_cost = route_cost(dest_with) - base_dest
                        net = insertion_cost - removal_gain
                        if net < best_net:
                            best_net = net
                            best_stop_val = stop
                            best_stop_pos = si
                            best_dest_idx = j
                            best_insert_pos = pos

            if best_stop_val is not None:
                routes[i] = routes[i][:best_stop_pos] + routes[i][best_stop_pos + 1:]
                routes[best_dest_idx] = (
                    routes[best_dest_idx][:best_insert_pos]
                    + [best_stop_val]
                    + routes[best_dest_idx][best_insert_pos:]
                )
                logger.info(
                    "rebalance_max_stops: moved stop %d from route %d→%d "
                    "(route %d now %d stops, net_cost %+.0f m)",
                    best_stop_val, i, best_dest_idx, i, len(routes[i]), best_net,
                )
                total_moves += 1
                changed = True
                break

    return [r for r in routes if r], total_moves


def _db_connect_kwargs() -> dict:
    """Parse DATABASE_URL into psycopg2 connect keyword arguments."""
    url = DATABASE_URL.strip()
    if url.startswith("postgres://") or url.startswith("postgresql://"):
        parsed = urllib.parse.urlparse(url)
        return dict(
            host=parsed.hostname,
            port=parsed.port or 5432,
            dbname=(parsed.path or "/").lstrip("/"),
            user=parsed.username,
            password=parsed.password,
        )
    return {"dsn": url}


# ── Connection pool ───────────────────────────────────────────────────────────
# Lazily initialised on first get_db() call so startup errors are surfaced
# through FastAPI's normal exception handling rather than at import time.
_db_pool: Optional[_psycopg2_pool.ThreadedConnectionPool] = None
_db_pool_lock = threading.Lock()


def _get_pool() -> _psycopg2_pool.ThreadedConnectionPool:
    """Return the shared connection pool, creating it on first call (thread-safe)."""
    global _db_pool
    if _db_pool is None:
        with _db_pool_lock:
            if _db_pool is None:
                _db_pool = _psycopg2_pool.ThreadedConnectionPool(
                    minconn=2, maxconn=15, **_db_connect_kwargs()
                )
    return _db_pool


class _PooledConn:
    """Thin wrapper around a psycopg2 connection borrowed from the pool.

    Intercepts .close() to return the connection to the pool instead of
    destroying it, so ALL existing call-sites work without modification.
    Rolls back any open transaction before returning so the next borrower
    always receives a clean connection.
    """
    __slots__ = ("_conn", "_pool", "_closed")

    def __init__(self, conn, pool: _psycopg2_pool.ThreadedConnectionPool):
        object.__setattr__(self, "_conn", conn)
        object.__setattr__(self, "_pool", pool)
        object.__setattr__(self, "_closed", False)

    # Pass attribute reads through to the underlying connection.
    def __getattr__(self, name):
        return getattr(object.__getattribute__(self, "_conn"), name)

    # Pass attribute writes through (e.g. conn.autocommit = False).
    def __setattr__(self, name, value):
        if name in _PooledConn.__slots__:
            object.__setattr__(self, name, value)
        else:
            setattr(object.__getattribute__(self, "_conn"), name, value)

    def close(self):
        """Return the connection to the pool (idempotent)."""
        if object.__getattribute__(self, "_closed"):
            return
        object.__setattr__(self, "_closed", True)
        conn = object.__getattribute__(self, "_conn")
        pool = object.__getattribute__(self, "_pool")
        try:
            # Roll back any open transaction so the next borrower gets a
            # clean connection regardless of what happened in this request.
            if not conn.closed:
                try:
                    conn.rollback()
                except Exception:
                    pass
            pool.putconn(conn)
        except Exception:
            # Last resort: destroy the connection rather than leak it.
            try:
                conn.close()
            except Exception:
                pass


def get_db() -> _PooledConn:
    """Borrow a connection from the pool.

    Supports both URL and key=value DSN formats.
    Falls back to a direct connection if the pool is temporarily exhausted,
    so the server degrades gracefully rather than returning 500.
    """
    try:
        pool = _get_pool()
        raw = pool.getconn()
        raw.autocommit = False
        return _PooledConn(raw, pool)
    except _psycopg2_pool.PoolError:
        # Pool exhausted — open a direct connection as a safety valve.
        logging.warning("DB pool exhausted — opening a direct connection")
        kwargs = _db_connect_kwargs()
        if "dsn" in kwargs:
            conn = psycopg2.connect(kwargs["dsn"])
        else:
            conn = psycopg2.connect(**kwargs)
        conn.autocommit = False
        return conn  # type: ignore[return-value]


def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS stores (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            address TEXT NOT NULL,
            lat DOUBLE PRECISION,
            lon DOUBLE PRECISION,
            map_url TEXT,
            geocode_status TEXT DEFAULT 'pending',
            time_window_from TEXT DEFAULT '09:00',
            time_window_to TEXT DEFAULT '18:00',
            unload_minutes INTEGER DEFAULT 15,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS route_sessions (
            id SERIAL PRIMARY KEY,
            date TEXT,
            num_vehicles INTEGER,
            total_km DOUBLE PRECISION,
            saved_km DOUBLE PRECISION,
            saved_rub INTEGER,
            num_points INTEGER DEFAULT 0,
            result_json TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("""
        ALTER TABLE route_sessions ADD COLUMN IF NOT EXISTS result_json TEXT
    """)
    cur.execute("""
        ALTER TABLE stores ADD COLUMN IF NOT EXISTS map_url TEXT
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS route_session_stores (
            id SERIAL PRIMARY KEY,
            session_id INTEGER REFERENCES route_sessions(id) ON DELETE CASCADE,
            store_id INTEGER REFERENCES stores(id) ON DELETE SET NULL,
            visit_order INTEGER
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # Performance indexes
    cur.execute("CREATE INDEX IF NOT EXISTS idx_stores_geocode ON stores(geocode_status)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_sessions_date ON route_sessions(date DESC)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_session_stores_session ON route_session_stores(session_id)")
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_users_username ON users(username)")
    # Company settings (single-row table for cost model params)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS company_settings (
            id SERIAL PRIMARY KEY,
            fuel_price DOUBLE PRECISION DEFAULT 67.0,
            fuel_consumption DOUBLE PRECISION DEFAULT 13.0,
            driver_salary DOUBLE PRECISION DEFAULT 0.0,
            cost_per_km DOUBLE PRECISION DEFAULT 8.71,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # Seed defaults if table is empty
    # driver_salary column kept for schema compatibility but no longer used in formula
    cur.execute("""
        INSERT INTO company_settings (fuel_price, fuel_consumption, cost_per_km)
        SELECT 67.0, 13.0, ROUND(CAST(67.0 * 13.0 / 100.0 AS numeric), 2)
        WHERE NOT EXISTS (SELECT 1 FROM company_settings)
    """)
    # Migration: recalculate cost_per_km for existing rows using fuel-only formula
    cur.execute("""
        UPDATE company_settings
           SET cost_per_km = ROUND(CAST(fuel_price * fuel_consumption / 100.0 AS numeric), 2)
         WHERE cost_per_km > fuel_price * fuel_consumption / 100.0 + 1
    """)
    # Historical cost_per_km per route session
    cur.execute("""
        ALTER TABLE route_sessions ADD COLUMN IF NOT EXISTS cost_per_km DOUBLE PRECISION
    """)
    # ── Multi-user isolation: owner_id columns ────────────────────────────────
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS is_active BOOLEAN DEFAULT TRUE")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS last_login_at TIMESTAMP")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS is_admin BOOLEAN DEFAULT FALSE")
    cur.execute("ALTER TABLE stores ADD COLUMN IF NOT EXISTS owner_id INTEGER REFERENCES users(id)")
    cur.execute("ALTER TABLE stores ADD COLUMN IF NOT EXISTS city TEXT DEFAULT ''")
    cur.execute("ALTER TABLE stores ADD COLUMN IF NOT EXISTS phone TEXT DEFAULT ''")
    cur.execute("ALTER TABLE stores ADD COLUMN IF NOT EXISTS client TEXT DEFAULT ''")
    cur.execute("ALTER TABLE stores ADD COLUMN IF NOT EXISTS external_id TEXT DEFAULT ''")
    cur.execute("ALTER TABLE stores ADD COLUMN IF NOT EXISTS source TEXT DEFAULT 'manual'")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_stores_ext_owner ON stores(owner_id, external_id) WHERE external_id != ''")
    # ── Import history ─────────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS order_import_history (
            id SERIAL PRIMARY KEY,
            owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            delivery_date DATE NOT NULL,
            filename TEXT DEFAULT '',
            total_rows INTEGER DEFAULT 0,
            matched_rows INTEGER DEFAULT 0,
            unmatched_rows INTEGER DEFAULT 0,
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_import_history_owner ON order_import_history(owner_id, delivery_date DESC)")
    cur.execute("ALTER TABLE order_import_history ADD COLUMN IF NOT EXISTS has_weight BOOLEAN DEFAULT TRUE")
    cur.execute("ALTER TABLE order_import_history ADD COLUMN IF NOT EXISTS total_weight_kg DOUBLE PRECISION DEFAULT 0")
    cur.execute("ALTER TABLE order_import_history ADD COLUMN IF NOT EXISTS total_volume_m3 DOUBLE PRECISION DEFAULT 0")
    cur.execute("ALTER TABLE order_import_history ADD COLUMN IF NOT EXISTS total_amount_rub DOUBLE PRECISION DEFAULT 0")
    cur.execute("ALTER TABLE route_sessions ADD COLUMN IF NOT EXISTS owner_id INTEGER REFERENCES users(id)")
    cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS owner_id INTEGER REFERENCES users(id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_stores_owner ON stores(owner_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_sessions_owner ON route_sessions(owner_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_settings_owner ON company_settings(owner_id)")
    # Driver directory. Assignments keep a phone snapshot so historical
    # WhatsApp links remain understandable after a directory edit.
    cur.execute("""
        CREATE TABLE IF NOT EXISTS drivers (
            id SERIAL PRIMARY KEY,
            owner_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            name TEXT NOT NULL,
            phone TEXT NOT NULL,
            vehicle_name TEXT DEFAULT '',
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_drivers_owner_active ON drivers(owner_id, is_active, name)")
    # Operational execution layer. Routing history stays immutable in
    # route_sessions; assignments and executions track the actual delivery.
    cur.execute("""
        CREATE TABLE IF NOT EXISTS route_assignments (
            id SERIAL PRIMARY KEY,
            owner_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            session_id INTEGER NOT NULL REFERENCES route_sessions(id) ON DELETE CASCADE,
            route_index INTEGER NOT NULL,
            driver_id INTEGER REFERENCES drivers(id) ON DELETE SET NULL,
            driver_name TEXT DEFAULT '',
            driver_phone TEXT DEFAULT '',
            vehicle_name TEXT DEFAULT '',
            route_yandex_url TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'planned',
            access_token_hash TEXT NOT NULL UNIQUE,
            token_created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            expires_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(session_id, route_index)
        )
    """)
    cur.execute("ALTER TABLE route_assignments ADD COLUMN IF NOT EXISTS driver_id INTEGER REFERENCES drivers(id) ON DELETE SET NULL")
    cur.execute("ALTER TABLE route_assignments ADD COLUMN IF NOT EXISTS driver_phone TEXT DEFAULT ''")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS route_executions (
            id SERIAL PRIMARY KEY,
            assignment_id INTEGER NOT NULL REFERENCES route_assignments(id) ON DELETE CASCADE,
            store_id INTEGER REFERENCES stores(id) ON DELETE SET NULL,
            visit_order INTEGER NOT NULL,
            store_name TEXT DEFAULT '',
            address TEXT DEFAULT '',
            lat DOUBLE PRECISION,
            lon DOUBLE PRECISION,
            products TEXT DEFAULT '',
            quantity DOUBLE PRECISION DEFAULT 0,
            actual_qty DOUBLE PRECISION DEFAULT 0,
            arrive_by TEXT DEFAULT '',
            yandex_url TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'planned',
            payment_method TEXT NOT NULL DEFAULT 'none',
            payment_status TEXT NOT NULL DEFAULT 'pending',
            driver_comment TEXT DEFAULT '',
            rescheduled_date DATE,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            delivered_at TIMESTAMP,
            UNIQUE(assignment_id, visit_order)
        )
    """)
    # Additive migrations for installations created by the first MVP draft.
    cur.execute("ALTER TABLE route_assignments ADD COLUMN IF NOT EXISTS route_yandex_url TEXT DEFAULT ''")
    cur.execute("ALTER TABLE route_assignments ADD COLUMN IF NOT EXISTS expires_at TIMESTAMP")
    cur.execute("ALTER TABLE route_executions ADD COLUMN IF NOT EXISTS rescheduled_date DATE")
    for _column, _definition in (
        ("store_name", "TEXT DEFAULT ''"),
        ("address", "TEXT DEFAULT ''"),
        ("lat", "DOUBLE PRECISION"),
        ("lon", "DOUBLE PRECISION"),
        ("products", "TEXT DEFAULT ''"),
        ("quantity", "DOUBLE PRECISION DEFAULT 0"),
        ("actual_qty", "DOUBLE PRECISION"),
        ("arrive_by", "TEXT DEFAULT ''"),
        ("yandex_url", "TEXT DEFAULT ''"),
        ("payment_status", "TEXT DEFAULT 'pending'"),
    ):
        cur.execute(
            f"ALTER TABLE route_executions ADD COLUMN IF NOT EXISTS {_column} {_definition}"
        )
    # Existing MVP executions had no actual quantity. Treat already delivered
    # points as fully delivered and all other historical points as zero.
    cur.execute(
        """UPDATE route_executions
           SET actual_qty = CASE WHEN status='delivered' THEN COALESCE(quantity, 0) ELSE 0 END
           WHERE actual_qty IS NULL"""
    )
    cur.execute("UPDATE route_executions SET payment_status='pending' WHERE payment_status IS NULL")
    cur.execute("ALTER TABLE route_executions ALTER COLUMN actual_qty SET DEFAULT 0")
    cur.execute("ALTER TABLE route_executions ALTER COLUMN actual_qty SET NOT NULL")
    cur.execute(
        """UPDATE route_assignments
           SET expires_at = COALESCE(expires_at, COALESCE(token_created_at, NOW()) + INTERVAL '48 hours')
           WHERE expires_at IS NULL"""
    )
    # Planned points have not been delivered yet. Older MVP rows incorrectly
    # copied the planned quantity into actual_qty, which made the dispatcher
    # show a false delivered amount before the driver started.
    cur.execute(
        """UPDATE route_executions
           SET actual_qty = 0
           WHERE status = 'planned' AND COALESCE(actual_qty, 0) <> 0"""
    )
    cur.execute(
        "UPDATE route_executions SET status='planned' WHERE status IN ('loaded', 'on_route')"
    )
    cur.execute("ALTER TABLE route_executions ALTER COLUMN payment_status SET DEFAULT 'pending'")
    cur.execute("ALTER TABLE route_executions ALTER COLUMN payment_status SET NOT NULL")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_assignments_owner ON route_assignments(owner_id, session_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_executions_assignment ON route_executions(assignment_id, visit_order)")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS driver_locations (
            id SERIAL PRIMARY KEY,
            assignment_id INTEGER NOT NULL REFERENCES route_assignments(id) ON DELETE CASCADE,
            lat DOUBLE PRECISION NOT NULL,
            lon DOUBLE PRECISION NOT NULL,
            accuracy DOUBLE PRECISION,
            captured_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_driver_locations_latest ON driver_locations(assignment_id)")
    # ── User plan & admin notes ────────────────────────────────────────────────
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS plan TEXT DEFAULT 'trial'")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS admin_note TEXT DEFAULT ''")
    # ── Admin audit log ────────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS admin_audit_log (
            id SERIAL PRIMARY KEY,
            admin_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
            target_user_id INTEGER,
            target_username TEXT,
            action TEXT NOT NULL,
            details TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_log_admin ON admin_audit_log(admin_user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_log_created ON admin_audit_log(created_at DESC)")
    # ── API Keys ───────────────────────────────────────────────────────────────
    # Bearer-token keys for machine-to-machine access (integrations, webhooks).
    # The full key is shown once on creation; only SHA-256 hash stored in DB.
    cur.execute("""
        CREATE TABLE IF NOT EXISTS api_keys (
            id SERIAL PRIMARY KEY,
            owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            name TEXT NOT NULL,
            key_prefix TEXT NOT NULL,
            key_hash TEXT NOT NULL UNIQUE,
            scopes TEXT[] NOT NULL DEFAULT '{}',
            is_active BOOLEAN DEFAULT TRUE,
            expires_at TIMESTAMP,
            last_used_at TIMESTAMP,
            last_used_ip TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_api_keys_owner ON api_keys(owner_id, is_active)")
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_api_keys_hash ON api_keys(key_hash)")
    # ── Persistent geocoding cache ─────────────────────────────────────────────
    # Replaces ephemeral in-memory dict; survives restarts, prevents redundant
    # Yandex API calls on re-import. Only successful results (lat/lon != NULL)
    # are stored. Admins can purge individual entries to force re-geocode.
    cur.execute("""
        CREATE TABLE IF NOT EXISTS geocode_cache (
            id SERIAL PRIMARY KEY,
            normalized_address TEXT NOT NULL UNIQUE,
            lat DOUBLE PRECISION NOT NULL,
            lon DOUBLE PRECISION NOT NULL,
            source TEXT DEFAULT 'unknown',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_geocode_cache_addr ON geocode_cache(normalized_address)")
    # ── Daily orders (заявки на день) ─────────────────────────────────────────
    # Stores per-store delivery quantities imported from Excel (1С / any source).
    # Rows are scoped by owner and delivery_date. weight_kg / volume_m3 drive VRP
    # capacity balancing when vehicle capacity_kg is set.
    cur.execute("""
        CREATE TABLE IF NOT EXISTS daily_orders (
            id SERIAL PRIMARY KEY,
            owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            store_id INTEGER REFERENCES stores(id) ON DELETE SET NULL,
            store_name_raw TEXT NOT NULL,
            order_number TEXT DEFAULT '',
            weight_kg DOUBLE PRECISION DEFAULT 0,
            volume_m3 DOUBLE PRECISION DEFAULT 0,
            amount_rub DOUBLE PRECISION DEFAULT 0,
            notes TEXT DEFAULT '',
            delivery_date DATE NOT NULL DEFAULT CURRENT_DATE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_daily_orders_owner_date ON daily_orders(owner_id, delivery_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_daily_orders_store ON daily_orders(store_id)")
    # Extended order fields (1C-style multi-row imports): raw address keeps the
    # delivery point identity even when unmatched; quantity + products are the
    # aggregated cargo breakdown shown to the driver (display-only).
    cur.execute("ALTER TABLE daily_orders ADD COLUMN IF NOT EXISTS address_raw TEXT DEFAULT ''")
    cur.execute("ALTER TABLE daily_orders ADD COLUMN IF NOT EXISTS quantity DOUBLE PRECISION DEFAULT 0")
    cur.execute("ALTER TABLE daily_orders ADD COLUMN IF NOT EXISTS products TEXT DEFAULT ''")
    # ── Integrations ───────────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS integrations (
            id SERIAL PRIMARY KEY,
            owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            type TEXT NOT NULL DEFAULT '1c',
            name TEXT NOT NULL DEFAULT '1C Интеграция',
            status TEXT DEFAULT 'setup',
            config JSONB DEFAULT '{}',
            last_sync_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_integrations_owner ON integrations(owner_id)")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS integration_sync_logs (
            id SERIAL PRIMARY KEY,
            integration_id INTEGER REFERENCES integrations(id) ON DELETE CASCADE,
            started_at TIMESTAMP DEFAULT NOW(),
            finished_at TIMESTAMP,
            duration_ms INTEGER DEFAULT 0,
            status TEXT DEFAULT 'success',
            orders_received INTEGER DEFAULT 0,
            stores_matched INTEGER DEFAULT 0,
            stores_unmatched INTEGER DEFAULT 0,
            errors_count INTEGER DEFAULT 0,
            error_detail TEXT DEFAULT '',
            meta JSONB DEFAULT '{}'
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_sync_logs_integration ON integration_sync_logs(integration_id, started_at DESC)")
    conn.commit()
    cur.close()
    conn.close()


def get_company_settings(user_id: int = None) -> dict:
    """Read cost model settings from DB (per-user row). Returns defaults if no row exists."""
    _defaults = {
        "fuel_price": 67.0,
        "fuel_consumption": 13.0,
        "cost_per_km": round(67.0 * 13.0 / 100.0, 2),
    }
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if user_id is not None:
            cur.execute(
                "SELECT fuel_price, fuel_consumption, cost_per_km FROM company_settings WHERE owner_id = %s LIMIT 1",
                (user_id,)
            )
        else:
            cur.execute("SELECT fuel_price, fuel_consumption, cost_per_km FROM company_settings ORDER BY id LIMIT 1")
        row = cur.fetchone()
        cur.close()
        conn.close()
        if row:
            return dict(row)
    except Exception as e:
        logger.warning("get_company_settings error: %s", e)
    return _defaults


def haversine_meters(c1: tuple, c2: tuple) -> int:
    R = 6371000
    lat1, lon1 = math.radians(c1[0]), math.radians(c1[1])
    lat2, lon2 = math.radians(c2[0]), math.radians(c2[1])
    dlat, dlon = lat2 - lat1, lon2 - lon1
    a = math.sin(dlat/2)**2 + math.cos(lat1)*math.cos(lat2)*math.sin(dlon/2)**2
    return int(R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a)))


def solve_vrp(all_coords: list, num_vehicles: int, capacities=None, demands=None,
              store_time_windows: list = None,
              max_stops_per_vehicle: int = None,
              optimize_by: str = "distance",
              capacities_m3=None, demands_m3=None) -> list:
    """
    Solve VRP with efficiency as the primary objective (min total km / time).

    Strategy
    ────────
    1. Build a full NxN Haversine distance matrix once (instant, no API).
       For small datasets (≤ GH_FREE_LIMIT) attempt GraphHopper first.

    2. Partition stores into `num_vehicles` contiguous geographic SECTORS using
       the polar-angle sweep around the depot.  Contiguous sectors (not
       round-robin) mean each vehicle is assigned a geographic wedge, so
       denser areas naturally get more stops and sparse areas fewer stops.
       This is the principal mechanism for unequal-but-efficient distribution.

    3. For each sector, extract the relevant sub-matrix and run a single-vehicle
       OR-Tools TSP to polish the visit order within the sector.
       If store_time_windows is provided, OR-Tools also enforces arrival time
       windows (TSPTW) so deliveries respect opening-hour constraints.

    Why this approach
    ─────────────────
    • All vehicles are used (each non-empty sector → one vehicle).
    • Stop counts per vehicle reflect geographic density, NOT a count target.
      A dense north cluster → 12 stops; a sparse south sector → 5 stops.
    • No GlobalSpanCostCoefficient — no artificial balancing penalty.
    • Fast: single-vehicle TSPs are much cheaper than full multi-vehicle VRP.

    Args:
        all_coords:          list of (lat, lon); index 0 is the depot.
        num_vehicles:        number of vehicles / routes to generate.
        capacities:          optional list of vehicle capacity (kg).
        demands:             optional list of store demands (index 0 = depot = 0).
        store_time_windows:  optional list of (tw_from_min, tw_to_min, service_min)
                             for each store, 0-indexed matching all_coords[1:].
                             When supplied, OR-Tools TSPTW is used per cluster.

    Fallback chain (transparent to caller):
      OR-Tools per-sector  →  sector order as-is  →  greedy round-robin
    """
    n = len(all_coords)
    store_count = n - 1  # node 0 is depot

    if store_count == 0:
        return [], "haversine"

    # ── No OR-Tools: geographic sectors, then greedy ordering ─────────────────
    if not ORTOOLS_AVAILABLE:
        if store_count >= 1:
            all_store_nodes = list(range(1, n))
            _use_cap_sweep_fb = (
                (capacities is not None and demands is not None
                 and any(c < 99999 for c in capacities))
                or (capacities_m3 is not None)
            )
            if _use_cap_sweep_fb:
                _cap_kg_fb = capacities if capacities is not None else [99999] * num_vehicles
                _dem_kg_fb = demands if demands is not None else [0] + [1] * store_count
                clusters = _cluster_by_weight_sweep(
                    all_store_nodes, all_coords, num_vehicles, _cap_kg_fb, _dem_kg_fb,
                    capacities_m3=capacities_m3, demands_m3=demands_m3)
            else:
                clusters = _cluster_by_sweep(all_store_nodes, all_coords, num_vehicles)
            return [c for c in clusters if c], "haversine"
        return _fallback_distribution(list(range(1, n)), num_vehicles), "haversine"

    # ── Step 1: full Haversine matrix (always-available baseline) ─────────────
    # Used as fallback when GraphHopper is unavailable for a given cluster.
    full_matrix = _build_haversine_matrix(all_coords)

    logger.info(
        "solve_vrp: %d stores / %d vehicles — building clusters",
        store_count, num_vehicles,
    )

    # ── Step 2: geographic sector partition (contiguous sweep) ────────────────
    # When weight demands and vehicle capacities are provided, use the
    # capacity-aware sweep that fills vehicles respecting their weight limits.
    # Otherwise fall back to the classic equal-angle geographic sweep.
    all_store_nodes = list(range(1, n))
    _use_cap_sweep = (
        (capacities is not None and demands is not None
         and any(c < 99999 for c in capacities))
        or (capacities_m3 is not None)
    )
    if _use_cap_sweep:
        _cap_kg = capacities if capacities is not None else [99999] * num_vehicles
        _dem_kg = demands if demands is not None else [0] + [1] * store_count
        logger.info(
            "solve_vrp: dual-capacity sweep clustering (kg_caps=%s, m3_caps=%s)",
            _cap_kg, capacities_m3,
        )
        clusters = _cluster_by_weight_sweep(
            all_store_nodes, all_coords, num_vehicles, _cap_kg, _dem_kg,
            capacities_m3=capacities_m3, demands_m3=demands_m3)
    else:
        clusters = _cluster_by_sweep(all_store_nodes, all_coords, num_vehicles)

    # ── Step 3: per-cluster road matrix → OR-Tools TSP ────────────────────────
    # Routing priority per cluster:
    #   GH   (paid plan, precise road distances)     → if GH key set + fits plan limit
    #   OSRM (free, real OSM roads, no key needed)   → public demo server fallback
    #   Haversine (instant math, always available)   → final fallback
    #
    # Matrix fetching is parallelised with ThreadPoolExecutor so that network-
    # bound GH/OSRM requests for different clusters overlap.  OR-Tools TSP
    # solving is run sequentially afterwards (CPU-bound, GIL-limited).
    t_vrp_start = time.time()

    non_empty_clusters = [c for c in clusters if c]
    # Phase A: fetch matrices for all clusters in parallel -----------------------
    def _fetch_matrix(cluster_nodes):
        """Return (cluster_nodes, sub_dist_matrix, sub_time_matrix, source) for one cluster.

        sub_time_matrix is the real travel-time matrix (seconds) from GH/OSRM,
        or None when only Haversine is available.  The caller decides whether to
        use it based on the optimize_by flag.
        """
        if len(cluster_nodes) == 1:
            return (cluster_nodes, None, None, "hv_single")
        group_indices = [0] + cluster_nodes
        group_coords = [all_coords[i] for i in group_indices]
        gh_result = get_cluster_matrix_gh(group_coords)
        if gh_result:
            sub_matrix, sub_time = gh_result
            return (cluster_nodes, sub_matrix, sub_time, "gh")
        osrm_result = get_cluster_matrix_osrm(group_coords)
        if osrm_result:
            sub_matrix, sub_time = osrm_result
            return (cluster_nodes, sub_matrix, sub_time, "osrm")
        group_indices2 = [0] + cluster_nodes
        sub_matrix = [[full_matrix[r][c] for c in group_indices2] for r in group_indices2]
        return (cluster_nodes, sub_matrix, None, "hv")

    max_workers = min(len(non_empty_clusters), 8)
    with concurrent.futures.ThreadPoolExecutor(max_workers=max(max_workers, 1)) as pool:
        matrix_futures = [pool.submit(_fetch_matrix, c) for c in non_empty_clusters]
        matrix_results = [f.result() for f in matrix_futures]

    # Phase B: OR-Tools TSP per cluster (sequential; CPU-bound) -----------------
    # Global 60 s OR-Tools budget spread equally across clusters so very large
    # datasets (300+ stops) stay within a predictable wall-clock time.
    # Each cluster gets at most (60 / num_clusters) seconds, further bounded by
    # the per-size adaptive limits computed inside _ortools_solve_group.
    GLOBAL_ORTOOLS_BUDGET_SECONDS = 60.0
    num_clusters = len([r for r in matrix_results if r[1] is not None])
    per_cluster_budget = (GLOBAL_ORTOOLS_BUDGET_SECONDS / max(num_clusters, 1))

    routes = []
    gh_clusters = 0
    osrm_clusters = 0
    hv_clusters = 0

    for cluster_nodes, sub_matrix, sub_time, source in matrix_results:
        if source == "hv_single":
            routes.append(cluster_nodes)
            hv_clusters += 1
            continue
        if source == "gh":
            gh_clusters += 1
        elif source == "osrm":
            osrm_clusters += 1
        else:
            hv_clusters += 1

        group_indices = [0] + cluster_nodes
        group_coords = [all_coords[i] for i in group_indices]
        cluster_tw = None
        if store_time_windows:
            try:
                cluster_tw = [store_time_windows[node - 1] for node in cluster_nodes]
            except IndexError:
                cluster_tw = None

        try:
            ordered = _ortools_solve_group(all_coords[0], cluster_nodes, group_coords, sub_matrix,
                                           time_windows=cluster_tw,
                                           time_limit_override=per_cluster_budget,
                                           time_matrix=sub_time,
                                           optimize_by=optimize_by)
        except Exception as cluster_exc:
            logger.warning(
                "solve_vrp: _ortools_solve_group raised for cluster of %d stops (%s) "
                "— keeping original sweep order",
                len(cluster_nodes), cluster_exc,
            )
            ordered = cluster_nodes

        # ── 2-opt polish within this route ────────────────────────────────────
        # Removes crossing edges and linear-chain artefacts that OR-Tools
        # occasionally leaves, using the same sub_matrix (real road distances
        # when OSRM/GH was available, otherwise Haversine).
        if sub_matrix is not None and len(ordered) >= 3:
            # Map global node indices back to local sub_matrix indices
            # group_indices = [0] + cluster_nodes → local[k] = group_indices[k]
            group_indices = [0] + cluster_nodes
            local_idx = {gn: li for li, gn in enumerate(group_indices)}
            local_route = [local_idx[gn] for gn in ordered]
            local_route = _two_opt_route(local_route, sub_matrix)
            ordered = [group_indices[li] for li in local_route]

        routes.append(ordered if ordered else cluster_nodes)

    # Determine reported matrix source
    total_clusters = gh_clusters + osrm_clusters + hv_clusters
    if osrm_clusters == total_clusters:
        matrix_source = "osrm"
    elif gh_clusters == total_clusters:
        matrix_source = "graphhopper"
    elif hv_clusters == total_clusters:
        matrix_source = "haversine"
    else:
        parts = []
        if gh_clusters:   parts.append(f"gh={gh_clusters}")
        if osrm_clusters: parts.append(f"osrm={osrm_clusters}")
        if hv_clusters:   parts.append(f"hv={hv_clusters}")
        matrix_source = "mixed (" + ", ".join(parts) + ")"

    logger.info(
        "solve_vrp clusters: total=%d, graphhopper=%d, osrm=%d, haversine=%d, "
        "elapsed=%.1fs, cache_hits=%d, cache_total=%d",
        total_clusters, gh_clusters, osrm_clusters, hv_clusters,
        time.time() - t_vrp_start,
        _matrix_cache_hits + _osrm_cache_hits, len(_matrix_cache),
    )

    if not routes:
        routes = _fallback_distribution(all_store_nodes, num_vehicles)

    # ── Step 4: fill unused vehicles by splitting oversized routes ────────────
    # If equal-angle sectors left some vehicles idle (empty angular wedge),
    # break the largest route in half by polar angle until all vehicles are used.
    while len(routes) < num_vehicles and any(len(r) > 1 for r in routes):
        largest_idx = max(range(len(routes)), key=lambda i: len(routes[i]))
        largest = routes[largest_idx]
        if len(largest) < 2:
            break
        # Sort by angle and split at midpoint — each half stays geographic
        depot = all_coords[0]
        def _angle(node):
            lat, lon = all_coords[node]
            return math.atan2(lon - depot[1], lat - depot[0])
        sorted_half = sorted(largest, key=_angle)
        mid = len(sorted_half) // 2
        routes[largest_idx] = sorted_half[:mid]
        routes.append(sorted_half[mid:])

    # Compute effective min-stops floor.
    # For large datasets (avg > MIN_STOPS_PER_VEHICLE) use 70% of avg so no
    # route drops below ≈70% of the mean — this caps ratio max/min at ~1.7x
    # while still allowing inter-route optimisation to fix misplaced stops.
    # For small datasets, fall back to avg-1 so no vehicle is left empty.
    avg_stops = store_count // max(len(routes), 1)
    if avg_stops <= MIN_STOPS_PER_VEHICLE:
        effective_min = max(1, avg_stops - 1)
    else:
        effective_min = max(MIN_STOPS_PER_VEHICLE, int(avg_stops * 0.70))

    # ── Step 5: inter-route relocate post-processing ──────────────────────────
    # Move individual stops between routes if it reduces total Haversine distance.
    # This equalises quality across all vehicles: TSP sectors are optimised in
    # isolation, so route 1 can end up much shorter than route 2.  Relocate fixes
    # this by shifting misplaced stops to the route where they cost the least.
    # The min_stops guard prevents Or-opt from draining routes below the floor.
    #
    # REMOVED: the old `store_count <= 80` gate that disabled this step entirely
    # for larger datasets — that was the root cause of quality degradation at
    # 100+ stops.  Instead we use adaptive iteration counts so that larger
    # problems still finish quickly (O(iters × stops × routes × cluster_size)).
    if len(routes) > 1:
        if store_count <= 80:
            relocate_iters = 5
        elif store_count <= 150:
            relocate_iters = 3
        elif store_count <= 300:
            relocate_iters = 2
        else:
            relocate_iters = 1

        before_km = sum(
            sum(full_matrix[r[k]][r[k + 1]] for k in range(len(r) - 1))
            + (full_matrix[0][r[0]] + full_matrix[r[-1]][0] if r else 0)
            for r in routes
        ) / 1000
        routes = _inter_route_relocate(routes, full_matrix,
                                       max_iter=relocate_iters,
                                       min_stops=effective_min,
                                       demands_kg=demands, demands_m3=demands_m3,
                                       capacities_kg=capacities, capacities_m3=capacities_m3)
        after_km = sum(
            sum(full_matrix[r[k]][r[k + 1]] for k in range(len(r) - 1))
            + (full_matrix[0][r[0]] + full_matrix[r[-1]][0] if r else 0)
            for r in routes
        ) / 1000
        logger.info(
            "inter_route_relocate: %.1f km → %.1f km (saved %.1f km, iters=%d, stores=%d)",
            before_km, after_km, before_km - after_km, relocate_iters, store_count,
        )

        # ── Step 5b: 2-opt re-polish after relocate ───────────────────────────
        # inter_route_relocate inserts stops at their best position but does NOT
        # re-run 2-opt after insertion, leaving crossing edges in the routes.
        # Confirmed by real data (Session 49): Route 9 retains 2.061 km (7.6%)
        # of 2-opt improvability after relocate; all routes combined: 3.038 km.
        # Running 2-opt here on full_matrix (Haversine) removes those crossings.
        # This is the primary cause of the 14-min gap vs Yandex on 37-stop routes.
        routes = [
            _two_opt_route(r, full_matrix) if len(r) >= 3 else r
            for r in routes
        ]
        after_2opt_km = sum(
            sum(full_matrix[r[k]][r[k + 1]] for k in range(len(r) - 1))
            + (full_matrix[0][r[0]] + full_matrix[r[-1]][0] if r else 0)
            for r in routes
        ) / 1000
        logger.info(
            "post-relocate 2-opt: %.1f km → %.1f km (saved %.1f km, stores=%d)",
            after_km, after_2opt_km, after_km - after_2opt_km, store_count,
        )

    # ── Step 5c: count-balance enforcement ───────────────────────────────────
    # After Or-opt + 2-opt the sector sweep can still leave a severe skew (e.g.
    # 5 vs 12 for 2 vehicles and 17 stores) because _inter_route_relocate only
    # moves stops that reduce total km — cross-sector transfers are often not
    # km-beneficial even when the stop count is badly unequal.
    # _rebalance_count_balance enforces max_diff ≤ 2 regardless of km impact,
    # choosing the geographically cheapest stop to transfer each time.
    # 2-opt re-polish runs afterwards to remove any crossing edges introduced.
    if len(routes) > 1:
        lens_before = sorted([len(r) for r in routes], reverse=True)
        routes = _rebalance_count_balance(routes, full_matrix, max_imbalance=2)
        lens_after = sorted([len(r) for r in routes], reverse=True)
        if lens_before != lens_after:
            routes = [
                _two_opt_route(r, full_matrix) if len(r) >= 3 else r
                for r in routes
            ]
            logger.info(
                "rebalance_count_balance: distribution %s → %s",
                lens_before, lens_after,
            )

    # ── Step 5d: capacity constraint enforcement ──────────────────────────────
    # After Or-opt + count-balance, stop-count rebalancers may have moved stores
    # across weight boundaries.  Run _enforce_capacity to move excess load from
    # over-capacity vehicles to those with remaining headroom.
    # Only active when real capacities were provided (not 99999 placeholders).
    if _use_cap_sweep and len(routes) > 1:
        _enf_dem = demands if demands is not None else ([0] + [1] * store_count)
        _enf_cap = capacities if capacities is not None else [99999] * num_vehicles
        routes = _enforce_capacity(routes, _enf_dem, _enf_cap, full_matrix,
                                   demands_m3=demands_m3, capacities_m3=capacities_m3)

    # ── Step 6: rebalance to minimum stops per vehicle ────────────────────────
    # After sector sweep + Or-opt some vehicles may still be underfull (< effective_min
    # stops).  Steal the cheapest stop (minimum distance penalty) from any donor
    # route that has more than effective_min stops and insert it optimally.
    # This step runs always (also when store_count > 80, where Or-opt is skipped).
    if len(routes) > 1 and effective_min >= 2:
        routes = _rebalance_min_stops(routes, full_matrix, effective_min,
                                      demands_kg=demands, demands_m3=demands_m3,
                                      capacities_kg=capacities, capacities_m3=capacities_m3)

    # ── Step 7: cap overloaded routes (optional) ──────────────────────────────
    # When max_stops_per_vehicle is specified, redistribute excess stops from
    # overloaded routes to less-loaded ones with minimum km penalty.
    # Benchmarked on 120 stores / 9 vehicles (Haversine):
    #   max=24: ratio 3.9x→2.7x (−30.8%), km −0.8km (−0.5%) ✅ passes criteria
    #   max=26: ratio 3.9x→2.9x (−25.6%), km −0.8km (−0.5%) ⚠️  below 30% threshold
    # 2-opt re-polish runs after to clean up any crossing edges introduced.
    if max_stops_per_vehicle is not None and len(routes) > 1:
        before_max_km = sum(
            (full_matrix[0][r[0]] + full_matrix[r[-1]][0]
             + sum(full_matrix[r[k]][r[k + 1]] for k in range(len(r) - 1)))
            for r in routes if r
        ) / 1000
        routes, moves = _rebalance_max_stops(routes, full_matrix, max_stops_per_vehicle,
                                             demands_kg=demands, demands_m3=demands_m3,
                                             capacities_kg=capacities, capacities_m3=capacities_m3)
        routes = [
            _two_opt_route(r, full_matrix) if len(r) >= 3 else r
            for r in routes
        ]
        after_max_km = sum(
            (full_matrix[0][r[0]] + full_matrix[r[-1]][0]
             + sum(full_matrix[r[k]][r[k + 1]] for k in range(len(r) - 1)))
            for r in routes if r
        ) / 1000
        logger.info(
            "rebalance_max_stops(cap=%d): %.1f km → %.1f km (Δ%+.1f km, moves=%d, "
            "max_stops=%d)",
            max_stops_per_vehicle, before_max_km, after_max_km,
            after_max_km - before_max_km, moves,
            max(len(r) for r in routes) if routes else 0,
        )

    return routes, matrix_source


def geocode_address_yandex(address: str) -> Optional[tuple]:
    """
    Geocode using Yandex Geocoder API (primary, no rate-limit delay needed).
    Returns (lat, lon) or None if unavailable / API key not configured.
    """
    if not YANDEX_GEOCODER_API_KEY:
        return None
    try:
        url = (
            "https://geocode-maps.yandex.ru/1.x/?"
            f"geocode={urllib.parse.quote(address)}"
            "&format=json"
            f"&apikey={YANDEX_GEOCODER_API_KEY}"
        )
        req = urllib.request.Request(url, headers={"User-Agent": "smartroute-app-1.0"})
        with urllib.request.urlopen(req, timeout=5) as r:
            data = json.loads(r.read().decode("utf-8"))
        members = data["response"]["GeoObjectCollection"]["featureMember"]
        if members:
            pos = members[0]["GeoObject"]["Point"]["pos"]
            lon_str, lat_str = pos.split()
            result = (float(lat_str), float(lon_str))
            logger.info("geocode_yandex OK: '%s' → (%.5f, %.5f)", address, result[0], result[1])
            return result
    except Exception as e:
        logger.warning("geocode_yandex FAILED for '%s': %s", address, e)
    return None


def geocode_address_nominatim(address: str) -> Optional[tuple]:
    """Geocode using Nominatim (fallback, 1 req/sec limit)."""
    try:
        url = (
            "https://nominatim.openstreetmap.org/search?"
            f"q={urllib.parse.quote(address)}"
            "&format=json&limit=1&accept-language=ru"
        )
        req = urllib.request.Request(url, headers={"User-Agent": "smartroute-app-1.0"})
        with urllib.request.urlopen(req, timeout=5) as r:
            data = json.loads(r.read().decode("utf-8"))
        if data:
            return (float(data[0]["lat"]), float(data[0]["lon"]))
    except Exception as e:
        logger.warning(f"Nominatim geocoding failed for '{address}': {e}")
    return None


def _geocache_db_lookup(cache_key: str) -> Optional[tuple]:
    """Check persistent DB geocoding cache. Returns (lat, lon) or None."""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT lat, lon FROM geocode_cache WHERE normalized_address = %s LIMIT 1", (cache_key,))
        row = cur.fetchone()
        cur.close(); conn.close()
        if row:
            return (float(row["lat"]), float(row["lon"]))
    except Exception as e:
        logger.debug("geocache_db_lookup error: %s", e)
    return None


def _geocache_db_store(cache_key: str, lat: float, lon: float, source: str = "unknown"):
    """Store successful geocoding result in persistent DB cache."""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            """INSERT INTO geocode_cache (normalized_address, lat, lon, source, updated_at)
               VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP)
               ON CONFLICT (normalized_address)
               DO UPDATE SET lat=EXCLUDED.lat, lon=EXCLUDED.lon,
                             source=EXCLUDED.source, updated_at=CURRENT_TIMESTAMP""",
            (cache_key, lat, lon, source)
        )
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        logger.debug("geocache_db_store error: %s", e)


def geocode_address(address: str) -> Optional[tuple]:
    """
    Geocode an address, trying Yandex Geocoder first (fast, no rate limit),
    then falling back to Nominatim.
    Lookup order: in-memory cache → persistent DB cache → Yandex → Nominatim.
    Only successful results are stored (never caches 'not found').
    """
    cache_key = address.strip().lower()

    # ── Level 1: in-memory cache (fastest, ephemeral) ────────────────────────
    if cache_key in geocode_cache:
        return geocode_cache[cache_key]

    # ── Level 2: persistent DB cache (survives restarts) ─────────────────────
    db_hit = _geocache_db_lookup(cache_key)
    if db_hit is not None:
        geocode_cache[cache_key] = db_hit
        return db_hit

    # ── Level 3: Yandex Geocoder ─────────────────────────────────────────────
    result = geocode_address_yandex(address)
    source = "yandex"

    # ── Level 4: Nominatim fallback ───────────────────────────────────────────
    if result is None:
        result = geocode_address_nominatim(address)
        source = "nominatim"

    # Cache successful results in memory; persist to DB only on success.
    # IMPORTANT: do NOT cache None — a transient Nominatim timeout would permanently
    # poison the in-memory cache for this server session, making retries futile.
    if result is not None:
        geocode_cache[cache_key] = result
        _geocache_db_store(cache_key, result[0], result[1], source)

    return result


def find_nearby_stores(lat: float, lon: float, radius_m: float = 20, exclude_id: int = None, owner_id: int = None) -> list:
    """Return stores within radius_m metres of (lat, lon), sorted by distance.
    Uses a degree-based bounding box pre-filter then exact Haversine check.
    If owner_id is provided, only looks within that user's stores."""
    try:
        delta = radius_m / 111320.0
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        owner_clause = "AND owner_id = %s" if owner_id is not None else ""
        if exclude_id is not None:
            params = [lat, lon, exclude_id, lat, delta, lon, delta]
            if owner_id is not None:
                params.append(owner_id)
            cur.execute(
                f"""SELECT id, name, address, lat, lon,
                    SQRT(POWER((lat-%s)*111320.0,2)
                         + POWER((lon-%s)*111320.0*COS(RADIANS(lat)),2)) AS dist_m
                   FROM stores
                   WHERE id != %s AND lat IS NOT NULL
                     AND ABS(lat-%s) < %s AND ABS(lon-%s) < %s
                     {owner_clause}
                   ORDER BY dist_m LIMIT 5""",
                params,
            )
        else:
            params = [lat, lon, lat, delta, lon, delta]
            if owner_id is not None:
                params.append(owner_id)
            cur.execute(
                f"""SELECT id, name, address, lat, lon,
                    SQRT(POWER((lat-%s)*111320.0,2)
                         + POWER((lon-%s)*111320.0*COS(RADIANS(lat)),2)) AS dist_m
                   FROM stores
                   WHERE lat IS NOT NULL
                     AND ABS(lat-%s) < %s AND ABS(lon-%s) < %s
                     {owner_clause}
                   ORDER BY dist_m LIMIT 5""",
                params,
            )
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return [dict(r) for r in rows if r["dist_m"] < radius_m]
    except Exception as e:
        logger.error("find_nearby_stores error: %s", e)
        return []


def parse_yandex_link(url: str) -> tuple:
    """
    Extract (lat, lon) from various Yandex Maps URL formats.
    Returns (None, None) if extraction fails.

    Supported formats:
    - whatshere[point]=lon,lat  (pin dropped by user — most accurate)
    - pt=lon,lat[,icontype]     (point marker — very common share format)
    - ll=lon,lat                (map centre — fallback, less precise)
    - rtext=lat,lon~...         (route first point)
    - /-/ short links (yandex.ru/maps/-/...) and ya.cc/... — follow redirect
    - maps.yandex.ru links — follow redirect
    """
    from urllib.parse import urlparse, parse_qs, unquote
    try:
        decoded = unquote(url)
        parsed = urlparse(decoded)
        params = parse_qs(parsed.query)

        # Format: whatshere[point]=lon,lat  (highest priority — exact pin)
        if "whatshere[point]" in params:
            parts = params["whatshere[point]"][0].split(",")
            if len(parts) >= 2:
                return float(parts[1]), float(parts[0])  # lat, lon

        # Format: pt=lon,lat[,icontype]  (point marker, comma-sep, first point)
        if "pt" in params:
            first_point = params["pt"][0].split("~")[0]  # handle multiple points
            parts = first_point.split(",")
            if len(parts) >= 2:
                return float(parts[1]), float(parts[0])  # lat, lon

        # Format: rtext=lat,lon~lat,lon (route first point — lat,lon order)
        if "rtext" in params:
            parts = params["rtext"][0].split("~")[0].split(",")
            if len(parts) >= 2:
                return float(parts[0]), float(parts[1])  # lat, lon

        # Format: ll=lon,lat (map centre — less precise, last resort for query params)
        if "ll" in params:
            parts = params["ll"][0].split(",")
            if len(parts) >= 2:
                return float(parts[1]), float(parts[0])  # lat, lon

        # Short links: yandex.ru/maps/-/..., ya.cc/..., maps.yandex.ru — follow redirect
        needs_redirect = (
            "/-/" in url
            or "ya.cc" in url
            or "maps.yandex" in url
            or (parsed.netloc in ("yandex.ru", "www.yandex.ru") and "/maps/" in parsed.path and not parsed.query)
        )
        if needs_redirect:
            req = urllib.request.Request(
                url, headers={"User-Agent": "Mozilla/5.0 SmartRoute/1.0"}
            )
            opener = urllib.request.build_opener(urllib.request.HTTPRedirectHandler())
            response = opener.open(req, timeout=10)
            final_url = response.geturl()
            if final_url != url:
                return parse_yandex_link(final_url)

        return None, None
    except Exception as e:
        logger.error("parse_yandex_link error for '%s': %s", url, e)
        return None, None


def reverse_geocode_nominatim(lat: float, lon: float) -> Optional[str]:
    """Get a human-readable address from coordinates using Nominatim."""
    try:
        url = (
            "https://nominatim.openstreetmap.org/reverse?"
            f"lat={lat}&lon={lon}&format=json&accept-language=ru"
        )
        req = urllib.request.Request(url, headers={"User-Agent": "smartroute-app-1.0"})
        with urllib.request.urlopen(req, timeout=5) as r:
            data = json.loads(r.read().decode("utf-8"))
        addr = data.get("address", {})
        road = addr.get("road", "")
        house = addr.get("house_number", "")
        city = addr.get("city") or addr.get("town") or addr.get("village") or ""
        parts = [p for p in [city, road, house] if p]
        return ", ".join(parts) if parts else data.get("display_name", "")
    except Exception as e:
        logger.warning("reverse_geocode_nominatim error: %s", e)
        return None


def calculate_savings(
    optimized_km: float,
    store_list: list,
    num_vehicles: int,
    depot_lat: float,
    depot_lon: float,
    settings: dict = None,
) -> dict:
    """
    Baseline: маршрут в порядке загрузки из Excel (как диспетчер без оптимизации),
    распределённый по машинам round-robin.  Сравниваем с оптимизированным SmartRoute.

    Методика прозрачна для клиента:
      «Если бы водители ехали в том же порядке, в котором магазины загружены в систему»
    """
    if settings is None:
        settings = get_company_settings()
    # Distribute stores round-robin across vehicles (same order as input)
    n = max(num_vehicles, 1)
    buckets: list = [[] for _ in range(n)]
    for idx, store in enumerate(store_list):
        buckets[idx % n].append(store)

    unoptimized_km = 0.0
    depot = (depot_lat, depot_lon)
    for bucket in buckets:
        if not bucket:
            continue
        prev = depot
        for s in bucket:
            if s.get("lat") and s.get("lon"):
                curr = (float(s["lat"]), float(s["lon"]))
                unoptimized_km += haversine_meters(prev, curr) / 1000.0
                prev = curr
        unoptimized_km += haversine_meters(prev, depot) / 1000.0

    unoptimized_km = max(float(unoptimized_km), optimized_km)
    saved_km = round(max(0.0, unoptimized_km - optimized_km), 1)
    saved_pct = round(saved_km / unoptimized_km * 100) if unoptimized_km > 0 else 0

    # ── Модель стоимости — параметры берём из настроек компании (БД) ─────────
    #
    # ROAD_FACTOR: географическая константа — Haversine → реальный пробег.
    # Оба маршрута (baseline и optimized) считаются через Haversine,
    # поэтому saved_km / saved_pct — честное сравнение.
    # Для монетарных метрик умножаем Haversine-км на ROAD_FACTOR.
    #
    ROAD_FACTOR: float = 1.4           # Haversine → реальный пробег (константа)
    fuel_l_per_100km: float = float(settings.get("fuel_consumption", 13.0))
    fuel_price_rub: float = float(settings.get("fuel_price", 67.0))
    cost_per_km_val: float = float(settings.get("cost_per_km", round(67.0 * 13.0 / 100.0, 2)))

    # Применяем ROAD_FACTOR только к монетарным метрикам, а не к saved_km,
    # чтобы не искажать честное сравнение маршрутов.
    saved_km_road = saved_km * ROAD_FACTOR
    saved_fuel_l = round(saved_km_road * fuel_l_per_100km / 100.0, 1)
    saved_fuel_cost_rub = round(saved_fuel_l * fuel_price_rub)
    # cost_per_km = fuel_price × consumption / 100 (только топливо)
    saved_rub_day = round(saved_km_road * cost_per_km_val)

    return {
        "optimized_km": round(optimized_km, 1),
        "unoptimized_km": round(unoptimized_km, 1),
        "saved_km": saved_km,
        "saved_pct": saved_pct,
        "saved_fuel_l": saved_fuel_l,
        "saved_fuel_cost_rub": saved_fuel_cost_rub,
        "saved_rub_day": saved_rub_day,
        "saved_rub_month": saved_rub_day * 30,
        "cost_per_km": round(cost_per_km_val, 2),
        "fuel_price": fuel_price_rub,
        "fuel_consumption": fuel_l_per_100km,
    }


# Яндекс.Навигатор (мобильное приложение) поддерживает максимум 20 промежуточных
# точек в одной ссылке rtext. При большем числе маршрут не строится.
YANDEX_NAV_MAX_STOPS = 20


def yandex_nav_url(coords_list: list) -> str:
    """Генерирует одну ссылку Яндекс.Навигатора.
    coords_list[0] = склад (depot) — первая точка rtext.
    Яндекс заменяет её GPS-позицией водителя, но все магазины (точки 2…N) сохраняются.
    """
    chunk = coords_list[:YANDEX_NAV_MAX_STOPS]
    points = "~".join(f"{lat},{lon}" for lat, lon in chunk)
    return f"https://yandex.ru/maps/?rtext={points}&rtt=auto"


def yandex_nav_urls(coords_list: list) -> list:
    """
    coords_list[0] = склад (depot), coords_list[1:] = магазины в порядке объезда.

    Разбивает маршрут на сегменты по YANDEX_NAV_MAX_STOPS точек (включая точку
    отправления). Каждый сегмент:
      - сегмент 1: [склад, магазин1 … магазин19]  (20 точек)
      - сегмент 2: [магазин19, магазин20 … магазин37] (20 точек)
      - …

    Яндекс Навигатор заменяет первую точку GPS-позицией водителя — это нормально:
    водитель стоит на складе в начале смены. Все магазины остаются на месте.
    """
    if not coords_list:
        return []

    stores = coords_list[1:]        # магазины (без склада)
    if not stores:
        return []

    max_stores = YANDEX_NAV_MAX_STOPS - 1  # 19 магазинов + 1 точка отправления = 20
    origin = coords_list[0]         # склад — точка отправления первого сегмента

    urls = []
    idx = 0
    while idx < len(stores):
        chunk = stores[idx: idx + max_stores]
        idx += max_stores

        segment = [origin] + chunk
        points = "~".join(f"{lat},{lon}" for lat, lon in segment)
        urls.append(f"https://yandex.ru/maps/?rtext={points}&rtt=auto")

        # Следующий сегмент стартует с последнего магазина текущего
        origin = chunk[-1]

    return urls


def whatsapp_url(vehicle_name: str, stores: list, total_km: float, yandex_urls: list) -> str:
    lines = [f"🚚 {vehicle_name} — маршрут на {total_km:.1f} км:"]
    for i, s in enumerate(stores, 1):
        lines.append(f"{i}. {s['store_name']} — {s['address']}")
    if len(yandex_urls) == 1:
        lines.append(f"\n🗺 Навигатор: {yandex_urls[0]}")
    elif len(yandex_urls) > 1:
        lines.append("")
        for idx, url in enumerate(yandex_urls, 1):
            lines.append(f"🗺 Маршрут {idx}: {url}")
    text = "\n".join(lines)
    return f"https://wa.me/?text={urllib.parse.quote(text)}"


def whatsapp_assignment_url(vehicle_name: str, route_url: str, driver_url: str) -> str:
    """Share both navigation and the token-scoped execution page."""
    lines = [
        f"SmartRoute: рейс {vehicle_name}",
        f"🗺 Навигация: {route_url}" if route_url else "",
        f"✅ Исполнение доставок: {driver_url}",
    ]
    return "https://wa.me/?text=" + urllib.parse.quote(
        "\n".join(line for line in lines if line)
    )


def _normalize_driver_phone(phone: str) -> str:
    """Return a WhatsApp-compatible phone value without punctuation."""
    return re.sub(r"\D", "", phone or "")


def whatsapp_driver_url(
    phone: str,
    delivery_date: str,
    vehicle_name: str,
    total_points: int,
    total_km: float,
    route_url: str,
    driver_url: str,
) -> str:
    """Build a prefilled WhatsApp link; delivery is always user-initiated."""
    text = "\n".join(line for line in [
        "SmartRoute — рейс",
        f"Дата: {delivery_date or 'не указана'}",
        f"Машина: {vehicle_name or 'не указана'}",
        f"Точек: {total_points}",
        f"Пробег: {round(float(total_km or 0), 1)} км",
        f"Яндекс Навигатор: {route_url}" if route_url else "",
        f"Исполнение доставок: {driver_url}",
    ] if line)
    normalized = _normalize_driver_phone(phone)
    target = f"https://wa.me/{normalized}" if normalized else "https://wa.me/"
    return f"{target}?text={urllib.parse.quote(text)}"


def _public_app_url(request: Request) -> str:
    """Resolve the public origin used inside driver links shared in WhatsApp."""
    configured = os.environ.get("PUBLIC_APP_URL", "").strip().rstrip("/")
    if configured:
        return configured
    forwarded_host = request.headers.get("x-forwarded-host") or request.headers.get("host")
    forwarded_proto = request.headers.get("x-forwarded-proto") or request.url.scheme
    if forwarded_host:
        return f"{forwarded_proto.split(',')[0].strip()}://{forwarded_host.split(',')[0].strip()}"
    return str(request.base_url).rstrip("/")


def store_row_to_dict(row) -> dict:
    return {
        "id": row["id"],
        "name": row["name"],
        "address": row["address"],
        "city": row.get("city") or "",
        "phone": row.get("phone") or "",
        "client": row.get("client") or "",
        "lat": row["lat"],
        "lon": row["lon"],
        "map_url": row.get("map_url"),
        "geocode_status": row["geocode_status"] or "pending",
        "time_window_from": row["time_window_from"] or "09:00",
        "time_window_to": row["time_window_to"] or "18:00",
        "unload_minutes": row["unload_minutes"] or 15,
        "created_at": str(row["created_at"]) if row["created_at"] else None,
    }


# ── Pydantic models ──────────────────────────────────────────────────────────

class StoreInput(BaseModel):
    name: str
    address: Optional[str] = None
    city: Optional[str] = None
    phone: Optional[str] = None
    client: Optional[str] = None
    yandex_url: Optional[str] = None
    lat: Optional[float] = None
    lon: Optional[float] = None
    map_url: Optional[str] = None
    time_window_from: Optional[str] = "09:00"
    time_window_to: Optional[str] = "18:00"
    unload_minutes: Optional[int] = 15


class StoreUpdate(BaseModel):
    name: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    phone: Optional[str] = None
    client: Optional[str] = None
    yandex_url: Optional[str] = None
    lat: Optional[float] = None
    lon: Optional[float] = None
    map_url: Optional[str] = None
    time_window_from: Optional[str] = None
    time_window_to: Optional[str] = None
    unload_minutes: Optional[int] = None


class VehicleInput(BaseModel):
    name: str
    capacity_kg: Optional[int] = None
    capacity_m3: Optional[float] = None
    average_speed: Optional[float] = None


class RouteRequest(BaseModel):
    store_ids: list[int]
    vehicles: list[VehicleInput]
    depot_lat: Optional[float] = None
    depot_lon: Optional[float] = None
    use_time_windows: Optional[bool] = False
    use_unload_time: Optional[bool] = False
    max_stops_per_vehicle: Optional[int] = None
    optimize_by: str = "distance"  # "distance" | "time" — backward-compat default
    delivery_date: Optional[str] = None  # "YYYY-MM-DD" — used to enrich stops with products/quantity


class CompanySettingsInput(BaseModel):
    fuel_price: float       # руб/литр
    fuel_consumption: float # л/100 км


EXECUTION_STATUSES = {"planned", "delivered", "partial", "failed", "rescheduled"}
PAYMENT_METHODS = {"cash", "card", "transfer", "none"}
PAYMENT_STATUSES = {"pending", "paid", "not_paid"}


class AssignmentCreate(BaseModel):
    route_index: int
    driver_name: str = ""
    vehicle_name: str = ""
    driver_id: Optional[int] = None


class AssignmentUpdate(BaseModel):
    driver_name: Optional[str] = None
    vehicle_name: Optional[str] = None
    driver_id: Optional[int] = None


class DriverCreate(BaseModel):
    name: str
    phone: str
    vehicle_name: str = ""


class DriverUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    vehicle_name: Optional[str] = None
    is_active: Optional[bool] = None


class DispatcherExecutionUpdate(BaseModel):
    rescheduled_date: Optional[str] = None


class DispatcherCreateRescheduledOrder(BaseModel):
    delivery_date: str


class ExecutionUpdate(BaseModel):
    status: str
    actual_qty: Optional[float] = None
    payment_method: Optional[str] = None
    payment_status: Optional[str] = None
    driver_comment: str = ""


class DriverLocationInput(BaseModel):
    lat: float
    lon: float
    accuracy: Optional[float] = None


_PRODUCT_QUANTITY_RE = re.compile(
    r"^\s*(?:(?P<prefix>\d+(?:[.,]\d+)?)\s+\S|(?P<suffix>\S.+?)\s+(?P<suffix_qty>\d+(?:[.,]\d+)?)"
    r"|(?P<multiplier>\S.+?)\s*[xх×]\s*(?P<multiplier_qty>\d+(?:[.,]\d+)?))",
    re.IGNORECASE,
)


def _quantity_from_products(products: object) -> Optional[float]:
    """Extract the total planned quantity from a products display string.

    Product summaries come from several importers, so accept both
    ``2 воды``/``воды 2`` and ``вода x2``/``Молоко×5``.  A bare product name
    does not provide enough information and intentionally returns None.
    """
    if products is None:
        return None
    text = str(products).strip()
    if not text:
        return None

    total = 0.0
    found = False
    for item in re.split(r"[,;]\s*", text):
        item = item.strip()
        if not item:
            continue
        match = _PRODUCT_QUANTITY_RE.match(item)
        if not match:
            continue
        raw_quantity = (
            match.group("prefix")
            or match.group("suffix_qty")
            or match.group("multiplier_qty")
        )
        if raw_quantity is None:
            continue
        quantity = float(raw_quantity.replace(",", "."))
        if math.isfinite(quantity) and quantity > 0:
            total += quantity
            found = True
    return total if found else None


def _execution_quantity(
    quantity: object,
    products: object,
    fallback_quantity: object = None,
) -> float:
    """Return a positive planned quantity without changing the DB schema."""
    try:
        parsed_quantity = float(quantity or 0)
    except (TypeError, ValueError):
        parsed_quantity = 0.0
    if math.isfinite(parsed_quantity) and parsed_quantity > 0:
        return parsed_quantity
    product_quantity = _quantity_from_products(products)
    if product_quantity is not None:
        return product_quantity
    try:
        parsed_fallback = float(fallback_quantity or 0)
    except (TypeError, ValueError):
        parsed_fallback = 0.0
    if math.isfinite(parsed_fallback) and parsed_fallback > 0:
        return parsed_fallback
    return 1.0


def _remaining_order_products(
    products: object,
    planned_qty: float,
    actual_qty: float,
    remaining_qty: float,
) -> tuple[str, str]:
    """Format residual cargo for one product, preserving multi-product text."""
    source = str(products or "").strip()
    parts = [part.strip() for part in re.split(r"[,;]\s*", source) if part.strip()]
    if len(parts) != 1:
        return source, "Требуется уточнение остатка по товарным позициям."

    item = parts[0]
    quantity_pattern = r"\d+(?:[.,]\d+)?"
    patterns = (
        re.compile(rf"^\s*(?P<qty>{quantity_pattern})\s+(?P<name>.+?)\s*$", re.IGNORECASE),
        re.compile(rf"^\s*(?P<name>.+?)\s*[xх×]\s*(?P<qty>{quantity_pattern})\s*$", re.IGNORECASE),
        re.compile(rf"^\s*(?P<name>.+?)\s+(?P<qty>{quantity_pattern})\s*$", re.IGNORECASE),
    )
    name = item
    for pattern in patterns:
        match = pattern.match(item)
        if match:
            name = match.group("name").strip()
            break
    if not name:
        return source, "Требуется уточнение остатка по товарным позициям."

    def format_number(value: float) -> str:
        numeric_value = float(value)
        return str(int(numeric_value)) if numeric_value.is_integer() else f"{numeric_value:g}"

    remaining_text = format_number(remaining_qty)
    comment = (
        f"Создано из остатка исполнения доставки. Было {format_number(planned_qty)}, "
        f"доставлено {format_number(actual_qty)}, осталось {remaining_text}."
    )
    return f"{name} × {remaining_text}", comment


def _validate_execution_quantities(status: str, planned_qty: float, actual_qty: float) -> None:
    """Keep delivered quantities bounded and status-consistent."""
    if not math.isfinite(planned_qty) or not math.isfinite(actual_qty):
        raise HTTPException(status_code=422, detail="Количество должно быть конечным числом")
    if planned_qty < 0:
        raise HTTPException(status_code=422, detail="Плановое количество не может быть отрицательным")
    if actual_qty < 0 or actual_qty > planned_qty:
        raise HTTPException(
            status_code=422,
            detail="Фактически доставленное количество должно быть от 0 до планового",
        )
    epsilon = 1e-9
    if status == "delivered" and abs(actual_qty - planned_qty) > epsilon:
        raise HTTPException(
            status_code=422,
            detail="Для статуса «Доставлено» фактическое количество должно совпадать с плановым",
        )
    if status == "partial" and (
        planned_qty <= epsilon
        or actual_qty <= epsilon
        or actual_qty >= planned_qty - epsilon
    ):
        raise HTTPException(
            status_code=422,
            detail="Для частичной доставки фактическое количество должно быть меньше планового",
        )


# ── Routes ───────────────────────────────────────────────────────────────────

@app.on_event("startup")
async def startup():
    # ── Routing chain status ──────────────────────────────────────────────────
    gh_status = f"enabled (plan_limit={_gh_plan_limit})" if GRAPHHOPPER_API_KEY else "disabled (no API key)"
    logger.info(
        "Routing chain: GH[%s] → OSRM[%s] → Haversine[always]",
        gh_status, OSRM_BASE_URL,
    )
    if not GRAPHHOPPER_API_KEY:
        logger.warning(
            "GRAPHHOPPER_API_KEY not set — GraphHopper disabled. "
            "OSRM (%s) will be used for real road distances.",
            OSRM_BASE_URL,
        )
    if not YANDEX_GEOCODER_API_KEY:
        logger.warning(
            "YANDEX_GEOCODER_API_KEY not set — Yandex Geocoder disabled, falling back to Nominatim"
        )
    init_db()
    # NOTE: migrate_moscow_stores() deliberately NOT called here.
    # It was a one-time dev migration that ran when the DB held old Moscow demo data.
    # Calling it on every startup is catastrophically dangerous for production clients
    # in any Russian city with lat > 50 (Moscow, SPb, Novosibirsk, Yekaterinburg, etc.).
    admin_id = seed_admin_user()  # migrates legacy NULL owner_id data to admin
    seed_demo_data(owner_id=admin_id)  # seeds only if admin has no stores

    # ── Periodic in-memory cleanup ────────────────────────────────────────────
    t = threading.Thread(target=_memory_cleanup_loop, daemon=True)
    t.start()


def _memory_cleanup_loop() -> None:
    """Daemon thread: purge stale in-memory entries every 10 minutes.

    Covers four growth vectors:
    1. _rl_store      — rate-limit buckets whose timestamps are all expired
    2. _login_attempts — failed-login buckets whose timestamps are all expired
    3. import_jobs / bulk_create_jobs — completed jobs older than 2 hours
    4. _matrix_cache   — already capped by MAX_MATRIX_CACHE_SIZE on write;
                         this adds a safety net in case of a burst
    """
    _JOB_TTL  = 2 * 3600   # 2 hours: keep finished jobs for polling
    _RATE_TTL = 3600        # 1 hour: dead rate-limit keys
    _INTERVAL = 600         # run every 10 minutes

    while True:
        time.sleep(_INTERVAL)
        now = time.time()

        # Rate-limit store: drop keys whose list is empty or entirely expired
        with _rl_lock:
            dead = [k for k, ts in list(_rl_store.items())
                    if not ts or now - max(ts) > _RATE_TTL]
            for k in dead:
                _rl_store.pop(k, None)

        # Login-attempt store: drop IPs with no recent attempts
        with _login_attempts_lock:
            dead = [ip for ip, ts in list(_login_attempts.items())
                    if not ts or now - max(ts) > LOGIN_WINDOW_SECONDS * 2]
            for ip in dead:
                _login_attempts.pop(ip, None)

        # Completed import/bulk-create jobs older than JOB_TTL
        for store in (import_jobs, bulk_create_jobs):
            done_old = [jid for jid, j in list(store.items())
                        if j.get("done")
                        and now - j.get("_created_at", now) > _JOB_TTL]
            for jid in done_old:
                store.pop(jid, None)

        # Matrix cache safety net (evict oldest 20% if still oversized)
        if len(_matrix_cache) > MAX_MATRIX_CACHE_SIZE:
            n_evict = len(_matrix_cache) - MAX_MATRIX_CACHE_SIZE
            for old_key in list(_matrix_cache.keys())[:n_evict]:
                _matrix_cache.pop(old_key, None)

        logger.debug(
            "Memory cleanup: rl_store=%d, login=%d, import_jobs=%d, "
            "bulk_jobs=%d, matrix_cache=%d",
            len(_rl_store), len(_login_attempts),
            len(import_jobs), len(bulk_create_jobs), len(_matrix_cache),
        )


def migrate_moscow_stores():
    """Удаляем старые московские демо-магазины (lat > 50), сохраняя всё остальное."""
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            "SELECT COUNT(*) as moscow_cnt "
            "FROM stores WHERE lat IS NOT NULL AND lat > 50"
        )
        row = cur.fetchone()
        moscow_cnt = int(row["moscow_cnt"])
        # Only remove if there are Moscow-like stores and it's a small demo-sized batch
        if moscow_cnt > 0 and moscow_cnt <= 15:
            logger.info("Removing %d Moscow demo stores (lat > 50)...", moscow_cnt)
            cur.execute("DELETE FROM stores WHERE lat IS NOT NULL AND lat > 50")
            conn.commit()
            logger.info("Migration done: %d Moscow stores removed.", moscow_cnt)
    except Exception as exc:
        logger.error("migrate_moscow_stores failed: %s", exc)
        conn.rollback()
    finally:
        cur.close()
        conn.close()


def seed_demo_data(owner_id: int = None):
    """Seed demo stores and route sessions for a specific user (owner_id).
    Only seeds if that user has no stores yet."""
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if owner_id is not None:
        cur.execute("SELECT COUNT(*) as cnt FROM stores WHERE owner_id = %s", (owner_id,))
    else:
        cur.execute("SELECT COUNT(*) as cnt FROM stores")
    row = cur.fetchone()
    # Seed if empty, or if only a handful remain after migration
    if row["cnt"] >= 3:
        cur.close()
        conn.close()
        return

    # No demo stores — new users start with onboarding flow (city-agnostic)

    # Seed some historical route sessions
    today = date.today()
    for i in range(14):
        d = today.replace(day=max(1, today.day - i))
        total_km = 80 + (i * 7) % 40
        saved_km = total_km * 0.23
        cur.execute(
            """INSERT INTO route_sessions (date, num_vehicles, total_km, saved_km, saved_rub, num_points, owner_id)
               VALUES (%s, %s, %s, %s, %s, %s, %s)""",
            (str(d), 2 + i % 3, round(total_km, 1), round(saved_km, 1), round(saved_km * 1.4 * 31), 8 + i % 6, owner_id)
        )

    conn.commit()
    cur.close()
    conn.close()


# ── Auth utilities ────────────────────────────────────────────────────────────

import bcrypt as _bcrypt_lib


def _truncate_password(password: str) -> bytes:
    """bcrypt silently truncates at 72 bytes — enforce explicitly."""
    return password.encode("utf-8")[:72]


def _hash_password(password: str) -> str:
    pw = _truncate_password(password)
    return _bcrypt_lib.hashpw(pw, _bcrypt_lib.gensalt(rounds=12)).decode("utf-8")


def _verify_password(plain: str, hashed: str) -> bool:
    try:
        pw = _truncate_password(plain)
        return _bcrypt_lib.checkpw(pw, hashed.encode("utf-8"))
    except Exception:
        return False


def _create_access_token(username: str) -> str:
    expire = datetime.utcnow() + timedelta(hours=JWT_TOKEN_TTL_HOURS)
    payload = {"sub": username, "exp": expire}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def _decode_token(token: str) -> Optional[str]:
    """Return username from token, or None if invalid/expired."""
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload.get("sub")
    except JWTError:
        return None


def _decode_token_with_exp(token: str) -> tuple:
    """Return (username, exp_datetime) or (None, None) if invalid/expired."""
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        username = payload.get("sub")
        exp_ts = payload.get("exp")
        exp_dt = datetime.utcfromtimestamp(exp_ts) if exp_ts else None
        return username, exp_dt
    except JWTError:
        return None, None


def seed_admin_user() -> Optional[int]:
    """Create the admin user from ADMIN_PASSWORD env var if not exists.
    Sets is_admin=TRUE. Migrates legacy (NULL owner_id) data to admin.
    Returns the admin user ID, or None on failure."""
    if not ADMIN_PASSWORD:
        logger.warning(
            "ADMIN_PASSWORD env var is not set — admin user will NOT be created. "
            "Set ADMIN_PASSWORD to enable login."
        )
        return None
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # Use a PostgreSQL advisory lock to prevent race conditions on concurrent startup
        cur.execute("SELECT pg_advisory_xact_lock(123456789)")
        cur.execute("SELECT id, is_admin FROM users WHERE username = %s", ("admin",))
        row = cur.fetchone()
        if row is None:
            hashed = _hash_password(ADMIN_PASSWORD)
            cur.execute(
                "INSERT INTO users (username, password_hash, is_admin, is_active) VALUES (%s, %s, TRUE, TRUE) RETURNING id",
                ("admin", hashed),
            )
            admin_id = cur.fetchone()["id"]
            conn.commit()
            logger.info("Admin user created (id=%d).", admin_id)
        else:
            admin_id = row["id"]
            # Ensure is_admin flag is set on existing admin account
            if not row.get("is_admin"):
                cur.execute("UPDATE users SET is_admin = TRUE, is_active = TRUE WHERE id = %s", (admin_id,))
                conn.commit()
            logger.info("Admin user already exists (id=%d).", admin_id)

        # One-time migration: assign stores/sessions/settings with NULL owner_id to admin
        cur.execute("UPDATE stores SET owner_id = %s WHERE owner_id IS NULL", (admin_id,))
        migrated_stores = cur.rowcount
        cur.execute("UPDATE route_sessions SET owner_id = %s WHERE owner_id IS NULL", (admin_id,))
        migrated_sessions = cur.rowcount
        cur.execute("UPDATE company_settings SET owner_id = %s WHERE owner_id IS NULL", (admin_id,))
        conn.commit()
        if migrated_stores > 0 or migrated_sessions > 0:
            logger.info("Migrated %d stores and %d sessions to admin (owner_id=%d).",
                        migrated_stores, migrated_sessions, admin_id)
        return admin_id
    except Exception as exc:
        logger.error("seed_admin_user failed: %s", exc)
        conn.rollback()
        return None
    finally:
        cur.close()
        conn.close()


# ── Auth middleware ───────────────────────────────────────────────────────────

def _hash_api_key(raw_key: str) -> str:
    """SHA-256 hash of a raw API key string (hex digest)."""
    return hashlib.sha256(raw_key.encode()).hexdigest()


def _generate_api_key() -> tuple[str, str]:
    """Return (full_key, key_prefix). Full key shown once; only hash stored."""
    rand = secrets.token_urlsafe(32)
    full_key = f"sr_live_{rand}"
    prefix = full_key[:16]  # "sr_live_" + 8 chars — safe to display
    return full_key, prefix


def _resolve_api_key(raw_key: str) -> dict | None:
    """Lookup an API key by hash; update last_used; return key row or None."""
    key_hash = _hash_api_key(raw_key)
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            """SELECT ak.id, ak.owner_id, ak.scopes, ak.is_active, ak.expires_at,
                      u.id as user_id, u.is_active as user_active, u.is_admin
               FROM api_keys ak
               JOIN users u ON u.id = ak.owner_id
               WHERE ak.key_hash = %s""",
            (key_hash,)
        )
        row = cur.fetchone()
        if row and row["is_active"] and row["user_active"]:
            if row["expires_at"] is None or row["expires_at"] > datetime.utcnow():
                # Update last_used (best-effort, no fail if it errors)
                try:
                    cur.execute(
                        "UPDATE api_keys SET last_used_at = NOW() WHERE id = %s",
                        (row["id"],)
                    )
                    conn.commit()
                except Exception:
                    pass
                cur.close(); conn.close()
                return dict(row)
        cur.close(); conn.close()
    except Exception as exc:
        logger.error("_resolve_api_key error: %s", exc)
    return None


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    # Always pass through OPTIONS (CORS pre-flight) and public paths
    if request.method == "OPTIONS":
        return await call_next(request)

    path = request.url.path

    # Webhook ingest: token-in-URL, authenticated inside the handler
    if path.startswith(_AUTH_WEBHOOK_PREFIX):
        return await call_next(request)
    # Driver execution endpoints authenticate with a scoped URL token.
    if path.startswith(_DRIVER_API_PREFIX):
        return await call_next(request)

    if path in _AUTH_PUBLIC_PATHS or not path.startswith("/api/"):
        return await call_next(request)

    # ── Helper: emit 401 in v1-envelope format for /api/v1/* paths ───────────
    _is_v1 = path.startswith("/api/v1/")

    def _auth_401(message: str, code: str = "UNAUTHORIZED"):
        if _is_v1:
            import uuid as _u
            body = {"error": {"code": code, "message": message},
                    "request_id": "req_" + _u.uuid4().hex[:12]}
            return JSONResponse(status_code=401, content=body)
        return JSONResponse(status_code=401, content={"detail": message})

    # ── 1. Cookie JWT (browser sessions) ─────────────────────────────────────
    token = request.cookies.get(JWT_COOKIE_NAME)
    if token:
        username, token_exp = _decode_token_with_exp(token)
        if not username:
            return _auth_401("Токен недействителен или истёк. Войдите снова.")
        try:
            _conn = get_db()
            _cur = _conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            _cur.execute("SELECT id, is_active, is_admin FROM users WHERE username = %s", (username,))
            _user_row = _cur.fetchone()
            _cur.close(); _conn.close()
        except Exception as _exc:
            logger.error("Auth middleware DB error: %s", _exc)
            _user_row = None
        if not _user_row:
            return _auth_401("Пользователь не найден. Войдите снова.")
        if not _user_row.get("is_active", True):
            return _auth_401("Аккаунт отключён. Обратитесь к администратору.")
        request.state.username = username
        request.state.user_id = _user_row["id"]
        request.state.is_admin = bool(_user_row.get("is_admin", False))
        request.state.api_key_scopes = None  # cookie auth = full access

        # ── Sliding session: silently refresh cookie when close to expiry ────
        # If token expires within JWT_REFRESH_THRESHOLD_HOURS, issue a new one.
        response = await call_next(request)
        if token_exp:
            hours_left = (token_exp - datetime.utcnow()).total_seconds() / 3600
            if hours_left < JWT_REFRESH_THRESHOLD_HOURS:
                new_token = _create_access_token(username)
                response.set_cookie(
                    key=JWT_COOKIE_NAME,
                    value=new_token,
                    httponly=True,
                    samesite=COOKIE_SAMESITE,
                    secure=COOKIE_SECURE,
                    max_age=JWT_TOKEN_TTL_HOURS * 3600,
                    path="/",
                )
        return response

    # ── 2. Bearer API Key (machine-to-machine) ────────────────────────────────
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        raw_key = auth_header[7:].strip()
        key_row = _resolve_api_key(raw_key)
        if not key_row:
            return _auth_401("Недействительный API-ключ.")
        request.state.username = f"api_key:{key_row['id']}"
        request.state.user_id = key_row["user_id"]
        request.state.is_admin = bool(key_row.get("is_admin", False))
        request.state.api_key_scopes = key_row.get("scopes") or []
        return await call_next(request)

    return _auth_401("Не авторизован. Укажите Bearer-токен в заголовке Authorization.")


# ── Auth endpoints ────────────────────────────────────────────────────────────

class LoginForm(BaseModel):
    username: str
    password: str


@app.post("/api/auth/login")
async def login(request: Request, response: Response):
    """Authenticate with username + password (form-encoded). Sets HttpOnly JWT cookie."""
    client_ip = _get_client_ip(request)
    _check_login_rate_limit(client_ip)

    content_type = request.headers.get("content-type", "")
    if "application/x-www-form-urlencoded" in content_type or "multipart/form-data" in content_type:
        form = await request.form()
        username = str(form.get("username", "")).strip()
        password = str(form.get("password", ""))
    else:
        try:
            body = await request.json()
            username = str(body.get("username", "")).strip()
            password = str(body.get("password", ""))
        except Exception:
            raise HTTPException(status_code=422, detail="Неверный формат запроса")

    if not username or not password:
        raise HTTPException(status_code=422, detail="Укажите логин и пароль")

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT password_hash, is_active, is_admin FROM users WHERE username = %s", (username,))
        row = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    if not row:
        _record_failed_login(client_ip)
        raise HTTPException(status_code=401, detail="Неверный логин или пароль")
    if not row.get("is_active", True):
        _record_failed_login(client_ip)
        raise HTTPException(status_code=401, detail="Аккаунт отключён. Обратитесь к администратору.")
    if not _verify_password(password, row["password_hash"]):
        _record_failed_login(client_ip)
        raise HTTPException(status_code=401, detail="Неверный логин или пароль")

    _clear_login_attempts(client_ip)

    # Update last_login_at
    try:
        _lconn = get_db()
        _lcur = _lconn.cursor()
        _lcur.execute("UPDATE users SET last_login_at = NOW() WHERE username = %s", (username,))
        _lconn.commit()
        _lcur.close()
        _lconn.close()
    except Exception:
        pass

    token = _create_access_token(username)
    resp = JSONResponse(content={"ok": True, "username": username, "is_admin": bool(row.get("is_admin", False))})
    resp.set_cookie(
        key=JWT_COOKIE_NAME,
        value=token,
        httponly=True,
        samesite=COOKIE_SAMESITE,
        secure=COOKIE_SECURE,
        max_age=JWT_TOKEN_TTL_HOURS * 3600,
        path="/",
    )
    return resp


@app.post("/api/auth/logout")
async def logout():
    """Clear the JWT cookie."""
    resp = JSONResponse(content={"ok": True})
    resp.delete_cookie(key=JWT_COOKIE_NAME, path="/", samesite=COOKIE_SAMESITE)
    return resp


@app.get("/api/auth/me")
async def me(request: Request):
    """Return the currently authenticated user."""
    username = getattr(request.state, "username", None)
    if not username:
        raise HTTPException(status_code=401, detail="Не авторизован")
    return {
        "username": username,
        "user_id": getattr(request.state, "user_id", None),
        "is_admin": getattr(request.state, "is_admin", False),
    }


# ── Auth helpers ──────────────────────────────────────────────────────────────

def get_user_id(request: Request) -> int:
    """Return the current user's ID from request state (set by auth middleware)."""
    uid = getattr(request.state, "user_id", None)
    if uid is None:
        raise HTTPException(status_code=401, detail="Не авторизован")
    return uid


def require_admin(request: Request) -> int:
    """Return current user ID if they are an admin, else raise 403.

    Admin endpoints require cookie-based session auth — Bearer API keys are
    intentionally excluded even when the underlying user is an admin.
    This prevents leaked/stolen API keys from being used to escalate privileges.
    """
    uid = get_user_id(request)
    # Reject Bearer-authenticated requests regardless of is_admin flag
    if getattr(request.state, "api_key_scopes", None) is not None:
        raise HTTPException(
            status_code=403,
            detail="Панель администратора доступна только через браузерную сессию, не через API-ключ."
        )
    if not getattr(request.state, "is_admin", False):
        raise HTTPException(status_code=403, detail="Доступ запрещён. Только для администраторов.")
    return uid


# ── Business routes ───────────────────────────────────────────────────────────

@app.get("/api/healthz")
def health_check():
    return {"status": "ok"}


@app.get("/api/stores")
def list_stores(request: Request):
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM stores WHERE owner_id = %s ORDER BY id", (uid,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [store_row_to_dict(r) for r in rows]


@app.post("/api/stores", status_code=201)
def create_store(request: Request, body: StoreInput, force: bool = Query(False, description="Пропустить предупреждение о дубликате")):
    uid = get_user_id(request)
    if not body.name or not body.name.strip():
        raise HTTPException(status_code=422, detail="Название магазина не может быть пустым")
    if not body.yandex_url and not body.address and not body.city and body.lat is None:
        raise HTTPException(status_code=422, detail="Укажите ссылку из Яндекс Карт или адрес")
    if body.lat is not None and not (-90 <= body.lat <= 90):
        raise HTTPException(status_code=422, detail="Широта должна быть от -90 до 90")
    if body.lon is not None and not (-180 <= body.lon <= 180):
        raise HTTPException(status_code=422, detail="Долгота должна быть от -180 до 180")

    raw_address = (body.address or "").strip()
    city = (body.city or "").strip()
    geocode_query = f"{city} {raw_address}".strip() if city and city not in raw_address else raw_address

    lat, lon, status = None, None, "not_found"
    address = raw_address or city

    # Priority 1: explicit lat/lon
    if body.lat is not None and body.lon is not None:
        lat, lon, status = body.lat, body.lon, "found"

    # Priority 2: parse Yandex Maps URL
    elif body.yandex_url:
        lat, lon = parse_yandex_link(body.yandex_url)
        if lat is not None and lon is not None:
            status = "found"
            if not address:
                address = reverse_geocode_nominatim(lat, lon) or f"{lat:.5f}, {lon:.5f}"
            logger.info("create_store: coords from yandex_url → (%.5f, %.5f)", lat, lon)
        elif geocode_query:
            coords = geocode_address(geocode_query)
            lat = coords[0] if coords else None
            lon = coords[1] if coords else None
            status = "found" if coords else "not_found"
            logger.info("create_store: yandex_url parse failed, geocoded '%s' → %s", geocode_query, status)

    # Priority 3: geocode address
    elif geocode_query:
        coords = geocode_address(geocode_query)
        lat = coords[0] if coords else None
        lon = coords[1] if coords else None
        status = "found" if coords else "not_found"
        logger.info("create_store: geocoded '%s' → %s", geocode_query, status)

    if not address:
        address = geocode_query or (f"{lat:.5f}, {lon:.5f}" if lat is not None else "Адрес не указан")

    # Duplicate detection (skip if force=True)
    if lat is not None and lon is not None and not force:
        nearby = find_nearby_stores(lat, lon, radius_m=20, owner_id=uid)
        if nearby:
            near = nearby[0]
            raise HTTPException(
                status_code=409,
                detail={
                    "type": "duplicate_warning",
                    "message": (
                        f"Найден похожий магазин в {near['dist_m']:.0f} м: "
                        f"«{near['name']}» ({near['address']}). "
                        "Создать всё равно?"
                    ),
                    "existing": {
                        "id": near["id"],
                        "name": near["name"],
                        "address": near["address"],
                        "dist_m": round(float(near["dist_m"]), 1),
                    },
                },
            )

    # Store yandex_url as map_url if no explicit map_url provided
    map_url = body.map_url or body.yandex_url

    # Prepend city to address for consistent "Город, адрес" format (enables city filter)
    if city and address and city not in address:
        address = f"{city}, {address}"
    elif city and not address:
        address = city

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """INSERT INTO stores (name, address, city, phone, client, lat, lon, map_url, geocode_status, time_window_from, time_window_to, unload_minutes, owner_id)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING *""",
        (body.name.strip(), address, city, (body.phone or "").strip(), (body.client or "").strip(),
         lat, lon, map_url,
         status, body.time_window_from, body.time_window_to, body.unload_minutes, uid)
    )
    row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    return store_row_to_dict(row)


@app.get("/api/geocode")
def geocode_endpoint(address: Optional[str] = None, yandex_url: Optional[str] = None):
    """Geocode an address or parse a Yandex Maps URL. Returns {lat, lon, display}."""
    if yandex_url and yandex_url.strip():
        lat, lon = parse_yandex_link(yandex_url.strip())
        if lat is not None and lon is not None:
            display = reverse_geocode_nominatim(lat, lon) or yandex_url.strip()
            return {"lat": lat, "lon": lon, "display": display}
        raise HTTPException(status_code=422, detail="Не удалось извлечь координаты из ссылки Яндекс Карт")
    if address and address.strip():
        result = geocode_address(address.strip())
        if result:
            lat, lon = result
            return {"lat": lat, "lon": lon, "display": address.strip()}
        raise HTTPException(status_code=404, detail="Адрес не найден. Попробуйте уточнить запрос.")
    raise HTTPException(status_code=400, detail="Укажите параметр address или yandex_url")


@app.get("/api/stores/template")
def download_stores_template():
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Магазины"

    # 7-column simplified template
    headers = [
        "Название",         # A — required
        "Ссылка Яндекс",    # B — recommended (coords parsed automatically)
        "Адрес",            # C — if no Yandex link
        "Город",            # D — optional, prepended to address
        "Телефон",          # E — optional
        "Клиент",           # F — optional
        "Разгрузка мин",    # G — optional
        "Время с",          # H — optional
        "Время до",         # I — optional
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Example row 1: with Yandex URL (no city-specific data)
    ws.append([
        "Супермаркет Центральный",
        "https://yandex.ru/maps/?whatshere[point]=37.6173,55.7558",
        "",
        "",
        "+7 900 000-00-00",
        "ООО Торг-Центр",
        15, "09:00", "18:00",
    ])
    # Example row 2: with address + city
    ws.append([
        "Магазин на Ленина",
        "",
        "ул. Ленина 15",
        "Ваш город",
        "+7 900 111-11-11",
        "ИП Иванов",
        20, "10:00", "17:00",
    ])

    col_widths = [28, 52, 36, 16, 18, 22, 16, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Hint row (skipped on import — starts with ←)
    note_row = [
        "← Название магазина",
        "← Ссылка из Яндекс: зажми место → Поделиться",
        "← Адрес если нет ссылки",
        "← Город",
        "← Телефон (необязательно)",
        "← Клиент/контрагент (необязательно)",
        "← Минут (число)",
        "← ЧЧ:ММ",
        "← ЧЧ:ММ",
    ]
    ws.append(note_row)
    for col_num in range(1, len(note_row) + 1):
        cell = ws.cell(row=ws.max_row, column=col_num)
        cell.font = Font(italic=True, color="888888")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    content = buf.read()

    import base64
    return {
        "data": base64.b64encode(content).decode("ascii"),
        "filename": "smartroute_template.xlsx",
    }


@app.get("/api/stores/export")
def export_stores(request: Request):
    """Export all user stores as Excel (base64 JSON). Same format as import template."""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT name, map_url, address, city, phone, client, lat, lon, unload_minutes, time_window_from, time_window_to "
        "FROM stores WHERE owner_id = %s ORDER BY id",
        (uid,),
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import base64

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Магазины"

    headers = [
        "Название",
        "Ссылка Яндекс",
        "Адрес",
        "Город",
        "Телефон",
        "Клиент",
        "Разгрузка мин",
        "Время с",
        "Время до",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([
            row.get("name") or "",
            row.get("map_url") or "",
            row.get("address") or "",
            row.get("city") or "",
            row.get("phone") or "",
            row.get("client") or "",
            row.get("unload_minutes") or 15,
            row.get("time_window_from") or "09:00",
            row.get("time_window_to") or "18:00",
        ])

    col_widths = [28, 52, 36, 16, 18, 22, 16, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    content = buf.read()

    from datetime import date
    filename = f"smartroute_stores_{date.today().isoformat()}.xlsx"
    return {
        "data": base64.b64encode(content).decode("ascii"),
        "filename": filename,
        "count": len(rows),
    }


@app.post("/api/stores/import")
async def import_stores(request: Request, file: UploadFile = File(...)):
    owner_id = get_user_id(request)
    _api_rate_limit(f"stores_import:{owner_id}", max_calls=10, window_seconds=60)
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Файл слишком большой (макс. 20 МБ)")
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # Read header row to detect column layout
    header_row = [str(c).strip().lower() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]

    # Map header names → column indices (0-based)
    def _col(candidates: list) -> Optional[int]:
        for name in candidates:
            for i, h in enumerate(header_row):
                if name in h:
                    return i
        return None

    c_name    = _col(["назван", "name", "store_name"])
    c_yandex  = _col(["ссылка яндекс", "яндекс", "yandex", "ссылка"])
    c_address = _col(["адрес", "address"])
    c_city    = _col(["город", "city"])
    c_phone   = _col(["телефон", "phone", "тел"])
    c_client  = _col(["клиент", "client", "контрагент"])
    c_lat     = _col(["широта", "lat", "latitude"])
    c_lon     = _col(["долгота", "lon", "longitude"])
    c_mapurl  = _col(["map_url", "ссылка на карт"])
    c_unload  = _col(["разгрузка", "unload"])
    c_from    = _col(["время с", "open_time", "с (", "time_from"])
    c_to      = _col(["время до", "close_time", "до (", "time_to"])

    # Fall back to positional mapping for old-format files (5-column):
    # Col0=name, Col1=address, Col2=tw_from, Col3=tw_to, Col4=unload
    if c_name is None:
        c_name = 0
    if c_address is None and c_yandex is None:
        c_address = 1

    def _get(row, idx, default=""):
        if idx is None or idx >= len(row):
            return default
        val = row[idx]
        return val if val is not None else default

    imported = 0
    failed = 0
    stores = []
    duplicates = []
    # Skip header row and any trailing note rows (those with col A starting with ←)
    all_data_rows = [
        r for r in ws.iter_rows(min_row=2, values_only=True)
        if r and r[0] and not str(r[0]).strip().startswith("←")
    ]
    total_rows = len(all_data_rows)
    logger.info("Excel import: %d data rows found (headers: %s)", total_rows, header_row)

    for i, row in enumerate(all_data_rows, start=1):
        name       = str(_get(row, c_name, "")).strip()
        yandex_url = str(_get(row, c_yandex, "")).strip() or None
        city       = str(_get(row, c_city, "")).strip()
        phone      = str(_get(row, c_phone, "")).strip()
        client     = str(_get(row, c_client, "")).strip()
        raw_addr   = str(_get(row, c_address, "")).strip()

        # Combine city + address: city goes FIRST so address.split(",")[0] == city
        if city and city not in raw_addr:
            address = f"{city}, {raw_addr}" if raw_addr else city
        else:
            address = raw_addr

        if not name or (not yandex_url and not address):
            logger.warning("Import row %d skipped: missing name or location", i)
            failed += 1
            continue

        # Parse optional lat/lon columns
        raw_lat = _get(row, c_lat)
        raw_lon = _get(row, c_lon)
        map_url = str(_get(row, c_mapurl, "")).strip() or None

        # Parse time window
        if c_from is None and len(row) > 2 and row[2] and ":" in str(row[2]):
            tw_from = str(row[2]).strip()
            tw_to   = str(row[3]).strip() if len(row) > 3 and row[3] else "18:00"
            unload  = int(row[4]) if len(row) > 4 and row[4] else 15
        else:
            tw_from = str(_get(row, c_from, "09:00")).strip() or "09:00"
            tw_to   = str(_get(row, c_to,   "18:00")).strip() or "18:00"
            try:
                unload = int(_get(row, c_unload, 15))
            except (ValueError, TypeError):
                unload = 15

        # ── Coord resolution (priority order) ────────────────────────────────
        lat, lon, status = None, None, "not_found"

        # 1. Explicit lat/lon columns
        try:
            prov_lat = float(raw_lat) if raw_lat not in (None, "", "None") else None
            prov_lon = float(raw_lon) if raw_lon not in (None, "", "None") else None
        except (ValueError, TypeError):
            prov_lat = prov_lon = None

        if prov_lat is not None and prov_lon is not None and (-90 <= prov_lat <= 90) and (-180 <= prov_lon <= 180):
            lat, lon, status = prov_lat, prov_lon, "found"
            logger.info("Import %d/%d — %s → explicit coords (%.4f, %.4f)", i, total_rows, name, lat, lon)

        # 2. Parse Yandex URL
        elif yandex_url:
            lat, lon = parse_yandex_link(yandex_url)
            if lat is not None and lon is not None:
                status = "found"
                if not address:
                    address = reverse_geocode_nominatim(lat, lon) or f"{lat:.5f}, {lon:.5f}"
                logger.info("Import %d/%d — %s → yandex_url (%.4f, %.4f)", i, total_rows, name, lat, lon)
            elif address:
                coords = geocode_address(address)
                lat = coords[0] if coords else None
                lon = coords[1] if coords else None
                status = "found" if coords else "not_found"
                if not YANDEX_GEOCODER_API_KEY:
                    time.sleep(1.1)
                logger.info("Import %d/%d — %s → geocoded (yandex failed) → %s", i, total_rows, name, status)

        # 3. Geocode address
        elif address:
            coords = geocode_address(address)
            lat = coords[0] if coords else None
            lon = coords[1] if coords else None
            status = "found" if coords else "not_found"
            logger.info("Import %d/%d — %s → geocoded '%s' → %s", i, total_rows, name, address, status)
            if not YANDEX_GEOCODER_API_KEY:
                time.sleep(1.1)

        if not address:
            address = (f"{lat:.5f}, {lon:.5f}" if lat is not None else "Адрес не указан")

        final_map_url = map_url or yandex_url

        # Duplicate detection (non-blocking — warn only)
        dup_warning = None
        if lat is not None and lon is not None:
            nearby = find_nearby_stores(lat, lon, radius_m=20, owner_id=owner_id)
            if nearby:
                near = nearby[0]
                dup_warning = {
                    "row": i,
                    "name": name,
                    "existing_id": near["id"],
                    "existing_name": near["name"],
                    "dist_m": round(float(near["dist_m"]), 1),
                }
                logger.info("Import row %d — possible duplicate: '%s' ≈ id=%d '%s' (%.1fm)",
                            i, name, near["id"], near["name"], near["dist_m"])

        try:
            conn = get_db()
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(
                """INSERT INTO stores (name, address, city, phone, client, lat, lon, map_url, geocode_status, time_window_from, time_window_to, unload_minutes, owner_id)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING *""",
                (name, address, city, phone, client, lat, lon, final_map_url, status, tw_from, tw_to, unload, owner_id)
            )
            db_row = cur.fetchone()
            conn.commit()
            cur.close()
            conn.close()
            stores.append(store_row_to_dict(db_row))
            imported += 1
            if dup_warning:
                duplicates.append(dup_warning)
        except Exception as e:
            logger.error(f"Failed to insert store row {i}: {e}")
            failed += 1

    return {
        "total": imported + failed,
        "imported": imported,
        "failed": failed,
        "stores": stores,
        "duplicates": duplicates,
    }


def _normalize_for_dedup(s) -> str:
    """Lowercase, trim, collapse whitespace — used as dedup key for (name, address)."""
    import re as _re_dedup
    return _re_dedup.sub(r'\s+', ' ', str(s or "").lower().strip())


# ── Extended keyword lists (SmartRoute + 1C terms) ─────────────────────────
_KWORDS_NAME    = ["контрагент", "покупатель", "store name", "назван", "name", "store_name", "магазин"]
_KWORDS_ADDRESS = ["адрес доставки", "адрес", "address"]
_KWORDS_CITY    = ["город", "city"]
_KWORDS_YANDEX  = ["ссылка яндекс", "яндекс", "yandex", "ссылка"]
_KWORDS_UNLOAD  = ["разгрузка", "unload"]
_KWORDS_FROM    = ["время с", "open_time", "с (", "time_from"]
_KWORDS_TO      = ["время до", "close_time", "до (", "time_to"]
_KWORDS_PHONE   = ["телефон", "phone", "тел.", "тел "]
# "контрагент" / "клиент" listed here but NOT in _KWORDS_NAME to avoid same-column clash
_KWORDS_CLIENT  = ["контрагент", "клиент", "client", "заказчик"]


def _detect_col(header_lower: list, candidates: list) -> Optional[int]:
    for kw in candidates:
        for i, h in enumerate(header_lower):
            if kw in h:
                return i
    return None


def _col_content_warning(col_idx: Optional[int], data_rows: list, field: str) -> Optional[str]:
    """Return a warning string if the column's sample data looks wrong for the given field.
    Samples up to 20 non-empty values from data_rows at col_idx."""
    if col_idx is None:
        return None
    samples = []
    for row in data_rows[:50]:
        v = str(row[col_idx]).strip() if col_idx < len(row) and row[col_idx] is not None else ""
        if v:
            samples.append(v)
        if len(samples) >= 20:
            break
    if not samples:
        return "Колонка пустая — нет данных для импорта"

    def _is_numeric(s: str) -> bool:
        return s.replace(" ", "").replace(",", "").replace(".", "").replace("-", "").replace("+", "").isdigit()

    numeric_count = sum(1 for s in samples if _is_numeric(s))
    numeric_ratio = numeric_count / len(samples)

    if field == "city":
        if numeric_ratio >= 0.5:
            return "Похоже, в этой колонке числа, а не названия городов"
        avg_len = sum(len(s) for s in samples) / len(samples)
        if avg_len > 35:
            return "Значения слишком длинные для города — возможно, это адреса"
        if avg_len < 2:
            return "Значения слишком короткие для названия города"

    elif field == "address":
        if numeric_ratio >= 0.7:
            return "Похоже, в этой колонке числа, а не адреса"
        avg_len = sum(len(s) for s in samples) / len(samples)
        if avg_len < 4:
            return "Значения слишком короткие для адресов"

    elif field == "name":
        if numeric_ratio >= 0.8:
            return "Похоже, в этой колонке числа, а не названия"

    elif field == "unload":
        non_numeric = sum(1 for s in samples if not _is_numeric(s))
        if non_numeric / len(samples) > 0.5:
            return "Ожидаются числа (минуты), но большинство значений — не числа"

    elif field == "phone":
        def _looks_like_phone(s: str) -> bool:
            digits = sum(c.isdigit() for c in s)
            return digits >= 7
        bad = sum(1 for s in samples if not _looks_like_phone(s))
        if bad / len(samples) > 0.5:
            return "Значения не похожи на номера телефонов"

    return None


@app.post("/api/stores/import/preview")
async def preview_import(request: Request, file: UploadFile = File(...)):
    """Read Excel file, return columns + first rows + auto-detected column mapping.
    Also checks how many rows already exist in the DB (by normalized name+address).
    Used by the frontend to show a mapping dialog before the actual import."""
    uid = get_user_id(request)
    _api_rate_limit(f"stores_preview:{uid}", max_calls=20, window_seconds=60)
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Файл слишком большой (макс. 20 МБ)")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Не удалось открыть Excel файл: {e}")
    ws = wb.active

    raw_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))
    columns = [str(c).strip() if c is not None else f"Колонка {i+1}" for i, c in enumerate(raw_headers)]
    header_lower = [c.lower() for c in columns]

    # Collect all data rows (skip ← hint rows)
    all_data_rows = [
        r for r in ws.iter_rows(min_row=2, values_only=True)
        if r and any(c is not None for c in r)
        and not str(r[0] or "").strip().startswith("←")
    ]
    total_rows = len(all_data_rows)

    # First 5 rows for preview table
    preview_rows = []
    for row in all_data_rows[:5]:
        preview_rows.append([
            str(c).strip() if c is not None else ""
            for c in list(row) + [""] * max(0, len(columns) - len(row))
        ][:len(columns)])

    # Auto-detect column mapping
    c_name    = _detect_col(header_lower, _KWORDS_NAME)
    c_address = _detect_col(header_lower, _KWORDS_ADDRESS)
    c_city    = _detect_col(header_lower, _KWORDS_CITY)
    c_yandex  = _detect_col(header_lower, _KWORDS_YANDEX)
    c_unload  = _detect_col(header_lower, _KWORDS_UNLOAD)
    c_from    = _detect_col(header_lower, _KWORDS_FROM)
    c_to      = _detect_col(header_lower, _KWORDS_TO)
    c_phone   = _detect_col(header_lower, _KWORDS_PHONE)
    c_client  = _detect_col(header_lower, _KWORDS_CLIENT)

    # ── Conflict resolution: no two fields may share the same column index ────
    # When "контрагент" matches both name and client, prefer name (primary key).
    if c_client is not None and c_client == c_name:
        c_client = None
    if c_phone is not None and c_phone in (c_name, c_address, c_city):
        c_phone = None

    # Count unique points after dedup (name + address) for info display
    seen: set = set()
    for row in all_data_rows:
        def _gv(idx, _row=row):
            if idx is None or idx >= len(_row): return ""
            return str(_row[idx] or "").strip()
        n = _normalize_for_dedup(_gv(c_name))
        a = _normalize_for_dedup(_gv(c_address))
        if n:
            seen.add((n, a))

    # ── Cross-check against existing DB stores for this user ─────────────────
    # For each unique file row, check multiple identity signals:
    #   1. name + address  → "name_address" (strong, likely duplicate)
    #   2. address only    → "address_only" (same building, different tenant — NOT auto-dup)
    #   3. Yandex URL      → "yandex_url"   (same link, strong signal)
    # Returns per-row matches so the UI can explain WHY each row is flagged.
    matches: list = []
    existing_count = 0
    try:
        conn_p = get_db()
        cur_p = conn_p.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur_p.execute("SELECT id, name, address, map_url FROM stores WHERE owner_id = %s", (uid,))
        db_rows = cur_p.fetchall()
        cur_p.close(); conn_p.close()

        # Build lookup maps
        db_by_name_addr: dict = {}   # (norm_name, norm_addr) → {id, name, address}
        db_by_addr: dict = {}        # norm_addr → list of {id, name, address}
        db_by_yandex: dict = {}      # normalized url → {id, name, address}
        for dbr in db_rows:
            dn = _normalize_for_dedup(dbr["name"] or "")
            da = _normalize_for_dedup(dbr["address"] or "")
            store_ref = {"id": dbr["id"], "name": dbr["name"], "address": dbr["address"]}
            if dn:
                db_by_name_addr[(dn, da)] = store_ref
            if da:
                db_by_addr.setdefault(da, []).append(store_ref)
            if dbr.get("map_url"):
                url_key = str(dbr["map_url"]).strip().lower()
                db_by_yandex[url_key] = store_ref

        matched_keys: set = set()
        for row in all_data_rows:
            def _gv2(idx, _row=row):
                if idx is None or idx >= len(_row): return ""
                return str(_row[idx] or "").strip()
            fn = _normalize_for_dedup(_gv2(c_name))
            fa_raw = _normalize_for_dedup(_gv2(c_address)) if c_address is not None else ""
            fy_url = _gv2(c_yandex).strip().lower() if c_yandex is not None else ""
            if not fn:
                continue
            key = (fn, fa_raw)
            if key in matched_keys:
                continue
            matched_keys.add(key)

            file_name  = _gv2(c_name)
            file_addr  = _gv2(c_address) if c_address is not None else ""

            # Signal 1: name + address exact match
            hit = db_by_name_addr.get((fn, fa_raw))
            if hit:
                matches.append({
                    "file_name": file_name, "file_address": file_addr,
                    "existing_id": hit["id"], "existing_name": hit["name"],
                    "existing_address": hit["address"],
                    "reason": "name_address",
                    "is_likely_duplicate": True,
                })
                continue

            # Signal 2: Yandex URL match
            if fy_url and fy_url in db_by_yandex:
                hit = db_by_yandex[fy_url]
                matches.append({
                    "file_name": file_name, "file_address": file_addr,
                    "existing_id": hit["id"], "existing_name": hit["name"],
                    "existing_address": hit["address"],
                    "reason": "yandex_url",
                    "is_likely_duplicate": True,
                })
                continue

            # Signal 3: same address, different name (same building, different tenant)
            if fa_raw and fa_raw in db_by_addr:
                candidates = db_by_addr[fa_raw]
                # Only flag if no name match (different-name tenants at same address)
                different_name_hits = [c for c in candidates if _normalize_for_dedup(c["name"]) != fn]
                if different_name_hits:
                    hit = different_name_hits[0]
                    matches.append({
                        "file_name": file_name, "file_address": file_addr,
                        "existing_id": hit["id"], "existing_name": hit["name"],
                        "existing_address": hit["address"],
                        "reason": "address_only",
                        "is_likely_duplicate": False,
                    })
                    continue

        existing_count = sum(1 for m in matches if m["is_likely_duplicate"])
    except Exception as e:
        logger.warning("preview_import: DB check failed: %s", e)

    new_count = len(seen) - existing_count

    return {
        "columns": columns,
        "rows": preview_rows,
        "total_rows": total_rows,
        "unique_count": len(seen),
        "existing_count": existing_count,
        "new_count": new_count,
        "matches": matches,
        "mapping": {
            "name":    c_name,
            "address": c_address,
            "city":    c_city,
            "yandex":  c_yandex,
            "unload":  c_unload,
            "tw_from": c_from,
            "tw_to":   c_to,
            "phone":   c_phone,
            "client":  c_client,
        },
    }


def _import_process_content_sync(content_bytes: bytes, job: dict, mapping: Optional[dict] = None, owner_id: int = None, import_mode: str = "new_only") -> None:
    """Run Excel import synchronously, updating job dict for progress tracking.
    Called from background thread by /api/stores/import/start endpoint.
    mapping: optional dict with column indices {name, address, city, yandex, unload, tw_from, tw_to}.
    import_mode: 'new_only' (skip existing by name+address), 'update' (update existing), 'all' (always insert)."""
    if not OPENPYXL_AVAILABLE:
        job["error"] = "openpyxl not installed"
        job["done"] = True
        return
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content_bytes))
    except Exception as e:
        job["error"] = f"Не удалось открыть Excel файл: {e}"
        job["done"] = True
        return

    ws = wb.active
    header_row = [str(c).strip().lower() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]

    # ── Column indices: use caller-supplied mapping, else auto-detect ──────────
    if mapping:
        c_name      = mapping.get("name")
        c_yandex    = mapping.get("yandex")
        c_addr      = mapping.get("address")
        c_city      = mapping.get("city")
        c_unload    = mapping.get("unload")
        c_from      = mapping.get("tw_from")
        c_to        = mapping.get("tw_to")
        c_phone     = mapping.get("phone")
        c_client    = mapping.get("client")
        default_city = str(mapping.get("default_city") or "").strip()
        # lat/lon/mapurl always auto-detected (not exposed in mapping UI)
        c_lat    = _detect_col(header_row, ["широта", "lat", "latitude"])
        c_lon    = _detect_col(header_row, ["долгота", "lon", "longitude"])
        c_mapurl = _detect_col(header_row, ["map_url", "ссылка на карт"])
    else:
        default_city = ""
        c_name   = _detect_col(header_row, _KWORDS_NAME)
        c_yandex = _detect_col(header_row, _KWORDS_YANDEX)
        c_addr   = _detect_col(header_row, _KWORDS_ADDRESS)
        c_city   = _detect_col(header_row, _KWORDS_CITY)
        c_lat    = _detect_col(header_row, ["широта", "lat", "latitude"])
        c_lon    = _detect_col(header_row, ["долгота", "lon", "longitude"])
        c_mapurl = _detect_col(header_row, ["map_url", "ссылка на карт"])
        c_unload = _detect_col(header_row, _KWORDS_UNLOAD)
        c_from   = _detect_col(header_row, _KWORDS_FROM)
        c_to     = _detect_col(header_row, _KWORDS_TO)
        c_phone  = _detect_col(header_row, _KWORDS_PHONE)
        c_client = _detect_col(header_row, _KWORDS_CLIENT)

    def _get(row, idx, default=""):
        if idx is None or idx >= len(row):
            return default
        v = row[idx]
        return v if v is not None else default

    all_rows = [
        r for r in ws.iter_rows(min_row=2, values_only=True)
        if r and r[0] and not str(r[0]).strip().startswith("←")
    ]

    # ── Deduplication by (normalize(name), normalize(address)) ───────────────
    # Typical 1C export: same store appears once per product line.
    # We collapse all duplicate (name+address) pairs to a single row BEFORE geocoding.
    seen_dedup: dict = {}
    deduped_rows = []
    for row in all_rows:
        n_key = _normalize_for_dedup(_get(row, c_name, ""))
        a_key = _normalize_for_dedup(_get(row, c_addr, ""))
        key = (n_key, a_key)
        if n_key and key not in seen_dedup:
            seen_dedup[key] = True
            deduped_rows.append(row)
    skipped_dedup = len(all_rows) - len(deduped_rows)
    if skipped_dedup:
        logger.info("Import dedup: removed %d duplicate rows, %d unique points remain", skipped_dedup, len(deduped_rows))

    total_rows = len(deduped_rows)
    job["total"] = total_rows
    job["deduped"] = skipped_dedup
    imported, failed, skipped_existing = 0, 0, 0
    stores_out: list = []
    duplicates: list = []
    # Geocoder source stats — how coordinates were obtained for each imported store
    geocode_stats = {"explicit": 0, "yandex_url": 0, "memory_cache": 0, "db_cache": 0, "yandex_api": 0, "nominatim": 0, "not_found": 0}

    # ── Pre-load existing store keys from DB (for import_mode check) ─────────
    # Maps normalized (name, address) → store_id for quick lookup.
    existing_db_keys: dict = {}
    if import_mode in ("new_only", "update") and owner_id is not None:
        try:
            conn_ex = get_db()
            cur_ex = conn_ex.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur_ex.execute("SELECT id, name, address FROM stores WHERE owner_id = %s", (owner_id,))
            for dbr in cur_ex.fetchall():
                dn = _normalize_for_dedup(dbr["name"] or "")
                da = _normalize_for_dedup(dbr["address"] or "")
                if dn:
                    existing_db_keys[(dn, da)] = dbr["id"]
            cur_ex.close(); conn_ex.close()
        except Exception as e:
            logger.warning("Import: failed to pre-load existing keys: %s", e)

    for i, row in enumerate(deduped_rows, start=1):
        name       = str(_get(row, c_name, "")).strip()
        yandex_url = str(_get(row, c_yandex, "")).strip() or None
        city       = str(_get(row, c_city, "")).strip() or default_city
        raw_addr   = str(_get(row, c_addr, "")).strip()
        address    = f"{city}, {raw_addr}" if city and city not in raw_addr else raw_addr
        if not address:
            address = city
        phone      = str(_get(row, c_phone, "")).strip() if c_phone is not None else ""
        client     = str(_get(row, c_client, "")).strip() if c_client is not None else ""

        if not name or (not yandex_url and not address):
            failed += 1
            job["processed"] = i; job["failed"] = failed
            continue

        # ── Check if this row already exists in DB ────────────────────────────
        # Try both raw_addr (as in file) and full address (city+raw_addr, as stored in DB)
        # to handle the case where city is in a separate column and was prepended on create.
        row_name_key = _normalize_for_dedup(name)
        row_addr_key_raw  = _normalize_for_dedup(raw_addr) if raw_addr else ""
        row_addr_key_full = _normalize_for_dedup(address)  if address  else row_addr_key_raw
        existing_store_id = (
            existing_db_keys.get((row_name_key, row_addr_key_raw)) or
            existing_db_keys.get((row_name_key, row_addr_key_full))
        )

        if existing_store_id and import_mode == "new_only":
            # Skip — store already exists and user wants new-only mode
            skipped_existing += 1
            job["processed"] = i
            job["skipped_existing"] = skipped_existing
            continue

        raw_lat = _get(row, c_lat)
        raw_lon = _get(row, c_lon)
        map_url = str(_get(row, c_mapurl, "")).strip() or None

        if c_from is None and len(row) > 2 and row[2] and ":" in str(row[2]):
            tw_from = str(row[2]).strip()
            tw_to   = str(row[3]).strip() if len(row) > 3 and row[3] else "18:00"
            unload  = int(row[4]) if len(row) > 4 and row[4] else 15
        else:
            tw_from = str(_get(row, c_from, "09:00")).strip() or "09:00"
            tw_to   = str(_get(row, c_to, "18:00")).strip() or "18:00"
            try:
                unload = int(_get(row, c_unload, 15))
            except (ValueError, TypeError):
                unload = 15

        lat, lon, status = None, None, "not_found"
        coord_source = "not_found"
        try:
            pv_lat = float(raw_lat) if raw_lat not in (None, "", "None") else None
            pv_lon = float(raw_lon) if raw_lon not in (None, "", "None") else None
        except (ValueError, TypeError):
            pv_lat = pv_lon = None

        def _geocode_with_tracking(addr: str) -> Optional[tuple]:
            """geocode_address with source tracking into geocode_stats."""
            nonlocal coord_source
            ck = addr.strip().lower()
            if ck in geocode_cache:
                coord_source = "memory_cache"
                return geocode_cache[ck]
            db_hit = _geocache_db_lookup(ck)
            if db_hit is not None:
                coord_source = "db_cache"
                geocode_cache[ck] = db_hit
                return db_hit
            result = geocode_address_yandex(addr)
            if result is not None:
                coord_source = "yandex_api"
            else:
                result = geocode_address_nominatim(addr)
                coord_source = "nominatim" if result is not None else "not_found"
            geocode_cache[ck] = result
            if result is not None:
                _geocache_db_store(ck, result[0], result[1], coord_source)
            return result

        if pv_lat is not None and pv_lon is not None and (-90 <= pv_lat <= 90) and (-180 <= pv_lon <= 180):
            lat, lon, status, coord_source = pv_lat, pv_lon, "found", "explicit"
        elif yandex_url:
            lat, lon = parse_yandex_link(yandex_url)
            if lat is not None:
                status, coord_source = "found", "yandex_url"
                if not address:
                    address = reverse_geocode_nominatim(lat, lon) or f"{lat:.5f}, {lon:.5f}"
            elif address:
                coords = _geocode_with_tracking(address)
                lat, lon = (coords[0], coords[1]) if coords else (None, None)
                status = "found" if coords else "not_found"
                if not YANDEX_GEOCODER_API_KEY and coord_source in ("nominatim", "not_found"):
                    time.sleep(1.1)
        elif address:
            coords = _geocode_with_tracking(address)
            lat, lon = (coords[0], coords[1]) if coords else (None, None)
            status = "found" if coords else "not_found"
            if not YANDEX_GEOCODER_API_KEY and coord_source in ("nominatim", "not_found"):
                time.sleep(1.1)

        geocode_stats[coord_source] = geocode_stats.get(coord_source, 0) + 1

        if not address:
            address = f"{lat:.5f}, {lon:.5f}" if lat is not None else "Адрес не указан"

        final_map_url = map_url or yandex_url

        dup_warning = None
        if lat is not None and lon is not None and import_mode != "update":
            nearby = find_nearby_stores(lat, lon, radius_m=20, owner_id=owner_id)
            if nearby:
                near = nearby[0]
                new_name_norm = _normalize_for_dedup(name)
                near_name_norm = _normalize_for_dedup(near.get("name") or "")
                near_addr_norm = _normalize_for_dedup(near.get("address") or "")
                new_addr_norm  = _normalize_for_dedup(address)
                same_name = new_name_norm == near_name_norm
                # Address similarity: one is substring of the other (handles city-prefix differences)
                same_addr = bool(near_addr_norm and new_addr_norm and (
                    near_addr_norm in new_addr_norm or new_addr_norm in near_addr_norm
                ))
                if same_name and same_addr:
                    match_reason = "name_address"
                    is_likely_duplicate = True
                elif same_name:
                    match_reason = "name_coords"
                    is_likely_duplicate = True
                else:
                    # Different names at the same location: different tenants in same building
                    # or geocoding imprecision (Nominatim centroid). NOT a duplicate — skip.
                    match_reason = "coords_only"
                    is_likely_duplicate = False
                # Only track as a duplicate when we actually believe it is one.
                # coords_only (different names, different addresses, just nearby coords)
                # is geocoding noise and must NOT be shown as a duplicate warning.
                if is_likely_duplicate:
                    dup_warning = {
                        "row": i, "name": name, "address": address,
                        "existing_id": near["id"], "existing_name": near["name"],
                        "existing_address": near.get("address") or "",
                        "dist_m": round(float(near["dist_m"]), 1),
                        "match_reason": match_reason,
                        "is_likely_duplicate": True,
                    }

        try:
            conn2 = get_db()
            cur2 = conn2.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

            if existing_store_id and import_mode == "update":
                # Update existing store in-place
                cur2.execute(
                    """UPDATE stores SET name=%s, address=%s, city=%s, phone=%s, client=%s,
                       lat=%s, lon=%s, map_url=%s,
                       geocode_status=%s, time_window_from=%s, time_window_to=%s, unload_minutes=%s
                       WHERE id=%s AND owner_id=%s RETURNING *""",
                    (name, address, city, phone, client, lat, lon, final_map_url, status,
                     tw_from, tw_to, unload, existing_store_id, owner_id),
                )
            else:
                cur2.execute(
                    """INSERT INTO stores (name, address, city, phone, client, lat, lon, map_url,
                       geocode_status, time_window_from, time_window_to, unload_minutes, owner_id)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING *""",
                    (name, address, city, phone, client, lat, lon, final_map_url,
                     status, tw_from, tw_to, unload, owner_id),
                )

            db_row = cur2.fetchone()
            conn2.commit(); cur2.close(); conn2.close()
            if db_row:
                stores_out.append(store_row_to_dict(db_row))
            imported += 1
            if dup_warning and db_row:
                dup_warning["new_store_id"] = db_row["id"]
                duplicates.append(dup_warning)
        except Exception as e:
            logger.error("Import job row %d failed: %s", i, e)
            failed += 1

        job["processed"] = i
        job["imported"] = imported
        job["failed"] = failed
        job["skipped_existing"] = skipped_existing
        job["duplicates"] = duplicates

    geocoded_found = sum(1 for s in stores_out if s.get("geocode_status") == "found")
    geocoded_not_found = sum(1 for s in stores_out if s.get("geocode_status") != "found")

    job["stores"] = stores_out
    job["geocoded_found"] = geocoded_found
    job["geocoded_not_found"] = geocoded_not_found
    job["geocode_stats"] = geocode_stats
    job["deduped"] = skipped_dedup
    job["skipped_existing"] = skipped_existing
    job["done"] = True
    logger.info("Import job done: %d imported (%d geocoded, %d no-coords), %d failed, %d true-duplicates, %d deduped, %d skipped_existing. Geocode sources: %s",
                imported, geocoded_found, geocoded_not_found, failed, len(duplicates), skipped_dedup, skipped_existing, geocode_stats)


@app.post("/api/stores/import/start", status_code=202)
async def start_import_stores(
    request: Request,
    file: UploadFile = File(...),
    mapping: Optional[str] = Form(None),
    import_mode: Optional[str] = Form("new_only"),
):
    """Start async background import. Returns job_id for progress polling.
    mapping: optional JSON string with column indices {name, address, city, yandex, unload, tw_from, tw_to}.
    import_mode: 'new_only' (default, skip existing), 'update' (update existing), 'all' (always insert)."""
    uid = get_user_id(request)
    _api_rate_limit(f"stores_import:{uid}", max_calls=10, window_seconds=60)
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Файл слишком большой (макс. 20 МБ)")

    parsed_mapping: Optional[dict] = None
    if mapping:
        try:
            parsed_mapping = json.loads(mapping)
        except Exception:
            raise HTTPException(status_code=422, detail="mapping must be valid JSON")

    safe_mode = import_mode if import_mode in ("new_only", "update", "all") else "new_only"

    job_id = _uuid.uuid4().hex[:8]
    job: dict = {
        "total": 0, "processed": 0, "imported": 0, "failed": 0,
        "done": False, "stores": [], "duplicates": [], "error": None, "deduped": 0,
        "skipped_existing": 0, "owner_id": uid, "_created_at": time.time(),
    }
    import_jobs[job_id] = job
    t = threading.Thread(target=_import_process_content_sync, args=(content, job, parsed_mapping, uid, safe_mode), daemon=True)
    t.start()
    return {"job_id": job_id}


@app.get("/api/stores/import/progress/{job_id}")
def get_import_progress(job_id: str, request: Request):
    """Poll import job progress. Returns current counters + done flag."""
    uid = get_user_id(request)
    if job_id not in import_jobs:
        raise HTTPException(status_code=404, detail="Import job not found")
    job = import_jobs[job_id]
    if job.get("owner_id") is not None and job["owner_id"] != uid:
        raise HTTPException(status_code=403, detail="Нет доступа к этому заданию импорта")
    return {
        "job_id": job_id,
        "total": job["total"],
        "processed": job["processed"],
        "imported": job["imported"],
        "failed": job["failed"],
        "done": job["done"],
        "duplicate_count": len(job.get("duplicates", [])),
        "error": job.get("error"),
    }


@app.get("/api/stores/import/result/{job_id}")
def get_import_result(job_id: str, request: Request):
    """Fetch final result of a completed import job."""
    uid = get_user_id(request)
    if job_id not in import_jobs:
        raise HTTPException(status_code=404, detail="Import job not found")
    job = import_jobs[job_id]
    if job.get("owner_id") is not None and job["owner_id"] != uid:
        raise HTTPException(status_code=403, detail="Нет доступа к этому заданию импорта")
    return {
        "total": job["total"],
        "imported": job["imported"],
        "failed": job["failed"],
        "stores": job.get("stores", []),
        "duplicates": job.get("duplicates", []),
        "deduped": job.get("deduped", 0),
        "skipped_existing": job.get("skipped_existing", 0),
        "geocoded_found": job.get("geocoded_found", 0),
        "geocoded_not_found": job.get("geocoded_not_found", 0),
        "geocode_stats": job.get("geocode_stats", {}),
        "done": job["done"],
        "error": job.get("error"),
    }


@app.get("/api/stores/{id}")
def get_store(id: int, request: Request):
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM stores WHERE id = %s AND owner_id = %s", (id, uid))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Store not found")
    return store_row_to_dict(row)


@app.put("/api/stores/{id}")
def update_store(id: int, body: StoreUpdate, request: Request):
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM stores WHERE id = %s AND owner_id = %s", (id, uid))
    existing = cur.fetchone()
    if not existing:
        cur.close()
        conn.close()
        raise HTTPException(status_code=404, detail="Store not found")

    fields = {}
    if body.name is not None:
        fields["name"] = body.name
    if body.address is not None:
        fields["address"] = body.address
        # Re-geocode if address changed
        coords = geocode_address(body.address)
        if coords:
            fields["lat"] = coords[0]
            fields["lon"] = coords[1]
            fields["geocode_status"] = "found"
        else:
            fields["lat"] = None
            fields["lon"] = None
            fields["geocode_status"] = "not_found"
    if body.yandex_url is not None:
        yurl = body.yandex_url.strip() if body.yandex_url else None
        fields["map_url"] = yurl
        if yurl:
            lat_y, lon_y = parse_yandex_link(yurl)
            if lat_y is not None and lon_y is not None:
                fields["lat"] = lat_y
                fields["lon"] = lon_y
                fields["geocode_status"] = "found"
    if body.city is not None:
        fields["city"] = body.city.strip()
    if body.phone is not None:
        fields["phone"] = body.phone.strip()
    if body.client is not None:
        fields["client"] = body.client.strip()
    if body.lat is not None:
        fields["lat"] = body.lat
    if body.lon is not None:
        fields["lon"] = body.lon
    if body.time_window_from is not None:
        fields["time_window_from"] = body.time_window_from
    if body.time_window_to is not None:
        fields["time_window_to"] = body.time_window_to
    if body.unload_minutes is not None:
        fields["unload_minutes"] = body.unload_minutes

    if fields:
        set_clause = ", ".join(f"{k} = %s" for k in fields)
        values = list(fields.values()) + [id, uid]
        cur.execute(f"UPDATE stores SET {set_clause} WHERE id = %s AND owner_id = %s RETURNING *", values)
        row = cur.fetchone()
        conn.commit()
    else:
        row = existing

    cur.close()
    conn.close()
    return store_row_to_dict(row)


@app.delete("/api/stores/{id}", status_code=204)
def delete_store(id: int, request: Request):
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM stores WHERE id = %s AND owner_id = %s", (id, uid))
    deleted = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    if deleted == 0:
        raise HTTPException(status_code=404, detail="Store not found")


class BulkDeleteRequest(BaseModel):
    ids: list[int]


@app.post("/api/stores/bulk-delete", status_code=200)
def bulk_delete_stores(request: Request, body: BulkDeleteRequest):
    """Delete multiple stores by ID. Only deletes stores owned by the current user."""
    uid = get_user_id(request)
    if not body.ids:
        return {"deleted": 0}
    if len(body.ids) > 5000:
        raise HTTPException(status_code=422, detail="Максимум 5000 магазинов за один запрос")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM stores WHERE id = ANY(%s) AND owner_id = %s",
        (body.ids, uid)
    )
    deleted = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    return {"deleted": deleted}


# ── Bulk store creation background job ────────────────────────────────────────

class BulkCreateStoreItem(BaseModel):
    name: str
    address: Optional[str] = None
    yandex_url: Optional[str] = None
    city: Optional[str] = None
    time_window_from: Optional[str] = "09:00"
    time_window_to: Optional[str] = "18:00"
    unload_minutes: Optional[int] = 15


class BulkCreateStartRequest(BaseModel):
    stores: list[BulkCreateStoreItem]
    delivery_date: Optional[str] = None


def _bulk_create_stores_sync(stores: list[dict], job: dict, uid: int):
    """Background thread: create stores one by one, update job dict in-place."""
    job["total"] = len(stores)
    job["created"] = 0
    job["failed"] = 0
    job["done"] = False
    job["results"] = []

    for store_data in stores:
        if job.get("cancelled"):
            break
        name = store_data.get("name", "").strip()
        if not name:
            job["failed"] += 1
            job["results"].append({"name": name or "(пусто)", "status": "failed", "reason": "Пустое название"})
            continue

        try:
            conn = get_db()
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

            address = store_data.get("address", "")
            yandex_url = store_data.get("yandex_url", "")
            city = store_data.get("city", "")
            time_from = store_data.get("time_window_from") or "09:00"
            time_to = store_data.get("time_window_to") or "18:00"
            unload = store_data.get("unload_minutes") or 15

            # Determine coordinates
            lat, lon, geocode_status = None, None, "not_found"
            if yandex_url:
                coords = parse_yandex_link(yandex_url)
                # IMPORTANT: parse_yandex_link may return (None, None) — tuple is always
                # truthy, so check the actual lat value, not just `if coords:`
                if coords and coords[0] is not None and coords[1] is not None:
                    lat, lon = coords
                    geocode_status = "found"
                    # Reverse geocode to get address if not provided
                    if not address:
                        rev = reverse_geocode_nominatim(lat, lon)
                        if rev:
                            address = rev

            if lat is None and address:
                geocode_query = f"{city}, {address}" if city else address
                coords = geocode_address(geocode_query)
                if coords:
                    lat, lon = coords
                    geocode_status = "found"
                else:
                    geocode_status = "not_found"
                # Nominatim rate-limit: 1 req/sec (skip if Yandex API key is set)
                if not YANDEX_GEOCODER_API_KEY:
                    time.sleep(1.1)

            # Build full address with city prefix
            full_address = address or ""
            if city and full_address and not full_address.startswith(city):
                full_address = f"{city}, {full_address}"
            elif city and not full_address:
                full_address = city
            # Guarantee non-empty address (DB NOT NULL constraint)
            if not full_address:
                full_address = "Адрес не указан"

            cur.execute(
                """INSERT INTO stores (owner_id, name, address, city, map_url, lat, lon,
                           geocode_status, time_window_from, time_window_to, unload_minutes)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                   RETURNING id""",
                (uid, name, full_address or None, city or None, yandex_url or None,
                 lat, lon, geocode_status, time_from, time_to, int(unload))
            )
            new_id = cur.fetchone()["id"]
            conn.commit()
            cur.close()
            conn.close()

            job["created"] += 1
            job["results"].append({
                "name": name,
                "address": address or "",
                "status": "created",
                "store_id": new_id,
                "geocode_status": geocode_status,
                "reason": None,
            })

        except Exception as exc:
            try:
                conn.rollback()
            except Exception:
                pass
            try:
                cur.close()
                conn.close()
            except Exception:
                pass
            reason = str(exc)[:200]
            job["failed"] += 1
            job["results"].append({"name": name, "address": address or "", "status": "failed", "reason": reason})

    job["done"] = True


@app.post("/api/stores/bulk-create/start", status_code=202)
def start_bulk_create_stores(request: Request, body: BulkCreateStartRequest):
    """Start a background job to bulk-create stores. Returns job_id for polling."""
    uid = get_user_id(request)
    if not body.stores:
        raise HTTPException(status_code=422, detail="Список магазинов пуст")
    if len(body.stores) > 500:
        raise HTTPException(status_code=422, detail="Максимум 500 магазинов за один запрос")

    job_id = _uuid.uuid4().hex[:8]
    job: dict = {
        "owner_id": uid,
        "total": len(body.stores),
        "created": 0,
        "failed": 0,
        "done": False,
        "cancelled": False,
        "results": [],
        "_created_at": time.time(),
    }
    bulk_create_jobs[job_id] = job

    stores_data = [s.model_dump() for s in body.stores]
    t = threading.Thread(target=_bulk_create_stores_sync, args=(stores_data, job, uid), daemon=True)
    t.start()

    return {"job_id": job_id}


@app.get("/api/stores/bulk-create/progress/{job_id}")
def get_bulk_create_progress(job_id: str, request: Request):
    """Poll progress of a bulk-create job."""
    uid = get_user_id(request)
    if job_id not in bulk_create_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    job = bulk_create_jobs[job_id]
    if job.get("owner_id") != uid:
        raise HTTPException(status_code=403, detail="Нет доступа")
    return {
        "job_id": job_id,
        "total": job["total"],
        "created": job["created"],
        "failed": job["failed"],
        "done": job["done"],
    }


@app.get("/api/stores/bulk-create/result/{job_id}")
def get_bulk_create_result(job_id: str, request: Request):
    """Fetch final results of a completed bulk-create job."""
    uid = get_user_id(request)
    if job_id not in bulk_create_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    job = bulk_create_jobs[job_id]
    if job.get("owner_id") != uid:
        raise HTTPException(status_code=403, detail="Нет доступа")
    return {
        "job_id": job_id,
        "total": job["total"],
        "created": job["created"],
        "failed": job["failed"],
        "done": job["done"],
        "results": job.get("results", []),
    }


@app.post("/api/stores/{id}/geocode")
def geocode_store(id: int, request: Request):
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM stores WHERE id = %s AND owner_id = %s", (id, uid))
    store = cur.fetchone()
    if not store:
        cur.close()
        conn.close()
        raise HTTPException(status_code=404, detail="Store not found")

    lat, lon, status = None, None, "not_found"

    # 1. Try yandex_url first (coordinate-precise, no geocoding needed)
    if store.get("map_url"):
        try:
            coords = parse_yandex_link(store["map_url"])
            if coords:
                lat, lon = coords
                status = "found"
        except Exception:
            pass

    # 2. Geocode by address
    if lat is None and store.get("address"):
        coords = geocode_address(store["address"])
        lat = coords[0] if coords else None
        lon = coords[1] if coords else None
        status = "found" if coords else "not_found"

    cur.execute(
        "UPDATE stores SET lat = %s, lon = %s, geocode_status = %s WHERE id = %s AND owner_id = %s RETURNING *",
        (lat, lon, status, id, uid)
    )
    row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    return store_row_to_dict(row)


@app.post("/api/stores/geocode-pending")
def geocode_pending_stores(request: Request, background_tasks: BackgroundTasks):
    """
    Background task: geocode all stores with geocode_status='not_found' or 'pending'
    for the current user. Returns immediately with a count.
    """
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT id FROM stores WHERE owner_id = %s AND (geocode_status = 'not_found' OR geocode_status = 'pending' OR lat IS NULL)",
        (uid,)
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    store_ids = [r["id"] for r in rows]

    def _geocode_all(ids: list, owner_id: int):
        for sid in ids:
            try:
                conn2 = get_db()
                cur2 = conn2.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                cur2.execute("SELECT * FROM stores WHERE id = %s AND owner_id = %s", (sid, owner_id))
                store = cur2.fetchone()
                if not store:
                    cur2.close(); conn2.close()
                    continue

                lat, lon, status = None, None, "not_found"

                # 1. Try parse_yandex_link from map_url first (most accurate for Russian addresses)
                if store.get("map_url"):
                    try:
                        y_lat, y_lon = parse_yandex_link(store["map_url"])
                        # parse_yandex_link returns (None, None) on failure — check explicitly
                        if y_lat is not None and y_lon is not None:
                            lat, lon = y_lat, y_lon
                            status = "found"
                    except Exception:
                        pass

                # 2. address geocoding (only if URL parsing didn't yield coords)
                if lat is None and store.get("address"):
                    coords = geocode_address(store["address"])
                    if coords:
                        lat, lon = coords
                        status = "found"
                    if not YANDEX_GEOCODER_API_KEY:
                        time.sleep(1.1)

                cur2.execute(
                    "UPDATE stores SET lat=%s, lon=%s, geocode_status=%s WHERE id=%s AND owner_id=%s",
                    (lat, lon, status, sid, owner_id)
                )
                conn2.commit()
                cur2.close(); conn2.close()
                logger.info("geocode_pending: store %d → %s (%.4f, %.4f)", sid, status, lat or 0, lon or 0)
            except Exception as e:
                logger.warning("geocode_pending: store %d error: %s", sid, e)

    background_tasks.add_task(_geocode_all, store_ids, uid)
    return {"queued": len(store_ids)}


# ══════════════════════════════════════════════════════════════════════════════
# Daily orders (заявки на день)
# ══════════════════════════════════════════════════════════════════════════════

# Keyword patterns for auto-detecting which Excel column maps to which field.
# Order matters: more specific patterns first.
_ORDER_COLUMN_PATTERNS: dict = {
    "store_name": ["торговая точка", "наименование контрагента", "название точки", "название магазина",
                   "магазин", "контрагент", "точка доставки", "название", "точка", "клиент", "наименование", "name"],
    "address":    ["адрес доставки", "адрес точки", "адрес", "address"],
    "product":    ["наименование товара", "номенклатура", "товар", "продукт", "product", "позиция"],
    "quantity":   ["количество", "кол-во", "колво", "кол во", "quantity", "qty", "штук", "шт"],
    "yandex_url": ["ссылка яндекс", "яндекс карты", "яндекс", "yandex", "ссылка на карту", "ссылка"],
    "weight_kg":  ["вес, кг", "вес (кг)", "вес,кг", "вес кг", "вес", "weight", "масса, кг",
                   "масса кг", "масса", "кг", "kg"],
    "volume_m3":  ["объём, м3", "объём (м3)", "объем, м3", "объем (м3)", "объём м3",
                   "объем м3", "объём", "объем", "volume", "м3", "m3", "куб"],
    "amount_rub": ["сумма, руб", "сумма (руб)", "сумма руб", "сумма заказа", "стоимость", "сумма", "amount", "руб"],
    "order_number":["номер заявки", "№ заявки", "заявка №", "номер накладной",
                    "накладная", "заявка", "номер", "заказ", "order", "number", "№"],
    "zone":       ["зона доставки", "маршрут водителя", "водитель", "зона", "маршрут", "zone", "driver"],
    "notes":      ["примечание", "комментарий", "notes", "comment", "note"],
    "time_from":  ["время с", "время от", "открытие", "open_time", "time_from", "с ("],
    "time_to":    ["время до", "время по", "закрытие", "close_time", "time_to", "до ("],
    "unload_minutes": ["разгрузка, мин", "разгрузка (мин)", "разгрузка мин", "разгрузка", "unload", "выгрузка мин"],
    "city":       ["город", "city"],
}


def _detect_column_mapping(headers: list[str]) -> dict[str, Optional[str]]:
    """Return best-guess mapping: field_name → header_name (or None if not detected)."""
    headers_lower = [h.lower().strip() for h in headers]
    mapping: dict[str, Optional[str]] = {k: None for k in _ORDER_COLUMN_PATTERNS}
    used: set[str] = set()

    for field, patterns in _ORDER_COLUMN_PATTERNS.items():
        for pattern in patterns:
            for orig, norm in zip(headers, headers_lower):
                if orig in used:
                    continue
                if pattern in norm or norm in pattern:
                    mapping[field] = orig
                    used.add(orig)
                    break
            if mapping[field] is not None:
                break

    return mapping


def _normalize_name(s: str) -> str:
    """Lowercase, replace punctuation with spaces, collapse whitespace.

    Replaces (not removes) punctuation so that hyphenated names like
    "Магазин-Приморский" become "магазин приморский" rather than the
    concatenated "магазинприморский" which would never match anything.
    """
    s = s.lower()
    s = re.sub(r"[^\w\s]", " ", s)   # punctuation → space (keeps word boundaries)
    s = re.sub(r"\s+", " ", s)        # collapse multiple spaces
    return s.strip()


# Common words that are too generic to be useful as matching signals.
# A store name consisting only of stop-words (e.g. "Продукты Центр") will NOT
# be fuzzy-matched — only exact-name match will work for it.
_MATCH_STOP_WORDS: frozenset = frozenset({
    "магазин", "супермаркет", "маркет", "мини", "центр", "аптека",
    "рынок", "базар", "торговый", "торговая", "дом", "склад", "точка",
    "продукты", "универсам", "универмаг", "павильон", "киоск",
    "салон", "бутик", "ларек", "ларёк", "лавка", "отдел", "гипермаркет",
    "ип", "ооо", "зао", "ао",          # юр. форма — не уникальна
})


def _significant_words(name: str) -> frozenset:
    """Return meaningful tokens: lowercase, no punctuation, no stop-words, len ≥ 3."""
    tokens = _normalize_name(name).split()
    return frozenset(
        w for w in tokens
        if w not in _MATCH_STOP_WORDS and len(w) >= 3
    )


def _match_store_to_db(raw_name: str, raw_address: str, db_stores: list[dict]) -> Optional[dict]:
    """Return the best matching store dict or None.

    A delivery point is identified by (name + address). The SAME chain name at a
    DIFFERENT address is a DIFFERENT delivery point and must NOT be merged — this
    is the core correctness rule for routing (each branch is its own stop).

    Passes:

    Pass 1 — Exact (name + address).
        Strongest signal. Always wins.

    Pass 2 — Exact name, address resolved.
        • No address in the file → match a same-name store only when it is the
          single same-name candidate (otherwise ambiguous → None).
        • Address in the file → match only a same-name store whose address is
          empty (catalog without address yet) or equal. Same name + different,
          non-empty address ⇒ NOT a match (different branch).

    Pass 3 — Jaccard ≥ 0.85 on significant words, with an address guard.
        Fuzzy name only when there is NO exact name match. If both file and
        candidate have a (non-empty) address and they differ, the candidate is
        rejected — prevents fuzzy-merging two branches of the same chain.

    Deliberately REMOVED (historic false positives): substring match and
    word-overlap ≥ 50%.
    """
    if not raw_name or not db_stores:
        return None

    norm = _normalize_name(raw_name)
    norm_addr = _normalize_for_dedup(raw_address) if raw_address else ""

    # ── Pass 1: exact name + address ─────────────────────────────────────────
    if norm_addr:
        for s in db_stores:
            if _normalize_name(s["name"]) != norm:
                continue
            db_addr = s.get("address") or ""
            db_addr_norm = _normalize_for_dedup(db_addr)
            if db_addr_norm == norm_addr:
                return s
            # City-prefix tolerance: bulk-create prepends city to the address
            # (e.g. "Москва, ул. Гагарина, 24" in DB vs "ул. Гагарина, 24"
            # in the order).  Strip the first comma-segment and retry.
            if "," in db_addr:
                db_addr_stripped = _normalize_for_dedup(db_addr.split(",", 1)[1])
                if db_addr_stripped == norm_addr:
                    return s

    # ── Pass 2: exact name, resolve by address ──────────────────────────────
    name_matches = [s for s in db_stores if _normalize_name(s["name"]) == norm]
    if name_matches:
        if not norm_addr:
            # No address to disambiguate: only safe if a single same-name store.
            return name_matches[0] if len(name_matches) == 1 else None
        # Address present but no exact (name+address) hit above.
        # 1. Accept a same-name store whose DB address is city-prefixed but
        #    otherwise equals the order address (Pass 1 handles the full
        #    comparison; here we recheck the city-stripped form for any
        #    same-name store that didn't appear in Pass 1).
        city_strip_matches = []
        for s in name_matches:
            db_addr = s.get("address") or ""
            if not db_addr:
                continue
            if "," in db_addr:
                db_addr_stripped = _normalize_for_dedup(db_addr.split(",", 1)[1])
                if db_addr_stripped == norm_addr:
                    city_strip_matches.append(s)
        if city_strip_matches:
            return city_strip_matches[0]
        # 2. Accept a same-name store that has no address yet.
        addr_less = [s for s in name_matches
                     if not _normalize_for_dedup(s.get("address") or "")]
        return addr_less[0] if addr_less else None

    # ── Pass 3: Jaccard ≥ 0.85 on significant words (address-guarded) ────────
    raw_sig = _significant_words(raw_name)
    if not raw_sig:
        # Only stop-words in the name → fuzzy matching too unreliable
        return None

    candidates: list[tuple[float, dict]] = []
    for s in db_stores:
        db_sig = _significant_words(s["name"])
        if not db_sig:
            continue
        shared = len(raw_sig & db_sig)
        if shared == 0:
            continue
        jaccard = shared / len(raw_sig | db_sig)
        if jaccard < 0.85:
            continue
        # Address guard: both sides have a (different) address → different point.
        s_addr = _normalize_for_dedup(s.get("address") or "")
        if norm_addr and s_addr and norm_addr != s_addr:
            continue
        candidates.append((jaccard, s))

    if not candidates:
        return None

    # Sort descending by score
    candidates.sort(key=lambda x: -x[0])

    # If top score is shared by multiple stores → ambiguous, refuse to guess
    if len(candidates) >= 2 and candidates[0][0] == candidates[1][0]:
        return None

    return candidates[0][1]


def _safe_float(val) -> float:
    """Convert any cell value to float, returning 0 on failure."""
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(",", ".").replace(" ", "").replace("\u00a0", "").strip())
    except (ValueError, TypeError):
        return 0.0


@app.get("/api/orders/template")
def download_orders_template():
    """Excel template for daily orders (заявки на день). Mirrors stores template UX."""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = [
        "Магазин",          # A — required, matched against catalog
        "Номер заявки",     # B — optional
        "Вес, кг",          # C — optional
        "Объём, м3",        # D — optional
        "Сумма, руб",       # E — optional
        "Комментарий",      # F — optional
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.append(["Супермаркет Каспий", "ЗАК-001", 120, 0.8, 45000, "Хрупкий груз"])
    ws.append(["Магазин Горный", "ЗАК-002", 60, 0.4, 18000, ""])

    col_widths = [28, 16, 12, 14, 14, 32]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    note_row = [
        "← Название магазина (как в разделе «Магазины»)",
        "← Номер заявки (необязательно)",
        "← Вес груза в кг (число)",
        "← Объём в м3 (число)",
        "← Сумма заказа в руб (число)",
        "← Комментарий для водителя (необязательно)",
    ]
    ws.append(note_row)
    for col_num in range(1, len(note_row) + 1):
        cell = ws.cell(row=ws.max_row, column=col_num)
        cell.font = Font(italic=True, color="888888")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    content = buf.read()

    import base64
    return {
        "data": base64.b64encode(content).decode("ascii"),
        "filename": "smartroute_orders_template.xlsx",
    }


class OrderImportRow(BaseModel):
    store_id: Optional[int] = None
    store_name_raw: str
    address_raw: str = ""
    order_number: str = ""
    weight_kg: float = 0.0
    volume_m3: float = 0.0
    amount_rub: float = 0.0
    quantity: float = 0.0
    products: str = ""
    notes: str = ""


class OrderImportRequest(BaseModel):
    delivery_date: str   # "YYYY-MM-DD"
    rows: list[OrderImportRow]
    clear_existing: bool = True  # replace today's orders on re-import
    filename: str = ""   # original Excel filename for history


@app.post("/api/orders/preview")
async def orders_preview(request: Request, file: UploadFile = File(...), mapping: Optional[str] = Form(None)):
    """
    Parse an Excel file and return:
    - detected column headers
    - auto-detected field mapping
    - first 200 rows as raw data
    - per-row store match results against caller's store base
    """
    uid = get_user_id(request)
    _api_rate_limit(f"orders_preview:{uid}", max_calls=20, window_seconds=60)

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Файл слишком большой (макс. 20 МБ)")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Не удалось открыть Excel файл: {e}")

    if not rows_raw:
        raise HTTPException(status_code=422, detail="Файл пустой")

    # Find header row: first row with ≥ 2 non-empty string cells
    header_row_idx = 0
    headers: list[str] = []
    for i, row in enumerate(rows_raw[:10]):
        str_cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if len(str_cells) >= 2:
            header_row_idx = i
            headers = [str(c).strip() if c is not None else f"Колонка {j+1}"
                       for j, c in enumerate(row)]
            break

    if not headers:
        raise HTTPException(status_code=422, detail="Не удалось найти строку заголовков в файле")

    # Remove entirely empty trailing columns
    while headers and headers[-1].startswith("Колонка "):
        headers.pop()

    detected_mapping = _detect_column_mapping(headers)

    # Apply user mapping override (sent when the dispatcher corrects an
    # auto-detected column). Only known fields + headers that exist are honored.
    if mapping:
        try:
            override = json.loads(mapping)
        except (ValueError, TypeError):
            override = None
        if isinstance(override, dict):
            for field, col in override.items():
                if field not in detected_mapping:
                    continue
                if col is None or col == "":
                    detected_mapping[field] = None
                elif col in headers:
                    detected_mapping[field] = col

    # Fetch owner's stores for matching
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, name, address FROM stores WHERE owner_id = %s", (uid,))
    db_stores = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()

    # Resolve the column index for each detected field once.
    def _col_idx(field: str) -> Optional[int]:
        h = detected_mapping.get(field)
        return headers.index(h) if h and h in headers else None

    idx = {f: _col_idx(f) for f in (
        "store_name", "address", "product", "quantity", "weight_kg",
        "volume_m3", "amount_rub", "order_number", "notes", "city",
        "yandex_url", "time_from", "time_to", "unload_minutes",
    )}

    def _cell(cells: list, field: str) -> str:
        i = idx.get(field)
        if i is None or i >= len(cells):
            return ""
        v = cells[i]
        return str(v).strip() if v is not None else ""

    # Parse ALL data rows and AGGREGATE by (name + address): a 1C-style file has
    # one product per row, so several rows describe ONE delivery point. Each
    # unique (name+address) collapses into a single point with summed
    # weight/volume/amount/quantity and a concatenated product list.
    MAX_DATA_ROWS = 5000
    data_rows = rows_raw[header_row_idx + 1:]
    file_total_rows = len(data_rows)          # raw count before cap
    truncated = file_total_rows > MAX_DATA_ROWS
    points: dict[tuple, dict] = {}
    order_seq: list[tuple] = []
    sample_rows: list[tuple] = []   # (cells, key) for the raw preview table
    total_data_rows = 0

    for row in data_rows[:MAX_DATA_ROWS]:
        cells = list(row) + [None] * max(0, len(headers) - len(row))
        name = _cell(cells, "store_name")
        if not name or name.lower() in ("none", "nan"):
            continue
        total_data_rows += 1
        address = _cell(cells, "address")
        key = (_normalize_for_dedup(name), _normalize_for_dedup(address))

        pt = points.get(key)
        if pt is None:
            pt = {
                "name": name, "address": address,
                "weight_kg": 0.0, "volume_m3": 0.0, "amount_rub": 0.0,
                "quantity": 0.0, "_products": [],
                "order_number": "", "notes": "", "city": "",
                "yandex_url": "", "time_from": "", "time_to": "",
                "unload_minutes": "", "order_lines": 0,
            }
            points[key] = pt
            order_seq.append(key)

        pt["weight_kg"] += _safe_float(_cell(cells, "weight_kg"))
        pt["volume_m3"] += _safe_float(_cell(cells, "volume_m3"))
        pt["amount_rub"] += _safe_float(_cell(cells, "amount_rub"))
        qv = _safe_float(_cell(cells, "quantity"))
        pt["quantity"] += qv
        prod = _cell(cells, "product")
        if prod:
            pt["_products"].append((prod, qv))
        for field in ("order_number", "notes", "city", "yandex_url",
                      "time_from", "time_to", "unload_minutes"):
            if not pt[field]:
                v = _cell(cells, field)
                if v:
                    pt[field] = v
        pt["order_lines"] += 1

        if len(sample_rows) < 50:
            sample_rows.append((cells, key))

    def _products_str(items: list) -> str:
        """'Молоко×4, Сахар×16' — merge duplicate products, keep first-seen order."""
        order, totals = [], {}
        for prod, q in items:
            if prod not in totals:
                totals[prod] = 0.0
                order.append(prod)
            totals[prod] += q
        out = []
        for prod in order:
            q = totals[prod]
            if q and q > 0:
                qstr = str(int(q)) if float(q).is_integer() else f"{q:g}"
                out.append(f"{prod}×{qstr}")
            else:
                out.append(prod)
        return ", ".join(out)[:500]

    # Match each aggregated point against the catalog (name + address aware).
    points_out: list[dict] = []
    match_by_key: dict[tuple, tuple] = {}
    matched_points = 0
    for key in order_seq:
        pt = points[key]
        m = _match_store_to_db(pt["name"], pt["address"], db_stores)
        mid = m["id"] if m else None
        mname = m["name"] if m else None
        if mid is not None:
            matched_points += 1
        match_by_key[key] = (mid, mname)
        points_out.append({
            "name": pt["name"], "address": pt["address"],
            "matched_store_id": mid, "matched_store_name": mname,
            "weight_kg": round(pt["weight_kg"], 3),
            "volume_m3": round(pt["volume_m3"], 4),
            "amount_rub": round(pt["amount_rub"], 2),
            "quantity": round(pt["quantity"], 3),
            "products": _products_str(pt["_products"]),
            "order_number": pt["order_number"][:100],
            "notes": pt["notes"][:500],
            "city": pt["city"], "yandex_url": pt["yandex_url"],
            "time_from": pt["time_from"], "time_to": pt["time_to"],
            "unload_minutes": pt["unload_minutes"],
            "order_lines": pt["order_lines"],
        })

    # Raw sample rows for the preview table, highlighted by their point's match.
    preview_rows = []
    for cells, key in sample_rows:
        row_dict = {headers[i]: cells[i] for i in range(len(headers))}
        mid, mname = match_by_key.get(key, (None, None))
        preview_rows.append({
            "cells": {k: (str(v) if v is not None else "") for k, v in row_dict.items()},
            "matched_store_id": mid,
            "matched_store_name": mname,
        })

    total_points = len(points_out)
    return {
        "headers": headers,
        "detected_mapping": detected_mapping,
        "rows": preview_rows,            # raw sample (≤50) for the preview table
        "points": points_out,            # aggregated delivery points (one per name+address)
        "total_points": total_points,
        "matched_points": matched_points,
        "unmatched_points": total_points - matched_points,
        "total_rows": total_data_rows,   # raw data lines parsed (after cap)
        "file_total_rows": file_total_rows,  # actual lines in the file (pre-cap)
        "truncated": truncated,          # True when file > MAX_DATA_ROWS
        # Backward-compatible aliases — now point-based (the correct semantics).
        "matched_stores": matched_points,
        "unmatched_stores": total_points - matched_points,
        "db_stores_count": len(db_stores),
    }


@app.post("/api/orders/import", status_code=201)
def orders_import(request: Request, body: OrderImportRequest):
    """Save confirmed orders for a delivery date."""
    uid = get_user_id(request)

    # Validate date format
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", body.delivery_date):
        raise HTTPException(status_code=422, detail="delivery_date должен быть в формате YYYY-MM-DD")

    if not body.rows:
        raise HTTPException(status_code=422, detail="Список заявок пуст")

    if len(body.rows) > 2000:
        raise HTTPException(status_code=422, detail="Слишком много заявок в одном импорте (максимум 2000 строк)")

    # Validate and cap
    for row in body.rows:
        row.weight_kg = max(0.0, row.weight_kg)
        row.volume_m3 = max(0.0, row.volume_m3)
        row.amount_rub = max(0.0, row.amount_rub)
        row.quantity = max(0.0, row.quantity)
        row.store_name_raw = (row.store_name_raw or "").strip()[:200]
        row.address_raw = (row.address_raw or "").strip()[:300]
        row.order_number = (row.order_number or "").strip()[:100]
        row.products = (row.products or "").strip()[:500]
        row.notes = (row.notes or "").strip()[:500]

    conn = get_db()
    cur = conn.cursor()
    try:
        if body.clear_existing:
            cur.execute(
                "DELETE FROM daily_orders WHERE owner_id = %s AND delivery_date = %s",
                (uid, body.delivery_date)
            )

        for row in body.rows:
            # Verify store belongs to this owner (if store_id provided)
            store_id = None
            if row.store_id is not None:
                cur.execute(
                    "SELECT id FROM stores WHERE id = %s AND owner_id = %s",
                    (row.store_id, uid)
                )
                if cur.fetchone():
                    store_id = row.store_id

            cur.execute(
                """INSERT INTO daily_orders
                   (owner_id, store_id, store_name_raw, address_raw, order_number,
                    weight_kg, volume_m3, amount_rub, quantity, products, notes,
                    delivery_date)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (uid, store_id, row.store_name_raw, row.address_raw, row.order_number,
                 row.weight_kg, row.volume_m3, row.amount_rub, row.quantity,
                 row.products, row.notes, body.delivery_date)
            )

        conn.commit()
    except Exception:
        conn.rollback()
        cur.close()
        conn.close()
        raise HTTPException(status_code=500, detail="Ошибка при сохранении заявок")

    # Return summary
    cur.execute(
        """SELECT COUNT(*) as cnt,
                  COALESCE(SUM(weight_kg),0) as total_weight,
                  COALESCE(SUM(volume_m3),0) as total_volume,
                  COALESCE(SUM(amount_rub),0) as total_amount,
                  COUNT(CASE WHEN store_id IS NOT NULL THEN 1 END) as matched_cnt,
                  COUNT(CASE WHEN store_id IS NULL THEN 1 END) as unmatched_cnt
             FROM daily_orders
            WHERE owner_id = %s AND delivery_date = %s""",
        (uid, body.delivery_date)
    )
    row = cur.fetchone()

    # Determine whether any weight / volume data was provided
    _has_weight = any(r.weight_kg > 0 for r in body.rows)
    _has_volume = any(r.volume_m3 > 0 for r in body.rows)
    _total_w = float(row[1])   # total_weight from the SELECT above
    _total_v = float(row[2])   # total_volume
    _total_a = float(row[3])   # total_amount

    # Save import history record
    try:
        cur.execute(
            """INSERT INTO order_import_history
               (owner_id, delivery_date, filename, total_rows, matched_rows, unmatched_rows,
                has_weight, total_weight_kg, total_volume_m3, total_amount_rub)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (uid, body.delivery_date, body.filename[:200] if body.filename else "",
             row[0], row[4], row[5], _has_weight, _total_w, _total_v, _total_a)
        )
        conn.commit()
    except Exception as _he:
        logger.warning("Failed to save import history: %s", _he)

    cur.close()
    conn.close()

    return {
        "delivery_date": body.delivery_date,
        "saved_count": row[0],
        "total_weight_kg": round(row[1], 2),
        "total_volume_m3": round(row[2], 3),
        "total_amount_rub": round(row[3], 2),
        "matched_count": row[4],
        "unmatched_count": row[5],
        "has_weight": _has_weight,
        "has_volume": _has_volume,
    }


@app.get("/api/orders")
def get_orders(request: Request, date: Optional[str] = None):
    """Return daily orders for a date (default: today). Joined with stores for display."""
    uid = get_user_id(request)
    target_date = date if date else str(datetime.now().date())

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """SELECT o.id, o.store_id, o.store_name_raw, o.address_raw, o.order_number,
                  o.weight_kg, o.volume_m3, o.amount_rub, o.quantity, o.products, o.notes,
                  o.delivery_date::text as delivery_date,
                  s.name as store_name_db, s.address as store_address
             FROM daily_orders o
             LEFT JOIN stores s ON s.id = o.store_id
            WHERE o.owner_id = %s AND o.delivery_date = %s
            ORDER BY o.id""",
        (uid, target_date)
    )
    rows = [dict(r) for r in cur.fetchall()]

    # Aggregate summary
    cur.execute(
        """SELECT COUNT(*) as cnt,
                  COALESCE(SUM(weight_kg),0) as total_weight,
                  COALESCE(SUM(volume_m3),0) as total_volume,
                  COALESCE(SUM(amount_rub),0) as total_amount
             FROM daily_orders
            WHERE owner_id = %s AND delivery_date = %s""",
        (uid, target_date)
    )
    summary = cur.fetchone()
    cur.close()
    conn.close()

    return {
        "delivery_date": target_date,
        "orders": rows,
        "total_count": summary["cnt"],
        "total_weight_kg": round(float(summary["total_weight"]), 2),
        "total_volume_m3": round(float(summary["total_volume"]), 3),
        "total_amount_rub": round(float(summary["total_amount"]), 2),
    }


@app.get("/api/orders/active-dates")
def get_orders_active_dates(request: Request):
    """Return list of dates (YYYY-MM-DD) that have at least one order for the current user."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """SELECT DISTINCT delivery_date::text
             FROM daily_orders
            WHERE owner_id = %s
            ORDER BY delivery_date""",
        (uid,)
    )
    dates = [row[0] for row in cur.fetchall()]
    cur.close()
    conn.close()
    return {"dates": dates}


@app.get("/api/orders/import-history")
def get_import_history(request: Request, limit: int = Query(50, ge=1, le=200)):
    """Return the last N import history records for the current user."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """SELECT id, delivery_date::text, filename, total_rows, matched_rows, unmatched_rows,
                  imported_at::text as imported_at,
                  COALESCE(has_weight, TRUE) as has_weight,
                  COALESCE(total_weight_kg, 0) as total_weight_kg,
                  COALESCE(total_volume_m3, 0) as total_volume_m3,
                  COALESCE(total_amount_rub, 0) as total_amount_rub
             FROM order_import_history
            WHERE owner_id = %s
            ORDER BY imported_at DESC
            LIMIT %s""",
        (uid, limit)
    )
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return {"imports": rows}


@app.get("/api/orders/import-history/{record_id}/details")
def get_import_history_details(record_id: int, request: Request):
    """Return summary + orders list for a specific import history record."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Fetch the import record to get delivery_date
    cur.execute(
        """SELECT id, delivery_date::text, filename, total_rows, matched_rows, unmatched_rows,
                  imported_at::text as imported_at,
                  COALESCE(total_weight_kg, 0) as total_weight_kg,
                  COALESCE(total_volume_m3, 0) as total_volume_m3,
                  COALESCE(total_amount_rub, 0) as total_amount_rub
             FROM order_import_history
            WHERE id = %s AND owner_id = %s""",
        (record_id, uid)
    )
    rec = cur.fetchone()
    if not rec:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Import history record not found")
    rec = dict(rec)

    # Fetch orders for this delivery_date
    cur.execute(
        """SELECT o.id, o.store_id, o.store_name_raw, o.order_number,
                  o.weight_kg, o.volume_m3, o.amount_rub, o.notes,
                  s.name as store_name_db, s.address as store_address
             FROM daily_orders o
             LEFT JOIN stores s ON s.id = o.store_id AND s.owner_id = %s
            WHERE o.owner_id = %s AND o.delivery_date = %s
            ORDER BY o.id
            LIMIT 500""",
        (uid, uid, rec["delivery_date"])
    )
    orders = [dict(r) for r in cur.fetchall()]

    # Unmatched stores: group by raw name where store_id IS NULL
    cur.execute(
        """SELECT store_name_raw, COUNT(*) as cnt,
                  COALESCE(SUM(weight_kg),0) as weight_kg,
                  COALESCE(SUM(volume_m3),0) as volume_m3
             FROM daily_orders
            WHERE owner_id = %s AND delivery_date = %s AND store_id IS NULL
            GROUP BY store_name_raw
            ORDER BY store_name_raw""",
        (uid, rec["delivery_date"])
    )
    unmatched = [dict(r) for r in cur.fetchall()]

    cur.close()
    conn.close()
    return {
        "record": rec,
        "orders": orders,
        "unmatched_stores": unmatched,
    }


@app.delete("/api/orders/import-history/{record_id}", status_code=204)
def delete_import_history_record(record_id: int, request: Request):
    """Delete a single import history record."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM order_import_history WHERE id = %s AND owner_id = %s",
        (record_id, uid)
    )
    deleted = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    if deleted == 0:
        raise HTTPException(status_code=404, detail="Record not found")


@app.delete("/api/orders/import-history", status_code=200)
def clear_import_history(request: Request):
    """Delete all import history records for the current user."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM order_import_history WHERE owner_id = %s", (uid,))
    deleted = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    return {"deleted": deleted}


@app.delete("/api/orders")
def delete_orders(request: Request, date: Optional[str] = None):
    """Delete all orders for a date (default: today)."""
    uid = get_user_id(request)
    target_date = date if date else str(datetime.now().date())

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM daily_orders WHERE owner_id = %s AND delivery_date = %s",
        (uid, target_date)
    )
    deleted = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    return {"deleted": deleted, "delivery_date": target_date}


class ManualOrderRequest(BaseModel):
    store_id: int
    delivery_date: str  # "YYYY-MM-DD"
    weight_kg: float = 0.0
    volume_m3: float = 0.0
    amount_rub: float = 0.0
    notes: str = ""
    order_number: str = ""
    products: str = ""


class ManualOrderUpdate(BaseModel):
    weight_kg: Optional[float] = None
    volume_m3: Optional[float] = None
    amount_rub: Optional[float] = None
    notes: Optional[str] = None
    order_number: Optional[str] = None
    products: Optional[str] = None


class ManualOrderBulkRequest(BaseModel):
    store_ids: list[int]
    delivery_date: str  # "YYYY-MM-DD"


@app.post("/api/orders/manual/bulk", status_code=201)
def create_manual_orders_bulk(request: Request, body: ManualOrderBulkRequest):
    """Create multiple daily order rows at once (skip duplicates). Returns created + skipped counts."""
    uid = get_user_id(request)
    if not body.store_ids:
        return {"created": [], "skipped": []}
    try:
        datetime.strptime(body.delivery_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=422, detail="Некорректный формат даты (ожидается YYYY-MM-DD)")

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # Fetch all requested stores in one query (ownership check included)
        cur.execute(
            "SELECT id, name FROM stores WHERE id = ANY(%s) AND owner_id = %s",
            (body.store_ids, uid)
        )
        stores_map = {r["id"]: r["name"] for r in cur.fetchall()}

        # Fetch already-existing orders for this date (to detect duplicates)
        cur.execute(
            "SELECT store_id FROM daily_orders WHERE owner_id = %s AND delivery_date = %s AND store_id = ANY(%s)",
            (uid, body.delivery_date, body.store_ids)
        )
        already_exists = {r["store_id"] for r in cur.fetchall()}

        created = []
        skipped = []
        for sid in body.store_ids:
            if sid not in stores_map:
                skipped.append({"store_id": sid, "reason": "not_found"})
                continue
            if sid in already_exists:
                skipped.append({"store_id": sid, "name": stores_map[sid], "reason": "duplicate"})
                continue
            cur.execute(
                """INSERT INTO daily_orders
                       (owner_id, store_id, store_name_raw, delivery_date,
                        weight_kg, volume_m3, amount_rub, notes, products, order_number)
                   VALUES (%s, %s, %s, %s, 0, 0, 0, '', '', '')
                   RETURNING id, store_id, store_name_raw, delivery_date::text""",
                (uid, sid, stores_map[sid], body.delivery_date)
            )
            created.append(dict(cur.fetchone()))

        conn.commit()
        return {"created": created, "skipped": skipped}
    finally:
        cur.close()
        conn.close()


@app.post("/api/orders/manual", status_code=201)
def create_manual_order(request: Request, body: ManualOrderRequest):
    """Create a single daily order row manually (no Excel). Returns the created order."""
    uid = get_user_id(request)

    # Validate numerics
    if body.weight_kg < 0:
        raise HTTPException(status_code=422, detail="Вес не может быть отрицательным")
    if body.volume_m3 < 0:
        raise HTTPException(status_code=422, detail="Объём не может быть отрицательным")
    if body.amount_rub < 0:
        raise HTTPException(status_code=422, detail="Сумма не может быть отрицательной")

    # Validate date
    try:
        datetime.strptime(body.delivery_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=422, detail="Некорректный формат даты (ожидается YYYY-MM-DD)")

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Verify store belongs to this user
    cur.execute("SELECT id, name FROM stores WHERE id = %s AND owner_id = %s", (body.store_id, uid))
    store = cur.fetchone()
    if not store:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Магазин не найден")

    # Duplicate check: same store_id + same date
    cur.execute(
        "SELECT id FROM daily_orders WHERE owner_id = %s AND store_id = %s AND delivery_date = %s",
        (uid, body.store_id, body.delivery_date)
    )
    if cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(
            status_code=409,
            detail=f"Магазин «{store['name']}» уже добавлен на {body.delivery_date}"
        )

    cur.execute(
        """INSERT INTO daily_orders
               (owner_id, store_id, store_name_raw, order_number,
                weight_kg, volume_m3, amount_rub, notes, products, delivery_date)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
           RETURNING id, store_id, store_name_raw, order_number,
                     weight_kg, volume_m3, amount_rub, notes, products,
                     delivery_date::text, created_at::text""",
        (uid, body.store_id, store["name"], body.order_number,
         max(0.0, body.weight_kg), max(0.0, body.volume_m3),
         max(0.0, body.amount_rub), body.notes, body.products or "", body.delivery_date)
    )
    row = dict(cur.fetchone())
    conn.commit()
    cur.close()
    conn.close()
    return row


@app.put("/api/orders/{order_id}")
def update_manual_order(request: Request, order_id: int, body: ManualOrderUpdate):
    """Update weight/volume/amount/notes/order_number of a daily order row."""
    uid = get_user_id(request)

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute(
        "SELECT id FROM daily_orders WHERE id = %s AND owner_id = %s",
        (order_id, uid)
    )
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Заявка не найдена")

    # Build dynamic SET clause from non-None fields
    fields, values = [], []
    if body.weight_kg is not None:
        if body.weight_kg < 0:
            cur.close(); conn.close()
            raise HTTPException(status_code=422, detail="Вес не может быть отрицательным")
        fields.append("weight_kg = %s"); values.append(max(0.0, body.weight_kg))
    if body.volume_m3 is not None:
        if body.volume_m3 < 0:
            cur.close(); conn.close()
            raise HTTPException(status_code=422, detail="Объём не может быть отрицательным")
        fields.append("volume_m3 = %s"); values.append(max(0.0, body.volume_m3))
    if body.amount_rub is not None:
        if body.amount_rub < 0:
            cur.close(); conn.close()
            raise HTTPException(status_code=422, detail="Сумма не может быть отрицательной")
        fields.append("amount_rub = %s"); values.append(max(0.0, body.amount_rub))
    if body.notes is not None:
        fields.append("notes = %s"); values.append(body.notes)
    if body.order_number is not None:
        fields.append("order_number = %s"); values.append(body.order_number)
    if body.products is not None:
        fields.append("products = %s"); values.append(body.products)

    if not fields:
        cur.close(); conn.close()
        raise HTTPException(status_code=422, detail="Нет полей для обновления")

    values.append(order_id)
    cur.execute(
        f"UPDATE daily_orders SET {', '.join(fields)} WHERE id = %s"
        " RETURNING id, store_id, store_name_raw, order_number,"
        " weight_kg, volume_m3, amount_rub, notes, products, delivery_date::text",
        values
    )
    row = dict(cur.fetchone())
    conn.commit()
    cur.close()
    conn.close()
    return row


@app.delete("/api/orders/{order_id}", status_code=204)
def delete_manual_order(request: Request, order_id: int):
    """Delete a single daily order row by id."""
    uid = get_user_id(request)

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM daily_orders WHERE id = %s AND owner_id = %s",
        (order_id, uid)
    )
    deleted = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    if deleted == 0:
        raise HTTPException(status_code=404, detail="Заявка не найдена")


@app.post("/api/orders/rematch")
def rematch_orders(request: Request, date: Optional[str] = None):
    """
    Re-run store matching for daily_orders with store_id=NULL.
    Called after new stores are created (e.g. bulk create from unmatched list).
    Returns count of newly matched orders.
    """
    uid = get_user_id(request)
    target_date = date if date else str(datetime.now().date())

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Get unmatched orders for this date
    cur.execute(
        "SELECT id, store_name_raw, address_raw FROM daily_orders WHERE owner_id = %s AND delivery_date = %s AND store_id IS NULL",
        (uid, target_date)
    )
    unmatched = [dict(r) for r in cur.fetchall()]

    if not unmatched:
        cur.close()
        conn.close()
        return {"matched_count": 0, "still_unmatched": 0}

    # Get current stores (fresh, after new stores were created)
    cur.execute("SELECT id, name, address FROM stores WHERE owner_id = %s", (uid,))
    db_stores = [dict(r) for r in cur.fetchall()]

    matched_count = 0
    cur2 = conn.cursor()
    for order in unmatched:
        match = _match_store_to_db(order["store_name_raw"], order.get("address_raw") or "", db_stores)
        if match:
            cur2.execute(
                "UPDATE daily_orders SET store_id = %s WHERE id = %s AND owner_id = %s",
                (match["id"], order["id"], uid)
            )
            matched_count += 1

    conn.commit()
    cur.close()
    cur2.close()
    conn.close()

    return {
        "matched_count": matched_count,
        "still_unmatched": len(unmatched) - matched_count,
    }


@app.post("/api/route/build")
def build_route(request: Request, body: RouteRequest):
    uid = get_user_id(request)
    _api_rate_limit(f"vrp:{uid}", max_calls=3, window_seconds=120)
    if not body.store_ids:
        raise HTTPException(status_code=400, detail="No stores selected")
    if not body.vehicles:
        raise HTTPException(status_code=400, detail="No vehicles provided")

    # Validate depot coordinates
    if body.depot_lat is not None and not (-90 <= body.depot_lat <= 90):
        raise HTTPException(status_code=422, detail="Широта склада должна быть от -90 до 90")
    if body.depot_lon is not None and not (-180 <= body.depot_lon <= 180):
        raise HTTPException(status_code=422, detail="Долгота склада должна быть от -180 до 180")

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    placeholders = ",".join(["%s"] * len(body.store_ids))
    # Filter by owner_id to prevent cross-user store access
    cur.execute(
        f"SELECT * FROM stores WHERE id IN ({placeholders}) AND owner_id = %s",
        (*body.store_ids, uid)
    )
    stores_rows = {r["id"]: r for r in cur.fetchall()}
    cur.close()
    conn.close()

    # Depot coordinates — required; no city-specific fallback
    if not body.depot_lat or not body.depot_lon:
        raise HTTPException(
            status_code=400,
            detail="Укажите адрес склада (депо) на странице построения маршрута перед запуском оптимизации."
        )
    depot_lat = body.depot_lat
    depot_lon = body.depot_lon

    # Build coordinate list: depot first, then stores (preserve input order for savings baseline)
    store_list = [stores_rows[sid] for sid in body.store_ids if sid in stores_rows and stores_rows[sid]["lat"]]
    if not store_list:
        raise HTTPException(status_code=400, detail="No geocoded stores found")

    num_vehicles = len(body.vehicles)

    # ── Vehicle count validation ──────────────────────────────────────────────
    if num_vehicles > len(store_list):
        raise HTTPException(
            status_code=422,
            detail=(
                f"Количество машин ({num_vehicles}) превышает количество магазинов "
                f"({len(store_list)}). Уменьшите число машин до {len(store_list)} или меньше."
            ),
        )

    all_coords = [(depot_lat, depot_lon)] + [(s["lat"], s["lon"]) for s in store_list]

    # ── Daily orders: look up weight / volume per store ───────────────────────
    # Loaded for the requested delivery_date (falls back to today when not set).
    # Used both for VRP capacity demands (when vehicle capacity_kg is set) and
    # for annotating result stores with actual load data.
    _orders_date = body.delivery_date if body.delivery_date else str(date.today())
    _store_weights: dict = {}
    _store_volumes: dict = {}
    try:
        _conn_w = get_db()
        _cur_w = _conn_w.cursor()
        _cur_w.execute(
            """SELECT store_id, COALESCE(SUM(weight_kg),0), COALESCE(SUM(volume_m3),0)
                 FROM daily_orders
                WHERE owner_id = %s AND delivery_date = %s AND store_id IS NOT NULL
                GROUP BY store_id""",
            (uid, _orders_date)
        )
        for _r in _cur_w.fetchall():
            _store_weights[int(_r[0])] = float(_r[1])
            _store_volumes[int(_r[0])] = float(_r[2])
        _cur_w.close()
        _conn_w.close()
        if _store_weights:
            logger.info(
                "build_route: loaded daily_orders weights for %d stores (max=%.1f kg)",
                len(_store_weights), max(_store_weights.values())
            )
    except Exception as _we:
        logger.warning("build_route: daily_orders weight lookup failed: %s", _we)

    capacities = None
    demands = None
    if any(v.capacity_kg for v in body.vehicles):
        capacities = [int(v.capacity_kg) if v.capacity_kg else 99999 for v in body.vehicles]
        if _store_weights:
            # Use actual weights as OR-Tools integer demands.
            # Scale down if values are very large (OR-Tools prefers smaller integers).
            _max_w = max(_store_weights.values(), default=1.0) or 1.0
            _fallback_w = _max_w / max(len(store_list), 1)
            _scale = 10 if _max_w > 10000 else 1
            demands = [0] + [
                max(1, int(_store_weights.get(s["id"], _fallback_w) / _scale))
                for s in store_list
            ]
            if _scale > 1:
                capacities = [max(1, int(c / _scale)) for c in capacities]
            logger.info("build_route: weight-based demands (scale=%d, max_demand=%d)",
                        _scale, max(demands[1:], default=1))
        else:
            demands = [0] + [1] * len(store_list)  # unit demands — no weight data today

    # ── Volume (м³) capacity pipeline ────────────────────────────────────────
    # Mirrors kg pipeline: build capacities_m3 / demands_m3 only when any
    # vehicle has capacity_m3 set.  demands_m3 are floats (no integer scaling
    # needed since m³ values are naturally small).
    capacities_m3 = None
    demands_m3 = None
    if any(v.capacity_m3 for v in body.vehicles):
        capacities_m3 = [float(v.capacity_m3) if v.capacity_m3 else 1e9
                         for v in body.vehicles]
        if _store_volumes:
            demands_m3 = [0.0] + [
                max(0.0, float(_store_volumes.get(s["id"], 0.0)))
                for s in store_list
            ]
            logger.info(
                "build_route: m³-based demands (max_demand=%.3f m³, vehicles=%s)",
                max(demands_m3[1:], default=0.0), capacities_m3,
            )
        else:
            demands_m3 = [0.0] + [0.0] * len(store_list)

    # ── Weight & volume capacity warnings ────────────────────────────────────
    # (inserted here — after demands/capacities and demands_m3/capacities_m3 are known)
    _any_weight_in_orders = any(v > 0 for v in _store_weights.values()) if _store_weights else False
    _any_volume_in_orders = any(v > 0 for v in _store_volumes.values()) if _store_volumes else False
    route_warnings: list[str] = []   # non-fatal issues surfaced to the frontend

    # Weight warnings
    if capacities is None and _any_weight_in_orders:
        # Weight data in orders but no vehicle kg limit set
        route_warnings.append(
            "Маршрут построен без учёта веса. "
            "В заявках есть данные по весу (кг), но ни одна машина не имеет ограничения "
            "по грузоподъёмности. Укажите «Грузоподъём. (кг)» в настройках транспорта, чтобы "
            "система учитывала вес при распределении маршрутов."
        )

    if capacities is not None and not _any_weight_in_orders:
        # Vehicle has kg limit set but orders have no weight data
        route_warnings.append(
            "Данные о весе не заполнены в заявках — "
            "ограничения по грузоподъёмности (кг) не применяются. "
            "Добавьте вес в заявки на доставку, чтобы система контролировала загрузку."
        )

    # Volume warnings
    if capacities_m3 is None and _any_volume_in_orders:
        # Volume data in orders but no vehicle m³ limit set
        route_warnings.append(
            "Маршрут построен без учёта объёма. "
            "В заявках есть данные по объёму (м³), но ни одна машина не имеет ограничения "
            "по объёму кузова. Укажите «Объём (м³)» в настройках транспорта, чтобы "
            "система учитывала объём при распределении маршрутов."
        )

    if capacities_m3 is not None and not _any_volume_in_orders:
        # Vehicle has m³ limit set but orders have no volume data
        route_warnings.append(
            "Данные об объёме не заполнены в заявках — "
            "ограничения по вместимости кузова (м³) не применяются. "
            "Добавьте объём в заявки на доставку, чтобы система контролировала загрузку кузова."
        )

    # ── Time windows (TSPTW) ─────────────────────────────────────────────────
    # When use_time_windows is True, pass (tw_from_min, tw_to_min, service_min)
    # per store to solve_vrp so OR-Tools enforces arrival constraints.
    store_time_windows = None
    if body.use_time_windows:
        store_time_windows = []
        invalid_tw_count = 0
        for s in store_list:
            tw_from = _parse_time_to_minutes(s.get("time_window_from") or "09:00")
            tw_to   = _parse_time_to_minutes(s.get("time_window_to")   or "18:00")
            service = int(s.get("unload_minutes") or 15) if body.use_unload_time else 0
            # Pre-validate: tw_from must be strictly less than tw_to, and window
            # must not close before depot departure (09:00 = 540 min).
            if tw_from >= tw_to or tw_to < 9 * 60:
                invalid_tw_count += 1
                tw_from, tw_to = 9 * 60, 23 * 60  # expand to full working day
            store_time_windows.append((tw_from, tw_to, service))
        if invalid_tw_count:
            route_warnings.append(
                f"{invalid_tw_count} магазин(ов) имели некорректные временные окна "
                f"(tw_from≥tw_to или закрытие раньше 09:00) — заменены на полный день."
            )
            logger.warning(
                "build_route: %d stores had invalid time windows (tw_from>=tw_to or "
                "tw_to<09:00) — replaced with full-day window",
                invalid_tw_count,
            )
        logger.info(
            "build_route: time windows enabled for %d stores (use_unload=%s)",
            len(store_time_windows), body.use_unload_time,
        )

    logger.info(
        "solve_vrp: %d stores, %d vehicles, capacities=%s, time_windows=%s",
        len(store_list), num_vehicles, capacities,
        "yes" if store_time_windows else "no",
    )

    # ── Pre-flight capacity check ─────────────────────────────────────────────
    # If the total order weight exceeds total vehicle capacity, routing is
    # physically impossible — tell the user BEFORE running the heavy solver.
    # Uses raw kg values (before OR-Tools integer scaling) for a clear message.
    if capacities is not None and demands is not None and _store_weights:
        _total_demand_kg = sum(_store_weights.get(s["id"], 0) for s in store_list)
        _total_capacity_kg = sum(
            int(v.capacity_kg) for v in body.vehicles if v.capacity_kg
        )
        _max_vehicle_cap = max(
            (int(v.capacity_kg) for v in body.vehicles if v.capacity_kg), default=99999
        )
        # Check if any single store's weight exceeds the largest vehicle capacity
        _oversized = [
            (s["name"], round(_store_weights[s["id"]]))
            for s in store_list
            if s["id"] in _store_weights and _store_weights[s["id"]] > _max_vehicle_cap
        ]
        if _oversized:
            names = ", ".join(f"{n} ({w} кг)" for n, w in _oversized[:3])
            raise HTTPException(
                status_code=422,
                detail=(
                    f"Вес {len(_oversized)} магазин(ов) превышает вместимость "
                    f"самой большой машины ({_max_vehicle_cap} кг): {names}"
                    + (" и др." if len(_oversized) > 3 else "") +
                    ". Увеличьте грузоподъёмность или разбейте заявки."
                )
            )
        if _total_demand_kg > _total_capacity_kg:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"Суммарный вес заявок ({round(_total_demand_kg)} кг) превышает "
                    f"суммарную грузоподъёмность транспорта ({_total_capacity_kg} кг). "
                    f"Добавьте машины или увеличьте грузоподъёмность."
                )
            )
        logger.info(
            "build_route: capacity pre-check OK — total_demand=%.0f kg, "
            "total_capacity=%d kg, utilisation=%.0f%%",
            _total_demand_kg, _total_capacity_kg,
            100 * _total_demand_kg / max(_total_capacity_kg, 1),
        )

    # ── Pre-flight m³ capacity check ──────────────────────────────────────────
    if capacities_m3 is not None and demands_m3 is not None and _store_volumes:
        _total_demand_m3 = sum(
            float(_store_volumes.get(s["id"], 0)) for s in store_list
        )
        _total_cap_m3 = sum(
            float(v.capacity_m3) for v in body.vehicles if v.capacity_m3
        )
        _max_cap_m3 = max(
            (float(v.capacity_m3) for v in body.vehicles if v.capacity_m3), default=1e9
        )
        _oversized_m3 = [
            (s["name"], round(float(_store_volumes[s["id"]]), 3))
            for s in store_list
            if s["id"] in _store_volumes and float(_store_volumes[s["id"]]) > _max_cap_m3
        ]
        if _oversized_m3:
            names_m3 = ", ".join(f"{n} ({v} м³)" for n, v in _oversized_m3[:3])
            raise HTTPException(
                status_code=422,
                detail=(
                    f"Объём {len(_oversized_m3)} магазин(ов) превышает вместимость "
                    f"самого большого кузова ({round(_max_cap_m3, 2)} м³): {names_m3}"
                    + (" и др." if len(_oversized_m3) > 3 else "") +
                    ". Увеличьте объём кузова или разбейте заявки."
                )
            )
        if _total_demand_m3 > _total_cap_m3:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"Суммарный объём заявок ({round(_total_demand_m3, 2)} м³) превышает "
                    f"суммарный объём кузовов ({round(_total_cap_m3, 2)} м³). "
                    f"Добавьте машины или увеличьте объём кузова."
                )
            )
        logger.info(
            "build_route: m³ pre-check OK — total_demand=%.2f m³, "
            "total_capacity=%.2f m³, utilisation=%.0f%%",
            _total_demand_m3, _total_cap_m3,
            100 * _total_demand_m3 / max(_total_cap_m3, 0.001),
        )

    # ── Degradation chain ────────────────────────────────────────────────────
    # Level 1: TW enabled (if requested)
    # Level 2: retry without TW on any solver exception
    # Level 3: greedy round-robin fallback if Level 2 also fails
    # Never return HTTP 500 — always produce a route or a clear 422 message.
    matrix_source = "haversine"
    # Validate max_stops_per_vehicle
    max_stops_cap = body.max_stops_per_vehicle
    if max_stops_cap is not None:
        if max_stops_cap < 1:
            raise HTTPException(
                status_code=422,
                detail="max_stops_per_vehicle должен быть ≥ 1"
            )
        avg_stops = len(store_list) / max(num_vehicles, 1)
        if max_stops_cap < math.ceil(avg_stops):
            raise HTTPException(
                status_code=422,
                detail=(
                    f"max_stops_per_vehicle={max_stops_cap} слишком мало: "
                    f"при {len(store_list)} магазинах и {num_vehicles} машинах "
                    f"минимальное значение = {math.ceil(avg_stops)} (среднее кол-во точек на машину)."
                )
            )
        logger.info(
            "build_route: max_stops_per_vehicle=%d requested (avg=%.1f per vehicle)",
            max_stops_cap, avg_stops,
        )

    # ── Auto-cap: apply ceil(avg × 1.5) when user did not specify a limit ─────
    # Prevents extreme imbalances like 34 / 8 / 7 stops caused by geographic
    # clustering without a ceiling.  Symmetric counterpart to the 0.70×avg floor
    # already applied by _rebalance_min_stops.  Silently logged; user can still
    # override by selecting one of the manual ≤N buttons in the UI.
    effective_max_stops = max_stops_cap
    if effective_max_stops is None and len(store_list) > 0 and num_vehicles > 0:
        _auto_avg = len(store_list) / num_vehicles
        effective_max_stops = math.ceil(_auto_avg * 1.5)
        logger.info(
            "build_route: auto max_stops_per_vehicle=%d "
            "(%.0f stores / %d vehicles, avg=%.1f × 1.5)",
            effective_max_stops, len(store_list), num_vehicles, _auto_avg,
        )

    # Validate optimize_by
    if body.optimize_by not in ("distance", "time"):
        raise HTTPException(
            status_code=422,
            detail="optimize_by должен быть 'distance' или 'time'"
        )
    if body.optimize_by == "time":
        logger.info("build_route: time-optimisation mode requested")

    try:
        vehicle_routes_indices, matrix_source = solve_vrp(
            all_coords, num_vehicles, capacities, demands, store_time_windows,
            max_stops_per_vehicle=effective_max_stops,
            optimize_by=body.optimize_by,
            capacities_m3=capacities_m3, demands_m3=demands_m3,
        )
    except Exception as vrp_exc_1:
        logger.error("solve_vrp (with TW) failed:\n%s", traceback.format_exc())
        if store_time_windows is not None:
            # Level 2: disable time windows and retry
            route_warnings.append(
                "Временные окна привели к неразрешимой задаче и были отключены. "
                "Маршрут построен без учёта временных окон."
            )
            logger.warning(
                "build_route: degrading to no-TW solve (%s stores, %s vehicles)",
                len(store_list), num_vehicles,
            )
            try:
                vehicle_routes_indices, matrix_source = solve_vrp(
                    all_coords, num_vehicles, capacities, demands, None,
                    max_stops_per_vehicle=effective_max_stops,
                    optimize_by=body.optimize_by,
                    capacities_m3=capacities_m3, demands_m3=demands_m3,
                )
            except Exception as vrp_exc_2:
                logger.error("solve_vrp (no TW) also failed:\n%s", traceback.format_exc())
                # Level 3: greedy round-robin — cannot fail
                route_warnings.append(
                    "Оптимизатор маршрутов недоступен. Магазины распределены по машинам "
                    "в порядке загрузки (round-robin). Результат не оптимален."
                )
                logger.warning("build_route: falling back to greedy round-robin distribution")
                store_nodes = list(range(1, len(all_coords)))
                vehicle_routes_indices = _fallback_distribution(store_nodes, num_vehicles)
                matrix_source = "haversine"
        else:
            # No TW was requested and basic VRP still failed — greedy fallback
            route_warnings.append(
                "Оптимизатор маршрутов недоступен. Магазины распределены по машинам "
                "в порядке загрузки (round-robin). Результат не оптимален."
            )
            logger.warning("build_route: falling back to greedy round-robin (no TW requested)")
            store_nodes = list(range(1, len(all_coords)))
            vehicle_routes_indices = _fallback_distribution(store_nodes, num_vehicles)
            matrix_source = "haversine"

    logger.info("solve_vrp result: %s routes", len(vehicle_routes_indices))

    # ── OSRM ETA: fetch real road travel times for each finalised route ────────
    # Separate post-solve calls — do NOT affect routing decisions.
    # We call OSRM once per route with the final ordered coordinates and read
    # consecutive leg times from the returned duration matrix diagonal.
    # Falls back to Haversine formula (ETA_ROAD_FACTOR) if OSRM unavailable.
    _route_leg_times: list = [None] * len(body.vehicles)
    _eta_coord_lists: list = []
    for _vi_eta, _v_eta in enumerate(body.vehicles):
        if _vi_eta >= len(vehicle_routes_indices) or not vehicle_routes_indices[_vi_eta]:
            _eta_coord_lists.append(None)
            continue
        _coords_eta = [(depot_lat, depot_lon)]
        for _idx_eta in vehicle_routes_indices[_vi_eta]:
            _sidx_eta = _idx_eta - 1
            if 0 <= _sidx_eta < len(store_list):
                _s_eta = store_list[_sidx_eta]
                _coords_eta.append((_s_eta["lat"], _s_eta["lon"]))
        _eta_coord_lists.append(_coords_eta if len(_coords_eta) >= 2 else None)

    _eta_workers = min(len([c for c in _eta_coord_lists if c is not None]), 8)
    if _eta_workers > 0:
        with concurrent.futures.ThreadPoolExecutor(max_workers=_eta_workers) as _eta_pool:
            _eta_futures = {
                _eta_pool.submit(_fetch_route_leg_times_osrm, _coords): _vi
                for _vi, _coords in enumerate(_eta_coord_lists)
                if _coords is not None
            }
            for _fut, _vi in _eta_futures.items():
                try:
                    _route_leg_times[_vi] = _fut.result()
                except Exception:
                    pass

    _osrm_eta_ok = sum(1 for t in _route_leg_times if t is not None)
    logger.info(
        "OSRM ETA prefetch: %d/%d routes have real road times (fallback=Haversine for rest)",
        _osrm_eta_ok, len(body.vehicles),
    )

    # Build result
    routes = []
    total_km = 0.0

    for vi, vehicle in enumerate(body.vehicles):
        if vi >= len(vehicle_routes_indices):
            break
        route_indices = vehicle_routes_indices[vi]
        if not route_indices:
            continue

        ROUTE_START_MINUTES = 9 * 60  # 09:00 departure from depot
        route_stores = []
        # ETA_ROAD_FACTOR: Haversine → road km conversion.  Used ONLY as fallback
        # when OSRM leg times are unavailable for this route.
        ETA_ROAD_FACTOR = 2.0
        route_coords = [(depot_lat, depot_lon)]
        dist_m = 0
        cumulative_min = 0  # elapsed minutes since 09:00
        prev_coord = (depot_lat, depot_lon)
        # Per-route OSRM leg times (seconds); None → use Haversine fallback
        leg_times = _route_leg_times[vi] if vi < len(_route_leg_times) else None
        # Sanity-check: any leg > 2 h (7 200 s) signals an unreachable/sea location
        # — discard the whole route's OSRM data and fall back to Haversine.
        if leg_times is not None and any(t > 7200 for t in leg_times):
            logger.warning(
                "OSRM ETA route %d: discarded — leg > 2h (likely unreachable coordinate)",
                vi,
            )
            leg_times = None
        leg_idx = 0  # cursor into leg_times list

        for order, idx in enumerate(route_indices, 1):
            store_idx = idx - 1  # node 0 = depot, node i = store_list[i-1]
            if store_idx < 0 or store_idx >= len(store_list):
                logger.warning("VRP returned out-of-range node %d (store_list len=%d) — skipping", idx, len(store_list))
                continue
            store = store_list[store_idx]
            curr_coord = (store["lat"], store["lon"])

            # Drive time from previous point
            leg_m = haversine_meters(prev_coord, curr_coord)
            dist_m += leg_m
            effective_speed = vehicle.average_speed if vehicle.average_speed else AVG_SPEED_KMH
            if leg_times is not None and leg_idx < len(leg_times):
                # Real OSRM road time — most accurate
                leg_drive_min = max(1, int(leg_times[leg_idx] / 60))
            else:
                # Haversine fallback (OSRM unavailable for this leg)
                leg_drive_min = max(1, int(leg_m * ETA_ROAD_FACTOR / 1000 / effective_speed * 60))
            leg_idx += 1
            cumulative_min += leg_drive_min

            # Estimated arrival time at this stop
            abs_min = ROUTE_START_MINUTES + cumulative_min
            arrive_hour = (abs_min // 60) % 24
            arrive_min_part = abs_min % 60
            arrive_by = f"{arrive_hour:02d}:{arrive_min_part:02d}"

            route_coords.append(curr_coord)
            route_stores.append({
                "order": order,
                "store_id": store["id"],
                "store_name": store["name"],
                "address": store["address"],
                "lat": store["lat"],
                "lon": store["lon"],
                "arrive_by": arrive_by,
                "weight_kg": _store_weights.get(store["id"], 0),
                "volume_m3": _store_volumes.get(store["id"], 0),
            })

            # Add unload time before driving to the next stop
            if body.use_unload_time:
                cumulative_min += store.get("unload_minutes", 15)

            prev_coord = curr_coord

        # Return to depot distance
        if len(route_coords) > 1:
            dist_m += haversine_meters(route_coords[-1], (depot_lat, depot_lon))

        km = dist_m / 1000.0
        total_km += km

        unload_min = sum(s["unload_minutes"] for s in store_list if s["id"] in [rs["store_id"] for rs in route_stores]) if body.use_unload_time else 0
        eff_spd = vehicle.average_speed if vehicle.average_speed else AVG_SPEED_KMH
        if leg_times is not None:
            # Sum of all OSRM leg times for this route (seconds → minutes)
            drive_min = max(1, int(sum(leg_times) / 60))
        else:
            drive_min = int(km * ETA_ROAD_FACTOR / eff_spd * 60)
        est_minutes = drive_min + unload_min

        # depot (route_coords[0]) остаётся первой точкой — Яндекс заменит его
        # GPS-позицией водителя, но все магазины сохранятся на своих местах.
        yurls = yandex_nav_urls(route_coords) if len(route_coords) > 1 else []
        yurl = yurls[0] if yurls else ""
        wurl = whatsapp_url(vehicle.name, route_stores, km, yurls)

        _route_weight = round(sum(_store_weights.get(rs["store_id"], 0) for rs in route_stores), 1)
        _vehicle_cap_kg = int(vehicle.capacity_kg) if vehicle.capacity_kg else 0
        _vehicle_cap_m3 = float(vehicle.capacity_m3) if vehicle.capacity_m3 else 0.0
        routes.append({
            "vehicle_name": vehicle.name,
            "stores": route_stores,
            "total_km": round(km, 1),
            "estimated_minutes": est_minutes,
            # ETA breakdown — allows frontend to display drive vs service separately
            "drive_minutes": drive_min,
            "service_minutes": unload_min,
            "yandex_url": yurl,
            "yandex_urls": yurls,
            "whatsapp_url": wurl,
            # Cargo summary from daily_orders (0 when no orders loaded for today)
            "total_weight_kg": _route_weight,
            "total_volume_m3": round(sum(_store_volumes.get(rs["store_id"], 0) for rs in route_stores), 3),
            # Vehicle capacity for frontend overload indicator (0 = not configured)
            "capacity_kg": _vehicle_cap_kg,
            "capacity_m3": _vehicle_cap_m3,
        })

    # ── Diagnostic: per-route weight/volume utilisation ──────────────────────
    for _ri, _r in enumerate(routes):
        _cap_kg  = _r["capacity_kg"]
        _cap_m3  = _r["capacity_m3"]
        _load_kg = _r["total_weight_kg"]
        _load_m3 = _r["total_volume_m3"]
        logger.info(
            "Route %d (%s): Weight: %.0f / %s kg | Volume: %.2f / %s m³ | Stops: %d",
            _ri + 1,
            _r["vehicle_name"],
            _load_kg,
            str(_cap_kg) if _cap_kg else "∞",
            _load_m3,
            f"{_cap_m3:.2f}" if _cap_m3 else "∞",
            len(_r["stores"]),
        )

    # ── Capacity overflow warning ─────────────────────────────────────────────
    # Generated AFTER routes are built so we can report actual per-vehicle loads.
    # Pre-flight 422 checks total feasibility; this catches bin-packing overflow
    # (e.g. 3 items × 600 kg can't fit in 2 bins × 1000 kg even though sum fits).
    if _store_weights:
        overloaded = [
            r for r in routes
            if r["capacity_kg"] > 0 and r["total_weight_kg"] > r["capacity_kg"]
        ]
        if overloaded:
            details = "; ".join(
                f"{r['vehicle_name']}: {r['total_weight_kg']:.0f} / {r['capacity_kg']} кг "
                f"(+{r['total_weight_kg'] - r['capacity_kg']:.0f} кг перегруза)"
                for r in overloaded
            )
            route_warnings.append(
                f"Перегруз: {details}. "
                f"Суммарный вес заявок вписывается в автопарк, но несколько крупных заявок "
                f"невозможно разделить между машинами так, чтобы ни одна не была перегружена. "
                f"Решение: добавьте ещё одну машину или разбейте крупные заявки на части."
            )
            logger.warning(
                "build_route: capacity overflow in %d route(s): %s",
                len(overloaded), details,
            )

    # ── Volume (m³) overflow warning ──────────────────────────────────────────
    if _store_volumes and capacities_m3:
        overloaded_m3 = [
            r for r in routes
            if r["capacity_m3"] > 0 and r["total_volume_m3"] > r["capacity_m3"]
        ]
        if overloaded_m3:
            details_m3 = "; ".join(
                f"{r['vehicle_name']}: {r['total_volume_m3']:.2f} / {r['capacity_m3']:.2f} м³ "
                f"(+{r['total_volume_m3'] - r['capacity_m3']:.2f} м³ перебор)"
                for r in overloaded_m3
            )
            route_warnings.append(
                f"Перебор по объёму: {details_m3}. "
                f"Суммарный объём заявок вписывается в автопарк, но несколько крупных "
                f"заявок невозможно распределить без переполнения кузова по объёму. "
                f"Решение: добавьте ещё одну машину или разбейте крупные заявки на части."
            )
            logger.warning(
                "build_route: m³ overflow in %d route(s): %s",
                len(overloaded_m3), details_m3,
            )

    _cost_settings = get_company_settings(user_id=uid)
    savings = calculate_savings(
        total_km,
        store_list,      # passed in original input order — used as baseline
        num_vehicles,
        depot_lat,
        depot_lon,
        settings=_cost_settings,
    )

    # Enrich stops with products/quantity from daily_orders if delivery_date provided
    if body.delivery_date:
        try:
            _all_store_ids = [s["store_id"] for r in routes for s in r["stores"]]
            if _all_store_ids:
                conn_enrich = get_db()
                cur_enrich = conn_enrich.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                cur_enrich.execute(
                    "SELECT store_id, products, quantity FROM daily_orders"
                    " WHERE owner_id = %s AND delivery_date = %s AND store_id = ANY(%s)",
                    (uid, body.delivery_date, _all_store_ids)
                )
                _orders_by_store = {r["store_id"]: r for r in cur_enrich.fetchall()}
                cur_enrich.close(); conn_enrich.close()
                for route in routes:
                    for stop in route["stores"]:
                        od = _orders_by_store.get(stop["store_id"])
                        if od:
                            stop["products"] = od.get("products") or ""
                            stop["quantity"] = _execution_quantity(
                                stop.get("quantity"),
                                od.get("products"),
                                od.get("quantity"),
                            )
        except Exception as _enrich_err:
            logger.warning(f"Failed to enrich stops with products: {_enrich_err}")

    result = {
        "routes": routes,
        "delivery_date": body.delivery_date,
        "savings": savings,
        "total_km": round(total_km, 1),
        "matrix_source": matrix_source,
        "geocoder_used": "yandex" if YANDEX_GEOCODER_API_KEY else "nominatim",
        "session_id": None,
        "warnings": route_warnings,  # non-fatal degradation notices for the frontend
        # Saved for historical replay — tells frontend whether estimated_minutes
        # includes service time (15 min/stop) or is drive-only
        "use_unload_time": bool(body.use_unload_time),
        "optimize_by": body.optimize_by,
    }

    # Save session to DB
    try:
        conn2 = get_db()
        cur2 = conn2.cursor()
        cur2.execute(
            """INSERT INTO route_sessions (date, num_vehicles, total_km, saved_km, saved_rub, num_points, cost_per_km, result_json, owner_id)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id""",
            (str(date.today()), num_vehicles, round(total_km, 1),
             savings["saved_km"], savings["saved_rub_day"], len(store_list),
             savings.get("cost_per_km"), json.dumps(result), uid)
        )
        session_id = cur2.fetchone()[0]
        result["session_id"] = session_id
        cur2.execute("UPDATE route_sessions SET result_json = %s WHERE id = %s",
                     (json.dumps(result), session_id))
        for rs in routes:
            for stop in rs["stores"]:
                cur2.execute(
                    "INSERT INTO route_session_stores (session_id, store_id, visit_order) VALUES (%s, %s, %s)",
                    (session_id, stop["store_id"], stop["order"])
                )
        conn2.commit()
        cur2.close()
        conn2.close()
    except Exception as e:
        logger.error(f"Failed to save route session: {e}")

    return result


def _serialize_driver(row: dict) -> dict:
    return {
        "id": int(row["id"]),
        "name": row.get("name") or "",
        "phone": row.get("phone") or "",
        "vehicle_name": row.get("vehicle_name") or "",
        "is_active": bool(row.get("is_active", True)),
        "created_at": str(row["created_at"]) if row.get("created_at") else None,
        "updated_at": str(row["updated_at"]) if row.get("updated_at") else None,
    }


@app.get("/api/drivers")
def list_drivers(request: Request, include_inactive: bool = False):
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """SELECT id, name, phone, vehicle_name, is_active, created_at, updated_at
                 FROM drivers
                WHERE owner_id=%s AND (%s OR is_active=TRUE)
                ORDER BY is_active DESC, name""",
            (uid, include_inactive),
        )
        return {"drivers": [_serialize_driver(dict(row)) for row in cur.fetchall()]}
    finally:
        cur.close(); conn.close()


@app.post("/api/drivers", status_code=201)
def create_driver(body: DriverCreate, request: Request):
    uid = get_user_id(request)
    name = body.name.strip()
    phone = body.phone.strip()
    if not name:
        raise HTTPException(status_code=422, detail="Укажите имя водителя")
    if len(_normalize_driver_phone(phone)) < 7:
        raise HTTPException(status_code=422, detail="Укажите корректный телефон водителя")
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """INSERT INTO drivers (owner_id, name, phone, vehicle_name)
               VALUES (%s,%s,%s,%s)
               RETURNING id, name, phone, vehicle_name, is_active, created_at, updated_at""",
            (uid, name, phone, body.vehicle_name.strip()),
        )
        row = dict(cur.fetchone())
        conn.commit()
        return _serialize_driver(row)
    finally:
        cur.close(); conn.close()


@app.patch("/api/drivers/{driver_id}")
def update_driver(driver_id: int, body: DriverUpdate, request: Request):
    uid = get_user_id(request)
    fields, values = [], []
    if body.name is not None:
        if not body.name.strip():
            raise HTTPException(status_code=422, detail="Имя водителя не может быть пустым")
        fields.append("name=%s"); values.append(body.name.strip())
    if body.phone is not None:
        if len(_normalize_driver_phone(body.phone)) < 7:
            raise HTTPException(status_code=422, detail="Укажите корректный телефон водителя")
        fields.append("phone=%s"); values.append(body.phone.strip())
    if body.vehicle_name is not None:
        fields.append("vehicle_name=%s"); values.append(body.vehicle_name.strip())
    if body.is_active is not None:
        fields.append("is_active=%s"); values.append(body.is_active)
    if not fields:
        raise HTTPException(status_code=422, detail="Нет изменений")
    fields.append("updated_at=NOW()")
    values.extend([driver_id, uid])
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            f"""UPDATE drivers SET {', '.join(fields)}
                    WHERE id=%s AND owner_id=%s
                RETURNING id, name, phone, vehicle_name, is_active, created_at, updated_at""",
            values,
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Водитель не найден")
        conn.commit()
        return _serialize_driver(dict(row))
    finally:
        cur.close(); conn.close()


@app.delete("/api/drivers/{driver_id}", status_code=204)
def archive_driver(driver_id: int, request: Request):
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE drivers SET is_active=FALSE, updated_at=NOW() WHERE id=%s AND owner_id=%s",
            (driver_id, uid),
        )
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Водитель не найден")
        conn.commit()
    finally:
        cur.close(); conn.close()


@app.get("/api/route/sessions/{id}")
def get_route_session(id: int, request: Request):
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT result_json FROM route_sessions WHERE id = %s AND owner_id = %s", (id, uid))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row or not row["result_json"]:
        raise HTTPException(status_code=404, detail="Route session not found")
    return json.loads(row["result_json"])


@app.get("/api/route/sessions/{session_id}/report.xlsx")
def download_route_report(session_id: int, request: Request):
    """Export the operational route report as a dispatcher-friendly workbook."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            "SELECT id, date, total_km, result_json FROM route_sessions WHERE id=%s AND owner_id=%s",
            (session_id, uid),
        )
        session = cur.fetchone()
        if not session:
            raise HTTPException(status_code=404, detail="Маршрут не найден")
        cur.execute(
            """SELECT a.vehicle_name, a.driver_name, a.driver_phone, a.route_index,
                      e.visit_order, e.store_name, e.address, e.status,
                      e.quantity, e.actual_qty, e.products, e.payment_method,
                      e.payment_status, e.driver_comment, e.rescheduled_date
                 FROM route_assignments a
                 LEFT JOIN route_executions e ON e.assignment_id=a.id
                WHERE a.session_id=%s AND a.owner_id=%s
                ORDER BY a.route_index, e.visit_order""",
            (session_id, uid),
        )
        rows = [dict(row) for row in cur.fetchall()]
    finally:
        cur.close(); conn.close()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт по рейсу"
    headers = [
        "Машина", "Водитель", "Телефон", "Дата", "№ точки", "Контрагент",
        "Адрес", "Статус", "План", "Доставлено", "Остаток", "Товар",
        "Способ оплаты", "Статус оплаты", "Комментарий", "Дата переноса",
        "Общий пробег, км",
    ]
    sheet.append(headers)
    header_fill = openpyxl.styles.PatternFill("solid", fgColor="1D4ED8")
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    total_km = float(session.get("total_km") or 0)
    for row in rows:
        planned = float(row.get("quantity") or 0)
        actual = float(row.get("actual_qty") or 0)
        sheet.append([
            row.get("vehicle_name") or "", row.get("driver_name") or "", row.get("driver_phone") or "",
            session.get("date") or "", row.get("visit_order") or "", row.get("store_name") or "",
            row.get("address") or "", row.get("status") or "planned", planned, actual,
            max(planned - actual, 0), row.get("products") or "", row.get("payment_method") or "none",
            row.get("payment_status") or "pending", row.get("driver_comment") or "",
            str(row["rescheduled_date"]) if row.get("rescheduled_date") else "", total_km,
        ])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [16, 20, 16, 13, 10, 30, 34, 16, 12, 14, 12, 30, 18, 18, 34, 16, 18]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(vertical="top", wrap_text=True)
    for col in (9, 10, 11, 17):
        for cell in list(sheet.iter_cols(min_col=col, max_col=col, min_row=2))[0]:
            cell.number_format = "0.0"
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"smartroute_route_{session_id}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _assignment_summary(row: dict) -> dict:
    """Serialize an assignment without exposing its one-time token."""
    assignment_status = row.get("status") or "planned"
    if assignment_status == "on_route":
        assignment_status = "planned"
    return {
        "id": row["id"],
        "session_id": row["session_id"],
        "route_index": row["route_index"],
        "driver_id": row.get("driver_id"),
        "driver_name": row.get("driver_name") or "",
        "driver_phone": row.get("driver_phone") or "",
        "vehicle_name": row.get("vehicle_name") or "",
        "route_yandex_url": row.get("route_yandex_url") or "",
        "status": assignment_status,
        "expires_at": str(row["expires_at"]) if row.get("expires_at") else None,
        "total_points": int(row.get("total_points") or 0),
        "completed_points": int(row.get("completed_points") or 0),
        "updated_at": str(row["updated_at"]) if row.get("updated_at") else None,
        "whatsapp_url": row.get("whatsapp_url"),
        "location_lat": row.get("location_lat"),
        "location_lon": row.get("location_lon"),
        "location_accuracy": row.get("location_accuracy"),
        "location_captured_at": row.get("location_captured_at"),
    }


def _load_assignment_for_token(token: str) -> tuple[dict, list[dict]]:
    """Load a driver assignment by a raw token; the raw token is never stored."""
    if not token or len(token) > 160:
        raise HTTPException(status_code=404, detail="Ссылка водителя недействительна")
    token_hash = hashlib.sha256(token.encode("utf-8")).hexdigest()
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """SELECT id, session_id, route_index, driver_id, driver_name, driver_phone, vehicle_name,
                  route_yandex_url, status, expires_at,
                  updated_at
           FROM route_assignments WHERE access_token_hash=%s""",
        (token_hash,),
    )
    assignment = cur.fetchone()
    if not assignment:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Ссылка водителя недействительна или отозвана")
    if assignment.get("expires_at") and assignment["expires_at"] <= datetime.utcnow():
        cur.close(); conn.close()
        raise HTTPException(status_code=410, detail="Срок действия ссылки водителя истёк. Попросите диспетчера выдать новую ссылку.")
    cur.execute(
        """SELECT id, store_id, visit_order, store_name, address, lat, lon,
                  products, quantity, actual_qty, arrive_by, yandex_url, status,
                  payment_method, payment_status, driver_comment, rescheduled_date,
                  updated_at, delivered_at
           FROM route_executions WHERE assignment_id=%s ORDER BY visit_order""",
        (assignment["id"],),
    )
    executions = [dict(r) for r in cur.fetchall()]
    # Repair legacy rows created before products-derived quantities existed.
    # This is intentionally a safe data update: it only changes rows whose
    # planned quantity is zero and leaves delivery status/actual quantity intact.
    repaired = []
    for execution in executions:
        if float(execution.get("quantity") or 0) > 0:
            continue
        planned_qty = _execution_quantity(0, execution.get("products"))
        cur.execute(
            """UPDATE route_executions
                  SET quantity=%s, updated_at=NOW()
                WHERE id=%s AND assignment_id=%s AND quantity <= 0""",
            (planned_qty, execution["id"], assignment["id"]),
        )
        if cur.rowcount:
            execution["quantity"] = planned_qty
            repaired.append(execution["id"])
    if repaired:
        conn.commit()
    cur.close(); conn.close()
    return dict(assignment), executions


def _serialize_execution(row: dict) -> dict:
    planned_qty = float(row.get("quantity") or 0)
    actual_raw = row.get("actual_qty")
    actual_qty = planned_qty if actual_raw is None else float(actual_raw)
    status = row.get("status") or "planned"
    if status in {"loaded", "on_route"}:
        status = "planned"
    if status == "planned":
        actual_qty = 0
    return {
        "id": row["id"],
        "store_id": row.get("store_id"),
        "visit_order": row["visit_order"],
        "store_name": row.get("store_name") or "",
        "address": row.get("address") or "",
        "lat": row.get("lat"),
        "lon": row.get("lon"),
        "products": row.get("products") or "",
        "quantity": planned_qty,
        "actual_qty": actual_qty,
        "shortfall_qty": max(
            planned_qty - actual_qty,
            0,
        ),
        "arrive_by": row.get("arrive_by") or "",
        "yandex_url": row.get("yandex_url") or "",
        "status": status,
        "payment_method": row.get("payment_method") or "none",
        "payment_status": row.get("payment_status") or "pending",
        "driver_comment": row.get("driver_comment") or "",
        "rescheduled_date": str(row["rescheduled_date"]) if row.get("rescheduled_date") else None,
        "remaining_order_date": str(row["remaining_order_date"]) if row.get("remaining_order_date") else None,
        "updated_at": str(row["updated_at"]) if row.get("updated_at") else None,
        "delivered_at": str(row["delivered_at"]) if row.get("delivered_at") else None,
    }


@app.get("/api/route/sessions/{session_id}/assignments")
def list_route_assignments(session_id: int, request: Request):
    """Dispatcher view of operational trips and point progress."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """SELECT a.id, a.session_id, a.route_index, a.driver_id, a.driver_name,
                  a.driver_phone, a.vehicle_name, a.route_yandex_url, a.status, a.expires_at, a.updated_at,
                  loc.lat AS location_lat, loc.lon AS location_lon,
                  loc.accuracy AS location_accuracy, loc.captured_at AS location_captured_at,
                  COUNT(e.id) AS total_points,
                  COUNT(e.id) FILTER (WHERE e.status IN
                    ('delivered','partial','failed','rescheduled')) AS completed_points
           FROM route_assignments a
           LEFT JOIN route_executions e ON e.assignment_id=a.id
           LEFT JOIN LATERAL (
             SELECT lat, lon, accuracy, captured_at
               FROM driver_locations
              WHERE assignment_id=a.id
              ORDER BY captured_at DESC
              LIMIT 1
           ) loc ON TRUE
           WHERE a.session_id=%s AND a.owner_id=%s
           GROUP BY a.id, loc.lat, loc.lon, loc.accuracy, loc.captured_at
           ORDER BY a.route_index""",
        (session_id, uid),
    )
    rows = [dict(row) for row in cur.fetchall()]
    items = [_assignment_summary(row) for row in rows]
    if rows:
        cur.execute(
            """SELECT e.assignment_id, e.id, e.store_id, e.visit_order, e.store_name,
                      e.address, e.lat, e.lon, e.products, e.quantity, e.actual_qty, e.arrive_by,
                      e.yandex_url, e.status, e.payment_method, e.payment_status,
                      e.driver_comment, e.rescheduled_date, e.updated_at, e.delivered_at,
                      (SELECT MAX(o.delivery_date)::text
                         FROM daily_orders o
                        WHERE o.owner_id = a.owner_id
                          AND o.store_id = e.store_id
                          AND (
                              o.notes LIKE 'Создано из остатка исполнения доставки%%'
                              OR o.notes = 'Требуется уточнение остатка по товарным позициям.'
                          )
                          AND o.created_at >= COALESCE(e.delivered_at, e.updated_at)
                      ) AS remaining_order_date
               FROM route_executions e
               JOIN route_assignments a ON a.id = e.assignment_id
               WHERE e.assignment_id = ANY(%s)
               ORDER BY e.assignment_id, e.visit_order""",
            ([row["id"] for row in rows],),
        )
        by_assignment = {}
        for execution in cur.fetchall():
            item = _serialize_execution(dict(execution))
            by_assignment.setdefault(execution["assignment_id"], []).append(item)
        for item in items:
            item["executions"] = by_assignment.get(item["id"], [])
            location = {
                "lat": item.pop("location_lat", None),
                "lon": item.pop("location_lon", None),
                "accuracy": item.pop("location_accuracy", None),
                "captured_at": str(item.pop("location_captured_at")) if item.get("location_captured_at") else None,
            }
            item["last_location"] = location if location["lat"] is not None else None
            next_stop = next((stop for stop in item["executions"] if stop.get("status") == "planned"), None)
            item["next_stop"] = {
                "visit_order": next_stop.get("visit_order"),
                "store_name": next_stop.get("store_name") or "",
                "address": next_stop.get("address") or "",
                "lat": next_stop.get("lat"),
                "lon": next_stop.get("lon"),
            } if next_stop else None
            if location["lat"] is not None and next_stop and next_stop.get("lat") is not None and next_stop.get("lon") is not None:
                distance_km = haversine_meters(
                    (float(location["lat"]), float(location["lon"])),
                    (float(next_stop["lat"]), float(next_stop["lon"])),
                ) / 1000
                item["next_stop_eta_minutes"] = max(1, round(distance_km / 30 * 60 * 1.6))
            else:
                item["next_stop_eta_minutes"] = None
    cur.close(); conn.close()
    return {"assignments": items}


@app.post("/api/route/sessions/{session_id}/assignments", status_code=201)
def create_route_assignment(session_id: int, body: AssignmentCreate, request: Request):
    """Create a trip assignment and return a fresh driver link."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            "SELECT result_json, date FROM route_sessions WHERE id=%s AND owner_id=%s",
            (session_id, uid),
        )
        session = cur.fetchone()
        if not session or not session["result_json"]:
            raise HTTPException(status_code=404, detail="Маршрут не найден")
        result = json.loads(session["result_json"])
        routes = result.get("routes") or []
        if body.route_index < 0 or body.route_index >= len(routes):
            raise HTTPException(status_code=422, detail="Указан неизвестный рейс маршрута")
        route = routes[body.route_index]
        # Older route sessions did not persist the delivery date and some
        # imported routes consequently lost quantity before the execution row
        # was created. Recover quantities from the route first, then from the
        # day's orders when the route carries a date.
        quantities_by_store: dict[int, float] = {}
        delivery_date = result.get("delivery_date") or (
            str(session["date"]) if session.get("date") else None
        )
        if delivery_date:
            cur.execute(
                """SELECT store_id, COALESCE(SUM(quantity), 0) AS quantity
                     FROM daily_orders
                    WHERE owner_id=%s AND delivery_date=%s AND store_id IS NOT NULL
                    GROUP BY store_id""",
                (uid, delivery_date),
            )
            quantities_by_store = {
                int(row["store_id"]): float(row["quantity"] or 0)
                for row in cur.fetchall()
            }
        raw_token = secrets.token_urlsafe(32)
        token_hash = hashlib.sha256(raw_token.encode("utf-8")).hexdigest()
        driver_id = body.driver_id
        driver_phone = ""
        driver_name = body.driver_name.strip()
        directory_vehicle = ""
        if driver_id is not None:
            cur.execute(
                "SELECT id, name, phone, vehicle_name FROM drivers WHERE id=%s AND owner_id=%s AND is_active=TRUE",
                (driver_id, uid),
            )
            driver = cur.fetchone()
            if not driver:
                raise HTTPException(status_code=422, detail="Выбранный водитель не найден или неактивен")
            driver_name = driver["name"]
            driver_phone = driver["phone"] or ""
            directory_vehicle = driver["vehicle_name"] or ""
        vehicle_name = body.vehicle_name.strip() or directory_vehicle or str(route.get("vehicle_name") or "")
        cur.execute(
            """INSERT INTO route_assignments
                 (owner_id, session_id, route_index, driver_id, driver_name, driver_phone, vehicle_name,
                  route_yandex_url, access_token_hash, token_created_at, expires_at)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),NOW() + INTERVAL '48 hours')
               ON CONFLICT (session_id, route_index)
               DO UPDATE SET driver_id=EXCLUDED.driver_id,
                             driver_name=EXCLUDED.driver_name,
                             driver_phone=EXCLUDED.driver_phone,
                             vehicle_name=EXCLUDED.vehicle_name,
                             route_yandex_url=EXCLUDED.route_yandex_url,
                             access_token_hash=EXCLUDED.access_token_hash,
                             token_created_at=NOW(),
                             expires_at=NOW() + INTERVAL '48 hours',
                             updated_at=NOW()
               RETURNING id, session_id, route_index, driver_id, driver_name, driver_phone, vehicle_name,
                         route_yandex_url, status, expires_at, updated_at""",
            (
                uid, session_id, body.route_index, driver_id, driver_name, driver_phone, vehicle_name,
                route.get("yandex_url") or "", token_hash,
            ),
        )
        assignment = dict(cur.fetchone())
        cur.execute(
            "SELECT COUNT(*) AS count FROM route_executions WHERE assignment_id=%s",
            (assignment["id"],),
        )
        if int(cur.fetchone()["count"]) == 0:
            for stop in route.get("stores") or []:
                cur.execute(
                    """INSERT INTO route_executions
                       (assignment_id, store_id, visit_order, store_name, address, lat, lon,
                        products, quantity, actual_qty, arrive_by, yandex_url,
                        payment_status)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (
                        assignment["id"], stop.get("store_id"), stop.get("order", 0),
                        stop.get("store_name", ""), stop.get("address", ""),
                        stop.get("lat"), stop.get("lon"), stop.get("products", ""),
                        _execution_quantity(
                            stop.get("quantity"),
                            stop.get("products"),
                            quantities_by_store.get(int(stop["store_id"]), 0)
                            if stop.get("store_id") is not None
                            else 0,
                        ),
                        0,
                        stop.get("arrive_by", ""),
                        stop.get("yandex_url", ""), "pending",
                    ),
                )
        conn.commit()
        # Relative path keeps the link on the public frontend origin when the
        # API is reached through Vite/Replit's development proxy.
        driver_url = f"/driver/{raw_token}"
        public_driver_url = f"{_public_app_url(request)}{driver_url}"
        assignment.update({
            "total_points": len(route.get("stores") or []),
            "completed_points": 0,
            "driver_url": driver_url,
            "whatsapp_url": whatsapp_driver_url(
                driver_phone, delivery_date or "", vehicle_name,
                len(route.get("stores") or []), route.get("total_km") or 0,
                route.get("yandex_url") or "", public_driver_url,
            ) if driver_phone else whatsapp_assignment_url(
                vehicle_name, route.get("yandex_url") or "", public_driver_url,
            ),
        })
        return assignment
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        logger.exception("create_route_assignment failed")
        raise HTTPException(status_code=500, detail="Не удалось создать рейс")
    finally:
        cur.close(); conn.close()


@app.post("/api/route/assignments/{assignment_id}/share")
def share_route_assignment(assignment_id: int, request: Request):
    """Issue a fresh driver link for an existing assignment for re-sharing."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """SELECT a.id, a.session_id, a.route_index, a.driver_name, a.driver_phone,
                      a.vehicle_name, a.route_yandex_url, s.date, s.result_json
                 FROM route_assignments a
                 JOIN route_sessions s ON s.id=a.session_id
                WHERE a.id=%s AND a.owner_id=%s""",
            (assignment_id, uid),
        )
        assignment = cur.fetchone()
        if not assignment:
            raise HTTPException(status_code=404, detail="Назначение не найдено")
        result = json.loads(assignment["result_json"] or "{}")
        routes = result.get("routes") or []
        route_index = int(assignment["route_index"])
        route = routes[route_index] if 0 <= route_index < len(routes) else {}
        cur.execute("SELECT COUNT(*) AS count FROM route_executions WHERE assignment_id=%s", (assignment_id,))
        total_points = int(cur.fetchone()["count"])
        raw_token = secrets.token_urlsafe(32)
        token_hash = hashlib.sha256(raw_token.encode("utf-8")).hexdigest()
        cur.execute(
            """UPDATE route_assignments
                  SET access_token_hash=%s, token_created_at=NOW(),
                      expires_at=NOW() + INTERVAL '48 hours', updated_at=NOW()
                WHERE id=%s AND owner_id=%s""",
            (token_hash, assignment_id, uid),
        )
        conn.commit()
        driver_url = f"/driver/{raw_token}"
        public_driver_url = f"{_public_app_url(request)}{driver_url}"
        route_url = assignment["route_yandex_url"] or route.get("yandex_url") or ""
        phone = assignment["driver_phone"] or ""
        return {
            "assignment_id": assignment_id,
            "driver_name": assignment["driver_name"] or "",
            "driver_phone": phone,
            "driver_url": driver_url,
            "whatsapp_url": whatsapp_driver_url(
                phone,
                result.get("delivery_date") or str(assignment["date"] or ""),
                assignment["vehicle_name"] or route.get("vehicle_name") or "",
                total_points,
                route.get("total_km") or 0,
                route_url,
                public_driver_url,
            ) if phone else whatsapp_assignment_url(
                assignment["vehicle_name"] or route.get("vehicle_name") or "",
                route_url,
                public_driver_url,
            ),
            "expires_at": str(datetime.utcnow() + timedelta(hours=48)),
        }
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        logger.exception("share_route_assignment failed")
        raise HTTPException(status_code=500, detail="Не удалось подготовить ссылку")
    finally:
        cur.close(); conn.close()


@app.patch("/api/route/assignments/{assignment_id}")
def update_route_assignment(assignment_id: int, body: AssignmentUpdate, request: Request):
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    fields, values = [], []
    if body.driver_name is not None:
        fields.append("driver_name=%s"); values.append(body.driver_name.strip())
    if body.driver_id is not None:
        cur.execute(
            "SELECT id, name, phone, vehicle_name FROM drivers WHERE id=%s AND owner_id=%s AND is_active=TRUE",
            (body.driver_id, uid),
        )
        driver = cur.fetchone()
        if not driver:
            raise HTTPException(status_code=422, detail="Выбранный водитель не найден или неактивен")
        fields.extend(["driver_id=%s", "driver_name=%s", "driver_phone=%s"])
        values.extend([driver["id"], driver["name"], driver["phone"] or ""])
        if body.vehicle_name is None and driver["vehicle_name"]:
            fields.append("vehicle_name=%s"); values.append(driver["vehicle_name"])
    if body.vehicle_name is not None:
        fields.append("vehicle_name=%s"); values.append(body.vehicle_name.strip())
    if not fields:
        raise HTTPException(status_code=422, detail="Нет изменений")
    fields.append("updated_at=NOW()")
    values.extend([assignment_id, uid])
    cur.execute(
        f"""UPDATE route_assignments SET {', '.join(fields)}
            WHERE id=%s AND owner_id=%s
            RETURNING id, session_id, route_index, driver_id, driver_name, driver_phone, vehicle_name,
                      route_yandex_url, status, expires_at, updated_at""",
        values,
    )
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Рейс не найден")
    conn.commit(); cur.close(); conn.close()
    return _assignment_summary(dict(row))


@app.patch("/api/route/assignments/{assignment_id}/executions/{execution_id}")
def update_dispatcher_execution(
    assignment_id: int,
    execution_id: int,
    body: DispatcherExecutionUpdate,
    request: Request,
):
    """Set the new delivery date for a rescheduled point."""
    uid = get_user_id(request)
    rescheduled_date = body.rescheduled_date
    if not rescheduled_date:
        raise HTTPException(status_code=422, detail="Укажите новую дату доставки")
    try:
        datetime.strptime(rescheduled_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=422, detail="Дата должна быть в формате YYYY-MM-DD")
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """UPDATE route_executions e
           SET rescheduled_date=%s, updated_at=NOW()
           FROM route_assignments a
           WHERE e.id=%s AND e.assignment_id=%s
             AND a.id=e.assignment_id AND a.owner_id=%s
             AND e.status='rescheduled'
           RETURNING e.id, e.store_id, e.visit_order, e.store_name, e.address,
                     e.lat, e.lon, e.products, e.quantity, e.actual_qty,
                     e.arrive_by, e.yandex_url, e.status, e.payment_method,
                     e.payment_status, e.driver_comment, e.rescheduled_date,
                     e.updated_at, e.delivered_at""",
        (rescheduled_date, execution_id, assignment_id, uid),
    )
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(
            status_code=404,
            detail="Перенесённая заявка не найдена или её статус уже изменён",
        )
    conn.commit(); cur.close(); conn.close()
    return {"execution": _serialize_execution(dict(row))}


@app.post("/api/route/assignments/{assignment_id}/executions/{execution_id}/rescheduled-order", status_code=201)
def create_rescheduled_order(
    assignment_id: int,
    execution_id: int,
    body: DispatcherCreateRescheduledOrder,
    request: Request,
):
    """Create a new daily order after a driver explicitly reschedules a point."""
    uid = get_user_id(request)
    try:
        datetime.strptime(body.delivery_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=422, detail="Дата должна быть в формате YYYY-MM-DD")

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """SELECT e.id, e.store_id, e.store_name, e.address, e.products,
                      e.quantity, e.actual_qty, e.status
                 FROM route_executions e
                 JOIN route_assignments a ON a.id=e.assignment_id
                WHERE e.id=%s AND e.assignment_id=%s AND a.owner_id=%s""",
            (execution_id, assignment_id, uid),
        )
        execution = cur.fetchone()
        if not execution:
            raise HTTPException(status_code=404, detail="Точка рейса не найдена")
        if execution["status"] != "rescheduled":
            raise HTTPException(status_code=422, detail="Создать новую заявку можно только для перенесённой точки")
        if not execution["store_id"]:
            raise HTTPException(status_code=422, detail="У точки нет связанного магазина")

        cur.execute(
            """SELECT id, store_id, store_name_raw, delivery_date::text
                 FROM daily_orders
                WHERE owner_id=%s AND store_id=%s AND delivery_date=%s
                LIMIT 1""",
            (uid, execution["store_id"], body.delivery_date),
        )
        existing = cur.fetchone()
        if existing:
            order = dict(existing)
        else:
            cur.execute(
                """INSERT INTO daily_orders
                       (owner_id, store_id, store_name_raw, address_raw,
                        delivery_date, quantity, products, notes)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                   RETURNING id, store_id, store_name_raw, delivery_date::text""",
                (
                    uid,
                    execution["store_id"],
                    execution["store_name"] or "",
                    execution["address"] or "",
                    body.delivery_date,
                    max(
                        float(execution["quantity"] or 0)
                        - float(execution["actual_qty"] or 0),
                        0,
                    ),
                    execution["products"] or "",
                    "Создано из переноса доставки",
                ),
            )
            order = dict(cur.fetchone())

        cur.execute(
            """UPDATE route_executions
                  SET rescheduled_date=%s, updated_at=NOW()
                WHERE id=%s AND assignment_id=%s
            RETURNING id, rescheduled_date""",
            (body.delivery_date, execution_id, assignment_id),
        )
        updated = dict(cur.fetchone())
        conn.commit()
        return {"order": order, "execution": updated}
    except HTTPException:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


@app.post("/api/route/assignments/{assignment_id}/executions/{execution_id}/remaining-order", status_code=201)
def create_remaining_order(
    assignment_id: int,
    execution_id: int,
    body: DispatcherCreateRescheduledOrder,
    request: Request,
):
    """Create a daily order containing only the quantity not delivered."""
    uid = get_user_id(request)
    try:
        datetime.strptime(body.delivery_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=422, detail="Дата должна быть в формате YYYY-MM-DD")

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """SELECT e.id, e.store_id, e.store_name, e.address, e.products,
                      e.quantity, e.actual_qty, e.status
                 FROM route_executions e
                 JOIN route_assignments a ON a.id=e.assignment_id
                WHERE e.id=%s AND e.assignment_id=%s AND a.owner_id=%s""",
            (execution_id, assignment_id, uid),
        )
        execution = cur.fetchone()
        if not execution:
            raise HTTPException(status_code=404, detail="Точка рейса не найдена")
        if execution["status"] not in {"partial", "failed"}:
            raise HTTPException(
                status_code=422,
                detail="Заявку на остаток можно создать только для частичной или недоставленной точки",
            )
        if not execution["store_id"]:
            raise HTTPException(status_code=422, detail="У точки нет связанного магазина")

        planned_qty = float(execution["quantity"] or 0)
        actual_qty = float(execution["actual_qty"] or 0)
        remaining_qty = max(planned_qty - actual_qty, 0)
        if remaining_qty <= 0:
            raise HTTPException(status_code=422, detail="У точки нет недоставленного количества")
        remaining_products, remaining_notes = _remaining_order_products(
            execution["products"],
            planned_qty,
            actual_qty,
            remaining_qty,
        )

        cur.execute(
            """INSERT INTO daily_orders
                   (owner_id, store_id, store_name_raw, address_raw,
                    delivery_date, quantity, products, notes)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
               RETURNING id, store_id, store_name_raw, delivery_date::text, quantity""",
            (
                uid,
                execution["store_id"],
                execution["store_name"] or "",
                execution["address"] or "",
                body.delivery_date,
                remaining_qty,
                remaining_products,
                remaining_notes,
            ),
        )
        order = dict(cur.fetchone())
        conn.commit()
        return {"order": order, "remaining_qty": remaining_qty}
    except HTTPException:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


@app.post("/api/driver/{token}/location")
def update_driver_location(token: str, body: DriverLocationInput):
    """Store the latest phone position for a token-scoped assignment."""
    _api_rate_limit(f"driver_location:{hashlib.sha256(token.encode()).hexdigest()}", 6, 60)
    if not (-90 <= body.lat <= 90 and -180 <= body.lon <= 180):
        raise HTTPException(status_code=422, detail="Некорректные координаты")
    if body.accuracy is not None and (body.accuracy < 0 or body.accuracy > 10000):
        raise HTTPException(status_code=422, detail="Некорректная точность геолокации")
    assignment, _ = _load_assignment_for_token(token)
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            """INSERT INTO driver_locations (assignment_id, lat, lon, accuracy, captured_at)
               VALUES (%s,%s,%s,%s,NOW())
               ON CONFLICT (assignment_id) DO UPDATE SET
                 lat=EXCLUDED.lat, lon=EXCLUDED.lon, accuracy=EXCLUDED.accuracy,
                 captured_at=EXCLUDED.captured_at""",
            (assignment["id"], body.lat, body.lon, body.accuracy),
        )
        conn.commit()
        return {"ok": True, "captured_at": datetime.utcnow().isoformat()}
    finally:
        cur.close(); conn.close()


@app.get("/api/driver/{token}")
def get_driver_assignment(token: str):
    """Public, token-scoped driver view. No account or cookie is accepted."""
    _api_rate_limit(f"driver_get:{hashlib.sha256(token.encode()).hexdigest()}", 120, 60)
    assignment, executions = _load_assignment_for_token(token)
    terminal = {"delivered", "partial", "failed", "rescheduled"}
    completed = sum(1 for row in executions if row.get("status") in terminal)
    next_stop = next((row for row in executions if row.get("status") == "planned"), None)
    return {
        "assignment": {
            "id": assignment["id"],
            "driver_name": assignment.get("driver_name") or "",
            "vehicle_name": assignment.get("vehicle_name") or "",
            "route_yandex_url": assignment.get("route_yandex_url") or "",
            "status": "planned" if assignment.get("status") == "on_route" else (assignment.get("status") or "planned"),
            "total_points": len(executions),
            "completed_points": completed,
            "next_stop": {
                "store_name": next_stop.get("store_name") or "",
                "address": next_stop.get("address") or "",
            } if next_stop else None,
        },
        "executions": [_serialize_execution(row) for row in executions],
    }


@app.patch("/api/driver/{token}/executions/{execution_id}")
def update_driver_execution(token: str, execution_id: int, body: ExecutionUpdate):
    """Update one delivery point through the scoped driver link."""
    if body.status not in EXECUTION_STATUSES:
        raise HTTPException(status_code=422, detail="Недопустимый статус доставки")
    assignment, executions = _load_assignment_for_token(token)
    current = next((row for row in executions if int(row["id"]) == execution_id), None)
    if not current:
        raise HTTPException(status_code=404, detail="Точка рейса не найдена")
    planned_qty = float(current.get("quantity") or 0)
    quantity_required_statuses = {"delivered", "partial"}
    if body.status in quantity_required_statuses and body.actual_qty is None:
        raise HTTPException(status_code=422, detail="Укажите фактически доставленное количество")
    actual_qty = float(
        0 if body.actual_qty is None and body.status not in quantity_required_statuses
        else current.get("actual_qty") or 0
        if body.actual_qty is None
        else body.actual_qty
    )
    # A few legacy executions were created with quantity=0 even though the
    # driver had a real order. Accept their first delivered quantity and
    # repair the execution row instead of returning a misleading 422.
    if body.status == "delivered" and planned_qty <= 0 and actual_qty > 0:
        planned_qty = actual_qty
    if body.status in quantity_required_statuses or body.actual_qty is not None:
        _validate_execution_quantities(body.status, planned_qty, actual_qty)
    if body.status == "rescheduled" and not body.driver_comment.strip():
        raise HTTPException(status_code=422, detail="Для переноса укажите причину")
    payment_method = body.payment_method or current.get("payment_method") or "none"
    payment_status = body.payment_status or current.get("payment_status") or "pending"
    if payment_method not in PAYMENT_METHODS:
        raise HTTPException(status_code=422, detail="Недопустимый способ оплаты")
    if payment_status not in PAYMENT_STATUSES:
        raise HTTPException(status_code=422, detail="Недопустимый статус оплаты")
    if payment_status == "paid" and payment_method == "none":
        raise HTTPException(
            status_code=422,
            detail="Для оплаченной доставки укажите способ оплаты",
        )
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """UPDATE route_executions
           SET status=%s, quantity=CASE WHEN quantity <= 0 AND %s='delivered'
                                        THEN %s ELSE quantity END,
               actual_qty=%s, payment_method=%s, payment_status=%s,
               driver_comment=%s,
               rescheduled_date=CASE WHEN %s='rescheduled' THEN rescheduled_date ELSE NULL END,
               updated_at=NOW(),
               delivered_at=CASE WHEN %s IN ('delivered','partial','failed')
                                 THEN COALESCE(delivered_at, NOW()) ELSE delivered_at END
           WHERE id=%s AND assignment_id=%s
           RETURNING id, store_id, visit_order, store_name, address, lat, lon,
                     products, quantity, actual_qty, arrive_by, yandex_url, status,
                     payment_method, payment_status, driver_comment, rescheduled_date,
                     updated_at, delivered_at""",
        (body.status, body.status, planned_qty, actual_qty, payment_method, payment_status, body.driver_comment.strip(),
         body.status,
         body.status, execution_id, assignment["id"]),
    )
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Точка рейса не найдена")
    cur.execute(
        """SELECT COUNT(*) AS total,
                  COUNT(*) FILTER (WHERE status IN
                    ('delivered','partial','failed','rescheduled')) AS completed,
                  COUNT(*) FILTER (WHERE status NOT IN
                    ('delivered','partial','failed','rescheduled','planned')) AS active
           FROM route_executions WHERE assignment_id=%s""",
        (assignment["id"],),
    )
    counts = cur.fetchone()
    total, completed = int(counts["total"]), int(counts["completed"])
    assignment_status = "completed" if total and completed == total else "planned"
    cur.execute(
        "UPDATE route_assignments SET status=%s, updated_at=NOW() WHERE id=%s",
        (assignment_status, assignment["id"]),
    )
    conn.commit(); cur.close(); conn.close()
    return {
        "execution": _serialize_execution(dict(row)),
        "assignment_status": assignment_status,
        "completed_points": completed,
        "total_points": total,
    }


@app.delete("/api/route/sessions/{id}")
def delete_route_session(id: int, request: Request):
    """Удалить сессию маршрута по ID."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM route_sessions WHERE id = %s AND owner_id = %s", (id, uid))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Route session not found")
        conn.commit()
        return {"ok": True}
    finally:
        cur.close()
        conn.close()


@app.get("/api/route/sessions")
def list_route_sessions(request: Request, page: int = 1, page_size: int = 20):
    """Список сессий маршрутов с пагинацией."""
    uid = get_user_id(request)
    if page < 1:
        page = 1
    if page_size < 1 or page_size > 200:
        page_size = 20
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT COUNT(*) as total FROM route_sessions WHERE owner_id = %s", (uid,))
        total = int(cur.fetchone()["total"])
        offset = (page - 1) * page_size
        cur.execute(
            """SELECT id, date, num_vehicles, total_km, saved_km, saved_rub, num_points, created_at
               FROM route_sessions
               WHERE owner_id = %s
               ORDER BY created_at DESC
               LIMIT %s OFFSET %s""",
            (uid, page_size, offset),
        )
        rows = cur.fetchall()
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "items": [
                {
                    "id": r["id"],
                    "date": r["date"],
                    "num_vehicles": r["num_vehicles"] or 0,
                    "total_km": round(float(r["total_km"] or 0), 1),
                    "saved_km": round(float(r["saved_km"] or 0), 1),
                    "saved_rub": int(r["saved_rub"] or 0),
                    "num_points": r["num_points"] or 0,
                    "created_at": str(r["created_at"]) if r["created_at"] else None,
                }
                for r in rows
            ],
        }
    finally:
        cur.close()
        conn.close()


@app.get("/api/settings")
def get_settings_endpoint(request: Request):
    """Получить текущие параметры расчёта стоимости км (для текущего пользователя)."""
    uid = get_user_id(request)
    return get_company_settings(user_id=uid)


@app.put("/api/settings")
def update_settings_endpoint(request: Request, body: CompanySettingsInput):
    """Обновить параметры расчёта стоимости км. cost_per_km = fuel_price × consumption / 100."""
    uid = get_user_id(request)
    if body.fuel_price <= 0 or body.fuel_consumption <= 0:
        raise HTTPException(status_code=400, detail="Все параметры должны быть положительными числами")
    cost_per_km = round(body.fuel_price * body.fuel_consumption / 100.0, 2)
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE company_settings
               SET fuel_price=%s, fuel_consumption=%s, cost_per_km=%s, updated_at=NOW()
             WHERE owner_id=%s
        """, (body.fuel_price, body.fuel_consumption, cost_per_km, uid))
        if cur.rowcount == 0:
            cur.execute("""
                INSERT INTO company_settings (fuel_price, fuel_consumption, cost_per_km, owner_id)
                VALUES (%s, %s, %s, %s)
            """, (body.fuel_price, body.fuel_consumption, cost_per_km, uid))
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return {
        "fuel_price": body.fuel_price,
        "fuel_consumption": body.fuel_consumption,
        "cost_per_km": cost_per_km,
    }


# ══════════════════════════════════════════════════════════════════════════════
# Integrations
# ══════════════════════════════════════════════════════════════════════════════

# 1C BSL module template v3.0 — placeholders {{BASE_URL}} and {{API_KEY}} are
# filled dynamically when the user downloads the personalised ZIP package.
# Single source of truth: this template is used for both the quick-setup ZIP
# and the /download-module endpoint.
_1C_BSL_MODULE = '''// ╔══════════════════════════════════════════════════════════════════════╗
// ║   SmartRoute — Модуль интеграции 1С:Предприятие 8.3+               ║
// ║   Версия: 3.1  |  Совместимость: 8.3.14+ (включая Учебную)         ║
// ╠══════════════════════════════════════════════════════════════════════╣
// ║  КАК УСТАНОВИТЬ В EPF:                                               ║
// ║   1. Конфигуратор → Файл → Новый → Внешняя обработка                ║
// ║   2. Имя объекта: SmartRoute                                         ║
// ║   3. Формы → Добавить → Произвольная форма → ОК                     ║
// ║   4. Разместите на форме три элемента:                               ║
// ║      а) Поле типа «Дата»    → Имя: ПолеДатаДоставки                ║
// ║      б) Кнопка «Проверить»  → Команда: СмартРоутПроверить           ║
// ║      в) Кнопка «Отправить»  → Команда: СмартРоутОтправить           ║
// ║   5. Форма → вкладка «Модуль» → вставьте содержимое этого файла     ║
// ║   6. Заполните функцию _ПолучитьНастройки() ниже                    ║
// ║   7. Файл → Сохранить как → Внешняя обработка (*.epf)               ║
// ║                                                                       ║
// ║  РЕГЛАМЕНТНОЕ ЗАДАНИЕ: скопируйте тело ОтправитьЗаявки() в модуль   ║
// ║  ОБЪЕКТА обработки (не формы), без директивы &НаСервере.             ║
// ╚══════════════════════════════════════════════════════════════════════╝

// ══════════════════════════════════════════════════════════════════
// ШАГ 1 — ЕДИНСТВЕННОЕ МЕСТО ДЛЯ НАСТРОЕК
// Откройте эту функцию и заполните значения.
// ══════════════════════════════════════════════════════════════════

&НаСервере
Функция _ПолучитьНастройки()
	Н = Новый Структура;

	// ── Подключение к SmartRoute ──────────────────────────────────────
	Н.Вставить("URL",   "{{BASE_URL}}");  // напр: https://xxx.smartroute.app
	Н.Вставить("Ключ",  "{{API_KEY}}");           // напр: sr_live_xxxx
	Н.Вставить("Город", "Махачкала");                    // город доставки по умолчанию

	// ── Документ заказа в 1С ──────────────────────────────────────────
	// Найти правильное имя: Конфигуратор → Документы → ПКМ → Свойства → поле «Имя»
	Н.Вставить("Документ",  "ЗаказКлиента");   // пример: ЗаказКлиента, ЗаказПокупателя

	// Имена реквизитов документа (столбец «Имя», НЕ «Синоним»)
	Н.Вставить("РеквДата",  "ДатаДоставки");    // реквизит с датой доставки
	Н.Вставить("РеквАдрес", "АдресДоставки");   // реквизит с адресом доставки
	Н.Вставить("РеквКол",   "КоличествоМест");  // реквизит с количеством мест
	Н.Вставить("РеквВес",   "ВесКГ");           // реквизит с весом (кг), 0 если нет

	Возврат Н;
КонецФункции

// ══════════════════════════════════════════════════════════════════
// ФОРМА — инициализация при открытии
// ══════════════════════════════════════════════════════════════════

&НаСервере
Процедура ПриСозданиеНаСервере(Отказ, СтандартнаяОбработка)
	Элементы.ПолеДатаДоставки.Значение = НачалоДня(ТекущаяДата());
КонецПроцедуры

// ══════════════════════════════════════════════════════════════════
// КНОПКА: ПРОВЕРИТЬ СОЕДИНЕНИЕ
// ══════════════════════════════════════════════════════════════════

&НаКлиенте
Процедура СмартРоутПроверить(Команда)
	Рез = ПроверитьСоединениеНаСервере();
	Предупреждение(Рез, , "SmartRoute — Проверка соединения");
КонецПроцедуры

&НаСервере
Функция ПроверитьСоединениеНаСервере()
	Н = _ПолучитьНастройки();

	Если ПустаяСтрока(Н.URL) Или Н.URL = "{{BASE_URL}}" Тогда
		Возврат "[НАСТРОЙКИ] Не заполнен URL сервера SmartRoute." + Символы.ПС
			+ "Откройте функцию _ПолучитьНастройки() и укажите адрес.";
	КонецЕсли;

	Если ПустаяСтрока(Н.Ключ) Или Н.Ключ = "{{API_KEY}}" Тогда
		Возврат "[НАСТРОЙКИ] Не заполнен API-ключ SmartRoute." + Символы.ПС
			+ "Вставьте ключ в поле Ключ функции _ПолучитьНастройки().";
	КонецЕсли;

	Попытка
		Соединение = _ПолучитьСоединение(Н.URL);
		Запрос = Новый HTTPЗапрос("/api/healthz", _Заголовки(Н.Ключ));
		Ответ = Соединение.Получить(Запрос);
	Исключение
		Возврат "[ОШИБКА СЕТИ] Нет связи с сервером SmartRoute." + Символы.ПС
			+ "Адрес: " + Н.URL + Символы.ПС + Символы.ПС
			+ "Проверьте:" + Символы.ПС
			+ "  1. Правильность адреса в _ПолучитьНастройки()" + Символы.ПС
			+ "  2. Доступ в интернет с этого компьютера (порт 443)" + Символы.ПС
			+ "  3. Брандмауэр не блокирует HTTPS" + Символы.ПС + Символы.ПС
			+ "Ошибка: " + ОписаниеОшибки();
	КонецПопытки;

	Если Ответ.КодСостояния = 200 Тогда
		КолМаг = "?";
		Попытка
			ЗапМаг = Новый HTTPЗапрос("/api/v1/stores?page_size=1", _Заголовки(Н.Ключ));
			ОтвМаг = Соединение.Получить(ЗапМаг);
			ДанМаг = _ЧтениеJSON(ОтвМаг.ПолучитьТелоКакСтроку());
			Если ТипЗнч(ДанМаг) = Тип("Соответствие") Тогда
				Мета = ДанМаг.Получить("meta");
				Если Мета <> Неопределено Тогда
					Итого = Мета.Получить("total");
					Если Итого <> Неопределено Тогда
						КолМаг = "" + Итого;
					КонецЕсли;
				КонецЕсли;
			КонецЕсли;
		Исключение
		КонецПопытки;

		Возврат "[OK] Соединение установлено!" + Символы.ПС + Символы.ПС
			+ "Адрес SmartRoute:             " + Н.URL + Символы.ПС
			+ "Магазинов в базе SmartRoute:  " + КолМаг + Символы.ПС + Символы.ПС
			+ "Можно настраивать регламентное задание.";

	ИначеЕсли Ответ.КодСостояния = 401 Тогда
		Возврат "[ОШИБКА 401] Неверный API-ключ." + Символы.ПС + Символы.ПС
			+ "Сгенерируйте новый пакет: SmartRoute -> Интеграции -> 1С -> Пересобрать пакет." + Символы.ПС
			+ "Скопируйте новый ключ в _ПолучитьНастройки().";

	ИначеЕсли Ответ.КодСостояния = 403 Тогда
		Возврат "[ОШИБКА 403] API-ключ не имеет права orders:write." + Символы.ПС
			+ "Создайте новый ключ через SmartRoute -> Интеграции -> 1С.";

	Иначе
		Возврат "[ОШИБКА " + Ответ.КодСостояния + "]" + Символы.ПС
			+ Лев(Ответ.ПолучитьТелоКакСтроку(), 300);
	КонецЕсли;
КонецФункции

// ══════════════════════════════════════════════════════════════════
// КНОПКА: ОТПРАВИТЬ ЗАЯВКИ (ручной запуск с формы)
// ══════════════════════════════════════════════════════════════════

&НаКлиенте
Процедура СмартРоутОтправить(Команда)
	ДатаДоставки = Элементы.ПолеДатаДоставки.Значение;
	Если НЕ ЗначениеЗаполнено(ДатаДоставки) Тогда
		Предупреждение("Укажите дату доставки!");
		Возврат;
	КонецЕсли;
	Рез = ОтправитьЗаявкиНаСервере(НачалоДня(ДатаДоставки));
	Предупреждение(Рез, , "SmartRoute — Результат отправки");
КонецПроцедуры

// ══════════════════════════════════════════════════════════════════
// РЕГЛАМЕНТНОЕ ЗАДАНИЕ
//
// Чтобы использовать как регламентное задание:
//   Скопируйте тело этой процедуры в МОДУЛЬ ОБЪЕКТА обработки
//   (Конфигуратор → обработка → Модуль объекта), уберите &НаСервере.
//   В регламентном задании укажите: Метод = ОтправитьЗаявки
// ══════════════════════════════════════════════════════════════════

&НаСервере
Процедура ОтправитьЗаявки() Экспорт
	ДатаДоставки = НачалоДня(ТекущаяДата());
	Рез = ОтправитьЗаявкиНаСервере(ДатаДоставки);
	ЗаписьЖурналаРегистрации(
		"SmartRoute.Синхронизация",
		УровеньЖурналаРегистрации.Информация,
		, ,
		"Дата: " + Формат(ДатаДоставки, "ДФ=дд.ММ.гггг") + " | " + Рез
	);
КонецПроцедуры

// ══════════════════════════════════════════════════════════════════
// ОСНОВНАЯ ФУНКЦИЯ ОТПРАВКИ ЗАЯВОК
// ══════════════════════════════════════════════════════════════════

&НаСервере
Функция ОтправитьЗаявкиНаСервере(ДатаДоставки)
	Н = _ПолучитьНастройки();

	Если ПустаяСтрока(Н.URL) Или Н.URL = "{{BASE_URL}}" Тогда
		Возврат "[НАСТРОЙКИ] Не заполнен URL. Откройте _ПолучитьНастройки().";
	КонецЕсли;
	Если ПустаяСтрока(Н.Ключ) Или Н.Ключ = "{{API_KEY}}" Тогда
		Возврат "[НАСТРОЙКИ] Не заполнен API-ключ. Откройте _ПолучитьНастройки().";
	КонецЕсли;

	// ── 1. Получить заказы из базы 1С ────────────────────────────────
	// counterparty_code (Контрагент.Код) — ключ совпадения в SmartRoute.
	// Не удаляйте это поле из запроса.
	ТекстЗапроса =
		"ВЫБРАТЬ"                                                     + Символы.ПС
		+ "|	З.Контрагент.Наименование  КАК Магазин,"               + Символы.ПС
		+ "|	З.Контрагент.Код           КАК КодКонтрагента,"        + Символы.ПС
		+ "|	З.Контрагент.Телефон       КАК Телефон,"               + Символы.ПС
		+ "|	З." + Н.РеквАдрес + "     КАК Адрес,"                 + Символы.ПС
		+ "|	З." + Н.РеквКол   + "     КАК Количество,"            + Символы.ПС
		+ "|	З." + Н.РеквВес   + "     КАК Вес,"                   + Символы.ПС
		+ "|	З.Комментарий              КАК Комментарий,"           + Символы.ПС
		+ "|	З.Номер                    КАК НомерДокумента"         + Символы.ПС
		+ "|ИЗ"                                                       + Символы.ПС
		+ "|	Документ." + Н.Документ + " КАК З"                    + Символы.ПС
		+ "|ГДЕ"                                                      + Символы.ПС
		+ "|	З." + Н.РеквДата + " = &ДатаДоставки"                  + Символы.ПС
		+ "|	И НЕ З.ПометкаУдаления"                               + Символы.ПС
		+ "|	И З.Проведён"                                          + Символы.ПС
		+ "|УПОРЯДОЧИТЬ ПО"                                          + Символы.ПС
		+ "|	З.Контрагент.Наименование";

	ОбъектЗапроса = Новый Запрос;
	ОбъектЗапроса.Текст = ТекстЗапроса;
	ОбъектЗапроса.УстановитьПараметр("ДатаДоставки", ДатаДоставки);

	Попытка
		РезультатЗапроса = ОбъектЗапроса.Выполнить().Выгрузить();
	Исключение
		Сообщ = "[ОШИБКА БД] Ошибка при выполнении запроса:" + Символы.ПС
			+ ОписаниеОшибки() + Символы.ПС + Символы.ПС
			+ "Проверьте имена реквизитов в _ПолучитьНастройки():" + Символы.ПС
			+ "  Документ  = " + Н.Документ  + Символы.ПС
			+ "  РеквДата  = " + Н.РеквДата  + Символы.ПС
			+ "  РеквАдрес = " + Н.РеквАдрес + Символы.ПС
			+ "  РеквКол   = " + Н.РеквКол   + Символы.ПС
			+ "  РеквВес   = " + Н.РеквВес;
		ЗаписьЖурналаРегистрации("SmartRoute.Ошибка",
			УровеньЖурналаРегистрации.Ошибка, , , Сообщ);
		Возврат Сообщ;
	КонецПопытки;

	Если РезультатЗапроса.Количество() = 0 Тогда
		Возврат "[ИНФО] Нет проведённых заказов на "
			+ Формат(ДатаДоставки, "ДФ=дд.ММ.гггг") + "." + Символы.ПС
			+ "Убедитесь, что документы проведены и реквизит "
			+ Н.РеквДата + " = " + Формат(ДатаДоставки, "ДФ=дд.ММ.гггг") + ".";
	КонецЕсли;

	// ── 2. Собрать массив заявок ──────────────────────────────────────
	ДатаСтрокой   = Формат(ДатаДоставки, "ДФ=гггг-ММ-дд");
	МассивЗаказов = Новый Массив;
	Пропущено     = 0;

	Для Каждого Строка Из РезультатЗапроса Цикл
		Если ПустаяСтрока(Строка.Магазин) Тогда
			Пропущено = Пропущено + 1;
			Продолжить;
		КонецЕсли;

		Кол = Строка.Количество;
		Если Кол <= 0 Тогда Кол = 1; КонецЕсли;

		Вес = Строка.Вес;
		Если НЕ ЗначениеЗаполнено(Вес) Тогда Вес = 0; КонецЕсли;

		Заказ = Новый Структура;
		Заказ.Вставить("store_name",        СокрЛП(Строка.Магазин));
		Заказ.Вставить("delivery_date",     ДатаСтрокой);
		Заказ.Вставить("counterparty_code", СокрЛП("" + Строка.КодКонтрагента));
		Заказ.Вставить("address",           СокрЛП("" + Строка.Адрес));
		Заказ.Вставить("city",              Н.Город);
		Заказ.Вставить("phone",             СокрЛП("" + Строка.Телефон));
		Заказ.Вставить("quantity",          Кол);
		Заказ.Вставить("weight_kg",         Вес);
		Заказ.Вставить("products",          СокрЛП("" + Строка.Комментарий));
		Заказ.Вставить("order_number",      СокрЛП("" + Строка.НомерДокумента));
		Заказ.Вставить("external_id",       СокрЛП("" + Строка.НомерДокумента));
		МассивЗаказов.Добавить(Заказ);
	КонецЦикла;

	Если МассивЗаказов.Количество() = 0 Тогда
		Возврат "[ИНФО] Нет заявок для отправки (все строки без контрагента — "
			+ Пропущено + " шт.)";
	КонецЕсли;

	// ── 3. Отправить в SmartRoute (3 попытки при сетевой ошибке) ─────
	Тело = Новый Структура;
	Тело.Вставить("orders",             МассивЗаказов);
	Тело.Вставить("replace_date",       Истина);
	Тело.Вставить("auto_create_stores", Истина);

	СтрокаJSON  = _ЗаписьJSON(Тело);
	МаксПопыток = 3;
	Соединение  = _ПолучитьСоединение(Н.URL);
	Ответ       = Неопределено;

	Для НомерПопытки = 1 По МаксПопыток Цикл
		Попытка
			ЗапросHTTP = Новый HTTPЗапрос("/api/v1/orders/batch", _Заголовки(Н.Ключ));
			ЗапросHTTP.УстановитьТелоИзСтроки(СтрокаJSON,
				КодировкаТекста.UTF8, БайтПорядокUnicode.НетМаркера);
			Ответ = Соединение.ОтправитьДляОбработки(ЗапросHTTP);
			Прервать; // успех — выходим из цикла повторов
		Исключение
			Если НомерПопытки >= МаксПопыток Тогда
				СообщОш = "[ОШИБКА СЕТИ] Нет связи с SmartRoute после "
					+ МаксПопыток + " попыток." + Символы.ПС
					+ "Проверьте доступ к интернету (порт 443)." + Символы.ПС
					+ "Ошибка: " + ОписаниеОшибки();
				ЗаписьЖурналаРегистрации("SmartRoute.Ошибка",
					УровеньЖурналаРегистрации.Ошибка, , , СообщОш);
				Возврат СообщОш;
			КонецЕсли;
			// Пауза ~3 сек между попытками (без ОбработкаПрерыванияПользователя —
			// эта функция недоступна в серверном контексте &НаСервере).
			Пауза = ТекущаяДата();
			Пока ТекущаяДата() - Пауза < 3 Цикл КонецЦикла;
		КонецПопытки;
	КонецЦикла;

	// ── 4. Разобрать ответ сервера ────────────────────────────────────
	Если Ответ.КодСостояния = 200 Или Ответ.КодСостояния = 201 Тогда
		Данные = _ЧтениеJSON(Ответ.ПолучитьТелоКакСтроку());

		// Соответствие — используем .Получить(ключ), НЕ .Свойство(ключ)
		// (.Свойство() — метод Структуры; у Соответствия его нет)
		ЗначСоздано  = Данные.Получить("created");
		ЗначСовпало  = Данные.Получить("matched");
		ЗначНовых    = Данные.Получить("auto_created_stores");
		ЗначПропущ   = Данные.Получить("skipped");

		Создано    = ?(ЗначСоздано  <> Неопределено, ЗначСоздано,  0);
		Совпало    = ?(ЗначСовпало  <> Неопределено, ЗначСовпало,  0);
		НовыхТочек = ?(ЗначНовых    <> Неопределено, ЗначНовых,    0);
		Пропущено2 = ?(ЗначПропущ   <> Неопределено, ЗначПропущ,   0);

		Результат = "[OK] Заявки переданы в SmartRoute!" + Символы.ПС + Символы.ПС
			+ "Загружено заявок:          " + Создано    + Символы.ПС
			+ "Совпало точек доставки:    " + Совпало;

		Если НовыхТочек > 0 Тогда
			Результат = Результат + Символы.ПС
				+ "Новых точек создано:       " + НовыхТочек + Символы.ПС
				+ "(требуется геокодирование: SmartRoute -> Магазины)";
		КонецЕсли;

		Если Пропущено2 > 0 Тогда
			Результат = Результат + Символы.ПС
				+ "Пропущено (нет адреса):    " + Пропущено2;
		КонецЕсли;

		ЗаписьЖурналаРегистрации("SmartRoute.Синхронизация",
			УровеньЖурналаРегистрации.Информация, , ,
			"Дата: " + Формат(ДатаДоставки, "ДФ=дд.ММ.гггг") + " | " + Результат);
		Возврат Результат;

	ИначеЕсли Ответ.КодСостояния = 401 Тогда
		Возврат "[ОШИБКА 401] Неверный API-ключ SmartRoute." + Символы.ПС
			+ "Сгенерируйте новый пакет в SmartRoute -> Интеграции -> 1С.";

	ИначеЕсли Ответ.КодСостояния = 403 Тогда
		Возврат "[ОШИБКА 403] Ключ не имеет права orders:write." + Символы.ПС
			+ "Создайте новый ключ через SmartRoute -> Интеграции -> 1С.";

	ИначеЕсли Ответ.КодСостояния = 422 Тогда
		Возврат "[ОШИБКА 422] Ошибка в данных:" + Символы.ПС
			+ Лев(Ответ.ПолучитьТелоКакСтроку(), 500) + Символы.ПС
			+ "Проверьте имена реквизитов в _ПолучитьНастройки().";

	ИначеЕсли Ответ.КодСостояния = 429 Тогда
		Возврат "[ПРЕДУПРЕЖДЕНИЕ 429] Слишком много запросов. Подождите минуту.";

	Иначе
		Возврат "[ОШИБКА " + Ответ.КодСостояния + "]" + Символы.ПС
			+ Лев(Ответ.ПолучитьТелоКакСтроку(), 300);
	КонецЕсли;
КонецФункции

// ══════════════════════════════════════════════════════════════════
// ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ — не изменять
// ══════════════════════════════════════════════════════════════════

// Создаёт HTTPСоединение для заданного URL.
//
// СОВМЕСТИМОСТЬ 8.3.27 / Учебная редакция:
//   ЗащищённоеСоединениеOpenSSL определяется ДИНАМИЧЕСКИ через Тип().
//   Это исключает ошибку компиляции "Тип не определён" в редакциях,
//   где OpenSSL-библиотека не поставляется с платформой 1С.
//   При отсутствии OpenSSL соединение создаётся без явного SSL-объекта;
//   на Windows платформа использует системный WinInet/Schannel для HTTPS.
&НаСервере
Функция _ПолучитьСоединение(URL)
	HTTPS    = (НРег(Лев(URL, 8)) = "https://");
	БезПреф  = СтрЗаменить(СтрЗаменить(URL, "https://", ""), "http://", "");
	ПозСлэш  = СтрНайти(БезПреф, "/");
	Если ПозСлэш > 0 Тогда
		Хост = Лев(БезПреф, ПозСлэш - 1);
	Иначе
		Хост = БезПреф;
	КонецЕсли;
	Порт = ?(HTTPS, 443, 80);

	Если HTTPS Тогда
		// Тип("...") возвращает Неопределено если тип не зарегистрирован —
		// это runtime-проверка, которая НЕ вызывает ошибок компиляции.
		ТипSSL = Тип("ЗащищённоеСоединениеOpenSSL");
		Если ТипSSL <> Неопределено Тогда
			// OpenSSL доступен. Новый(ТипSSL) = Новый ЗащищённоеСоединениеOpenSSL()
			// без параметров: проверка цепочки сертификатов включена (по умолчанию).
			ЗащитаSSL = Новый(ТипSSL);
			Возврат Новый HTTPСоединение(Хост, Порт, "", "", Неопределено, 30, ЗащитаSSL);
		Иначе
			// OpenSSL недоступен (Учебная редакция или платформа без OpenSSL).
			// На Windows 1С использует системный TLS (WinInet/Schannel).
			// Если и это не работает — смените порт на 80 или используйте HTTP-прокси.
			Возврат Новый HTTPСоединение(Хост, Порт, "", "", Неопределено, 30);
		КонецЕсли;
	Иначе
		Возврат Новый HTTPСоединение(Хост, Порт, "", "", Неопределено, 30);
	КонецЕсли;
КонецФункции

&НаСервере
Функция _Заголовки(Ключ)
	З = Новый Соответствие;
	З.Вставить("Authorization", "Bearer " + Ключ);
	З.Вставить("Content-Type",  "application/json; charset=utf-8");
	Возврат З;
КонецФункции

// Сериализует Структура / Массив в JSON-строку.
// Использует глобальную функцию ЗаписатьJSON (доступна с 8.3.6+).
// НЕ использует СериализаторXDTO — он предназначен для XDTO-объектов,
// а не для обычных Структур и Массивов.
&НаСервере
Функция _ЗаписьJSON(Объект)
	ЗаписьJSON = Новый ЗаписьJSON;
	ЗаписатьJSON(ЗаписьJSON, Объект);
	Возврат ЗаписьJSON.Закрыть();
КонецФункции

// Разбирает JSON-строку. Объекты читаются как Соответствие, массивы — как Массив.
// Для доступа к полям используйте .Получить("ключ"), НЕ .Свойство("ключ").
// .Свойство() — метод Структуры; Соответствие его не имеет.
&НаСервере
Функция _ЧтениеJSON(Строка)
	ЧтениеJSON = Новый ЧтениеJSON;
	ЧтениеJSON.УстановитьСтроку(Строка);
	Возврат ПрочитатьJSON(ЧтениеJSON, Истина);
КонецФункции

'''


class IntegrationCreate(BaseModel):
    type: str = "1c"
    name: str = "1C Интеграция"
    config: dict = {}


class IntegrationUpdate(BaseModel):
    name: Optional[str] = None
    config: Optional[dict] = None
    status: Optional[str] = None


def _integration_row_to_dict(row: dict) -> dict:
    """Normalize an integration DB row to API response shape."""
    config = row.get("config") or {}
    if isinstance(config, str):
        try:
            config = json.loads(config)
        except Exception:
            config = {}
    return {
        "id": row["id"],
        "type": row["type"],
        "name": row["name"],
        "status": row["status"] or "setup",
        "config": config,
        "last_sync_at": row["last_sync_at"].isoformat() if row.get("last_sync_at") else None,
        "created_at": row["created_at"].isoformat() if row.get("created_at") else None,
    }


def _record_integration_sync(request: Request, orders_received: int,
                              stores_matched: int, stores_unmatched: int,
                              errors_count: int, error_detail: str = "") -> None:
    """When /api/v1/orders/batch is called via API key, write a sync log
    if that key is associated with an integration."""
    username = getattr(request.state, "username", "")
    if not username.startswith("api_key:"):
        return
    try:
        key_id_str = username.split(":")[1]
        key_id = int(key_id_str)
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        # Find integration linked to this API key
        cur.execute(
            "SELECT id FROM integrations WHERE (config->>'api_key_id')::int = %s AND status != 'disabled'",
            (key_id,)
        )
        row = cur.fetchone()
        if row:
            integration_id = row["id"]
            if errors_count > 0 and orders_received == 0:
                status = "error"
            elif errors_count > 0:
                status = "partial"
            else:
                status = "success"
            cur.execute("""
                INSERT INTO integration_sync_logs
                    (integration_id, started_at, finished_at, duration_ms, status,
                     orders_received, stores_matched, stores_unmatched, errors_count, error_detail)
                VALUES (%s, NOW(), NOW(), 0, %s, %s, %s, %s, %s, %s)
            """, (integration_id, status, orders_received, stores_matched,
                  stores_unmatched, errors_count, error_detail[:500]))
            cur.execute(
                "UPDATE integrations SET last_sync_at=NOW(), status=%s WHERE id=%s",
                (status if status != "partial" else "active", integration_id),
            )
            conn.commit()
        cur.close()
        conn.close()
    except Exception as exc:
        logger.warning("_record_integration_sync error: %s", exc)


@app.get("/api/integrations")
def list_integrations(request: Request):
    """Список интеграций текущего пользователя."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT id, type, name, status, config, last_sync_at, created_at "
        "FROM integrations WHERE owner_id=%s ORDER BY created_at DESC",
        (uid,)
    )
    rows = [_integration_row_to_dict(dict(r)) for r in cur.fetchall()]
    cur.close(); conn.close()
    return rows


def _generate_1c_zip(base_url: str, api_key: str) -> bytes:
    """Build a personalised ZIP package for the 1C specialist.

    Contents:
      SmartRoute.epf        — connector file (real or placeholder)
      Инструкция.txt        — plain-text instructions (two sections: director + specialist)
      SmartRoute.bsl        — BSL source (always included for manual setup fallback)
    """
    import datetime as _dt
    today = _dt.date.today().strftime("%d.%m.%Y")

    # Try to read the real EPF from disk; fall back to a stub placeholder
    try:
        with open(_EPF_PATH, "rb") as _ef:
            epf_bytes = _ef.read()
        epf_note = ""
    except OSError:
        epf_bytes = b"SmartRoute.epf placeholder - replace with real file"
        epf_note = (
            "ВНИМАНИЕ: SmartRoute.epf в этом архиве является заглушкой.\n"
            "Используйте вместо него SmartRoute.bsl для ручной сборки EPF.\n\n"
        )

    bsl = _1C_BSL_MODULE.replace("{{BASE_URL}}", base_url).replace("{{API_KEY}}", api_key)

    readme = (
        "SmartRoute — Пакет подключения к 1С:Предприятие\n"
        "=================================================\n"
        f"Дата создания: {today}\n"
        "\n"
        "ЧТО В ЭТОМ АРХИВЕ:\n"
        "  SmartRoute.epf       — файл подключения SmartRoute\n"
        "  SmartRoute.bsl       — исходный код (для специалиста по 1С)\n"
        "  Инструкция.txt       — эта инструкция\n"
        "\n"
        + epf_note +
        "────────────────────────────────────────────────────\n"
        "ДЛЯ РУКОВОДИТЕЛЯ / ЛОГИСТА\n"
        "────────────────────────────────────────────────────\n"
        "Передайте файл SmartRoute.epf и эту инструкцию специалисту по 1С.\n"
        "Установка займёт 15–30 минут.\n"
        "После установки заказы будут передаваться\n"
        "в SmartRoute автоматически каждое утро в 07:30.\n"
        "\n"
        "────────────────────────────────────────────────────\n"
        "ДЛЯ СПЕЦИАЛИСТА ПО 1С\n"
        "────────────────────────────────────────────────────\n"
        "\n"
        "ВАРИАНТ А — установка готового EPF (рекомендуется)\n"
        "\n"
        "1. ОТКРОЙТЕ ФАЙЛ В 1С\n"
        "   Запустите 1С:Предприятие (не Конфигуратор).\n"
        "   Файл → Открыть → найдите SmartRoute.epf\n"
        "\n"
        "2. ВВЕДИТЕ НАСТРОЙКИ\n"
        f"   Адрес SmartRoute: {base_url}\n"
        f"   API-ключ:          {api_key}\n"
        "\n"
        "3. НАЖМИТЕ «Проверить соединение»\n"
        "   Ожидаемый результат: «✅ Соединение успешно. SmartRoute подключён.»\n"
        "\n"
        "4. НАСТРОЙТЕ РАСПИСАНИЕ\n"
        "   Меню → Сервис → Регламентные задания → Добавить\n"
        "   Метод: ОтправитьЗаявкиВSmartRoute\n"
        "   Расписание: ежедневно в 07:30\n"
        "\n"
        "────────────────────────────────────────────────────\n"
        "ВАРИАНТ Б — сборка EPF из исходного кода (если EPF не открывается)\n"
        "\n"
        "1. Откройте 1С:Предприятие в режиме Конфигуратора.\n"
        "   Файл → Новый → Внешняя обработка. Имя: SmartRoute\n"
        "\n"
        "2. Формы → Добавить → Произвольная форма → ОК\n"
        "   Перейдите на вкладку «Модуль»\n"
        "\n"
        "3. Откройте SmartRoute.bsl из этого архива.\n"
        "   Скопируйте всё содержимое и вставьте в модуль формы.\n"
        "   (URL и API-ключ уже встроены в код — менять не нужно)\n"
        "\n"
        "4. Файл → Сохранить как... → тип «Внешняя обработка (*.epf)»\n"
        "   Имя: SmartRoute\n"
        "\n"
        "5. Откройте SmartRoute.epf в режиме Предприятия.\n"
        "   Нажмите «Проверить соединение» → ожидается ✅\n"
        "\n"
        "6. Настройте регламентное задание как в Варианте А, шаг 4.\n"
        "\n"
        "────────────────────────────────────────────────────\n"
        "ПАРАМЕТРЫ ПОДКЛЮЧЕНИЯ (уже встроены в SmartRoute.bsl):\n"
        f"  URL:      {base_url}\n"
        f"  API-ключ: {api_key}\n"
        "\n"
        "ПОДДЕРЖКА: support@smartroute.app\n"
    )

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("SmartRoute.epf", epf_bytes)
        zf.writestr("SmartRoute.bsl", bsl.encode("utf-8-sig"))
        zf.writestr("Инструкция.txt", readme.encode("utf-8-sig"))
    return buf.getvalue()


@app.post("/api/integrations/quick-setup", status_code=201)
def quick_setup_integration(request: Request):
    """Одношаговая настройка интеграции 1С.

    Создаёт (или обновляет) запись интеграции и новый API-ключ с нужными правами.
    Полный ключ возвращается один раз — сохраните его в BSL-модуле.

    Требует сессионной аутентификации (cookie). Вызовы через Bearer API-ключ отклоняются,
    чтобы предотвратить возможность создания новых ключей через ранее выданный ключ.
    """
    # Reject API-key-based auth — only browser sessions may mint new keys
    username = getattr(request.state, "username", "")
    if username.startswith("api_key:"):
        raise HTTPException(
            status_code=403,
            detail="Этот endpoint доступен только через браузерную сессию, не через API-ключ."
        )
    uid = get_user_id(request)
    base_url = str(request.base_url).rstrip("/")

    full_key, prefix = _generate_api_key()
    key_hash = _hash_api_key(full_key)

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Create a dedicated API key for this integration
    import datetime as _dt
    key_name = "1С — SmartRoute (" + _dt.date.today().strftime("%d.%m.%Y") + ")"
    cur.execute(
        """INSERT INTO api_keys (owner_id, name, key_prefix, key_hash, scopes)
           VALUES (%s, %s, %s, %s, %s) RETURNING id""",
        (uid, key_name, prefix, key_hash, ["orders:write", "webhooks:receive"])
    )
    api_key_id = cur.fetchone()["id"]

    # Check for existing 1C integration
    cur.execute(
        "SELECT id, config FROM integrations WHERE owner_id=%s AND type='1c' ORDER BY created_at DESC LIMIT 1",
        (uid,)
    )
    existing = cur.fetchone()

    if existing:
        integration_id = existing["id"]

        # Revoke the previous integration API key so old credentials stop working
        old_config = existing["config"] or {}
        if isinstance(old_config, str):
            try:
                old_config = json.loads(old_config)
            except Exception:
                old_config = {}
        old_key_id = old_config.get("api_key_id")
        if old_key_id and old_key_id != api_key_id:
            cur.execute(
                "UPDATE api_keys SET is_active = FALSE WHERE id = %s AND owner_id = %s",
                (old_key_id, uid)
            )

        # Update config with new key and set status to setup (waiting for first sync)
        cur.execute(
            """UPDATE integrations
               SET config = config || %s::jsonb, status='setup'
               WHERE id=%s
               RETURNING id, type, name, status, config, last_sync_at, created_at""",
            (json.dumps({"api_key_id": api_key_id, "base_url": base_url}), integration_id)
        )
    else:
        cur.execute(
            """INSERT INTO integrations (owner_id, type, name, status, config)
               VALUES (%s, '1c', '1С:Предприятие', 'setup', %s)
               RETURNING id, type, name, status, config, last_sync_at, created_at""",
            (uid, json.dumps({"api_key_id": api_key_id, "base_url": base_url}))
        )

    row = dict(cur.fetchone())
    conn.commit()
    cur.close()
    conn.close()

    # Build the ZIP package (key embedded) and return as base64 — generated once
    import base64 as _b64
    package_bytes = _generate_1c_zip(base_url, full_key)
    package_b64 = _b64.b64encode(package_bytes).decode()

    return {
        **_integration_row_to_dict(row),
        "api_key_id": api_key_id,
        "key_prefix": prefix,
        "full_key": full_key,      # shown ONCE on screen
        "base_url": base_url,
        "package_b64": package_b64,  # ready-to-download ZIP
    }


@app.get("/api/integrations/1c/epf-info")
def get_epf_info(request: Request):
    """Return metadata about the SmartRoute.epf connector file."""
    get_user_id(request)
    available = os.path.exists(_EPF_PATH)
    return {
        "available": available,
        "is_placeholder": _EPF_IS_PLACEHOLDER,
        "version": _epf_meta.get("version", "unknown"),
        "min_1c_version": _epf_meta.get("min_1c_version", "8.3"),
    }


@app.post("/api/integrations")
def create_integration(request: Request, body: IntegrationCreate):
    """Создать интеграцию."""
    uid = get_user_id(request)
    if body.type not in ("1c", "moysklad", "bitrix24", "google_sheets"):
        raise HTTPException(status_code=400, detail="Неизвестный тип интеграции")
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """INSERT INTO integrations (owner_id, type, name, status, config)
           VALUES (%s, %s, %s, 'setup', %s) RETURNING id, type, name, status, config, last_sync_at, created_at""",
        (uid, body.type, body.name, json.dumps(body.config))
    )
    row = dict(cur.fetchone())
    conn.commit(); cur.close(); conn.close()
    return _integration_row_to_dict(row)


@app.get("/api/integrations/{integration_id}")
def get_integration(request: Request, integration_id: int):
    """Получить одну интеграцию с агрегированной статистикой."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT id, type, name, status, config, last_sync_at, created_at "
        "FROM integrations WHERE id=%s AND owner_id=%s",
        (integration_id, uid)
    )
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Интеграция не найдена")
    result = _integration_row_to_dict(dict(row))
    # Aggregate stats from sync logs
    cur.execute("""
        SELECT
            COUNT(*)                           AS total_syncs,
            COALESCE(SUM(orders_received), 0)  AS total_orders,
            COALESCE(SUM(stores_matched), 0)   AS total_matched,
            COALESCE(SUM(errors_count), 0)     AS total_errors,
            MAX(started_at)                    AS last_sync
        FROM integration_sync_logs WHERE integration_id=%s
    """, (integration_id,))
    stats_row = cur.fetchone()
    cur.close(); conn.close()
    result["stats"] = {
        "total_syncs": int(stats_row["total_syncs"] or 0),
        "total_orders": int(stats_row["total_orders"] or 0),
        "total_matched": int(stats_row["total_matched"] or 0),
        "total_errors": int(stats_row["total_errors"] or 0),
    }
    # Count stores auto-created via this integration that still need geocoding
    conn2 = get_db()
    cur2 = conn2.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur2.execute(
        "SELECT COUNT(*) AS cnt FROM stores WHERE owner_id=%s AND source='1c' AND geocode_status='pending'",
        (uid,)
    )
    pending_row = cur2.fetchone()
    cur2.close(); conn2.close()
    result["pending_stores"] = int(pending_row["cnt"] or 0)
    return result


@app.put("/api/integrations/{integration_id}")
def update_integration(request: Request, integration_id: int, body: IntegrationUpdate):
    """Обновить настройки интеграции."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT id, config FROM integrations WHERE id=%s AND owner_id=%s",
        (integration_id, uid)
    )
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Интеграция не найдена")
    # Merge config
    existing_config = row["config"] or {}
    if isinstance(existing_config, str):
        try:
            existing_config = json.loads(existing_config)
        except Exception:
            existing_config = {}
    if body.config is not None:
        existing_config.update(body.config)
    fields = []
    params = []
    if body.name is not None:
        fields.append("name=%s"); params.append(body.name)
    if body.status is not None:
        allowed = {"setup", "active", "error", "disabled"}
        if body.status not in allowed:
            raise HTTPException(status_code=400, detail="Недопустимый статус")
        fields.append("status=%s"); params.append(body.status)
    fields.append("config=%s"); params.append(json.dumps(existing_config))
    params.extend([integration_id, uid])
    cur.execute(
        f"UPDATE integrations SET {', '.join(fields)} WHERE id=%s AND owner_id=%s "
        "RETURNING id, type, name, status, config, last_sync_at, created_at",
        params,
    )
    updated = dict(cur.fetchone())
    conn.commit(); cur.close(); conn.close()
    return _integration_row_to_dict(updated)


@app.delete("/api/integrations/{integration_id}")
def delete_integration(request: Request, integration_id: int):
    """Удалить интеграцию и все её логи."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM integrations WHERE id=%s AND owner_id=%s",
        (integration_id, uid)
    )
    deleted = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    if deleted == 0:
        raise HTTPException(status_code=404, detail="Интеграция не найдена")
    return {"ok": True}


@app.post("/api/integrations/{integration_id}/test")
def test_integration(request: Request, integration_id: int):
    """Проверить API-ключ, привязанный к интеграции."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT config FROM integrations WHERE id=%s AND owner_id=%s",
        (integration_id, uid)
    )
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Интеграция не найдена")
    config = row["config"] or {}
    if isinstance(config, str):
        try:
            config = json.loads(config)
        except Exception:
            config = {}
    api_key_id = config.get("api_key_id")
    if not api_key_id:
        return {"ok": False, "message": "API-ключ не настроен. Завершите настройку интеграции."}
    # Verify the key still exists and is active
    conn2 = get_db()
    cur2 = conn2.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur2.execute(
        "SELECT id, name, is_active, expires_at, scopes FROM api_keys WHERE id=%s AND owner_id=%s",
        (api_key_id, uid)
    )
    key_row = cur2.fetchone()
    cur2.close(); conn2.close()
    if not key_row:
        return {"ok": False, "message": "❌ API-ключ не найден. Создайте новый ключ в Настройках → API."}
    if not key_row["is_active"]:
        return {"ok": False, "message": "❌ API-ключ отозван. Создайте новый ключ в Настройках → API."}
    expires_at = key_row.get("expires_at")
    if expires_at and expires_at < datetime.utcnow():
        return {"ok": False, "message": "❌ API-ключ истёк. Создайте новый ключ в Настройках → API."}
    scopes = key_row.get("scopes") or []
    if "orders:write" not in scopes:
        return {"ok": False, "message": f"❌ Ключ '{key_row['name']}' не имеет разрешения orders:write."}
    return {"ok": True, "message": f"✅ Соединение успешно. Ключ «{key_row['name']}» активен."}


@app.post("/api/integrations/{integration_id}/sync")
def sync_integration(request: Request, integration_id: int):
    """Создать запись о ручной синхронизации (для тестирования)."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT id, status FROM integrations WHERE id=%s AND owner_id=%s",
        (integration_id, uid)
    )
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Интеграция не найдена")
    # Check last received orders count for context
    cur.execute(
        "SELECT COUNT(*) as cnt FROM daily_orders WHERE owner_id=%s AND created_at > NOW() - INTERVAL '1 hour'",
        (uid,)
    )
    recent = cur.fetchone()
    recent_count = int(recent["cnt"] or 0)
    # Write a manual sync log entry
    cur.execute("""
        INSERT INTO integration_sync_logs
            (integration_id, started_at, finished_at, duration_ms, status,
             orders_received, stores_matched, stores_unmatched, errors_count, error_detail)
        VALUES (%s, NOW(), NOW(), 0, 'success', %s, %s, 0, 0, 'Ручная проверка')
        RETURNING id
    """, (integration_id, recent_count, recent_count))
    if row["status"] == "setup":
        cur.execute("UPDATE integrations SET status='active' WHERE id=%s", (integration_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "orders_checked": recent_count}


@app.get("/api/integrations/{integration_id}/logs")
def get_integration_logs(
    request: Request,
    integration_id: int,
    limit: int = Query(50, ge=1, le=200),
):
    """Журнал синхронизаций интеграции."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT id FROM integrations WHERE id=%s AND owner_id=%s",
        (integration_id, uid)
    )
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Интеграция не найдена")
    cur.execute("""
        SELECT id, started_at, finished_at, duration_ms, status,
               orders_received, stores_matched, stores_unmatched, errors_count, error_detail
        FROM integration_sync_logs
        WHERE integration_id=%s
        ORDER BY started_at DESC
        LIMIT %s
    """, (integration_id, limit))
    logs = []
    for r in cur.fetchall():
        d = dict(r)
        d["started_at"] = d["started_at"].isoformat() if d.get("started_at") else None
        d["finished_at"] = d["finished_at"].isoformat() if d.get("finished_at") else None
        logs.append(d)
    cur.close(); conn.close()
    return logs


@app.get("/api/integrations/{integration_id}/download-module")
def download_integration_module(request: Request, integration_id: int):
    """Скачать готовый BSL-модуль для 1С с заполненными настройками (base64)."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT config FROM integrations WHERE id=%s AND owner_id=%s",
        (integration_id, uid)
    )
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Интеграция не найдена")
    config = row["config"] or {}
    if isinstance(config, str):
        try:
            config = json.loads(config)
        except Exception:
            config = {}
    base_url = config.get("base_url", "https://YOUR-SMARTROUTE-URL")

    # The full API key is never stored server-side (only a hash+prefix).
    # We return the BSL with the URL already filled in, but with a clearly
    # visible placeholder for the key. The user must paste the key that was
    # shown ONCE during quick-setup. If lost, they must re-run quick-setup.
    api_key_placeholder = "ВСТАВЬТЕ_ЗДЕСЬ_API_КЛЮЧ_ИЗ_SMARTROUTE"

    bsl_content = _1C_BSL_MODULE.replace("{{BASE_URL}}", base_url)
    bsl_content = bsl_content.replace("{{API_KEY}}", api_key_placeholder)

    import base64 as _b64
    encoded = _b64.b64encode(bsl_content.encode("utf-8-sig")).decode("ascii")
    return {
        "data": encoded,
        "filename": "SmartRoute_1C_Integration.bsl",
        "note": (
            "URL уже встроен в файл. "
            "Замените ВСТАВЬТЕ_ЗДЕСЬ_API_КЛЮЧ_ИЗ_SMARTROUTE на ваш API-ключ. "
            "Если ключ утерян — перегенерируйте его через «Настройки → Пересобрать пакет»."
        ),
    }


@app.get("/api/analytics/summary")
def get_analytics_summary(request: Request):
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT
            COUNT(*) as total_routes,
            COALESCE(SUM(total_km), 0) as total_km,
            COALESCE(SUM(saved_km), 0) as saved_km,
            COALESCE(SUM(saved_rub), 0) as saved_rub,
            COALESCE(AVG(num_points), 0) as avg_points_per_route
        FROM route_sessions
        WHERE owner_id = %s
    """, (uid,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    return {
        "total_routes": row["total_routes"] or 0,
        "total_km": round(float(row["total_km"] or 0), 1),
        "saved_km": round(float(row["saved_km"] or 0), 1),
        "saved_rub": int(row["saved_rub"] or 0),
        "avg_points_per_route": round(float(row["avg_points_per_route"] or 0), 1),
    }


@app.get("/api/analytics/daily")
def get_analytics_daily(request: Request, date_from: Optional[str] = None, date_to: Optional[str] = None):
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    params: list = [uid]
    conditions: list = ["owner_id = %s"]
    if date_from:
        conditions.append("date >= %s")
        params.append(date_from)
    else:
        conditions.append("date >= (CURRENT_DATE - INTERVAL '30 days')::TEXT")
    if date_to:
        conditions.append("date <= %s")
        params.append(date_to)
    where = "WHERE " + " AND ".join(conditions)
    cur.execute(f"""
        SELECT
            date,
            COUNT(*) as routes,
            COALESCE(SUM(total_km), 0) as total_km,
            COALESCE(SUM(saved_km), 0) as saved_km,
            COALESCE(SUM(saved_rub), 0) as saved_rub
        FROM route_sessions
        {where}
        GROUP BY date
        ORDER BY date
    """, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [
        {
            "date": r["date"],
            "routes": r["routes"],
            "total_km": round(float(r["total_km"]), 1),
            "saved_km": round(float(r["saved_km"]), 1),
            "saved_rub": int(r["saved_rub"]),
        }
        for r in rows
    ]


@app.get("/api/analytics/monthly")
def get_analytics_monthly(request: Request, date_from: Optional[str] = None, date_to: Optional[str] = None):
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    params: list = [uid]
    conditions: list = ["owner_id = %s"]
    if date_from:
        conditions.append("date >= %s")
        params.append(date_from)
    else:
        conditions.append("created_at >= NOW() - INTERVAL '12 months'")
    if date_to:
        conditions.append("date <= %s")
        params.append(date_to)
    where = "WHERE " + " AND ".join(conditions)
    cur.execute(f"""
        SELECT
            TO_CHAR(created_at, 'YYYY-MM') as month,
            COUNT(*) as routes,
            COALESCE(SUM(total_km), 0) as total_km,
            COALESCE(SUM(saved_rub), 0) as saved_rub
        FROM route_sessions
        {where}
        GROUP BY month
        ORDER BY month
    """, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [
        {
            "month": r["month"],
            "routes": r["routes"],
            "total_km": round(float(r["total_km"]), 1),
            "saved_rub": int(r["saved_rub"]),
        }
        for r in rows
    ]


@app.get("/api/analytics/vehicle-load")
def get_analytics_vehicle_load(request: Request, date_from: Optional[str] = None, date_to: Optional[str] = None):
    """Среднее количество точек на машину по дням."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    params: list = [uid]
    conditions = ["owner_id = %s", "num_vehicles > 0", "num_points > 0"]
    if date_from:
        conditions.append("date >= %s")
        params.append(date_from)
    else:
        conditions.append("date >= (CURRENT_DATE - INTERVAL '30 days')::TEXT")
    if date_to:
        conditions.append("date <= %s")
        params.append(date_to)
    where = "WHERE " + " AND ".join(conditions)
    cur.execute(f"""
        SELECT
            date,
            ROUND(AVG(num_points::float / NULLIF(num_vehicles, 0))::numeric, 1) as avg_points_per_vehicle,
            SUM(num_points) as total_points,
            SUM(num_vehicles) as total_vehicles
        FROM route_sessions
        {where}
        GROUP BY date
        ORDER BY date
    """, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [
        {
            "date": r["date"],
            "avg_points_per_vehicle": float(r["avg_points_per_vehicle"] or 0),
            "total_points": int(r["total_points"] or 0),
            "total_vehicles": int(r["total_vehicles"] or 0),
        }
        for r in rows
    ]


@app.get("/api/analytics/top-stores")
def get_top_stores(request: Request):
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT
            s.id as store_id,
            s.name as store_name,
            COUNT(rss.id) as visit_count
        FROM stores s
        LEFT JOIN route_session_stores rss ON rss.store_id = s.id
        WHERE s.owner_id = %s
        GROUP BY s.id, s.name
        ORDER BY visit_count DESC
        LIMIT 10
    """, (uid,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [
        {
            "store_id": r["store_id"],
            "store_name": r["store_name"],
            "visit_count": r["visit_count"],
        }
        for r in rows
    ]


# ── Admin endpoints ────────────────────────────────────────────────────────────

class AdminUserCreate(BaseModel):
    username: str
    password: str
    is_admin: bool = False
    plan: str = "trial"
    admin_note: str = ""

class AdminUserUpdate(BaseModel):
    password: Optional[str] = None
    is_admin: Optional[bool] = None
    is_active: Optional[bool] = None
    plan: Optional[str] = None
    admin_note: Optional[str] = None


def _count_active_admins() -> int:
    """Count total active admins in the system."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM users WHERE is_admin=TRUE AND is_active=TRUE")
    count = cur.fetchone()[0]
    cur.close()
    conn.close()
    return count


def _audit_log(conn, admin_user_id: int, target_user_id: int, target_username: str, action: str, details: str = ""):
    """Write one entry to admin_audit_log using the given (open) connection."""
    try:
        cur = conn.cursor()
        cur.execute(
            """INSERT INTO admin_audit_log (admin_user_id, target_user_id, target_username, action, details)
               VALUES (%s, %s, %s, %s, %s)""",
            (admin_user_id, target_user_id, target_username, action, details)
        )
        cur.close()
    except Exception as e:
        logger.warning("audit_log write failed: %s", e)


@app.get("/api/admin/users")
def admin_list_users(request: Request):
    require_admin(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT u.id, u.username, u.is_admin, u.is_active, u.created_at, u.last_login_at,
               u.plan, u.admin_note,
               COUNT(DISTINCT s.id) as stores_count,
               COUNT(DISTINCT rs.id) as sessions_count
        FROM users u
        LEFT JOIN stores s ON s.owner_id = u.id
        LEFT JOIN route_sessions rs ON rs.owner_id = u.id
        GROUP BY u.id, u.username, u.is_admin, u.is_active, u.created_at, u.last_login_at,
                 u.plan, u.admin_note
        ORDER BY u.created_at DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [
        {
            "id": r["id"],
            "username": r["username"],
            "is_admin": bool(r["is_admin"]),
            "is_active": bool(r["is_active"]),
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            "last_login_at": r["last_login_at"].isoformat() if r["last_login_at"] else None,
            "stores_count": int(r["stores_count"] or 0),
            "sessions_count": int(r["sessions_count"] or 0),
            "plan": r["plan"] or "trial",
            "admin_note": r["admin_note"] or "",
        }
        for r in rows
    ]


_VALID_PLANS = {"trial", "basic", "pro", "enterprise"}

@app.post("/api/admin/users", status_code=201)
def admin_create_user(request: Request, body: AdminUserCreate):
    require_admin(request)
    if not body.username.strip():
        raise HTTPException(status_code=422, detail="Логин не может быть пустым")
    if len(body.password) < 4:
        raise HTTPException(status_code=422, detail="Пароль должен быть не менее 4 символов")
    plan = body.plan if body.plan in _VALID_PLANS else "trial"
    hashed = _bcrypt_lib.hashpw(body.password.encode(), _bcrypt_lib.gensalt()).decode()
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(
            """INSERT INTO users (username, password_hash, is_admin, is_active, plan, admin_note)
               VALUES (%s, %s, %s, %s, %s, %s)
               RETURNING id, username, is_admin, is_active, created_at, plan, admin_note""",
            (body.username.strip(), hashed, body.is_admin, True, plan, body.admin_note or "")
        )
        row = cur.fetchone()
        admin_uid = get_user_id(request)
        _audit_log(conn, admin_uid, row["id"], row["username"], "user_created",
                   f"plan={row['plan']}, is_admin={row['is_admin']}")
        conn.commit()
        return {
            "id": row["id"],
            "username": row["username"],
            "is_admin": bool(row["is_admin"]),
            "is_active": bool(row["is_active"]),
            "created_at": row["created_at"].isoformat() if row["created_at"] else None,
            "last_login_at": None,
            "stores_count": 0,
            "sessions_count": 0,
            "plan": row["plan"] or "trial",
            "admin_note": row["admin_note"] or "",
        }
    except Exception as e:
        conn.rollback()
        if "unique" in str(e).lower() or "duplicate" in str(e).lower():
            raise HTTPException(status_code=409, detail="Пользователь с таким логином уже существует")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()


@app.patch("/api/admin/users/{user_id}")
def admin_update_user(user_id: int, request: Request, body: AdminUserUpdate):
    require_admin(request)
    current_uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, username, is_admin, is_active, plan, admin_note FROM users WHERE id = %s", (user_id,))
    target = cur.fetchone()
    if not target:
        cur.close()
        conn.close()
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    # Self-protection: cannot modify own admin/active status
    if user_id == current_uid:
        if body.is_active is False:
            cur.close(); conn.close()
            raise HTTPException(status_code=400, detail="Нельзя деактивировать свой аккаунт")
        if body.is_admin is False:
            cur.close(); conn.close()
            raise HTTPException(status_code=400, detail="Нельзя снять права администратора у себя")

    # Last-admin protection: cannot deactivate or strip the last active admin
    if target["is_admin"] and target["is_active"]:
        if body.is_active is False or body.is_admin is False:
            admin_count = _count_active_admins()
            if admin_count <= 1:
                cur.close(); conn.close()
                raise HTTPException(
                    status_code=400,
                    detail="Нельзя деактивировать или лишить прав последнего администратора системы"
                )

    sets = []
    params = []
    if body.password is not None:
        if len(body.password) < 4:
            cur.close(); conn.close()
            raise HTTPException(status_code=422, detail="Пароль должен быть не менее 4 символов")
        hashed = _bcrypt_lib.hashpw(body.password.encode(), _bcrypt_lib.gensalt()).decode()
        sets.append("password_hash = %s")
        params.append(hashed)
    if body.is_admin is not None:
        sets.append("is_admin = %s")
        params.append(body.is_admin)
    if body.is_active is not None:
        sets.append("is_active = %s")
        params.append(body.is_active)
    if body.plan is not None:
        plan = body.plan if body.plan in _VALID_PLANS else "trial"
        sets.append("plan = %s")
        params.append(plan)
    if body.admin_note is not None:
        sets.append("admin_note = %s")
        params.append(body.admin_note)
    if not sets:
        cur.close(); conn.close()
        raise HTTPException(status_code=422, detail="Нет полей для обновления")
    params.append(user_id)
    cur.execute(
        f"UPDATE users SET {', '.join(sets)} WHERE id = %s RETURNING id, username, is_admin, is_active, plan, admin_note",
        params
    )
    row = cur.fetchone()
    # Audit each changed field
    if body.password is not None:
        _audit_log(conn, current_uid, user_id, target["username"], "password_reset_by_admin", "")
    if body.is_admin is not None:
        _audit_log(conn, current_uid, user_id, target["username"],
                   "admin_granted" if body.is_admin else "admin_removed", "")
    if body.is_active is not None:
        _audit_log(conn, current_uid, user_id, target["username"],
                   "user_unblocked" if body.is_active else "user_blocked", "")
    if body.plan is not None:
        _audit_log(conn, current_uid, user_id, target["username"], "plan_changed",
                   f"{target['plan']} → {row['plan']}")
    if body.admin_note is not None:
        _audit_log(conn, current_uid, user_id, target["username"], "note_changed", "")
    conn.commit()
    cur.close()
    conn.close()
    return {
        "id": row["id"],
        "username": row["username"],
        "is_admin": bool(row["is_admin"]),
        "is_active": bool(row["is_active"]),
        "plan": row["plan"] or "trial",
        "admin_note": row["admin_note"] or "",
    }


@app.delete("/api/admin/users/{user_id}", status_code=200)
def admin_delete_user(user_id: int, request: Request):
    require_admin(request)
    current_uid = get_user_id(request)
    if user_id == current_uid:
        raise HTTPException(status_code=400, detail="Нельзя удалить свой аккаунт")
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, username, is_admin, is_active FROM users WHERE id = %s", (user_id,))
    target = cur.fetchone()
    if not target:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    # Last-admin protection
    if target["is_admin"] and target["is_active"]:
        admin_count = _count_active_admins()
        if admin_count <= 1:
            cur.close(); conn.close()
            raise HTTPException(
                status_code=400,
                detail="Нельзя удалить последнего администратора системы"
            )
    # Cascade: delete user's data manually (FK is NO ACTION)
    cur2 = conn.cursor()
    cur2.execute("DELETE FROM route_session_stores WHERE session_id IN (SELECT id FROM route_sessions WHERE owner_id=%s)", (user_id,))
    cur2.execute("DELETE FROM route_sessions WHERE owner_id = %s", (user_id,))
    cur2.execute("DELETE FROM stores WHERE owner_id = %s", (user_id,))
    cur2.execute("DELETE FROM company_settings WHERE owner_id = %s", (user_id,))
    cur2.execute("DELETE FROM users WHERE id = %s", (user_id,))
    _audit_log(conn, current_uid, user_id, target["username"], "user_deleted",
               f"stores={target.get('stores_count',0)}, was_admin={target['is_admin']}")
    conn.commit()
    cur.close()
    cur2.close()
    conn.close()
    return {"ok": True, "username": target["username"]}


@app.get("/api/admin/audit-log")
def admin_audit_log(request: Request, limit: int = 100):
    require_admin(request)
    limit = max(1, min(limit, 500))
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT l.id, l.admin_user_id, u.username AS admin_username,
               l.target_user_id, l.target_username, l.action, l.details, l.created_at
        FROM admin_audit_log l
        LEFT JOIN users u ON u.id = l.admin_user_id
        ORDER BY l.created_at DESC
        LIMIT %s
    """, (limit,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [
        {
            "id": r["id"],
            "admin_user_id": r["admin_user_id"],
            "admin_username": r["admin_username"] or "—",
            "target_user_id": r["target_user_id"],
            "target_username": r["target_username"] or "—",
            "action": r["action"],
            "details": r["details"] or "",
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
        }
        for r in rows
    ]


@app.get("/api/admin/geocode-cache")
def admin_list_geocode_cache(request: Request, limit: int = 100, offset: int = 0):
    """List entries in the persistent geocoding cache. Admin-only."""
    require_admin(request)
    limit = max(1, min(limit, 500))
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT COUNT(*) FROM geocode_cache")
    total = cur.fetchone()["count"]
    cur.execute(
        "SELECT id, normalized_address, lat, lon, source, created_at, updated_at FROM geocode_cache ORDER BY updated_at DESC LIMIT %s OFFSET %s",
        (limit, max(0, offset))
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return {
        "total": total,
        "items": [
            {
                "id": r["id"],
                "address": r["normalized_address"],
                "lat": float(r["lat"]),
                "lon": float(r["lon"]),
                "source": r["source"],
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
            }
            for r in rows
        ],
    }


@app.delete("/api/admin/geocode-cache/{entry_id}", status_code=200)
def admin_delete_geocode_cache_entry(entry_id: int, request: Request):
    """Delete a single geocoding cache entry so the address will be re-geocoded. Admin-only."""
    require_admin(request)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT normalized_address FROM geocode_cache WHERE id = %s", (entry_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Запись не найдена")
    addr = row[0]
    cur.execute("DELETE FROM geocode_cache WHERE id = %s", (entry_id,))
    conn.commit(); cur.close(); conn.close()
    # Also evict from in-memory cache so next geocode call goes to API
    geocode_cache.pop(addr, None)
    return {"ok": True, "deleted_address": addr}


@app.delete("/api/admin/geocode-cache", status_code=200)
def admin_clear_geocode_cache(request: Request):
    """Purge the entire geocoding cache. Admin-only. Use with caution."""
    require_admin(request)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM geocode_cache")
    count = cur.fetchone()[0]
    cur.execute("DELETE FROM geocode_cache")
    conn.commit(); cur.close(); conn.close()
    geocode_cache.clear()
    return {"ok": True, "deleted_count": count}


@app.delete("/api/admin/api-keys/cleanup-test", status_code=200)
def admin_cleanup_test_api_keys(request: Request):
    """Hard-delete API keys whose name looks like a test key.

    Matches: names starting with 'test_', 'smoke_', 'rc_', 'fin', 'release_check',
    'rl', 'rl2', 'rl3', 'final_adm', 'limited', 'u2', 'iso', or containing '_test'.
    Admin cookie required.
    """
    require_admin(request)
    _TEST_PATTERNS = [
        "test_%", "smoke_%", "rc_%", "release_%", "rl_%",
        "%_test%", "%_final%", "fin", "final_adm", "limited", "u2",
        "iso%", "rl", "rl2", "rl3", "rl_final",
    ]
    conn = get_db()
    cur = conn.cursor()
    total = 0
    for pattern in _TEST_PATTERNS:
        cur.execute("DELETE FROM api_keys WHERE name ILIKE %s", (pattern,))
        total += cur.rowcount
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "deleted": total}


@app.post("/api/admin/cleanup-test-data", status_code=200)
def admin_cleanup_test_data(request: Request, user_id: int | None = Query(None)):
    """Delete all test data for a specific user (or all users if user_id=None).

    Removes stores, daily_orders, route_sessions (cascade) that were created
    during integration testing. Use ONLY in dev/staging — not in production.

    Admin cookie required. API keys cannot access this endpoint.
    """
    require_admin(request)
    conn = get_db()
    cur = conn.cursor()
    try:
        uid_filter = "AND owner_id = %s" if user_id else ""
        uid_args = (user_id,) if user_id else ()

        # Route sessions → cascade removes route_session_stores
        if user_id:
            cur.execute(
                "DELETE FROM route_session_stores WHERE session_id IN "
                "(SELECT id FROM route_sessions WHERE owner_id = %s)", uid_args
            )
            cur.execute("DELETE FROM route_sessions WHERE owner_id = %s", uid_args)
        else:
            cur.execute("DELETE FROM route_session_stores")
            cur.execute("DELETE FROM route_sessions")
        routes_deleted = cur.rowcount

        # Daily orders
        if user_id:
            cur.execute("DELETE FROM daily_orders WHERE owner_id = %s", uid_args)
        else:
            cur.execute("DELETE FROM daily_orders")
        orders_deleted = cur.rowcount

        # Stores
        if user_id:
            cur.execute("DELETE FROM stores WHERE owner_id = %s", uid_args)
        else:
            cur.execute("DELETE FROM stores")
        stores_deleted = cur.rowcount

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close(); conn.close()

    return {
        "ok": True,
        "user_id": user_id,
        "deleted": {
            "stores": stores_deleted,
            "orders": orders_deleted,
            "route_sessions": routes_deleted,
        },
    }


# ── API Key management endpoints ─────────────────────────────────────────────

class ApiKeyCreate(BaseModel):
    name: str
    scopes: list[str] = []
    expires_days: int | None = None  # None = never expires


@app.get("/api/auth/api-keys")
def list_api_keys(request: Request):
    """List all API keys for the current user (without the secret part)."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """SELECT id, name, key_prefix, scopes, is_active, expires_at,
                  last_used_at, created_at
           FROM api_keys WHERE owner_id = %s ORDER BY created_at DESC""",
        (uid,)
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


@app.post("/api/auth/api-keys", status_code=201)
def create_api_key(request: Request, body: ApiKeyCreate):
    """Create a new API key. Returns the full key ONCE — store it safely."""
    uid = get_user_id(request)
    if not body.name.strip():
        raise HTTPException(status_code=422, detail="Укажите название ключа")

    # Validate scopes
    _VALID_SCOPES = {
        "stores:read", "stores:write",
        "orders:read", "orders:write",
        "routes:read", "routes:build", "routes:write",
        "analytics:read",
        "settings:read", "settings:write",
        "webhooks:receive",
    }
    invalid = [s for s in body.scopes if s not in _VALID_SCOPES]
    if invalid:
        raise HTTPException(status_code=422, detail=f"Недопустимые scopes: {invalid}")

    full_key, prefix = _generate_api_key()
    key_hash = _hash_api_key(full_key)
    expires_at = None
    if body.expires_days:
        expires_at = datetime.utcnow() + timedelta(days=body.expires_days)

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """INSERT INTO api_keys (owner_id, name, key_prefix, key_hash, scopes, expires_at)
           VALUES (%s, %s, %s, %s, %s, %s) RETURNING id, name, key_prefix, scopes, is_active, created_at""",
        (uid, body.name.strip(), prefix, key_hash, body.scopes, expires_at)
    )
    row = dict(cur.fetchone())
    conn.commit(); cur.close(); conn.close()

    row["key"] = full_key  # shown ONCE
    return row


@app.delete("/api/auth/api-keys/{key_id}", status_code=200)
def revoke_api_key(key_id: int, request: Request, permanent: bool = Query(False)):
    """Revoke (deactivate) an API key.

    By default keeps the record for audit trail (is_active=False).
    Pass ?permanent=true to hard-delete the record entirely.
    Only the key owner can revoke/delete their own keys.
    """
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor()
    if permanent:
        cur.execute(
            "DELETE FROM api_keys WHERE id = %s AND owner_id = %s",
            (key_id, uid)
        )
    else:
        cur.execute(
            "UPDATE api_keys SET is_active = FALSE WHERE id = %s AND owner_id = %s",
            (key_id, uid)
        )
    if cur.rowcount == 0:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Ключ не найден")
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "permanent": permanent}


@app.delete("/api/auth/api-keys", status_code=200)
def purge_revoked_api_keys(request: Request):
    """Hard-delete all revoked (is_active=False) API keys for the current user.

    Active keys are NOT affected. Use to clean up the audit-trail list.
    """
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM api_keys WHERE owner_id = %s AND is_active = FALSE",
        (uid,)
    )
    deleted = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "deleted": deleted}


@app.post("/api/auth/api-keys/{key_id}/rotate", status_code=201)
def rotate_api_key(key_id: int, request: Request):
    """Rotate: deactivate old key, create new one with same name+scopes."""
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT name, scopes, expires_at FROM api_keys WHERE id = %s AND owner_id = %s AND is_active = TRUE",
        (key_id, uid)
    )
    old = cur.fetchone()
    if not old:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Ключ не найден или уже отозван")

    # Deactivate old
    cur.execute("UPDATE api_keys SET is_active = FALSE WHERE id = %s", (key_id,))

    # Create new
    full_key, prefix = _generate_api_key()
    key_hash = _hash_api_key(full_key)
    cur.execute(
        """INSERT INTO api_keys (owner_id, name, key_prefix, key_hash, scopes, expires_at)
           VALUES (%s, %s, %s, %s, %s, %s) RETURNING id, name, key_prefix, scopes, is_active, created_at""",
        (uid, old["name"], prefix, key_hash, old["scopes"], old["expires_at"])
    )
    row = dict(cur.fetchone())
    conn.commit(); cur.close(); conn.close()
    row["key"] = full_key
    return row


# ── Universal Webhook Ingest endpoint ─────────────────────────────────────────
# Accepts delivery orders from any external system (1С, МойСклад, Bitrix24, etc.)
# in a single universal format. Authentication via token-in-URL (api_key with
# scope "webhooks:receive"). Connector-specific adapters call this endpoint
# after transforming their native format to UniversalOrderModel.

class WebhookOrderItem(BaseModel):
    store_name: str
    address: str | None = None
    city: str = ""               # Город (используется при авто-создании магазина)
    delivery_date: str           # YYYY-MM-DD
    weight_kg: float = 0.0
    volume_m3: float = 0.0
    quantity: float = 0.0
    amount_rub: float = 0.0
    products: str = ""
    order_number: str = ""
    notes: str = ""
    time_window_from: str | None = None
    time_window_to: str | None = None
    latitude: float | None = None
    longitude: float | None = None
    external_id: str = ""        # ID заказа в источнике (идемпотентность)
    source: str = "webhook"      # "1c" | "moysklad" | "bitrix24" | "api" | ...
    counterparty_code: str = ""  # Код контрагента в 1С → stores.external_id
    phone: str = ""              # Телефон точки доставки


class WebhookIngestRequest(BaseModel):
    orders: list[WebhookOrderItem]
    replace_date: bool = False       # if True — delete existing orders for dates in batch first
    auto_create_stores: bool = True  # Авто-создавать магазины из незнакомых заказов


def _auto_create_store_if_missing(
    uid: int,
    store_name: str,
    address: str,
    counterparty_code: str,
    phone: str,
    city: str,
    db_stores: list[dict],
    cur,
    auto_create: bool = True,
) -> tuple:
    """Match or auto-create a store. Returns (store_id, was_auto_created).

    Priority:
    1. Match by external_id (counterparty_code) — fastest, most reliable for 1C
    2. Fuzzy name match (_match_store_to_db)
    3. Auto-create with geocode_status='pending' if auto_create=True and enough data

    Modifies db_stores in-place when creating so subsequent orders in the same
    batch find the newly created store.
    """
    # 1. Match by counterparty_code (external_id)
    if counterparty_code:
        ext_match = next(
            (s for s in db_stores if s.get("external_id") == counterparty_code),
            None,
        )
        if ext_match:
            return ext_match["id"], False

    # 2. Fuzzy name match
    name_match = _match_store_to_db(store_name, address, db_stores)
    if name_match:
        # Backfill external_id if we now have one and the store doesn't
        if counterparty_code and not name_match.get("external_id"):
            try:
                cur.execute(
                    "UPDATE stores SET external_id=%s, source='1c' WHERE id=%s AND owner_id=%s",
                    (counterparty_code, name_match["id"], uid),
                )
                name_match["external_id"] = counterparty_code
            except Exception:
                pass
        return name_match["id"], False

    # 3. Auto-create
    if not auto_create or not store_name.strip():
        return None, False
    if not address and not counterparty_code:
        return None, False  # not enough data to create a useful store

    try:
        resolved_city = city.strip()
        if not resolved_city and address and "," in address:
            resolved_city = address.split(",", 1)[0].strip()

        cur.execute(
            """INSERT INTO stores
               (owner_id, name, address, city, phone, external_id, source, geocode_status)
               VALUES (%s, %s, %s, %s, %s, %s, '1c', 'pending')
               RETURNING id, name, address, city, phone, external_id""",
            (uid, store_name.strip(), address or "",
             resolved_city, phone or "", counterparty_code or ""),
        )
        new_store = dict(cur.fetchone())
        db_stores.append(new_store)
        return new_store["id"], True
    except Exception as exc:
        logger.warning("auto-create store failed for '%s': %s", store_name, exc)
        return None, False


@app.post("/api/v1/webhooks/ingest/{token}")
def webhook_ingest(token: str, body: WebhookIngestRequest, request: Request):
    """
    Universal order ingest endpoint.
    Auth: token is a raw API key with scope 'webhooks:receive'.
    Accepts a batch of orders in universal format.
    """
    # Authenticate via token (key_hash lookup)
    key_row = _resolve_api_key(token)
    if not key_row:
        raise HTTPException(status_code=401, detail="Недействительный webhook-токен")
    if "webhooks:receive" not in (key_row.get("scopes") or []):
        raise HTTPException(status_code=403, detail="Ключ не имеет права webhooks:receive")

    uid = key_row["user_id"]

    if not body.orders:
        return {"created": 0, "matched": 0, "skipped": 0, "errors": []}

    # Validate all dates
    for item in body.orders:
        try:
            datetime.strptime(item.delivery_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(
                status_code=422,
                detail=f"Неверный формат даты '{item.delivery_date}'. Используйте YYYY-MM-DD"
            )

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Load stores for matching (include external_id for 1C counterparty matching)
    cur.execute("SELECT id, name, address, city, phone, external_id FROM stores WHERE owner_id = %s", (uid,))
    db_stores = [dict(r) for r in cur.fetchall()]

    # Optional: delete existing orders for replaced dates
    if body.replace_date:
        dates_to_replace = list({item.delivery_date for item in body.orders})
        for d in dates_to_replace:
            cur.execute(
                "DELETE FROM daily_orders WHERE owner_id = %s AND delivery_date = %s",
                (uid, d)
            )

    created = matched = skipped = auto_created = 0
    errors = []

    for item in body.orders:
        try:
            if not item.store_name.strip():
                skipped += 1
                continue

            # Match or auto-create store
            store_id, was_created = _auto_create_store_if_missing(
                uid, item.store_name, item.address or "",
                item.counterparty_code, item.phone, item.city,
                db_stores, cur, auto_create=body.auto_create_stores,
            )
            if store_id is not None:
                if was_created:
                    auto_created += 1
                else:
                    matched += 1

            cur.execute(
                """INSERT INTO daily_orders
                   (owner_id, store_id, store_name_raw, address_raw, delivery_date,
                    weight_kg, volume_m3, quantity, amount_rub, products,
                    order_number, notes)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (uid, store_id, item.store_name.strip(), item.address or "",
                 item.delivery_date, item.weight_kg, item.volume_m3,
                 item.quantity, item.amount_rub, item.products,
                 item.order_number, item.notes)
            )
            created += 1
        except Exception as exc:
            errors.append({"store": item.store_name, "error": str(exc)})
            logger.error("webhook_ingest row error: %s", exc)

    conn.commit(); cur.close(); conn.close()
    return {
        "created": created, "matched": matched, "skipped": skipped,
        "auto_created_stores": auto_created, "errors": errors[:20],
    }


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC API v1
# ══════════════════════════════════════════════════════════════════════════════
#
# Design principles:
#   • All routes live under /api/v1/*
#   • Auth: Bearer <api_key>  OR  HttpOnly cookie (same middleware already handles both)
#   • Every handler calls _v1_require_scope(request, "scope:action") first
#   • Responses always wrapped in {"data":…,"meta":…,"request_id":…} envelope
#   • Errors always {"error":{"code":…,"message":…},"request_id":…}
#   • Per-key rate limiting with X-RateLimit-* headers
#   • NO business logic duplication — thin wrappers over existing functions
# ══════════════════════════════════════════════════════════════════════════════

import uuid as _uuid

_V1_VALID_SCOPES = {
    "stores:read", "stores:write",
    "orders:read", "orders:write",
    "routes:read", "routes:build", "routes:write",
    "analytics:read",
    "settings:read", "settings:write",
    "webhooks:receive",
    "keys:read",
}

# ── Per-key rate limiter (sliding window, shared with _rl_store/_rl_lock) ─────

def _v1_key_rate_limit(key_hash: str, limit: int = 60, window: int = 60) -> tuple[int, int]:
    """Return (remaining, reset_unix). Raises _V1Error 429 if limit exceeded."""
    bucket = f"v1:{key_hash}"
    now = time.time()
    with _rl_lock:
        ts = _rl_store.get(bucket, [])
        ts = [t for t in ts if now - t < window]
        remaining = max(0, limit - len(ts) - 1)
        reset_ts = int(now) + window
        if len(ts) >= limit:
            retry = max(1, int(window - (now - ts[0])) + 1)
            err = _V1Error(
                code="RATE_LIMITED",
                message=f"Превышен лимит {limit} запросов в минуту. Повторите через {retry} сек.",
                status=429,
                details={"retry_after": retry, "limit": limit, "window_seconds": window},
            )
            # Attach rate-limit headers so they reach the client via _v1_error_handler
            err._extra_headers = {
                "Retry-After": str(retry),
                "X-RateLimit-Limit": str(limit),
                "X-RateLimit-Remaining": "0",
                "X-RateLimit-Reset": str(reset_ts),
            }
            raise err
        ts.append(now)
        _rl_store[bucket] = ts
    return remaining, reset_ts


def _v1_rl_headers(request: Request, limit: int = 60, window: int = 60) -> dict:
    """Apply per-key rate limiting; return X-RateLimit-* headers dict."""
    scopes = getattr(request.state, "api_key_scopes", None)
    if scopes is None:
        # Cookie auth — no per-key rate limit, return informational headers only
        return {"X-RateLimit-Limit": str(limit), "X-RateLimit-Remaining": str(limit)}
    # Identify key by username set by middleware ("api_key:<id>")
    username = getattr(request.state, "username", "unknown")
    remaining, reset_ts = _v1_key_rate_limit(username, limit=limit, window=window)
    return {
        "X-RateLimit-Limit": str(limit),
        "X-RateLimit-Remaining": str(remaining),
        "X-RateLimit-Reset": str(reset_ts),
    }


# ── Scope enforcement ─────────────────────────────────────────────────────────

def _v1_require_scope(request: Request, scope: str) -> None:
    """Raise _V1Error 403 if the authenticated API key doesn't have *scope*.
    Cookie-authenticated users (api_key_scopes is None) always pass."""
    key_scopes = getattr(request.state, "api_key_scopes", None)
    if key_scopes is None:
        return  # cookie auth = full access
    if "*" in key_scopes:
        return  # wildcard key
    if scope not in key_scopes:
        raise _V1Error(
            code="FORBIDDEN",
            message=f"API ключ не имеет права «{scope}». Добавьте этот scope при создании ключа.",
            status=403,
            details={"required_scope": scope, "key_scopes": key_scopes},
        )


# ── Envelope helpers ──────────────────────────────────────────────────────────

def _v1_request_id() -> str:
    return "req_" + _uuid.uuid4().hex[:12]


def _v1_ok(data, *, meta: dict | None = None, request_id: str | None = None) -> dict:
    # Always include `meta` key (null for single-resource responses) for consistent envelope
    return {"data": data, "meta": meta, "request_id": request_id or _v1_request_id()}


class _V1Error(Exception):
    """Custom exception for v1 API errors — bypasses FastAPI detail wrapping."""
    def __init__(self, code: str, message: str, status: int, details: dict | None = None):
        self.code = code
        self.message = message
        self.status = status
        self.details = details
        self.request_id = _v1_request_id()


def _v1_err(code: str, message: str, status: int, details: dict | None = None) -> "_V1Error":
    return _V1Error(code, message, status, details)


@app.exception_handler(_V1Error)
async def _v1_error_handler(request: Request, exc: "_V1Error"):
    body: dict = {
        "error": {"code": exc.code, "message": exc.message},
        "request_id": exc.request_id,
    }
    if exc.details:
        body["error"]["details"] = exc.details
    headers = {}
    try:
        headers = _v1_rl_headers(request)
    except Exception:
        pass
    # Merge any extra headers (e.g. Retry-After from rate limiter)
    extra = getattr(exc, "_extra_headers", {})
    headers.update(extra)
    return JSONResponse(status_code=exc.status, content=body, headers=headers)


# ── Response helper that adds rate-limit headers ──────────────────────────────

def _v1_serialize(obj):
    """Recursively make a dict/list JSON-safe (convert datetime, Decimal, etc.)."""
    import decimal
    if isinstance(obj, dict):
        return {k: _v1_serialize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_v1_serialize(v) for v in obj]
    if isinstance(obj, datetime):
        return obj.isoformat()
    if isinstance(obj, decimal.Decimal):
        return float(obj)
    return obj


def _v1_response(data, request: Request, *, meta: dict | None = None,
                 status_code: int = 200) -> JSONResponse:
    headers = _v1_rl_headers(request)
    return JSONResponse(
        content=_v1_ok(_v1_serialize(data), meta=meta),
        status_code=status_code,
        headers=headers,
    )


# ══════════════════════════════════════════════════════════════════════════════
# v1 — STORES
# ══════════════════════════════════════════════════════════════════════════════

class V1StoreCreate(BaseModel):
    name: str
    address: str | None = None
    city: str | None = None
    lat: float | None = None
    lon: float | None = None
    yandex_url: str | None = None
    phone: str | None = None
    client: str | None = None
    time_window_from: str | None = None
    time_window_to: str | None = None
    unload_minutes: int | None = None


class V1StoreUpdate(BaseModel):
    name: str | None = None
    address: str | None = None
    city: str | None = None
    lat: float | None = None
    lon: float | None = None
    yandex_url: str | None = None
    phone: str | None = None
    client: str | None = None
    time_window_from: str | None = None
    time_window_to: str | None = None
    unload_minutes: int | None = None


class V1StoreBatchItem(BaseModel):
    name: str
    address: str | None = None
    city: str | None = None
    lat: float | None = None
    lon: float | None = None
    yandex_url: str | None = None
    phone: str | None = None
    client: str | None = None
    time_window_from: str | None = None
    time_window_to: str | None = None
    unload_minutes: int | None = None
    external_id: str | None = None  # for idempotency tracking


class V1StoreBatchRequest(BaseModel):
    stores: list[V1StoreBatchItem]


@app.get("/api/v1/stores",
         summary="Список магазинов",
         tags=["v1-stores"])
def v1_list_stores(
    request: Request,
    page: int = Query(1, description="Номер страницы (≥1; отрицательные → 1)"),
    page_size: int = Query(50, description="Размер страницы (1–500; вне диапазона → ближайшая граница)"),
    q: str | None = Query(None, description="Поиск по имени или адресу"),
    city: str | None = Query(None),
    geocode_status: str | None = Query(None, description="found | not_found | pending"),
):
    """Список точек доставки с пагинацией и фильтрами.

    **Auth**: `stores:read`
    """
    _v1_require_scope(request, "stores:read")
    # Clamp instead of rejecting — friendlier for API consumers
    page = max(1, page)
    page_size = max(1, min(500, page_size))
    uid = get_user_id(request)
    conditions = ["owner_id = %s"]
    params: list = [uid]
    if q:
        conditions.append("(name ILIKE %s OR address ILIKE %s)")
        params += [f"%{q}%", f"%{q}%"]
    if city:
        conditions.append("city ILIKE %s")
        params.append(f"%{city}%")
    if geocode_status:
        conditions.append("geocode_status = %s")
        params.append(geocode_status)
    where = "WHERE " + " AND ".join(conditions)
    offset = (page - 1) * page_size
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"SELECT COUNT(*) as total FROM stores {where}", params)
    total = int(cur.fetchone()["total"])
    cur.execute(
        f"SELECT * FROM stores {where} ORDER BY id LIMIT %s OFFSET %s",
        params + [page_size, offset],
    )
    rows = [store_row_to_dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    meta = {"total": total, "page": page, "page_size": page_size,
            "pages": max(1, (total + page_size - 1) // page_size)}
    return _v1_response(rows, request, meta=meta)


@app.get("/api/v1/stores/{store_id}",
         summary="Магазин по ID",
         tags=["v1-stores"])
def v1_get_store(store_id: int, request: Request):
    """Получить одну точку доставки по ID.

    **Auth**: `stores:read`
    """
    _v1_require_scope(request, "stores:read")
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM stores WHERE id = %s AND owner_id = %s", (store_id, uid))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        raise _v1_err("STORE_NOT_FOUND", f"Магазин с ID {store_id} не найден", 404,
                      {"store_id": store_id})
    return _v1_response(store_row_to_dict(row), request)


@app.post("/api/v1/stores",
          status_code=201,
          summary="Создать магазин",
          tags=["v1-stores"])
def v1_create_store(request: Request, body: V1StoreCreate):
    """Создать точку доставки. Геокодинг выполняется автоматически.

    **Auth**: `stores:write`
    """
    _v1_require_scope(request, "stores:write")
    uid = get_user_id(request)
    if not body.name or not body.name.strip():
        raise _v1_err("VALIDATION_ERROR", "Поле name обязательно", 422)

    # Resolve coordinates
    lat, lon, status = body.lat, body.lon, "not_found"
    address = (body.address or "").strip()
    city = (body.city or "").strip()
    geocode_query = f"{city} {address}".strip() if city and city not in address else address

    if lat is not None and lon is not None:
        status = "found"
    elif body.yandex_url:
        lat, lon = parse_yandex_link(body.yandex_url)
        if lat is not None:
            status = "found"
            if not address:
                address = reverse_geocode_nominatim(lat, lon) or f"{lat:.5f},{lon:.5f}"
        elif geocode_query:
            coords = geocode_address(geocode_query)
            lat, lon = (coords[0], coords[1]) if coords else (None, None)
            status = "found" if coords else "not_found"
    elif geocode_query:
        coords = geocode_address(geocode_query)
        lat, lon = (coords[0], coords[1]) if coords else (None, None)
        status = "found" if coords else "not_found"

    if not address:
        address = geocode_query or (f"{lat:.5f},{lon:.5f}" if lat else "")

    if city and address and city not in address:
        address = f"{city}, {address}"

    map_url = body.yandex_url or None
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """INSERT INTO stores (name, address, city, phone, client, lat, lon, map_url,
           geocode_status, time_window_from, time_window_to, unload_minutes, owner_id)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
        (body.name.strip(), address, city, (body.phone or "").strip(),
         (body.client or "").strip(), lat, lon, map_url, status,
         body.time_window_from, body.time_window_to, body.unload_minutes, uid),
    )
    row = cur.fetchone()
    conn.commit(); cur.close(); conn.close()
    return _v1_response(store_row_to_dict(row), request, status_code=201)


@app.put("/api/v1/stores/{store_id}",
         summary="Обновить магазин",
         tags=["v1-stores"])
def v1_update_store(store_id: int, request: Request, body: V1StoreUpdate):
    """Частичное обновление точки доставки (PATCH-семантика).

    **Auth**: `stores:write`
    """
    _v1_require_scope(request, "stores:write")
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM stores WHERE id = %s AND owner_id = %s", (store_id, uid))
    existing = cur.fetchone()
    if not existing:
        cur.close(); conn.close()
        raise _v1_err("STORE_NOT_FOUND", f"Магазин с ID {store_id} не найден", 404)

    fields: dict = {}
    if body.name is not None:
        fields["name"] = body.name.strip()
    if body.address is not None:
        fields["address"] = body.address
        coords = geocode_address(body.address)
        fields["lat"], fields["lon"] = (coords[0], coords[1]) if coords else (None, None)
        fields["geocode_status"] = "found" if coords else "not_found"
    if body.yandex_url is not None:
        fields["map_url"] = body.yandex_url or None
        if body.yandex_url:
            ly, loy = parse_yandex_link(body.yandex_url)
            if ly is not None:
                fields["lat"] = ly; fields["lon"] = loy; fields["geocode_status"] = "found"
    if body.city is not None:
        fields["city"] = body.city.strip()
    if body.phone is not None:
        fields["phone"] = body.phone.strip()
    if body.client is not None:
        fields["client"] = body.client.strip()
    if body.lat is not None:
        fields["lat"] = body.lat
    if body.lon is not None:
        fields["lon"] = body.lon
    if body.time_window_from is not None:
        fields["time_window_from"] = body.time_window_from
    if body.time_window_to is not None:
        fields["time_window_to"] = body.time_window_to
    if body.unload_minutes is not None:
        fields["unload_minutes"] = body.unload_minutes

    if fields:
        set_clause = ", ".join(f"{k} = %s" for k in fields)
        cur.execute(
            f"UPDATE stores SET {set_clause} WHERE id = %s AND owner_id = %s RETURNING *",
            list(fields.values()) + [store_id, uid],
        )
        row = cur.fetchone()
        conn.commit()
    else:
        row = existing
    cur.close(); conn.close()
    return _v1_response(store_row_to_dict(row), request)


@app.delete("/api/v1/stores/{store_id}",
            summary="Удалить магазин",
            tags=["v1-stores"])
def v1_delete_store(store_id: int, request: Request):
    """Удалить точку доставки.

    **Auth**: `stores:write`
    """
    _v1_require_scope(request, "stores:write")
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM stores WHERE id = %s AND owner_id = %s", (store_id, uid))
    deleted = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    if deleted == 0:
        raise _v1_err("STORE_NOT_FOUND", f"Магазин с ID {store_id} не найден", 404)
    return _v1_response({"ok": True, "deleted": 1}, request)


@app.post("/api/v1/stores/batch",
          summary="Создать/обновить магазины пакетом",
          tags=["v1-stores"])
def v1_batch_stores(request: Request, body: V1StoreBatchRequest):
    """Upsert магазинов пакетом. Если магазин с таким именем и городом уже существует —
    обновляется; иначе создаётся новый. Максимум 500 за запрос.

    **Auth**: `stores:write`
    """
    _v1_require_scope(request, "stores:write")
    uid = get_user_id(request)
    if not body.stores:
        raise _v1_err("VALIDATION_ERROR", "Список stores не может быть пустым", 422)
    if len(body.stores) > 1000:
        raise _v1_err("VALIDATION_ERROR", "Максимум 1000 магазинов за один запрос", 422)

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    # Load existing stores for upsert matching
    cur.execute("SELECT id, name, city FROM stores WHERE owner_id = %s", (uid,))
    existing_map = {(r["name"].strip().lower(), (r["city"] or "").strip().lower()): r["id"]
                    for r in cur.fetchall()}

    created = updated = errors = 0
    result_ids: list[int] = []

    for item in body.stores:
        if not item.name or not item.name.strip():
            errors += 1
            continue
        try:
            lat, lon, status = item.lat, item.lon, "not_found"
            address = (item.address or "").strip()
            city = (item.city or "").strip()
            geocode_query = f"{city} {address}".strip() if city and city not in address else address

            if lat is not None and lon is not None:
                status = "found"
            elif item.yandex_url:
                lat, lon = parse_yandex_link(item.yandex_url)
                status = "found" if lat is not None else "not_found"
            elif geocode_query:
                coords = geocode_address(geocode_query)
                lat, lon = (coords[0], coords[1]) if coords else (None, None)
                status = "found" if coords else "not_found"

            if city and address and city not in address:
                address = f"{city}, {address}"

            key = (item.name.strip().lower(), city.lower())
            existing_id = existing_map.get(key)

            # address column is NOT NULL in DB — fallback to city or name
            safe_address = address or city or item.name.strip()
            if existing_id:
                cur.execute(
                    """UPDATE stores SET address=%s, city=%s, phone=%s, client=%s,
                       lat=%s, lon=%s, geocode_status=%s,
                       time_window_from=%s, time_window_to=%s, unload_minutes=%s
                       WHERE id=%s AND owner_id=%s RETURNING id""",
                    (safe_address, city or "", (item.phone or "").strip(),
                     (item.client or "").strip(), lat, lon, status,
                     item.time_window_from, item.time_window_to, item.unload_minutes,
                     existing_id, uid),
                )
                result_ids.append(existing_id)
                updated += 1
            else:
                cur.execute(
                    """INSERT INTO stores (name, address, city, phone, client, lat, lon,
                       geocode_status, time_window_from, time_window_to, unload_minutes, owner_id)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                    (item.name.strip(), safe_address, city or "",
                     (item.phone or "").strip(), (item.client or "").strip(),
                     lat, lon, status,
                     item.time_window_from, item.time_window_to, item.unload_minutes, uid),
                )
                new_id = cur.fetchone()["id"]
                existing_map[key] = new_id
                result_ids.append(new_id)
                created += 1
        except Exception as exc:
            logger.error("v1_batch_stores item error: %s", exc, exc_info=True)
            errors += 1

    conn.commit(); cur.close(); conn.close()
    return _v1_response(
        {"created": created, "updated": updated, "errors": errors, "ids": result_ids},
        request, status_code=200,
    )


class V1BulkDeleteRequest(BaseModel):
    ids: list[int]


@app.post("/api/v1/stores/bulk-delete",
          summary="Удалить магазины пакетом",
          tags=["v1-stores"])
def v1_bulk_delete_stores(request: Request, body: V1BulkDeleteRequest):
    """Удалить несколько магазинов за один запрос.

    Удаляет только магазины текущего пользователя (по Bearer-ключу).
    IDs чужих магазинов молча игнорируются.

    **Auth**: `stores:write`
    """
    _v1_require_scope(request, "stores:write")
    uid = get_user_id(request)
    if not body.ids:
        return _v1_response({"deleted": 0}, request)
    if len(body.ids) > 5000:
        raise _v1_err("VALIDATION_ERROR", "Максимум 5000 ID за один запрос", 422,
                      {"limit": 5000, "received": len(body.ids)})
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM stores WHERE id = ANY(%s) AND owner_id = %s",
        (body.ids, uid),
    )
    deleted = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    return _v1_response({"deleted": deleted, "requested": len(body.ids)}, request)


# ══════════════════════════════════════════════════════════════════════════════
# v1 — ORDERS
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/orders",
         summary="Заявки на доставку",
         tags=["v1-orders"])
def v1_get_orders(
    request: Request,
    date: str | None = Query(None, description="YYYY-MM-DD (default: сегодня)"),
):
    """Заявки на доставку за указанную дату.

    **Auth**: `orders:read`
    """
    _v1_require_scope(request, "orders:read")
    uid = get_user_id(request)
    target_date = date if date else str(datetime.now().date())
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """SELECT o.id, o.store_id, o.store_name_raw, o.address_raw, o.order_number,
                  o.weight_kg, o.volume_m3, o.amount_rub, o.quantity, o.products, o.notes,
                  o.delivery_date::text as delivery_date,
                  s.name as store_name_db, s.address as store_address
             FROM daily_orders o
             LEFT JOIN stores s ON s.id = o.store_id
            WHERE o.owner_id = %s AND o.delivery_date = %s
            ORDER BY o.id""",
        (uid, target_date),
    )
    orders = [dict(r) for r in cur.fetchall()]
    cur.execute(
        """SELECT COUNT(*) as cnt, COALESCE(SUM(weight_kg),0) as tw,
                  COALESCE(SUM(volume_m3),0) as tv, COALESCE(SUM(amount_rub),0) as ta
             FROM daily_orders WHERE owner_id=%s AND delivery_date=%s""",
        (uid, target_date),
    )
    s = cur.fetchone()
    cur.close(); conn.close()
    data = {
        "delivery_date": target_date,
        "orders": orders,
        "total_count": s["cnt"],
        "total_weight_kg": round(float(s["tw"]), 2),
        "total_volume_m3": round(float(s["tv"]), 3),
        "total_amount_rub": round(float(s["ta"]), 2),
    }
    return _v1_response(data, request)


@app.delete("/api/v1/orders",
            summary="Удалить заявки за дату",
            tags=["v1-orders"])
def v1_delete_orders(
    request: Request,
    date: str = Query(..., description="YYYY-MM-DD — дата для удаления"),
):
    """Удалить все заявки за указанную дату.

    **Auth**: `orders:write`
    """
    _v1_require_scope(request, "orders:write")
    uid = get_user_id(request)
    try:
        datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        raise _v1_err("VALIDATION_ERROR", "Неверный формат даты. Используйте YYYY-MM-DD", 422)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM daily_orders WHERE owner_id=%s AND delivery_date=%s", (uid, date))
    deleted = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    return _v1_response({"ok": True, "deleted": deleted, "date": date}, request)


@app.post("/api/v1/orders/batch",
          summary="Загрузить заявки пакетом",
          tags=["v1-orders"])
def v1_orders_batch(request: Request, body: WebhookIngestRequest):
    """Загрузить заявки на доставку пакетом. Принимает тот же формат, что и
    `POST /api/v1/webhooks/ingest`, но использует стандартный Bearer-заголовок.

    Поле `replace_date=true` удаляет существующие заявки за указанные даты перед вставкой.

    **Auth**: `orders:write`
    """
    _v1_require_scope(request, "orders:write")
    uid = get_user_id(request)

    if not body.orders:
        return _v1_response({"created": 0, "matched": 0, "skipped": 0, "errors": []}, request)

    for item in body.orders:
        try:
            datetime.strptime(item.delivery_date, "%Y-%m-%d")
        except ValueError:
            raise _v1_err(
                "VALIDATION_ERROR",
                f"Неверный формат даты '{item.delivery_date}'. Используйте YYYY-MM-DD",
                422,
            )

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, name, address, city, phone, external_id FROM stores WHERE owner_id=%s", (uid,))
    db_stores = [dict(r) for r in cur.fetchall()]

    if body.replace_date:
        for d in {item.delivery_date for item in body.orders}:
            cur.execute("DELETE FROM daily_orders WHERE owner_id=%s AND delivery_date=%s", (uid, d))

    created = matched = skipped = auto_created = 0
    errors: list = []
    for item in body.orders:
        try:
            if not item.store_name.strip():
                skipped += 1
                continue
            store_id, was_created = _auto_create_store_if_missing(
                uid, item.store_name, item.address or "",
                item.counterparty_code, item.phone, item.city,
                db_stores, cur, auto_create=body.auto_create_stores,
            )
            if store_id is not None:
                if was_created:
                    auto_created += 1
                else:
                    matched += 1
            cur.execute(
                """INSERT INTO daily_orders
                   (owner_id, store_id, store_name_raw, address_raw, delivery_date,
                    weight_kg, volume_m3, quantity, amount_rub, products, order_number, notes)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (uid, store_id, item.store_name.strip(), item.address or "",
                 item.delivery_date, item.weight_kg, item.volume_m3,
                 item.quantity, item.amount_rub, item.products,
                 item.order_number, item.notes),
            )
            created += 1
        except Exception as exc:
            errors.append({"store": item.store_name, "error": str(exc)})
    conn.commit(); cur.close(); conn.close()
    _record_integration_sync(request, created, matched,
                              len(body.orders) - created - skipped,
                              len(errors),
                              "; ".join(e.get("error", "") for e in errors[:3]))
    return _v1_response(
        {"created": created, "matched": matched, "skipped": skipped,
         "auto_created_stores": auto_created, "errors": errors[:20]},
        request,
    )


# ══════════════════════════════════════════════════════════════════════════════
# v1 — ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/routes",
         summary="История маршрутов",
         tags=["v1-routes"])
def v1_list_routes(
    request: Request,
    page: int = Query(1, description="Номер страницы (≥1)"),
    page_size: int = Query(20, description="Размер страницы (1–100)"),
):
    """Список построенных маршрутов с пагинацией.

    **Auth**: `routes:read`
    """
    _v1_require_scope(request, "routes:read")
    page = max(1, page)
    page_size = max(1, min(100, page_size))
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT COUNT(*) as total FROM route_sessions WHERE owner_id=%s", (uid,))
    total = int(cur.fetchone()["total"])
    offset = (page - 1) * page_size
    cur.execute(
        """SELECT id, date, num_vehicles, total_km, saved_km, saved_rub, num_points, created_at
           FROM route_sessions WHERE owner_id=%s ORDER BY created_at DESC LIMIT %s OFFSET %s""",
        (uid, page_size, offset),
    )
    items = [
        {
            "id": r["id"],
            "date": r["date"],
            "num_vehicles": r["num_vehicles"] or 0,
            "total_km": round(float(r["total_km"] or 0), 1),
            "saved_km": round(float(r["saved_km"] or 0), 1),
            "saved_rub": int(r["saved_rub"] or 0),
            "num_points": r["num_points"] or 0,
            "created_at": str(r["created_at"]) if r["created_at"] else None,
        }
        for r in cur.fetchall()
    ]
    cur.close(); conn.close()
    meta = {"total": total, "page": page, "page_size": page_size,
            "pages": max(1, (total + page_size - 1) // page_size)}
    return _v1_response(items, request, meta=meta)


@app.get("/api/v1/routes/{route_id}",
         summary="Маршрут по ID",
         tags=["v1-routes"])
def v1_get_route(route_id: int, request: Request):
    """Полный результат маршрута по ID.

    **Auth**: `routes:read`
    """
    _v1_require_scope(request, "routes:read")
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT result_json FROM route_sessions WHERE id=%s AND owner_id=%s", (route_id, uid))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row or not row["result_json"]:
        raise _v1_err("ROUTE_NOT_FOUND", f"Маршрут с ID {route_id} не найден", 404,
                      {"route_id": route_id})
    return _v1_response(json.loads(row["result_json"]), request)


@app.delete("/api/v1/routes/{route_id}",
            summary="Удалить маршрут",
            tags=["v1-routes"])
def v1_delete_route(route_id: int, request: Request):
    """Удалить маршрут из истории.

    **Auth**: `routes:write`
    """
    _v1_require_scope(request, "routes:write")
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM route_sessions WHERE id=%s AND owner_id=%s", (route_id, uid))
    deleted = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    if deleted == 0:
        raise _v1_err("ROUTE_NOT_FOUND", f"Маршрут с ID {route_id} не найден", 404)
    return _v1_response({"ok": True, "deleted": 1}, request)


@app.post("/api/v1/routes/build",
          summary="Построить маршрут",
          tags=["v1-routes"])
def v1_build_route(request: Request, body: RouteRequest):
    """Запустить VRP-оптимизацию маршрутов.

    Поведение идентично `POST /api/route/build`. Depot (склад) задаётся через
    `depot_lat` + `depot_lon` — оба поля обязательны.

    **Auth**: `routes:build`
    """
    _v1_require_scope(request, "routes:build")
    result = build_route(request, body)
    # build_route returns a JSONResponse (binary-safe); extract body and wrap in envelope
    if isinstance(result, JSONResponse):
        data = json.loads(result.body)
        return _v1_response(data, request, status_code=result.status_code)
    return _v1_response(result, request)


# ══════════════════════════════════════════════════════════════════════════════
# v1 — ANALYTICS
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/analytics/summary",
         summary="Сводная аналитика",
         tags=["v1-analytics"])
def v1_analytics_summary(request: Request):
    """Агрегированные показатели за всё время.

    **Auth**: `analytics:read`
    """
    _v1_require_scope(request, "analytics:read")
    return _v1_response(get_analytics_summary(request), request)


@app.get("/api/v1/analytics/daily",
         summary="Аналитика по дням",
         tags=["v1-analytics"])
def v1_analytics_daily(
    request: Request,
    date_from: str | None = Query(None, description="YYYY-MM-DD"),
    date_to: str | None = Query(None, description="YYYY-MM-DD"),
):
    """Пробег, экономия по дням.

    **Auth**: `analytics:read`
    """
    _v1_require_scope(request, "analytics:read")
    return _v1_response(get_analytics_daily(request, date_from, date_to), request)


@app.get("/api/v1/analytics/monthly",
         summary="Аналитика по месяцам",
         tags=["v1-analytics"])
def v1_analytics_monthly(
    request: Request,
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
):
    """Помесячная статистика.

    **Auth**: `analytics:read`
    """
    _v1_require_scope(request, "analytics:read")
    return _v1_response(get_analytics_monthly(request, date_from, date_to), request)


@app.get("/api/v1/analytics/vehicle-load",
         summary="Загрузка машин",
         tags=["v1-analytics"])
def v1_analytics_vehicle_load(
    request: Request,
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
):
    """Среднее число точек на машину по дням.

    **Auth**: `analytics:read`
    """
    _v1_require_scope(request, "analytics:read")
    return _v1_response(get_analytics_vehicle_load(request, date_from, date_to), request)


@app.get("/api/v1/analytics/top-stores",
         summary="Топ магазинов",
         tags=["v1-analytics"])
def v1_analytics_top_stores(request: Request):
    """10 магазинов с наибольшим числом доставок.

    **Auth**: `analytics:read`
    """
    _v1_require_scope(request, "analytics:read")
    return _v1_response(get_top_stores(request), request)


# ══════════════════════════════════════════════════════════════════════════════
# v1 — SETTINGS
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/settings",
         summary="Настройки компании",
         tags=["v1-settings"])
def v1_get_settings(request: Request):
    """Параметры расчёта стоимости км.

    **Auth**: `settings:read`
    """
    _v1_require_scope(request, "settings:read")
    uid = get_user_id(request)
    return _v1_response(get_company_settings(user_id=uid), request)


@app.put("/api/v1/settings",
         summary="Обновить настройки",
         tags=["v1-settings"])
def v1_update_settings(request: Request, body: CompanySettingsInput):
    """Обновить цену топлива и расход. `cost_per_km` рассчитывается автоматически.

    **Auth**: `settings:write`
    """
    _v1_require_scope(request, "settings:write")
    return _v1_response(update_settings_endpoint(request, body), request)


# ══════════════════════════════════════════════════════════════════════════════
# v1 — WEBHOOKS (Bearer variant)
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/api/v1/webhooks/ingest",
          summary="Webhook ingest (Bearer)",
          tags=["v1-webhooks"])
def v1_webhook_ingest_bearer(request: Request, body: WebhookIngestRequest):
    """Универсальный приём заявок через стандартный Bearer-заголовок.

    Идентичен `POST /api/v1/webhooks/ingest/{token}`, но токен передаётся
    в заголовке `Authorization: Bearer <key>` — более удобен для REST-клиентов.

    **Auth**: `webhooks:receive`
    """
    _v1_require_scope(request, "webhooks:receive")
    uid = get_user_id(request)

    if not body.orders:
        return _v1_response({"created": 0, "matched": 0, "skipped": 0, "errors": []}, request)

    for item in body.orders:
        try:
            datetime.strptime(item.delivery_date, "%Y-%m-%d")
        except ValueError:
            raise _v1_err(
                "VALIDATION_ERROR",
                f"Неверный формат даты '{item.delivery_date}'. Используйте YYYY-MM-DD", 422,
            )

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, name, address, city FROM stores WHERE owner_id=%s", (uid,))
    db_stores = [dict(r) for r in cur.fetchall()]

    if body.replace_date:
        for d in {item.delivery_date for item in body.orders}:
            cur.execute("DELETE FROM daily_orders WHERE owner_id=%s AND delivery_date=%s", (uid, d))

    created = matched = skipped = 0
    errors: list = []
    for item in body.orders:
        try:
            if not item.store_name.strip():
                skipped += 1; continue
            store_id = None
            ms = _match_store_to_db(item.store_name, item.address or "", db_stores)
            if ms:
                store_id = ms["id"]; matched += 1
            cur.execute(
                """INSERT INTO daily_orders
                   (owner_id, store_id, store_name_raw, address_raw, delivery_date,
                    weight_kg, volume_m3, quantity, amount_rub, products, order_number, notes)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (uid, store_id, item.store_name.strip(), item.address or "",
                 item.delivery_date, item.weight_kg, item.volume_m3,
                 item.quantity, item.amount_rub, item.products,
                 item.order_number, item.notes),
            )
            created += 1
        except Exception as exc:
            errors.append({"store": item.store_name, "error": str(exc)})
    conn.commit(); cur.close(); conn.close()
    return _v1_response(
        {"created": created, "matched": matched, "skipped": skipped, "errors": errors[:20]},
        request,
    )


# ══════════════════════════════════════════════════════════════════════════════
# v1 — KEYS (self-service)
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/keys/me",
         summary="Метаданные текущего ключа",
         tags=["v1-keys"])
def v1_key_me(request: Request):
    """Вернуть метаданные API ключа, которым выполнен запрос.
    Секретная часть ключа не возвращается никогда.

    **Auth**: любой API ключ (scope не требуется)
    """
    scopes = getattr(request.state, "api_key_scopes", None)
    if scopes is None:
        raise _v1_err(
            "COOKIE_AUTH",
            "Этот endpoint предназначен для API-ключей, а не для cookie-сессий.",
            400,
        )
    username = getattr(request.state, "username", "")
    try:
        key_id = int(username.split(":")[-1]) if ":" in username else None
    except (ValueError, TypeError):
        key_id = None
    if not key_id:
        raise _v1_err("KEY_NOT_FOUND", "Не удалось определить ID ключа", 404)
    uid = get_user_id(request)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """SELECT id, name, key_prefix, scopes, is_active, expires_at, last_used_at, created_at
           FROM api_keys WHERE id=%s AND owner_id=%s""",
        (key_id, uid),
    )
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        raise _v1_err("KEY_NOT_FOUND", "Ключ не найден", 404)
    return _v1_response(dict(row), request)


# ══════════════════════════════════════════════════════════════════════════════
# v1 — OpenAPI docs (separate sub-app for clean /api/v1/docs URL)
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/openapi.json",
         include_in_schema=False)
def v1_openapi_json():
    """Return OpenAPI spec filtered to v1 tags only."""
    from fastapi.openapi.utils import get_openapi
    full = get_openapi(
        title="SmartRoute Public API v1",
        version="1.0.0",
        description=(
            "## SmartRoute Public API\n\n"
            "Публичный REST API для интеграции с внешними системами.\n\n"
            "### Аутентификация\n\n"
            "Все запросы требуют заголовок:\n"
            "```\nAuthorization: Bearer sr_live_<ваш_ключ>\n```\n\n"
            "API-ключи создаются в настройках SmartRoute → раздел «API-ключи».\n\n"
            "### Rate Limiting\n\n"
            "60 запросов в минуту на ключ. Заголовки:\n"
            "- `X-RateLimit-Limit` — лимит\n"
            "- `X-RateLimit-Remaining` — осталось запросов\n"
            "- `X-RateLimit-Reset` — Unix timestamp сброса\n\n"
            "### Формат ответов\n\n"
            "```json\n"
            '{"data": ..., "meta": {...}, "request_id": "req_abc123"}\n'
            "```\n\n"
            "### Формат ошибок\n\n"
            "```json\n"
            '{"error": {"code": "STORE_NOT_FOUND", "message": "..."}, '
            '"request_id": "req_abc123"}\n'
            "```"
        ),
        routes=[r for r in app.routes if any(
            tag.startswith("v1-") for tag in getattr(r, "tags", [])
        )],
        tags=[
            {"name": "v1-stores",    "description": "Точки доставки (магазины)"},
            {"name": "v1-orders",    "description": "Заявки на доставку"},
            {"name": "v1-routes",    "description": "Маршруты и VRP-оптимизация"},
            {"name": "v1-analytics", "description": "Аналитика"},
            {"name": "v1-settings",  "description": "Настройки компании"},
            {"name": "v1-webhooks",  "description": "Webhook ingest"},
            {"name": "v1-keys",      "description": "API-ключи (self-service)"},
        ],
    )
    # Inject Bearer security scheme
    full["components"] = full.get("components", {})
    full["components"]["securitySchemes"] = {
        "bearerAuth": {
            "type": "http",
            "scheme": "bearer",
            "bearerFormat": "SmartRoute API Key (sr_live_...)",
        }
    }
    full["security"] = [{"bearerAuth": []}]
    return full


@app.get("/api/v1/docs",
         include_in_schema=False)
def v1_swagger_ui():
    """Swagger UI for Public API v1."""
    html = """<!DOCTYPE html>
<html>
<head>
  <title>SmartRoute API v1</title>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <link rel="stylesheet" type="text/css"
    href="https://unpkg.com/swagger-ui-dist@5/swagger-ui.css">
</head>
<body>
<div id="swagger-ui"></div>
<script src="https://unpkg.com/swagger-ui-dist@5/swagger-ui-bundle.js"></script>
<script>
SwaggerUIBundle({
  url: "/api/v1/openapi.json",
  dom_id: '#swagger-ui',
  presets: [SwaggerUIBundle.presets.apis, SwaggerUIBundle.SwaggerUIStandalonePreset],
  layout: "BaseLayout",
  deepLinking: true,
  persistAuthorization: true,
})
</script>
</body>
</html>"""
    from fastapi.responses import HTMLResponse
    return HTMLResponse(html)


# ── Production static file serving ────────────────────────────────────────────
# In production (Docker / Railway), FastAPI serves the Vite build output from
# ./static/. In development, the Vite dev server handles this via the proxy.
# This block must come AFTER all /api/* routes so the catch-all doesn't
# shadow them. FastAPI matches routes in registration order.
from pathlib import Path as _Path  # noqa: E402

_STATIC_DIR = _Path(__file__).parent / "static"
if _STATIC_DIR.exists() and _STATIC_DIR.is_dir():
    from fastapi.staticfiles import StaticFiles as _StaticFiles  # noqa: E402
    from fastapi.responses import FileResponse as _FileResponse  # noqa: E402

    # Vite outputs JS/CSS bundles under /assets/ — mount separately so
    # FileResponse header (Content-Type) is set correctly.
    _assets_dir = _STATIC_DIR / "assets"
    if _assets_dir.exists():
        app.mount("/assets", _StaticFiles(directory=str(_assets_dir)), name="static-assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    async def _serve_spa(full_path: str):
        """Serve exact static file if it exists, otherwise fall back to SPA index.html."""
        if full_path and ".." not in full_path:
            candidate = _STATIC_DIR / full_path
            if candidate.is_file():
                return _FileResponse(str(candidate))
        return _FileResponse(str(_STATIC_DIR / "index.html"))

    logger.info("Static dir found at %s — serving frontend from FastAPI", _STATIC_DIR)


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8080))
    uvicorn.run(app, host="0.0.0.0", port=port)
